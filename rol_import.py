"""Importação e utilitários do Rol ANS (Anexo I, Anexo II/DUTs, correlação TUSS)."""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Sequence

from sqlalchemy.dialects.mysql import insert as mysql_insert

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # noqa: BLE001
    PdfReader = None


SEGMENT_KEYS = ('od', 'amb', 'hco', 'hso', 'ref')
_TRUTHY = frozenset({'X', 'S', 'SIM', '1', 'TRUE', 'VERDADEIRO', 'YES'})


@dataclass
class RolProcedimentoRow:
    descricao: str
    rn_alteracao: str | None = None
    vigencia: str | None = None
    seg_od: bool = False
    seg_amb: bool = False
    seg_hco: bool = False
    seg_hso: bool = False
    seg_ref: bool = False
    pac: bool = False
    dut_numero: str | None = None
    subgrupo: str | None = None
    grupo: str | None = None
    capitulo: str | None = None


@dataclass
class TussCorrelacaoRow:
    codigo: str
    descricao: str | None
    consta_rol: bool
    rol_descricao: str | None = None


@dataclass
class RolImportBundle:
    procedimentos: list[RolProcedimentoRow]
    correlacoes: list[TussCorrelacaoRow]
    formato: str
    erros: list[str] = field(default_factory=list)


@dataclass
class RolDutRow:
    numero: str
    titulo: str
    texto_completo: str
    resumo: str = ''
    resumo_tipo: str = 'automatico'


@dataclass
class ImportStats:
    processados: int = 0
    criados: int = 0
    atualizados: int = 0
    erros: list[str] = field(default_factory=list)


def norm_header(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r'[^0-9A-Za-z]+', '', text.upper())
    return text


def norm_descricao(value: str | None) -> str:
    if not value:
        return ''
    text = unicodedata.normalize('NFKC', str(value)).strip()
    text = re.sub(r'\s+', ' ', text)
    return text.upper()


def parse_flag(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().upper()
    if not text or text in {'-', '—', 'NA', 'N/A', 'NÃO', 'NAO', '0', 'FALSE', 'N'}:
        return False
    if text in _TRUTHY:
        return True
    # Colunas do Anexo I às vezes repetem o nome da segmentação (AMB, HCO...)
    if text in {'OD', 'AMB', 'HCO', 'HSO', 'REF', 'PAC'}:
        return True
    return bool(text)


def parse_dut_numero(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {'-', '—', 'NA', 'N/A'}:
        return None
    match = re.search(r'\d+', text)
    return match.group(0) if match else None


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return str(value).strip()


def format_rn_display(raw: str | None) -> str | None:
    if not raw:
        return None
    text = str(raw).strip()
    if not text or text in {'---', '—', '-'}:
        return None
    if text.upper().startswith('RN'):
        return text
    return f'RN {text}'


def format_vigencia_display(raw: str | None) -> str | None:
    if not raw:
        return None
    text = str(raw).strip()
    if not text or text in {'---', '—', '-'}:
        return None
    if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', text):
        return text
    try:
        num = float(text.replace(',', '.'))
        if 1000 <= num <= 120000:
            parsed = date(1899, 12, 30) + timedelta(days=int(num))
            return parsed.strftime('%d/%m/%Y')
    except (ValueError, OverflowError):
        pass
    return text


def _parse_vigencia_cell(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime('%d/%m/%Y')
    text = _cell_str(raw)
    return format_vigencia_display(text) or text or None


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows):
        headers = [norm_header(c) for c in row]
        if 'PROCEDIMENTO' in headers:
            mapping = {h: i for i, h in enumerate(headers) if h}
            return idx, mapping
    return None


def parse_sim_nao(raw_value: Any) -> bool | None:
    if raw_value is None:
        return None
    value = str(raw_value).strip()
    if not value:
        return None
    value_norm = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode().upper()
    if value_norm in {'SIM', 'S', '1', 'TRUE', 'T', 'YES', 'Y'}:
        return True
    if value_norm in {'NAO', 'N', '0', 'FALSE', 'F', 'NO'}:
        return False
    return None


def _col(mapping: dict[str, int], *names: str) -> int | None:
    for name in names:
        key = norm_header(name)
        if key in mapping:
            return mapping[key]
    return None


def _col_prefix(mapping: dict[str, int], prefix: str) -> int | None:
    prefix_norm = norm_header(prefix)
    for key, idx in mapping.items():
        if key.startswith(prefix_norm):
            return idx
    return None


def _is_tuss_rol_planilha(mapping: dict[str, int]) -> bool:
    return 'CODIGO' in mapping and (
        'CORRELACAOSIMNAO' in mapping
        or _col_prefix(mapping, 'CORRELACAO') is not None
        or _col_prefix(mapping, 'TERMINOLOGIA') is not None
    )


def _normalize_tuss_codigo(raw: Any) -> str | None:
    if raw is None:
        return None
    code = str(raw).strip()
    if not code:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if float(raw).is_integer():
            code = str(int(raw))
    if re.fullmatch(r'\d+\.0+', code):
        code = code.split('.', 1)[0]
    code = code.upper()
    if code.isdigit() and len(code) < 8:
        code = code.zfill(8)
    return code


def parse_anexo_i_rows(
    rows: Sequence[Sequence[Any]],
) -> tuple[list[RolProcedimentoRow], list[TussCorrelacaoRow], list[str]]:
    rows_list = [tuple(r) for r in rows]
    header_info = _find_header_row(rows_list)
    if not header_info:
        return [], [], ['Cabeçalho com coluna PROCEDIMENTO não encontrado.']

    header_idx, mapping = header_info
    proc_idx = mapping.get('PROCEDIMENTO')
    if proc_idx is None:
        return [], [], ['Coluna PROCEDIMENTO não identificada.']

    planilha_tuss_rol = _is_tuss_rol_planilha(mapping)
    idx_rn = _col(mapping, 'RN', 'RNALTERACAO', 'ALTERACAO') or _col_prefix(mapping, 'RESOLUCAONORMATIVA')
    idx_vig = _col(mapping, 'VIGENCIA')
    idx_od = _col(mapping, 'OD')
    idx_amb = _col(mapping, 'AMB')
    idx_hco = _col(mapping, 'HCO')
    idx_hso = _col(mapping, 'HSO')
    idx_ref = _col(mapping, 'REF')
    idx_pac = _col(mapping, 'PAC')
    idx_dut = _col(mapping, 'DUT')
    idx_sub = _col(mapping, 'SUBGRUPO')
    idx_grupo = _col(mapping, 'GRUPO')
    idx_cap = _col(mapping, 'CAPITULO')
    idx_codigo = _col(mapping, 'CODIGO')
    idx_terminologia = _col_prefix(mapping, 'TERMINOLOGIA')
    idx_correlacao = _col(mapping, 'SECONSTA', 'CONSTA') or _col_prefix(mapping, 'CORRELACAO')

    def _get(row: tuple[Any, ...], index: int | None) -> str:
        if index is None or index >= len(row):
            return ''
        return _cell_str(row[index])

    parsed: list[RolProcedimentoRow] = []
    correlacoes: list[TussCorrelacaoRow] = []
    procedimentos_vistos: set[str] = set()
    erros: list[str] = []
    for offset, row in enumerate(rows_list[header_idx + 1:], start=header_idx + 2):
        descricao = _get(row, proc_idx)
        if planilha_tuss_rol and idx_codigo is not None:
            codigo_raw = _get(row, idx_codigo)
            codigo_norm = _normalize_tuss_codigo(codigo_raw)
            if codigo_norm:
                desc_tuss = _get(row, idx_terminologia) or None
                consta_raw = _get(row, idx_correlacao)
                consta_val = parse_sim_nao(consta_raw)
                if consta_val is None and consta_raw:
                    erros.append(f'Linha {offset}: correlação inválida ({consta_raw!r}).')
                correlacoes.append(TussCorrelacaoRow(
                    codigo=codigo_norm,
                    descricao=(desc_tuss or '').strip() or None,
                    consta_rol=bool(consta_val) if consta_val is not None else False,
                    rol_descricao=(descricao or '').strip() or None,
                ))

        if not descricao:
            continue
        upper = descricao.upper()
        if upper.startswith('LEGENDA') or 'ROL DE PROCEDIMENTOS' in upper:
            continue
        if upper in {'PROCEDIMENTO', 'OD', 'AMB', 'HCO', 'HSO', 'REF', 'PAC', 'DUT'}:
            continue

        desc_key = norm_descricao(descricao)
        if desc_key in procedimentos_vistos:
            continue
        procedimentos_vistos.add(desc_key)

        parsed.append(RolProcedimentoRow(
            descricao=descricao.strip(),
            rn_alteracao=_get(row, idx_rn) or None,
            vigencia=_parse_vigencia_cell(row[idx_vig]) if idx_vig is not None and idx_vig < len(row) else None,
            seg_od=parse_flag(_get(row, idx_od)),
            seg_amb=parse_flag(_get(row, idx_amb)),
            seg_hco=parse_flag(_get(row, idx_hco)),
            seg_hso=parse_flag(_get(row, idx_hso)),
            seg_ref=parse_flag(_get(row, idx_ref)),
            pac=parse_flag(_get(row, idx_pac)),
            dut_numero=parse_dut_numero(_get(row, idx_dut)),
            subgrupo=_get(row, idx_sub) or None,
            grupo=_get(row, idx_grupo) or None,
            capitulo=_get(row, idx_cap) or None,
        ))

    correlacoes = dedupe_tuss_correlacoes(correlacoes)
    if not parsed and not correlacoes:
        erros.append('Nenhum procedimento ou correlação TUSS válida encontrada na planilha.')
    return parsed, correlacoes, erros


def parse_rol_spreadsheet_file(data: bytes, suffix: str) -> RolImportBundle:
    headers, data_rows = read_spreadsheet_rows(data, suffix)
    all_rows = [headers] + data_rows
    procedimentos, correlacoes, erros = parse_anexo_i_rows(all_rows)
    formato = 'tuss_rol_planilha' if correlacoes else 'anexo_i'
    return RolImportBundle(
        procedimentos=procedimentos,
        correlacoes=correlacoes,
        formato=formato,
        erros=erros,
    )


def parse_anexo_i_file(data: bytes, suffix: str) -> RolImportBundle:
    """Compatível com import Anexo I puro ou planilha oficial TUSS × Rol ANS."""
    return parse_rol_spreadsheet_file(data, suffix)


def read_spreadsheet_rows(data: bytes, suffix: str) -> tuple[list[Any], list[list[Any]]]:
    ext = (suffix or '').lower()
    if ext in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        if load_workbook is None:
            raise ValueError('Biblioteca openpyxl não disponível para ler arquivos XLSX.')
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            raise ValueError('Planilha XLSX sem conteúdo.')
        headers = list(rows[0])
        data_rows = [list(r) for r in rows[1:]]
        return headers, data_rows

    import csv
    try:
        text_data = data.decode('utf-8-sig')
    except UnicodeDecodeError:
        text_data = data.decode('latin-1', errors='ignore')
    sio = io.StringIO(text_data)
    try:
        sample = sio.read(2048)
        sio.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=';,')
    except Exception:
        sio.seek(0)
        dialect = csv.excel
    reader = csv.reader(sio, dialect)
    headers = next(reader, [])
    data_rows = [row for row in reader]
    return headers, data_rows


def extract_pdf_text(data: bytes) -> str:
    if PdfReader is None:
        raise ValueError('Biblioteca pypdf não disponível para ler PDF.')
    reader = PdfReader(io.BytesIO(data))
    chunks: list[str] = []
    for page in reader.pages:
        chunks.append(page.extract_text() or '')
    return '\n'.join(chunks)


def generate_dut_resumo(texto_completo: str, titulo: str | None = None) -> str:
    text = (texto_completo or '').strip()
    if not text:
        return titulo or ''

    normalized = re.sub(r'\s+', ' ', text)
    criterios = re.findall(
        r'(\d+[\.\)]\s*(?:[^\d\n]|(?!\d+[\.\)]).)*?(?=\d+[\.\)]|$))',
        text,
        flags=re.DOTALL,
    )
    criterios = [re.sub(r'\s+', ' ', c).strip() for c in criterios if len(c.strip()) > 10]

    if criterios:
        partes = [c[:220].rstrip() + ('…' if len(c) > 220 else '') for c in criterios[:3]]
        resumo = ' '.join(partes)
        if len(criterios) > 3:
            resumo += f' (+{len(criterios) - 3} critérios adicionais)'
        return resumo[:900]

    if len(normalized) <= 420:
        return normalized

    snippet = normalized[:420]
    last_period = snippet.rfind('.')
    if last_period >= 120:
        return snippet[: last_period + 1]
    return snippet.rstrip() + '…'


_DUT_START_DASH = re.compile(
    r'(?:^|\n)\s*(?:DUT\s*N?[º°o\.]?\s*)?(\d{1,3}(?:\.\d+)*)\s*[-–—]\s*(.+?)(?:\n|$)',
    re.IGNORECASE | re.MULTILINE,
)
_DUT_START_DOT = re.compile(
    r'^\s*(\d{1,3})\.\s+(.+?)\s*$',
    re.MULTILINE,
)
_DUT_START_SUB = re.compile(
    r'^\s*(\d{1,3}\.\d+(?:\.\d+)?)\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ][^\n]{10,}?)\s*$',
    re.MULTILINE | re.UNICODE,
)

# PDF consolidado publicado pela ANS (Anexo II da RN 465/2021 e alterações).
ANS_ANEXO_II_PDF_URL = (
    'https://www.gov.br/ans/pt-br/acesso-a-informacao/participacao-da-sociedade/'
    'atualizacao-do-rol-de-procedimentos/'
    'Anexo_II_DUT_2021_RN_465.2021_RN628.2025_RN629.2025.pdf/@@download/file'
)


def _dut_sort_key(numero: str) -> tuple[int, ...]:
    parts = re.split(r'\.', numero.strip())
    key: list[int] = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
    return tuple(key) if key else (9999,)


def _title_upper_ratio(title: str) -> float:
    letters = [c for c in title if c.isalpha()]
    if not letters:
        return 0.0
    return sum(1 for c in letters if c.isupper()) / len(letters)


def _is_valid_ans_dut_title(title: str) -> bool:
    title = (title or '').strip()
    if len(title) < 3:
        return False
    if len(title) < 12 and not (title.isupper() and title.replace(' ', '').isalpha()):
        return False
    if '....' in title:
        return False
    if title.lower().startswith('cobertura'):
        return False
    if re.match(r'^\d+\s*[-–]\s*\d+', title):
        return False
    return _title_upper_ratio(title) >= 0.55


def _looks_like_reference(title: str) -> bool:
    lower = title.lower()
    markers = (
        'http://', 'https://', 'et al', 'doi.', 'pubmed', 'bookshelf',
        'disponível em', 'acessado em', 'national institute', 'gene review',
        'binderup', 'devarajan', 'weitzel', 'portaria sas', 'nice.',
    )
    return any(m in lower for m in markers)


def _score_dut_header(text: str, pos: int, title: str) -> int:
    score = 0
    ratio = _title_upper_ratio(title)
    if ratio >= 0.75:
        score += 50
    elif ratio >= 0.55:
        score += 20
    if len(title) >= 35:
        score += 15
    if _looks_like_reference(title):
        score -= 120
    snippet = text[pos: pos + 600]
    if re.search(r'\n\s*1\.\s+Cobertura\s+obrigat', snippet, re.IGNORECASE):
        score += 100
    return score


def _pick_best_dut_occurrence(text: str, candidates: list[tuple[int, str]]) -> tuple[int, str]:
    """Prefere cabeçalho seguido de critérios de cobertura, não referências bibliográficas."""
    if not candidates:
        raise ValueError('lista vazia')
    scored = [(_score_dut_header(text, pos, title), pos, title) for pos, title in candidates]
    scored.sort(key=lambda item: (-item[0], item[1]))
    best_score = scored[0][0]
    if best_score < 0:
        return candidates[-1]
    top = [item for item in scored if item[0] == best_score]
    top.sort(key=lambda item: item[1])
    _, pos, title = top[0]
    return pos, title


def _find_ans_pdf_dut_headers(text: str) -> list[tuple[int, str, str, int]]:
    headers: dict[str, list[tuple[int, str]]] = {}

    for match in _DUT_START_DOT.finditer(text):
        numero = match.group(1).strip()
        titulo = re.sub(r'\s+', ' ', match.group(2).strip())
        if not _is_valid_ans_dut_title(titulo):
            continue
        headers.setdefault(numero, []).append((match.start(), titulo))

    for match in _DUT_START_SUB.finditer(text):
        numero = match.group(1).strip()
        titulo = re.sub(r'\s+', ' ', match.group(2).strip())
        if not _is_valid_ans_dut_title(titulo):
            continue
        headers.setdefault(numero, []).append((match.start(), titulo))

    resolved: list[tuple[int, str, str, int]] = []
    for numero, occs in headers.items():
        pos, titulo = _pick_best_dut_occurrence(text, occs)
        resolved.append((pos, numero, titulo, pos + len(titulo)))

    resolved.sort(key=lambda item: (item[0], _dut_sort_key(item[1])))
    return resolved


def _parse_anexo_ii_dash_format(cleaned: str) -> list[RolDutRow]:
    matches = list(_DUT_START_DASH.finditer(cleaned))
    parsed: list[RolDutRow] = []
    for i, match in enumerate(matches):
        numero = match.group(1).strip()
        titulo = re.sub(r'\s+', ' ', match.group(2).strip())
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(cleaned)
        corpo = cleaned[start:end].strip()
        texto_completo = f'{titulo}\n\n{corpo}'.strip() if corpo else titulo
        resumo = generate_dut_resumo(texto_completo, titulo=titulo)
        parsed.append(RolDutRow(
            numero=numero,
            titulo=titulo[:500],
            texto_completo=texto_completo,
            resumo=resumo,
            resumo_tipo='automatico',
        ))
    return parsed


def _parse_anexo_ii_ans_pdf_format(cleaned: str) -> list[RolDutRow]:
    headers = _find_ans_pdf_dut_headers(cleaned)
    parsed: list[RolDutRow] = []
    for i, (_pos, numero, titulo, _end) in enumerate(headers):
        start = _pos
        end = headers[i + 1][0] if i + 1 < len(headers) else len(cleaned)
        block = cleaned[start:end].strip()
        # Remove a linha do título duplicada no início do bloco
        corpo = re.sub(rf'^\s*{re.escape(numero)}(?:\.\s+|\s+).+?\n', '', block, count=1, flags=re.M)
        texto_completo = f'{titulo}\n\n{corpo.strip()}'.strip() if corpo.strip() else titulo
        resumo = generate_dut_resumo(texto_completo, titulo=titulo)
        parsed.append(RolDutRow(
            numero=numero,
            titulo=titulo[:500],
            texto_completo=texto_completo,
            resumo=resumo,
            resumo_tipo='automatico',
        ))
    return parsed


def download_ans_anexo_ii_pdf(url: str | None = None, timeout: int = 120) -> bytes:
    import urllib.request

    target = url or ANS_ANEXO_II_PDF_URL
    req = urllib.request.Request(
        target,
        headers={'User-Agent': 'sistema_preco/1.0 (importação Rol ANS)'},
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = resp.read()
    if not data:
        raise ValueError('Download do Anexo II retornou arquivo vazio.')
    return data


def _looks_like_ans_consolidated_pdf(text: str) -> bool:
    upper = text[:8000].upper()
    return (
        'SUMÁRIO' in upper or 'SUMARIO' in upper
        or 'DIRETRIZES DE UTILIZAÇÃO' in upper
        or 'DIRETRIZES DE UTILIZACAO' in upper
    )


def parse_anexo_ii_text(text: str) -> tuple[list[RolDutRow], list[str]]:
    if not text or not text.strip():
        return [], ['PDF do Anexo II sem texto extraível.']

    cleaned = re.sub(r'\r\n?', '\n', text)
    ans_headers = _find_ans_pdf_dut_headers(cleaned)
    dash_matches = list(_DUT_START_DASH.finditer(cleaned))

    use_ans = len(ans_headers) >= 10 and (
        len(ans_headers) >= len(dash_matches)
        or _looks_like_ans_consolidated_pdf(cleaned)
    )
    if use_ans:
        parsed = _parse_anexo_ii_ans_pdf_format(cleaned)
    elif dash_matches:
        parsed = _parse_anexo_ii_dash_format(cleaned)
    elif ans_headers:
        parsed = _parse_anexo_ii_ans_pdf_format(cleaned)
    else:
        return [], ['Nenhuma DUT identificada no Anexo II (formatos "NN - Título" ou "NN. TÍTULO").']

    if not parsed:
        return [], ['Nenhuma DUT identificada no Anexo II.']

    # Última ocorrência vence se o parser capturou duplicatas residuais
    by_numero: dict[str, RolDutRow] = {}
    for row in parsed:
        by_numero[row.numero] = row
    ordered = sorted(by_numero.values(), key=lambda r: _dut_sort_key(r.numero))
    return ordered, []


def fetch_and_parse_ans_anexo_ii(url: str | None = None) -> tuple[list[RolDutRow], list[str]]:
    """Baixa o PDF oficial da ANS e extrai todas as DUTs."""
    data = download_ans_anexo_ii_pdf(url=url)
    return parse_anexo_ii_pdf(data)


def parse_anexo_ii_pdf(data: bytes) -> tuple[list[RolDutRow], list[str]]:
    text = extract_pdf_text(data)
    return parse_anexo_ii_text(text)


def segmentacao_list(row: RolProcedimentoRow | dict[str, Any]) -> list[str]:
    if isinstance(row, dict):
        flags = {k: bool(row.get(k)) for k in ('seg_od', 'seg_amb', 'seg_hco', 'seg_hso', 'seg_ref')}
    else:
        flags = {
            'seg_od': row.seg_od,
            'seg_amb': row.seg_amb,
            'seg_hco': row.seg_hco,
            'seg_hso': row.seg_hso,
            'seg_ref': row.seg_ref,
        }
    labels = {
        'seg_od': 'OD',
        'seg_amb': 'AMB',
        'seg_hco': 'HCO',
        'seg_hso': 'HSO',
        'seg_ref': 'REF',
    }
    return [labels[k] for k, v in flags.items() if v]


def serialize_rol_procedimento(proc: Any, dut: Any | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {
        'id': proc.id,
        'descricao': proc.descricao,
        'rn_alteracao': proc.rn_alteracao,
        'rn_label': format_rn_display(proc.rn_alteracao),
        'vigencia': proc.vigencia,
        'vigencia_label': format_vigencia_display(proc.vigencia),
        'segmentacao': segmentacao_list({
            'seg_od': proc.seg_od,
            'seg_amb': proc.seg_amb,
            'seg_hco': proc.seg_hco,
            'seg_hso': proc.seg_hso,
            'seg_ref': proc.seg_ref,
        }),
        'pac': bool(proc.pac),
        'dut_numero': proc.dut_numero,
        'capitulo': proc.capitulo.nome if getattr(proc, 'capitulo', None) else None,
        'grupo': proc.grupo.nome if getattr(proc, 'grupo', None) else None,
        'subgrupo': proc.subgrupo.nome if getattr(proc, 'subgrupo', None) else None,
    }
    if dut:
        payload['dut'] = serialize_rol_dut(dut, include_full_text=False)
    elif proc.dut_numero:
        payload['dut'] = {'numero': proc.dut_numero}
    return payload


def serialize_rol_dut(dut: Any, *, include_full_text: bool = True) -> dict[str, Any]:
    data = {
        'numero': dut.numero,
        'titulo': dut.titulo,
        'resumo': dut.resumo,
        'resumo_tipo': dut.resumo_tipo,
    }
    if include_full_text:
        data['texto_completo'] = dut.texto_completo
    return data


def import_anexo_i_to_db(
    session: Any,
    models: dict[str, Any],
    rows: Sequence[RolProcedimentoRow],
    *,
    versao_label: str | None = None,
) -> ImportStats:
    RolCapitulo = models['RolCapitulo']
    RolGrupo = models['RolGrupo']
    RolSubgrupo = models['RolSubgrupo']
    RolProcedimento = models['RolProcedimento']

    stats = ImportStats()
    capitulo_cache: dict[str, Any] = {}
    grupo_cache: dict[str, Any] = {}
    subgrupo_cache: dict[str, Any] = {}

    def _get_capitulo(nome: str | None):
        if not nome:
            return None
        key = norm_descricao(nome)
        if key in capitulo_cache:
            return capitulo_cache[key]
        row = RolCapitulo.query.filter_by(nome=nome.strip()).first()
        if not row:
            row = RolCapitulo(nome=nome.strip())
            session.add(row)
            session.flush()
            stats.criados += 1
        capitulo_cache[key] = row
        return row

    def _get_grupo(capitulo, nome: str | None):
        if not capitulo or not nome:
            return None
        key = f'{capitulo.id}:{norm_descricao(nome)}'
        if key in grupo_cache:
            return grupo_cache[key]
        row = RolGrupo.query.filter_by(capitulo_id=capitulo.id, nome=nome.strip()).first()
        if not row:
            row = RolGrupo(capitulo_id=capitulo.id, nome=nome.strip())
            session.add(row)
            session.flush()
            stats.criados += 1
        grupo_cache[key] = row
        return row

    def _get_subgrupo(grupo, nome: str | None):
        if not grupo or not nome:
            return None
        key = f'{grupo.id}:{norm_descricao(nome)}'
        if key in subgrupo_cache:
            return subgrupo_cache[key]
        row = RolSubgrupo.query.filter_by(grupo_id=grupo.id, nome=nome.strip()).first()
        if not row:
            row = RolSubgrupo(grupo_id=grupo.id, nome=nome.strip())
            session.add(row)
            session.flush()
            stats.criados += 1
        subgrupo_cache[key] = row
        return row

    for item in rows:
        stats.processados += 1
        capitulo = _get_capitulo(item.capitulo)
        grupo = _get_grupo(capitulo, item.grupo)
        subgrupo = _get_subgrupo(grupo, item.subgrupo)

        desc_key = norm_descricao(item.descricao)
        proc = RolProcedimento.query.filter_by(descricao_norm=desc_key).first()
        is_new = proc is None
        if is_new:
            proc = RolProcedimento(descricao=item.descricao.strip(), descricao_norm=desc_key)
            session.add(proc)
            stats.criados += 1
        else:
            stats.atualizados += 1

        proc.descricao = item.descricao.strip()
        proc.descricao_norm = desc_key
        proc.capitulo_id = capitulo.id if capitulo else None
        proc.grupo_id = grupo.id if grupo else None
        proc.subgrupo_id = subgrupo.id if subgrupo else None
        proc.rn_alteracao = item.rn_alteracao
        proc.vigencia = item.vigencia
        proc.seg_od = item.seg_od
        proc.seg_amb = item.seg_amb
        proc.seg_hco = item.seg_hco
        proc.seg_hso = item.seg_hso
        proc.seg_ref = item.seg_ref
        proc.pac = item.pac
        proc.dut_numero = item.dut_numero
        if versao_label:
            proc.versao_label = versao_label

    return stats


def dedupe_tuss_correlacoes(rows: Sequence[TussCorrelacaoRow]) -> list[TussCorrelacaoRow]:
    """Mantém a última ocorrência de cada código TUSS (planilha ANS pode repetir linhas)."""
    by_codigo: dict[str, TussCorrelacaoRow] = {}
    for row in rows:
        if row.codigo:
            by_codigo[row.codigo] = row
    return list(by_codigo.values())


def import_tuss_correlacoes_to_db(
    session: Any,
    models: dict[str, Any],
    rows: Sequence[TussCorrelacaoRow],
) -> ImportStats:
    TussRolCorrelacao = models['TussRolCorrelacao']
    RolProcedimento = models['RolProcedimento']

    stats = ImportStats()
    rows = dedupe_tuss_correlacoes(rows)
    if not rows:
        return stats

    rol_by_desc = {
        p.descricao_norm: p
        for p in RolProcedimento.query.all()
        if p.descricao_norm
    }

    bind = session.get_bind()
    use_mysql_upsert = bind is not None and bind.dialect.name == 'mysql'

    for item in rows:
        stats.processados += 1
        if not item.codigo:
            continue

        rol_proc = None
        if item.rol_descricao:
            rol_proc = rol_by_desc.get(norm_descricao(item.rol_descricao))

        payload = {
            'codigo': item.codigo,
            'descricao': item.descricao,
            'consta_rol': bool(item.consta_rol),
            'rol_procedimento_id': rol_proc.id if rol_proc else None,
        }

        if use_mysql_upsert:
            stmt = mysql_insert(TussRolCorrelacao.__table__).values(**payload)
            stmt = stmt.on_duplicate_key_update(
                descricao=payload['descricao'],
                consta_rol=payload['consta_rol'],
                rol_procedimento_id=payload['rol_procedimento_id'],
            )
            result = session.execute(stmt)
            if result.rowcount == 1:
                stats.criados += 1
            elif result.rowcount == 2:
                stats.atualizados += 1
            continue

        row = TussRolCorrelacao.query.filter_by(codigo=item.codigo).first()
        if row:
            alterou = False
            if (row.descricao or '') != (item.descricao or ''):
                row.descricao = item.descricao
                alterou = True
            if bool(row.consta_rol) != bool(item.consta_rol):
                row.consta_rol = bool(item.consta_rol)
                alterou = True
            new_link = rol_proc.id if rol_proc else None
            if row.rol_procedimento_id != new_link:
                row.rol_procedimento_id = new_link
                alterou = True
            if alterou:
                stats.atualizados += 1
        else:
            session.add(TussRolCorrelacao(**payload))
            stats.criados += 1

    return stats


def import_duts_to_db(session: Any, models: dict[str, Any], rows: Sequence[RolDutRow]) -> ImportStats:
    RolDut = models['RolDut']
    stats = ImportStats()

    for item in rows:
        stats.processados += 1
        dut = RolDut.query.filter_by(numero=item.numero).first()
        if dut:
            stats.atualizados += 1
        else:
            dut = RolDut(numero=item.numero)
            session.add(dut)
            stats.criados += 1

        dut.titulo = item.titulo
        dut.texto_completo = item.texto_completo
        if not dut.resumo or dut.resumo_tipo != 'manual':
            dut.resumo = item.resumo or generate_dut_resumo(item.texto_completo, item.titulo)
            dut.resumo_tipo = item.resumo_tipo

    return stats


def link_tuss_correlacoes(session: Any, models: dict[str, Any]) -> int:
    TussRolCorrelacao = models['TussRolCorrelacao']
    RolProcedimento = models['RolProcedimento']

    procedimentos = RolProcedimento.query.all()
    by_desc = {p.descricao_norm: p for p in procedimentos if p.descricao_norm}

    linked = 0
    for corr in TussRolCorrelacao.query.filter(TussRolCorrelacao.rol_procedimento_id.is_(None)).all():
        key = norm_descricao(corr.descricao)
        proc = by_desc.get(key)
        if proc:
            corr.rol_procedimento_id = proc.id
            linked += 1
    return linked
