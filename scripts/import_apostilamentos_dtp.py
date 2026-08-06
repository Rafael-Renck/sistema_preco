#!/usr/bin/env python3
"""Importa apostilamentos DTP (XLSX) para o banco configurado em DATABASE_URL."""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from app import app, db, Procedimento, Tabela  # noqa: E402


def _norm_header(s: str) -> str:
    s = (s or "").strip()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.lower().replace(" ", "").replace("-", "").replace("_", "")
    return s


def _parse_money(v) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        return Decimal("0")
    s = s.replace("R$", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _clean_prestador(raw: str | None, fallback: str) -> str:
    """Usa coluna A; se vier nome de PDF, limpa para ficar legível."""
    name = (raw or "").strip()
    if not name:
        name = fallback
    # Remove extensão .pdf e prefixo "Apostilamento N .... - "
    name = re.sub(r"(?i)\.pdf$", "", name).strip()
    name = re.sub(r"(?i)^Apostilamento\s+N\.?\s*\d+[\./]\d+\s*[-–]?\s*", "", name).strip()
    # Remove sufixo " (2)" etc.
    name = re.sub(r"\s*\(\d+\)\s*$", "", name).strip()
    return name or fallback


def read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    keys = [_norm_header(str(h) if h is not None else "") for h in rows[0]]
    items = []
    for r in rows[1:]:
        item = {keys[i]: (r[i] if i < len(keys) else None) for i in range(len(keys))}
        items.append(item)
    return items


def collect_files(root: Path) -> list[Path]:
    return sorted(root.rglob("*.xlsx"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa APOSTILAMENTOS para tabela DTP (MPF)")
    parser.add_argument(
        "--dir",
        type=Path,
        default=ROOT / "APOSTILAMENTOS",
        help="Pasta com subpastas MAIO/JUNHO/JULHO",
    )
    parser.add_argument(
        "--tabela-id",
        type=int,
        default=41,
        help="ID da tabela DTP destino (default: 41 = DIÁRIAS, TAXAS E PACOTES)",
    )
    parser.add_argument(
        "--operadora-id",
        type=int,
        default=1,
        help="Operadora (default: 1 = MPF)",
    )
    parser.add_argument(
        "--substituir-prestador",
        action="store_true",
        help="Se o prestador já existir na tabela, remove itens dele antes de inserir",
    )
    parser.add_argument("--dry-run", action="store_true", help="Só simula, não grava")
    args = parser.parse_args()

    files = collect_files(args.dir)
    if not files:
        print(f"Nenhum XLSX em {args.dir}")
        return 1

    with app.app_context():
        tab = Tabela.query.get(args.tabela_id)
        if not tab:
            print(f"Tabela id={args.tabela_id} não encontrada")
            return 1
        if tab.tipo_tabela != "diarias_taxas_pacotes":
            print(f"Tabela {tab.id} não é DTP: {tab.tipo_tabela}")
            return 1
        if tab.id_operadora != args.operadora_id:
            print(
                f"Atenção: tabela {tab.id} é operadora {tab.id_operadora}, "
                f"mas importando com operadora_id={args.operadora_id}"
            )

        print(f"Destino: [{tab.id}] {tab.nome} | operadora_id={args.operadora_id}")
        print(f"Arquivos: {len(files)} | dry_run={args.dry_run} | substituir={args.substituir_prestador}")
        print("=" * 70)

        total_insert = 0
        total_skip = 0
        total_deleted = 0
        by_prest: dict[str, int] = {}

        for path in files:
            fallback = path.stem
            items = read_xlsx(path)
            parsed = []
            for item in items:
                # Coluna A = arquivo_origem (prestador)
                prest_raw = (
                    item.get("arquivoorigem")
                    or item.get("prestador")
                    or item.get("fornecedor")
                    or item.get("credenciado")
                )
                prest = _clean_prestador(
                    str(prest_raw) if prest_raw is not None else None,
                    fallback,
                )
                codigo = item.get("codigo") or item.get("cod")
                descricao = item.get("descricao") or item.get("descricao")
                # Preferir valor_numero quando existir
                valor_num = item.get("valornumero")
                if valor_num is not None and str(valor_num).strip() != "":
                    valor = _parse_money(valor_num)
                else:
                    valor = _parse_money(item.get("valor"))
                if not codigo or not descricao:
                    total_skip += 1
                    continue
                parsed.append((prest, str(codigo).strip(), str(descricao).strip(), valor))

            if not parsed:
                print(f"  VAZIO: {path.relative_to(args.dir)}")
                continue

            prestadores = sorted({p[0] for p in parsed})
            print(f"\n{path.parent.name}/{path.name}")
            print(f"  linhas={len(parsed)} prestadores={prestadores}")

            if args.dry_run:
                for p in prestadores:
                    by_prest[p] = by_prest.get(p, 0) + sum(1 for x in parsed if x[0] == p)
                total_insert += len(parsed)
                continue

            if args.substituir_prestador:
                for p in prestadores:
                    deleted = (
                        db.session.query(Procedimento)
                        .filter(
                            Procedimento.id_tabela == tab.id,
                            Procedimento.operadora_id == args.operadora_id,
                            Procedimento.prestador == p,
                        )
                        .delete(synchronize_session=False)
                    )
                    total_deleted += deleted
                    if deleted:
                        print(f"  removidos {deleted} itens de '{p}'")

            for prest, codigo, descricao, valor in parsed:
                db.session.add(
                    Procedimento(
                        codigo=codigo,
                        descricao=descricao[:500],
                        valor=valor,
                        prestador=prest,
                        uf=tab.uf,
                        id_tabela=tab.id,
                        operadora_id=args.operadora_id,
                    )
                )
                by_prest[prest] = by_prest.get(prest, 0) + 1
                total_insert += 1

            db.session.commit()
            print(f"  OK inseridos {len(parsed)}")

        print("\n" + "=" * 70)
        print(f"Total inseridos: {total_insert}")
        print(f"Total ignorados: {total_skip}")
        print(f"Total removidos (substituir): {total_deleted}")
        print("Por prestador:")
        for p, n in sorted(by_prest.items(), key=lambda x: (-x[1], x[0])):
            print(f"  {n:5d}  {p}")

        if args.dry_run:
            print("\n[DRY-RUN] Nada foi gravado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
