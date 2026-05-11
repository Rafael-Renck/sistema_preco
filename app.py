import os
import time
import csv
import math
import re
import threading
from pathlib import Path
from datetime import datetime, timedelta

import click
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    jsonify,
    session,
    send_file,
    g,
    abort,
    flash,
    has_request_context,
)
import json
from flask_sqlalchemy import SQLAlchemy
import pymysql
from dotenv import load_dotenv
from functools import wraps
from sqlalchemy import text, or_, and_, func, false, case, literal, cast, Numeric, Unicode
from sqlalchemy.dialects.mysql import insert as mysql_insert
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import unicodedata
import html
import hashlib
from datetime import date, datetime, timedelta
from uuid import uuid4
from types import SimpleNamespace
from dataclasses import dataclass
from enum import Enum
from typing import Optional, Sequence
from flask import make_response
import io
import tempfile
import shutil
import xlsxwriter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from werkzeug.security import generate_password_hash, check_password_hash
from cachetools import TTLCache
try:
    from pypdf import PdfReader
except Exception:  # noqa: BLE001
    PdfReader = None

try:
    import pytesseract
except Exception:  # noqa: BLE001
    pytesseract = None

try:
    from pdf2image import convert_from_path
except Exception:  # noqa: BLE001
    convert_from_path = None

try:
    from PIL import Image as PILImage
    from PIL import ImageOps
except Exception:  # noqa: BLE001
    PILImage = None
    ImageOps = None
# --- 1. CONFIGURAÇÃO INICIAL ---
# Inicializa a aplicação Flask
load_dotenv()

app = Flask(__name__, template_folder='templates', static_folder='static')

# Chave de sessão (ajuste em produção via variável de ambiente)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key')

# Configuração da conexão com o banco de dados MySQL
# Formato: mysql+pymysql://<usuario>:<senha>@<host>/<nome_do_banco>
# Para o XAMPP padrão, o usuário é 'root' e a senha é vazia.
# DATABASE_URL pode vir do Docker Compose. Fallback para dev local.
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv(
    'DATABASE_URL', 'mysql+pymysql://root:@localhost/operadora_saude'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Otimizações de connection pool para múltiplos acessos
engine_options = dict(app.config.get('SQLALCHEMY_ENGINE_OPTIONS') or {})
connect_args = dict(engine_options.get('connect_args') or {})
database_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
if not str(database_uri).startswith('sqlite:'):
    connect_args.setdefault('local_infile', 1)
engine_options['connect_args'] = connect_args

# Pool otimizado para ~25 usuários simultâneos
engine_options.setdefault('pool_size', 10)           # Conexões mantidas no pool
engine_options.setdefault('max_overflow', 20)        # Conexões extras em picos (total: 30)
engine_options.setdefault('pool_recycle', 3600)      # Recicla conexões após 1h
engine_options.setdefault('pool_pre_ping', True)     # Verifica conexão antes de usar
engine_options.setdefault('pool_timeout', 30)        # Timeout ao aguardar conexão

app.config['SQLALCHEMY_ENGINE_OPTIONS'] = engine_options

# Inicializa o SQLAlchemy para interagir com o banco de dados
db = SQLAlchemy(app)

# Cache com limite de memória (TTL Cache) para múltiplos acessos
# maxsize limita quantidade de itens, ttl limita tempo de vida
_insumo_cache_ttl = 300  # 5 minutos (para compatibilidade)
_insumo_cache = TTLCache(maxsize=1000, ttl=_insumo_cache_ttl)
_teto_cache = TTLCache(maxsize=500, ttl=600)     # 500 itens, 10 minutos
_rol_cache = TTLCache(maxsize=2000, ttl=900)     # 2000 itens, 15 minutos

def _safe_int_env(var_name: str, default: int) -> int:
    try:
        return int(os.getenv(var_name, default))
    except (TypeError, ValueError):
        return default


def _bras_clamped_batch(var_name: str, default: int) -> int:
    """Lotes pequenos e com teto evitam pico de CPU/memória e `IN` gigante (arquivos 10k+ linhas). Máx. 4000."""
    v = _safe_int_env(var_name, default)
    return max(200, min(4000, v))


def _bras_throttle_between_batches() -> None:
    """Pausa opcional entre lotes (ms) para aliviar MySQL/CPU em servidores fracos. 0 = desligado. Máx. 5s."""
    ms = min(5000, max(0, _safe_int_env('BRAS_BETWEEN_BATCHES_MS', 0)))
    if ms:
        time.sleep(ms / 1000.0)

_cbhpm_api_cache_ttl = _safe_int_env('CBHPM_API_CACHE_TTL', 300)
_cbhpm_api_cache = TTLCache(maxsize=1000, ttl=_cbhpm_api_cache_ttl)
_cbhpm_api_detail_cache = TTLCache(maxsize=2000, ttl=_cbhpm_api_cache_ttl)

# Importação Brasíndice: padrão ~2000 linhas/lote (ex.: 14k linhas ≈ 7 passadas), teto 4000.
# Arquivos grandes: lotes pequenos + BRAS_BETWEEN_BATCHES_MS deixam o servidor mais “respirar”.
BRAS_MATERIALIZE_BATCH = _bras_clamped_batch('BRAS_MATERIALIZE_BATCH_SIZE', 2000)
BRAS_INDEX_SYNC_BATCH = _bras_clamped_batch('BRAS_INDEX_SYNC_BATCH_SIZE', 2000)
BRAS_RAW_CSV_BATCH = _bras_clamped_batch('BRAS_RAW_CSV_BATCH', 2000)
# SIMPRO: INSERT único gigante pode estourar pacote/servidor; lotes por simpro_item_preco.id.
SIMPRO_INDEX_SYNC_BATCH = _bras_clamped_batch('SIMPRO_INDEX_SYNC_BATCH_SIZE', 4000)
# SIMPRO JSON: materialização em lotes (evita lista gigante em RAM e permite heartbeat no job).
SIMPRO_JSON_MATERIALIZE_BATCH = _bras_clamped_batch('SIMPRO_JSON_MATERIALIZE_BATCH_SIZE', 500)
# SIMPRO split: tamanho máximo do IN (...) por query (MySQL/plan cache).
SIMPRO_SPLIT_IN_CHUNK = _bras_clamped_batch('SIMPRO_SPLIT_FILTER_CHUNK_SIZE', 2000)

def _load_public_api_tokens() -> set[str]:
    tokens: set[str] = set()
    raw = (
        os.getenv('PUBLIC_API_TOKENS')
        or os.getenv('PUBLIC_API_TOKEN')
        or os.getenv('API_BEARER_TOKEN')
        or ''
    )
    for token in raw.split(','):
        candidate = (token or '').strip()
        if candidate:
            tokens.add(candidate)
    return tokens

_PUBLIC_API_TOKENS = _load_public_api_tokens()

def _clear_insumo_cache():
    """Limpa o cache de insumos (chamar após importações)"""
    global _insumo_cache, _CATALOGO_COUNT_CACHE
    _insumo_cache.clear()
    try:
        _CATALOGO_COUNT_CACHE.clear()
    except Exception:
        pass


def _catalog_count_cache_ttl_seconds() -> float:
    """TTL do cache de COUNT na busca de insumos (segundos). Após import, subir vía INSUMOS_COUNT_CACHE_TTL reduz picos."""
    try:
        return max(30.0, float(os.getenv('INSUMOS_COUNT_CACHE_TTL', '120')))
    except (TypeError, ValueError):
        return 120.0


def _schedule_post_import_analyze_tables() -> None:
    """
    Opcional: ANALYZE TABLE em segundo plano após import grande ajuda o MySQL/MariaDB
    a atualizar estatísticas e evitar planos lentos em COUNT/JOIN.

    POST_IMPORT_ANALYZE_TABLES=lista separada por vírgula, ex.:
    insumos_index,simpro_item_cadastro,simpro_item_preco
    """
    raw = (os.getenv('POST_IMPORT_ANALYZE_TABLES') or '').strip()
    if not raw or raw.lower() in {'0', 'false', 'no'}:
        return
    tables = [t.strip() for t in raw.split(',') if t.strip()]
    if not tables:
        return

    def _run() -> None:
        try:
            with app.app_context():
                bind = db.session.bind
                dialect = (bind.dialect.name if bind is not None else '').lower()
                if dialect not in {'mysql', 'mariadb'}:
                    return
                for tbl in tables[:40]:
                    safe = ''.join(c for c in tbl if c.isalnum() or c == '_')
                    if not safe or safe != tbl:
                        app.logger.warning('POST_IMPORT ANALYZE: nome de tabela ignorado (%r)', tbl)
                        continue
                    db.session.execute(text(f'ANALYZE TABLE `{safe}`'))
                db.session.commit()
                app.logger.info('POST_IMPORT ANALYZE TABLE concluído: %s', tables)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning('POST_IMPORT ANALYZE falhou: %s', exc)

    threading.Thread(target=_run, name='PostImportAnalyze', daemon=True).start()

PASSWORD_EXPIRATION_DAYS = 90
PASSWORD_HISTORY_SIZE = int(os.getenv('PASSWORD_HISTORY_SIZE', '5') or '5')
PASSWORD_MIN_LENGTH = int(os.getenv('PASSWORD_MIN_LENGTH', '10') or '10')
MAX_FAILED_LOGIN_ATTEMPTS = int(os.getenv('MAX_FAILED_LOGIN_ATTEMPTS', '5') or '5')
ACCOUNT_LOCK_MINUTES = int(os.getenv('ACCOUNT_LOCK_MINUTES', '15') or '15')
SESSION_LIFETIME_MINUTES = int(os.getenv('SESSION_LIFETIME_MINUTES', '120') or '120')

app.permanent_session_lifetime = timedelta(minutes=SESSION_LIFETIME_MINUTES)

usuario_operadoras = db.Table(
    'usuario_operadoras',
    db.Column('usuario_id', db.Integer, db.ForeignKey('usuarios.id'), primary_key=True),
    db.Column('operadora_id', db.Integer, db.ForeignKey('operadoras.id'), primary_key=True),
)


def _is_password_hashed(value: Optional[str]) -> bool:
    if not value:
        return False
    return value.startswith(('pbkdf2:', 'scrypt:', 'argon2:', 'bcrypt:'))


def _hash_password(raw_password: str) -> str:
    return generate_password_hash(raw_password.strip(), method='pbkdf2:sha256', salt_length=16)


def _verify_password(stored: Optional[str], candidate: str) -> bool:
    if not stored or candidate is None:
        return False
    if _is_password_hashed(stored):
        try:
            return check_password_hash(stored, candidate)
        except ValueError:
            return False
    return stored == candidate


def _now_utc() -> datetime:
    return datetime.utcnow()


# --- 1.1 Autorização/Session helpers ---
def _api_error(code: str, message: str, status: int):
    return jsonify({'error': {'code': code, 'message': message}}), status


def _extract_bearer_token(auth_header: str | None) -> str | None:
    if not auth_header:
        return None
    parts = auth_header.split()
    if len(parts) != 2:
        return None
    scheme, token = parts
    if scheme.lower() != 'bearer':
        return None
    token = (token or '').strip()
    return token or None


def _xhr_wants_json() -> bool:
    """Requisições com fetch costumam enviar este header; evita redirect HTML que quebra JSON no browser."""
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            if _xhr_wants_json() and (request.path or '').startswith('/insumos/'):
                return (
                    jsonify(
                        {
                            'status': 'error',
                            'message': 'Sessão expirada ou não autenticado. Atualize a página e faça login novamente.',
                        }
                    ),
                    401,
                )
            if os.getenv('DEV_AUTO_LOGIN', '0') == '1':
                # Auto-login for dev to unblock access when auth is unstable
                usuario = Usuario.query.order_by(Usuario.id.asc()).first()
                if usuario:
                    session.clear()
                    session.permanent = True
                    session['user_id'] = usuario.id
                    session['perfil'] = usuario.perfil
                    session['nome'] = usuario.nome
                    nomes = [op.nome for op in usuario.operadoras]
                    ids = [op.id for op in usuario.operadoras]
                    session['operadora_ids'] = ids
                    session['operadora_id'] = ids[0] if ids else None
                    session['operadora_nomes'] = nomes
                    session['operadora_nome'] = ', '.join(nomes) if nomes else None
                    session['feature_insumos'] = bool(usuario.acesso_insumos) or (usuario.perfil == 'adm')
                    session['feature_consulta'] = bool(usuario.acesso_consulta) or (usuario.perfil == 'adm')
                    session['feature_contratos'] = bool(getattr(usuario, 'acesso_contratos', True)) or (usuario.perfil in {'adm', 'adm de contrato', 'operadora'})
                    session['feature_tuss_rol'] = bool(usuario.acesso_tuss_rol) or (usuario.perfil == 'adm')
                    session['login_time'] = _now_utc().isoformat()
                    session['session_nonce'] = uuid4().hex
                    session['login_ip'] = _get_remote_addr()
                    session['password_changed_at'] = usuario.senha_atualizada_em.isoformat() if usuario.senha_atualizada_em else None
                    session['must_change_senha'] = False
                    session.modified = True
                    g.current_user = usuario
                    return f(*args, **kwargs)
            session_cookie_name = app.config.get('SESSION_COOKIE_NAME', 'session')
            has_cookie = bool(request.cookies.get(session_cookie_name))
            app.logger.warning('Auth redirect: sem user_id. path=%s has_cookie=%s', request.path, has_cookie)
            return redirect(url_for('login'))
        try:
            usuario = Usuario.query.get(user_id)
        except Exception:
            usuario = None
        if not usuario:
            app.logger.warning('Auth redirect: usuario inexistente. user_id=%s path=%s', user_id, request.path)
            session.clear()
            session.modified = True
            return redirect(url_for('login'))
        g.current_user = usuario

        login_time = _parse_iso_datetime(session.get('login_time'))
        if login_time is None:
            login_time = _now_utc()
            session['login_time'] = login_time.isoformat()
            session.modified = True
        now = _now_utc()
        if usuario.last_logout_at and login_time < usuario.last_logout_at:
            app.logger.warning(
                'Auth redirect: last_logout_at. user_id=%s login_time=%s last_logout_at=%s path=%s',
                user_id,
                login_time,
                usuario.last_logout_at,
                request.path,
            )
            session.clear()
            session.modified = True
            _safe_flash('Sua sessão expirou. Faça login novamente.', 'warning')
            return redirect(url_for('login'))
        if usuario.senha_atualizada_em and login_time < usuario.senha_atualizada_em:
            app.logger.warning(
                'Auth redirect: senha_atualizada. user_id=%s login_time=%s senha_atualizada_em=%s path=%s',
                user_id,
                login_time,
                usuario.senha_atualizada_em,
                request.path,
            )
            session.clear()
            session.modified = True
            _safe_flash('Sua sessão foi invalidada após a troca de senha. Faça login novamente.', 'warning')
            return redirect(url_for('login'))
        if usuario.locked_until and usuario.locked_until > now:
            app.logger.warning(
                'Auth redirect: locked_until. user_id=%s locked_until=%s path=%s',
                user_id,
                usuario.locked_until,
                request.path,
            )
            session.clear()
            session.modified = True
            _safe_flash('Conta temporariamente bloqueada. Faça login novamente após o desbloqueio.', 'danger')
            return redirect(url_for('login'))

        must_change_flag = bool(usuario.must_reset_senha)
        if session.get('must_change_senha') != must_change_flag:
            session['must_change_senha'] = must_change_flag
            session.modified = True
        if session.get('must_change_senha'):
            allow_endpoints = {'alterar_senha', 'logout', 'static'}
            if request.endpoint not in allow_endpoints:
                return redirect(url_for('alterar_senha'))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if session.get('perfil') != 'adm':
            if _xhr_wants_json() and (request.path or '').startswith('/insumos/'):
                return (
                    jsonify(
                        {
                            'status': 'error',
                            'message': 'Acesso negado. A importação/rotina requer perfil de administrador.',
                        }
                    ),
                    403,
                )
            return redirect(url_for('consulta_comparar'))
        return f(*args, **kwargs)
    return wrapper


def public_api_key_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not _PUBLIC_API_TOKENS:
            return _api_error('service_unavailable', 'API key não configurada.', 503)
        token = _extract_bearer_token(request.headers.get('Authorization'))
        if not token:
            return _api_error('unauthorized', 'Token ausente ou inválido.', 401)
        if token not in _PUBLIC_API_TOKENS:
            return _api_error('forbidden', 'Token inválido.', 403)
        return f(*args, **kwargs)
    return wrapper


def _feature_enabled(feature_key: str) -> bool:
    if session.get('perfil') == 'adm':
        return True
    flag = session.get(f'feature_{feature_key}')
    if flag is None:
        return False
    return bool(flag)


def feature_required(feature_key: str):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not _feature_enabled(feature_key):
                if _xhr_wants_json() and (request.path or '').startswith('/insumos/'):
                    return (
                        jsonify(
                            {
                                'status': 'error',
                                'message': 'Sem permissão para acessar Insumos/Brasíndice. Solicite acesso ao administrador.',
                            }
                        ),
                        403,
                    )
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return decorator


def _safe_flash(message: str, category: str = 'info') -> None:
    """Flash a message when in a request context, otherwise log it."""
    if has_request_context():
        flash(message, category)
        return
    log_fn = app.logger.warning if category in {'danger', 'warning'} else app.logger.info
    log_fn('flash[%s]: %s', category, message)




def _store_history_entry(entry: dict):
    entry = dict(entry or {})
    if not entry:
        return
    entry.setdefault('type', 'generic')
    entry.setdefault('label', 'Consulta recente')
    entry.setdefault('timestamp', datetime.now().strftime('%d/%m %H:%M'))
    if 'id' not in entry:
        entry['id'] = uuid4().hex[:8]
    if 'url_fragment' not in entry:
        entry['url_fragment'] = f"sim_hist={entry['id']}"
    if 'signature' not in entry:
        entry['signature'] = f"{entry.get('type', 'generic')}:{entry['id']}"
    history = session.get('sim_history') or []
    history = [h for h in history if h.get('signature') != entry['signature']]
    history.insert(0, entry)
    session['sim_history'] = history[:5]
    session.pop('ultima_simulacao_url', None)
    session.pop('ultima_simulacao_label', None)
    session.modified = True


@app.context_processor
def inject_session():
    history_raw = session.get('sim_history') or []
    history = []
    for item in history_raw:
        entry = dict(item or {})
        if 'url_fragment' not in entry and entry.get('url'):
            entry['url_fragment'] = entry['url']
        if not entry.get('label'):
            entry['label'] = 'Consulta recente'
        if 'id' not in entry:
            entry['id'] = uuid4().hex[:8]
        if 'signature' not in entry:
            entry['signature'] = f"legacy:{entry.get('url_fragment', entry['id'])}"
        history.append(entry)
    if not history:
        legacy_url = session.get('ultima_simulacao_url')
        if legacy_url:
            history = [{
                'id': uuid4().hex[:8],
                'type': 'compare',
                'url_fragment': legacy_url,
                'label': session.get('ultima_simulacao_label') or 'Consulta recente',
                'timestamp': '',
                'signature': f"legacy:{legacy_url}",
            }]
    last = history[0] if history else {}
    return {
        "session_perfil": session.get('perfil'),
        "session_nome": session.get('nome'),
        "session_last_simulation_url": last.get('url_fragment'),
        "session_last_simulation_label": last.get('label'),
        "session_sim_history": history,
        "session_operadora_nome": session.get('operadora_nome'),
        "session_operadora_id": session.get('operadora_id'),
        "session_operadora_nomes": session.get('operadora_nomes') or [],
        "session_operadora_ids": session.get('operadora_ids') or [],
        "session_feature_insumos": (session.get('feature_insumos') if session.get('feature_insumos') is not None else (session.get('perfil') == 'adm')),
        "session_feature_consulta": (session.get('feature_consulta') if session.get('feature_consulta') is not None else (session.get('perfil') == 'adm')),
        "session_feature_contratos": (session.get('feature_contratos') if session.get('feature_contratos') is not None else (session.get('perfil') in {'adm', 'adm de contrato', 'operadora'})),
        "session_feature_tuss_rol": (session.get('feature_tuss_rol') if session.get('feature_tuss_rol') is not None else (session.get('perfil') == 'adm')),
        "security_password_min_length": PASSWORD_MIN_LENGTH,
        "security_password_history_size": PASSWORD_HISTORY_SIZE,
    }


# --- 2. DEFINIÇÃO DOS MODELOS (TABELAS) ---
# Cada classe representa uma tabela no banco de dados.

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), nullable=False)
    email = db.Column(db.String(255), unique=True, nullable=False)
    senha = db.Column(db.String(255), nullable=False)
    perfil = db.Column(db.String(50), nullable=False)
    operadoras = db.relationship(
        'Operadora',
        secondary='usuario_operadoras',
        lazy='joined',
        backref=db.backref('usuarios', lazy='dynamic')
    )
    acesso_insumos = db.Column(db.Boolean, nullable=False, default=True, server_default=text('1'))
    acesso_consulta = db.Column(db.Boolean, nullable=False, default=True, server_default=text('1'))
    acesso_contratos = db.Column(db.Boolean, nullable=False, default=True, server_default=text('1'))
    acesso_tuss_rol = db.Column(db.Boolean, nullable=False, default=True, server_default=text('1'))
    must_reset_senha = db.Column(db.Boolean, nullable=False, default=True, server_default=text('1'))
    senha_atualizada_em = db.Column(db.DateTime, nullable=True)
    failed_login_attempts = db.Column(db.Integer, nullable=False, default=0, server_default=text('0'))
    locked_until = db.Column(db.DateTime, nullable=True)
    last_logout_at = db.Column(db.DateTime, nullable=True)
    senhas_historico = db.relationship(
        'UsuarioSenhaHistorico',
        backref='usuario',
        lazy='dynamic',
        cascade='all, delete-orphan'
    )


class UsuarioSenhaHistorico(db.Model):
    __tablename__ = 'usuario_senhas_historico'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id', ondelete='CASCADE'), nullable=False, index=True)
    senha_hash = db.Column(db.String(255), nullable=False)
    criada_em = db.Column(db.DateTime, nullable=False, default=_now_utc)


class ContractSummary(db.Model):
    __tablename__ = 'contratos_resumo'

    id = db.Column(db.Integer, primary_key=True)
    prestador = db.Column(db.String(255), nullable=False)
    tabela_honorarios = db.Column(db.String(255), nullable=True)
    tabela_portes = db.Column(db.String(255), nullable=True)
    valor_uco = db.Column(db.Numeric(12, 4), nullable=True)
    inflator_deflator = db.Column(db.String(120), nullable=True)
    filme_radiologico = db.Column(db.String(120), nullable=True)
    observacoes = db.Column(db.Text, nullable=True)
    # Multi-operadora: cada contrato pertence a uma operadora
    operadora_id = db.Column(db.Integer, db.ForeignKey('operadoras.id', ondelete='CASCADE'),
                             nullable=False, default=1, index=True)
    created_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )


class ReembolsoDocumento(db.Model):
    __tablename__ = 'reembolso_documentos'

    id = db.Column(db.BigInteger, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id', ondelete='SET NULL'), nullable=True, index=True)
    tipo_documento = db.Column(
        db.Enum('NOTA_FISCAL', 'RECIBO', 'COMPROVANTE', 'DESCONHECIDO', name='reembolso_tipo_documento'),
        nullable=False,
        default='DESCONHECIDO',
        server_default='DESCONHECIDO',
    )
    original_filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    storage_path = db.Column(db.String(512), nullable=False)
    mime_type = db.Column(db.String(120), nullable=True)
    is_pdf_native = db.Column(db.Boolean, nullable=False, default=False, server_default=text('0'))
    texto_extraido = db.Column(db.Text, nullable=True)
    dados_extraidos = db.Column(db.JSON, nullable=True)
    dados_validado = db.Column(db.JSON, nullable=True)
    status = db.Column(
        db.Enum('PENDENTE', 'VALIDADO', 'REJEITADO', name='reembolso_status'),
        nullable=False,
        default='PENDENTE',
        server_default='PENDENTE',
    )
    ocr_status = db.Column(db.String(40), nullable=True)
    ocr_message = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=_now_utc)
    updated_at = db.Column(db.DateTime, nullable=False, default=_now_utc, onupdate=_now_utc)

    usuario = db.relationship('Usuario', backref=db.backref('reembolsos', lazy='dynamic'))

    __table_args__ = (
        db.Index('idx_reembolso_usuario', 'usuario_id'),
        db.Index('idx_reembolso_created_at', 'created_at'),
    )


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.BigInteger, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id', ondelete='SET NULL'), nullable=True, index=True)
    email_alvo = db.Column(db.String(255), nullable=True)
    evento = db.Column(db.String(64), nullable=False)
    ip = db.Column(db.String(64), nullable=True)
    detalhes = db.Column(db.Text, nullable=True)
    criado_em = db.Column(db.DateTime, nullable=False, default=_now_utc, index=True)
    usuario = db.relationship('Usuario', lazy='joined')


_PASSWORD_COMPLEXITY_TESTS = [
    (re.compile(r'[A-Z]'), 'ao menos uma letra maiúscula'),
    (re.compile(r'[a-z]'), 'ao menos uma letra minúscula'),
    (re.compile(r'[0-9]'), 'ao menos um dígito'),
    (re.compile(r'[^A-Za-z0-9]'), 'ao menos um caractere especial'),
]


def _password_policy_error(password: str) -> Optional[str]:
    if len(password or '') < PASSWORD_MIN_LENGTH:
        return f'A senha deve conter pelo menos {PASSWORD_MIN_LENGTH} caracteres.'
    for pattern, description in _PASSWORD_COMPLEXITY_TESTS:
        if not pattern.search(password or ''):
            return f'A senha deve conter {description}.'
    return None


def _password_was_used_recently(usuario: Usuario, password: str) -> bool:
    if not usuario or not password:
        return False
    historico = (
        UsuarioSenhaHistorico.query.filter_by(usuario_id=usuario.id)
        .order_by(UsuarioSenhaHistorico.id.desc())
        .limit(PASSWORD_HISTORY_SIZE)
        .all()
    )
    for registro in historico:
        try:
            if check_password_hash(registro.senha_hash, password):
                return True
        except ValueError:
            continue
    if usuario.senha and _is_password_hashed(usuario.senha):
        try:
            return check_password_hash(usuario.senha, password)
        except ValueError:
            return False
    return usuario.senha == password


def _append_password_history(usuario: Usuario, senha_hash: str) -> None:
    if not usuario or not senha_hash:
        return
    registro = UsuarioSenhaHistorico(usuario_id=usuario.id, senha_hash=senha_hash)
    db.session.add(registro)
    db.session.flush()
    excess = (
        UsuarioSenhaHistorico.query
        .filter(UsuarioSenhaHistorico.usuario_id == usuario.id)
        .order_by(UsuarioSenhaHistorico.id.desc())
        .offset(PASSWORD_HISTORY_SIZE)
    ).all()
    ids_to_delete = [r.id for r in excess]
    if ids_to_delete:
        UsuarioSenhaHistorico.query.filter(UsuarioSenhaHistorico.id.in_(ids_to_delete)).delete(synchronize_session=False)


def _get_remote_addr() -> Optional[str]:
    if not has_request_context():
        return None
    forwarded = request.headers.get('X-Forwarded-For')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr


def _register_audit(evento: str, usuario: Optional[Usuario] = None, email_alvo: Optional[str] = None, detalhes: Optional[dict | str] = None) -> None:
    try:
        detalhes_str = None
        if isinstance(detalhes, dict):
            detalhes_str = json.dumps(detalhes, ensure_ascii=False)
        elif detalhes is not None:
            detalhes_str = str(detalhes)
        registro = AuditLog(
            usuario_id=usuario.id if usuario else None,
            email_alvo=email_alvo,
            evento=evento,
            ip=_get_remote_addr(),
            detalhes=detalhes_str,
        )
        db.session.add(registro)
    except Exception as exc:
        app.logger.warning('Falha ao registrar auditoria %s: %s', evento, exc)


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except (ValueError, TypeError):
        return None


class Operadora(db.Model):
    __tablename__ = 'operadoras'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), nullable=False)
    uf = db.Column(db.String(2), nullable=True)
    cnpj = db.Column(db.String(20), nullable=True)
    status = db.Column(db.String(50), nullable=False)
    # Relacionamento: uma operadora pode ter várias tabelas de preços
    tabelas = db.relationship('Tabela', backref='operadora', lazy=True)

class Tabela(db.Model):
    __tablename__ = 'tabelas'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), nullable=False)
    data_vigencia = db.Column(db.Date, nullable=True)
    prestador = db.Column(db.String(255), nullable=True)
    tipo_tabela = db.Column(db.String(50), nullable=True)
    uf = db.Column(db.String(2), nullable=True)
    uco_valor = db.Column(db.Numeric(12, 2), nullable=True)
    # Chave estrangeira para ligar à tabela de operadoras
    id_operadora = db.Column(db.Integer, db.ForeignKey('operadoras.id'), nullable=False)
    # Relacionamento: uma tabela contém vários procedimentos
    procedimentos = db.relationship('Procedimento', backref='tabela', lazy=True)

class Procedimento(db.Model):
    __tablename__ = 'procedimentos'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.String(500), nullable=False)
    valor = db.Column(db.Numeric(10, 2), nullable=False)
    prestador = db.Column(db.String(255), nullable=True)
    uf = db.Column(db.String(2), nullable=True)
    # Chave estrangeira para ligar à tabela de preços
    id_tabela = db.Column(db.Integer, db.ForeignKey('tabelas.id'), nullable=False)
    # Multi-operadora: cada procedimento/DTP pode ter valores específicos por operadora
    operadora_id = db.Column(db.Integer, db.ForeignKey('operadoras.id', ondelete='CASCADE'),
                             nullable=False, default=1)

    # Relacionamento
    operadora = db.relationship('Operadora', backref='procedimentos')


class CBHPMItem(db.Model):
    __tablename__ = 'cbhpm_itens'
    id = db.Column(db.Integer, primary_key=True)
    # básicos
    codigo = db.Column(db.String(100), nullable=False)
    procedimento = db.Column(db.String(500), nullable=False)
    uf = db.Column(db.String(2), nullable=True)
    # porte cirúrgico
    porte = db.Column(db.String(50), nullable=True)
    fracao_porte = db.Column(db.Numeric(10, 2), nullable=True)
    valor_porte = db.Column(db.Numeric(12, 2), nullable=True)
    total_porte = db.Column(db.Numeric(12, 2), nullable=True)
    # incidências e filme
    incidencias = db.Column(db.String(255), nullable=True)
    filme = db.Column(db.Numeric(12, 2), nullable=True)
    total_filme = db.Column(db.Numeric(12, 2), nullable=True)
    # uco
    uco = db.Column(db.Numeric(12, 2), nullable=True)
    total_uco = db.Column(db.Numeric(12, 2), nullable=True)
    # anestesia
    porte_anestesico = db.Column(db.String(50), nullable=True)
    valor_porte_anestesico = db.Column(db.Numeric(12, 2), nullable=True)
    total_porte_anestesico = db.Column(db.Numeric(12, 2), nullable=True)
    # auxiliares
    numero_auxiliares = db.Column(db.Integer, nullable=True)
    total_auxiliares = db.Column(db.Numeric(12, 2), nullable=True)
    total_1_aux = db.Column(db.Numeric(12, 2), nullable=True)
    total_2_aux = db.Column(db.Numeric(12, 2), nullable=True)
    total_3_aux = db.Column(db.Numeric(12, 2), nullable=True)
    total_4_aux = db.Column(db.Numeric(12, 2), nullable=True)
    # subtotal
    subtotal = db.Column(db.Numeric(12, 2), nullable=True)
    # vínculo
    id_tabela = db.Column(db.Integer, db.ForeignKey('tabelas.id'), nullable=False)


class TussRolCorrelacao(db.Model):
    __tablename__ = 'tuss_rol_correlacoes'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(30), unique=True, nullable=False)
    descricao = db.Column(db.String(500), nullable=True)
    consta_rol = db.Column(db.Boolean, nullable=False, default=False)
    atualizado_em = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
        server_onupdate=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_tuss_rol_codigo', 'codigo'),
        db.Index('idx_tuss_rol_flag', 'consta_rol'),
    )


class CbhpmTeto(db.Model):
    __tablename__ = 'cbhpm_teto'

    codigo = db.Column(db.String(20), primary_key=True)
    operadora_id = db.Column(db.Integer, db.ForeignKey('operadoras.id', ondelete='CASCADE'), primary_key=True, nullable=False)
    descricao = db.Column(db.String(255), nullable=False)
    valor_total = db.Column(db.Numeric(15, 2), nullable=False)
    updated_at = db.Column(
        db.TIMESTAMP,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
        server_onupdate=text('CURRENT_TIMESTAMP'),
    )

    # Relacionamento
    operadora = db.relationship('Operadora', backref='tetos_cbhpm')

    __table_args__ = (
        db.Index('idx_cbhpm_teto_descricao', 'descricao'),
        db.Index('idx_cbhpm_teto_operadora', 'operadora_id'),
    )


class LoteStatus(Enum):
    PENDENTE = 'PENDENTE'
    VALIDADO = 'VALIDADO'
    REPROVADO = 'REPROVADO'
    PUBLICADO = 'PUBLICADO'


class UfAliquota(db.Model):
    __tablename__ = 'uf_aliquota'

    uf = db.Column(db.String(2), primary_key=True)
    valid_from = db.Column(db.Date, primary_key=True)
    aliquota_bp = db.Column(db.Integer, nullable=False)
    valid_to = db.Column(db.Date, nullable=True)
    is_current = db.Column(db.Boolean, nullable=False, server_default=text('1'), default=True)
    created_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    updated_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'), server_onupdate=text('CURRENT_TIMESTAMP'))

    __table_args__ = (
        db.CheckConstraint('aliquota_bp >= 0', name='ck_uf_aliquota_non_negative'),
        db.Index('idx_uf_aliquota_current', 'uf', 'is_current'),
    )


class Lote(db.Model):
    __tablename__ = 'lote'

    id = db.Column(db.BigInteger, primary_key=True)
    fornecedor = db.Column(db.String(50), nullable=False)
    aliquota_bp = db.Column(db.Integer, nullable=False)
    periodo = db.Column(db.String(6), nullable=False)
    sequencia = db.Column(db.SmallInteger, nullable=False)
    arquivo_label = db.Column(db.String(255), nullable=False)
    hash_arquivo = db.Column(db.String(128), nullable=True)
    total_itens = db.Column(db.Integer, nullable=True)
    status = db.Column(db.Enum(LoteStatus, name='lote_status'), nullable=False, default=LoteStatus.PENDENTE, server_default=LoteStatus.PENDENTE.value)
    validado_em = db.Column(db.TIMESTAMP, nullable=True)
    publicado_em = db.Column(db.TIMESTAMP, nullable=True)
    created_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    updated_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'), server_onupdate=text('CURRENT_TIMESTAMP'))

    __table_args__ = (
        db.UniqueConstraint('fornecedor', 'aliquota_bp', 'periodo', 'sequencia', name='uq_lote_identidade'),
        db.Index('idx_lote_status', 'status'),
        db.Index('idx_lote_arquivo_label', 'arquivo_label'),
    )


class Publicacao(db.Model):
    __tablename__ = 'publicacao'

    id = db.Column(db.BigInteger, primary_key=True)
    fornecedor = db.Column(db.String(50), nullable=False)
    aliquota_bp = db.Column(db.Integer, nullable=False)
    periodo = db.Column(db.String(6), nullable=False)
    sequencia = db.Column(db.SmallInteger, nullable=False)
    lote_id = db.Column(db.BigInteger, db.ForeignKey('lote.id', ondelete='CASCADE'), nullable=False)
    publicado_em = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    etag_versao = db.Column(db.String(128), nullable=False)
    criado_em = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'))

    lote = db.relationship('Lote', backref=db.backref('publicacoes', lazy='dynamic'))

    __table_args__ = (
        db.UniqueConstraint('fornecedor', 'aliquota_bp', 'periodo', 'sequencia', name='uq_publicacao_identidade'),
        db.Index('idx_publicacao_fornecedor', 'fornecedor'),
    )


class LinhaHash(db.Model):
    __tablename__ = 'linha_hash'

    id = db.Column(db.BigInteger, primary_key=True)
    lote_id = db.Column(db.BigInteger, db.ForeignKey('lote.id', ondelete='CASCADE'), nullable=False)
    item_chave = db.Column(db.String(255), nullable=False)
    hash_linha = db.Column(db.String(128), nullable=False)
    payload_snapshot = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    updated_at = db.Column(db.TIMESTAMP, nullable=False, server_default=text('CURRENT_TIMESTAMP'), server_onupdate=text('CURRENT_TIMESTAMP'))

    lote = db.relationship('Lote', backref=db.backref('linhas', lazy='dynamic', cascade='all, delete-orphan'))

    __table_args__ = (
        db.UniqueConstraint('lote_id', 'item_chave', name='uq_linha_hash_item'),
        db.Index('idx_linha_hash_lote', 'lote_id'),
    )


class ImportJobStatus(Enum):
    PENDING = 'PENDING'
    RUNNING = 'RUNNING'
    SUCCESS = 'SUCCESS'
    FAILED = 'FAILED'


class ImportPauseRequested(RuntimeError):
    """Sinaliza que a importação foi pausada sob solicitação do usuário."""


class ImportJob(db.Model):
    __tablename__ = 'insumo_import_jobs'

    id = db.Column(db.String(36), primary_key=True)
    origem = db.Column(db.Enum('BRAS', 'SIMPRO', name='insumo_import_job_origem'), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    data_path = db.Column(db.String(512), nullable=False)
    status = db.Column(
        db.Enum('PENDING', 'RUNNING', 'SUCCESS', 'FAILED', name='insumo_import_job_status'),
        nullable=False,
        default=ImportJobStatus.PENDING.value,
        server_default=ImportJobStatus.PENDING.value,
    )
    message = db.Column(db.String(500), nullable=True)
    total_linhas = db.Column(db.Integer, nullable=True)
    linhas_materializadas = db.Column(db.Integer, nullable=True)
    versao = db.Column(db.String(50), nullable=True)
    aliquota = db.Column(db.Numeric(12, 4), nullable=True)
    uf_list = db.Column(db.String(255), nullable=True)
    params = db.Column(db.JSON, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)

    __table_args__ = (
        db.Index('idx_import_jobs_created_at', 'created_at'),
        db.Index('idx_import_jobs_status', 'status'),
    )


def _job_message_trim(message: str | None, *, limit: int = 500) -> str | None:
    if not message:
        return None
    msg = message.strip()
    if not msg:
        return None
    if len(msg) <= limit:
        return msg
    return msg[: limit - 3] + '...'


def _touch_import_job_progress(
    job_id: str | None,
    *,
    message: str | None = None,
    total_linhas: int | None = None,
    linhas_materializadas: int | None = None,
) -> None:
    """Atualiza linha do job durante import longo (SIMPRO JSON, etc.). Falhas são ignoradas (não abortam import)."""
    if not job_id:
        return
    try:
        job = ImportJob.query.get(job_id)
        if not job:
            return
        if message is not None:
            job.message = _job_message_trim(message)
        if total_linhas is not None:
            job.total_linhas = total_linhas
        if linhas_materializadas is not None:
            job.linhas_materializadas = linhas_materializadas
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        app.logger.warning('Falha ao atualizar progresso do job %s: %s', job_id, exc)


_IMPORT_JOB_CONTEXT = threading.local()
_IMPORT_PAUSE_REQUESTS_LOCK = threading.Lock()
_IMPORT_PAUSE_REQUESTS: set[str] = set()


def _set_current_import_job(job_id: str | None) -> None:
    _IMPORT_JOB_CONTEXT.job_id = job_id


def _get_current_import_job() -> str | None:
    return getattr(_IMPORT_JOB_CONTEXT, 'job_id', None)


def _request_import_pause(job_id: str) -> None:
    with _IMPORT_PAUSE_REQUESTS_LOCK:
        _IMPORT_PAUSE_REQUESTS.add(job_id)


def _clear_import_pause_request(job_id: str) -> None:
    with _IMPORT_PAUSE_REQUESTS_LOCK:
        _IMPORT_PAUSE_REQUESTS.discard(job_id)


def _is_import_pause_requested(job_id: str | None = None) -> bool:
    target = job_id or _get_current_import_job()
    if not target:
        return False
    with _IMPORT_PAUSE_REQUESTS_LOCK:
        return target in _IMPORT_PAUSE_REQUESTS


def _raise_if_import_paused(job_id: str | None = None) -> None:
    target = job_id or _get_current_import_job()
    if target and _is_import_pause_requested(target):
        raise ImportPauseRequested('Importação pausada pelo usuário.')

class BrasRaw(db.Model):
    __tablename__ = 'bras_raw'

    id = db.Column(db.BigInteger, primary_key=True)
    arquivo = db.Column(db.String(255), nullable=False)
    linha_num = db.Column(db.Integer, nullable=False)
    col01 = db.Column(db.String(255))
    col02 = db.Column(db.String(255))
    col03 = db.Column(db.String(255))
    col04 = db.Column(db.String(255))
    col05 = db.Column(db.String(255))
    col06 = db.Column(db.String(255))
    col07 = db.Column(db.String(255))
    col08 = db.Column(db.String(255))
    col09 = db.Column(db.String(255))
    col10 = db.Column(db.String(255))
    col11 = db.Column(db.String(255))
    col12 = db.Column(db.String(255))
    col13 = db.Column(db.String(255))
    col14 = db.Column(db.String(255))
    col15 = db.Column(db.String(255))
    col16 = db.Column(db.String(255))
    col17 = db.Column(db.String(255))
    col18 = db.Column(db.String(255))
    col19 = db.Column(db.String(255))
    col20 = db.Column(db.String(255))
    col21 = db.Column(db.String(255))
    col22 = db.Column(db.String(255))
    col23 = db.Column(db.String(255))
    imported_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_bras_raw_arquivo', 'arquivo'),
        db.Index('idx_bras_raw_col17', 'col17'),
        db.Index('idx_bras_raw_col03', 'col03'),
        db.Index('idx_bras_raw_col06', 'col06'),
    )


class BrasFixedStage(db.Model):
    __tablename__ = 'bras_fixed_stage'

    id = db.Column(db.BigInteger, primary_key=True)
    arquivo = db.Column(db.String(255), nullable=False)
    linha_num = db.Column(db.Integer, nullable=False)
    linha = db.Column(db.Text, nullable=False)
    imported_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_bras_fixed_arquivo', 'arquivo'),
    )


class BrasItemNormalized(db.Model):
    __tablename__ = 'bras_item_n'

    id = db.Column(db.BigInteger, primary_key=True)
    arquivo = db.Column(db.String(255), nullable=False)
    linha_num = db.Column(db.Integer, nullable=False)
    laboratorio_codigo = db.Column(db.String(50), nullable=True)
    laboratorio_nome = db.Column(db.String(255), index=True, nullable=True)
    produto_codigo = db.Column(db.String(50), index=True, nullable=True)
    produto_nome = db.Column(db.String(255), index=True, nullable=True)
    apresentacao_codigo = db.Column(db.String(50), nullable=True)
    apresentacao_descricao = db.Column(db.String(255), index=True, nullable=True)
    ean = db.Column(db.String(20), index=True, nullable=True)
    registro_anvisa = db.Column(db.String(50), index=True, nullable=True)
    edicao = db.Column(db.String(50), index=True, nullable=True)
    preco_pmc_pacote = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pfb_pacote = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pmc_unit = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pfb_unit = db.Column(db.Numeric(15, 4), nullable=True)
    aliquota_ou_ipi = db.Column(db.Numeric(15, 4), nullable=True)
    quantidade_embalagem = db.Column(db.Integer, nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True)

    __table_args__ = (
        db.Index('idx_bras_item_n_ean', 'ean'),
        db.Index('idx_bras_item_n_prod', 'produto_codigo'),
        db.Index('idx_bras_item_n_desc', 'produto_nome', 'apresentacao_descricao'),
        db.Index('idx_bras_item_n_anvisa', 'registro_anvisa'),
        db.Index('idx_bras_item_n_edicao', 'edicao'),
    )


class BrasCatalogSnapshot(db.Model):
    __tablename__ = 'bras_catalog_snapshot'

    id = db.Column(db.BigInteger, primary_key=True)
    versao = db.Column(db.String(50), nullable=False, index=True)
    source_file = db.Column(db.String(255), nullable=True)
    item_key = db.Column(db.String(255), nullable=False)
    key_kind = db.Column(db.String(32), nullable=False)
    laboratorio_codigo = db.Column(db.String(50), nullable=True)
    laboratorio_nome = db.Column(db.String(255), nullable=True)
    produto_codigo = db.Column(db.String(50), nullable=True)
    produto_nome = db.Column(db.String(255), nullable=True)
    apresentacao_codigo = db.Column(db.String(50), nullable=True)
    apresentacao_descricao = db.Column(db.String(255), nullable=True)
    codigo_composto = db.Column(db.String(100), nullable=True)
    ean = db.Column(db.String(32), nullable=True, index=True)
    codigo_interno = db.Column(db.String(50), nullable=True)
    tuss = db.Column(db.String(32), nullable=True, index=True)
    row_hash = db.Column(db.String(64), nullable=False)
    imported_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.UniqueConstraint('versao', 'item_key', name='uq_bras_catalog_snapshot_versao_item'),
        db.Index('idx_bras_catalog_snapshot_item', 'item_key'),
        db.Index('idx_bras_catalog_snapshot_version_hash', 'versao', 'row_hash'),
    )


class BrasItemCadastro(db.Model):
    """
    Identidade do item (cadastro) por edição — alinhado a uma única linha lógica por (edição, EAN).
    Preços por alíquota ficam em `BrasItemPreco` (carga 1x cadastro, depois só preços leves).
    """

    __tablename__ = 'bras_item_cadastro'

    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    edicao = db.Column(db.String(50), nullable=False, index=True)
    ean = db.Column(db.String(20), index=True, nullable=True)
    laboratorio_codigo = db.Column(db.String(50), nullable=True)
    laboratorio_nome = db.Column(db.String(255), nullable=True)
    produto_codigo = db.Column(db.String(50), nullable=True)
    produto_nome = db.Column(db.String(255), nullable=True)
    apresentacao_codigo = db.Column(db.String(50), nullable=True)
    apresentacao_descricao = db.Column(db.String(255), nullable=True)
    registro_anvisa = db.Column(db.String(50), nullable=True)
    quantidade_embalagem = db.Column(db.Integer, nullable=True)
    linha_num = db.Column(db.Integer, nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True, server_default=text('CURRENT_TIMESTAMP'))

    __table_args__ = (
        db.UniqueConstraint('edicao', 'ean', name='uq_bras_cadastro_edicao_ean'),
        db.Index('idx_bras_cad_edicao', 'edicao', 'ean'),
    )


class BrasItemPreco(db.Model):
    """Preços de um item por alíquota (Uma linha por (cadastro, alíquota))."""

    __tablename__ = 'bras_item_preco'

    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    cadastro_id = db.Column(db.BigInteger, db.ForeignKey('bras_item_cadastro.id'), nullable=False, index=True)
    aliquota = db.Column(db.Numeric(6, 2), nullable=False, index=True)
    preco_pmc_pacote = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pfb_pacote = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pmc_unit = db.Column(db.Numeric(15, 4), nullable=True)
    preco_pfb_unit = db.Column(db.Numeric(15, 4), nullable=True)
    arquivo_fonte = db.Column(db.String(255), nullable=True)
    imported_at = db.Column(
        db.DateTime,
        nullable=True,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.UniqueConstraint('cadastro_id', 'aliquota', name='uq_bras_preco_cadastro_aliquota'),
        db.Index('idx_bras_preco_cad', 'cadastro_id', 'aliquota'),
    )


class SimproItem(db.Model):
    __tablename__ = 'simpro_item'

    id = db.Column(db.Integer, primary_key=True)
    tuss = db.Column(db.String(50), index=True, nullable=True)
    tiss = db.Column(db.String(50), index=True, nullable=True)
    anvisa = db.Column(db.String(50), index=True, nullable=True)
    descricao = db.Column(db.String(500), nullable=False, index=True)
    preco = db.Column(db.Numeric(12, 4), nullable=True)
    aliquota = db.Column(db.Numeric(12, 4), nullable=True)
    fabricante = db.Column(db.String(255), nullable=True)
    versao_tabela = db.Column(db.String(100), nullable=True)
    data_atualizacao = db.Column(db.Date, nullable=True)
    uf_referencia = db.Column(db.String(64), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
        server_onupdate=text('CURRENT_TIMESTAMP'),
    )


class SimproItemCadastro(db.Model):
    """Cadastro SIMPRO por identidade (versão + item-chave)."""

    __tablename__ = 'simpro_item_cadastro'

    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    versao = db.Column(db.String(100), nullable=False, index=True)
    item_key = db.Column(db.String(64), nullable=False, index=True)
    tuss_numero = db.Column(db.String(16), nullable=True, index=True)
    codigo = db.Column(db.String(20), nullable=True, index=True)
    codigo_interno = db.Column(db.String(20), nullable=True)
    codigo_alt = db.Column(db.String(20), nullable=True)
    descricao = db.Column(db.String(255), nullable=True, index=True)
    fabricante = db.Column(db.String(80), nullable=True)
    referencia = db.Column(db.String(120), nullable=True)
    anvisa = db.Column(db.String(20), nullable=True, index=True)
    ean = db.Column(db.String(32), nullable=True, index=True)
    unidade = db.Column(db.String(16), nullable=True)
    qtd_unidade = db.Column(db.Integer, nullable=True)
    fracionavel = db.Column(db.String(1), nullable=True)
    status_final = db.Column(db.String(8), nullable=True)
    data_ref = db.Column(db.Date, nullable=True)
    linha_num = db.Column(db.Integer, nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True, server_default=text('CURRENT_TIMESTAMP'))

    __table_args__ = (
        db.UniqueConstraint('versao', 'item_key', name='uq_simpro_cadastro_versao_item'),
        db.Index('idx_simpro_cad_versao_item', 'versao', 'item_key'),
    )


class SimproItemPreco(db.Model):
    """Preços SIMPRO por alíquota para cada cadastro."""

    __tablename__ = 'simpro_item_preco'

    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    cadastro_id = db.Column(db.BigInteger, db.ForeignKey('simpro_item_cadastro.id'), nullable=False, index=True)
    aliquota = db.Column(db.Numeric(6, 2), nullable=False, index=True)
    preco1 = db.Column(db.Numeric(15, 4), nullable=True)
    preco2 = db.Column(db.Numeric(15, 4), nullable=True)
    preco3 = db.Column(db.Numeric(15, 4), nullable=True)
    preco4 = db.Column(db.Numeric(15, 4), nullable=True)
    arquivo_fonte = db.Column(db.String(255), nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True, server_default=text('CURRENT_TIMESTAMP'))

    __table_args__ = (
        db.UniqueConstraint('cadastro_id', 'aliquota', name='uq_simpro_preco_cadastro_aliquota'),
        db.Index('idx_simpro_preco_cad', 'cadastro_id', 'aliquota'),
    )


def _cleanup_simpro_cadastro_orphans(cadastro_ids: Sequence[int] | None = None) -> int:
    query = SimproItemCadastro.query
    if cadastro_ids is not None:
        ids = [int(cid) for cid in cadastro_ids if cid is not None]
        if not ids:
            return 0
        query = query.filter(SimproItemCadastro.id.in_(ids))
    rows = query.with_entities(SimproItemCadastro.id).all()
    if not rows:
        return 0
    target_ids = [int(row.id) for row in rows if row.id is not None]
    if not target_ids:
        return 0

    referenced = {
        int(row.cadastro_id)
        for row in db.session.query(SimproItemPreco.cadastro_id)
        .filter(SimproItemPreco.cadastro_id.in_(target_ids))
        .all()
        if row.cadastro_id is not None
    }
    orphan_ids = [cid for cid in target_ids if cid not in referenced]
    if not orphan_ids:
        return 0
    deleted = (
        SimproItemCadastro.query
        .filter(SimproItemCadastro.id.in_(orphan_ids))
        .delete(synchronize_session=False)
    ) or 0
    return int(deleted)


class SimproFixedStage(db.Model):
    __tablename__ = 'simpro_fixed_stage'

    id = db.Column(db.BigInteger, primary_key=True)
    arquivo = db.Column(db.String(255), nullable=False)
    linha_num = db.Column(db.Integer, nullable=False)
    linha = db.Column(db.Text, nullable=False)
    imported_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_simpro_fixed_arquivo', 'arquivo'),
    )


class SimproItemNormalized(db.Model):
    __tablename__ = 'simpro_item_norm'

    id = db.Column(db.BigInteger, primary_key=True)
    arquivo = db.Column(db.String(255), nullable=False)
    linha_num = db.Column(db.Integer, nullable=False)
    codigo_interno = db.Column(db.String(20), nullable=True)
    codigo = db.Column(db.String(20), index=True, nullable=False)
    codigo_alt = db.Column(db.String(20), index=True, nullable=True)
    descricao = db.Column(db.String(255), index=True, nullable=False)
    data_ref = db.Column(db.Date, nullable=True)
    tipo_reg = db.Column(db.String(4), nullable=True)
    preco1 = db.Column(db.Numeric(15, 4), nullable=True)
    preco2 = db.Column(db.Numeric(15, 4), nullable=True)
    preco3 = db.Column(db.Numeric(15, 4), nullable=True)
    preco4 = db.Column(db.Numeric(15, 4), nullable=True)
    unidade = db.Column(db.String(16), nullable=True)
    qtd_unidade = db.Column(db.Integer, nullable=True)
    fabricante = db.Column(db.String(80), nullable=True)
    referencia = db.Column(db.String(120), nullable=True)
    anvisa = db.Column(db.String(20), index=True, nullable=True)
    validade_anvisa = db.Column(db.Date, nullable=True)
    ean = db.Column(db.String(32), index=True, nullable=True)
    situacao = db.Column(db.String(40), nullable=True)
    fracionavel = db.Column(db.String(1), nullable=True)
    versao = db.Column(db.String(100), nullable=True)
    uf_referencia = db.Column(db.String(64), nullable=True)
    tuss_prefix = db.Column(db.String(4), nullable=True)
    tuss_numero = db.Column(db.String(16), nullable=True)
    status_final = db.Column(db.String(8), nullable=True)
    imported_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_simpro_item_norm_cod_interno', 'codigo_interno'),
        db.Index('idx_simpro_item_norm_desc', 'descricao'),
        db.Index('idx_simpro_item_norm_ean', 'ean'),
        db.Index('idx_simpro_item_norm_anvisa', 'anvisa'),
        db.Index('idx_simpro_item_norm_versao', 'versao'),
        db.Index('idx_simpro_item_norm_tuss_numero', 'tuss_numero'),
    )


def _strip_json_comments(text: str) -> str:
    result_chars: list[str] = []
    in_string = False
    escape = False
    idx = 0
    length = len(text)
    while idx < length:
        ch = text[idx]
        if in_string:
            result_chars.append(ch)
            if escape:
                escape = False
            elif ch == '\\':
                escape = True
            elif ch == '"':
                in_string = False
            idx += 1
            continue
        if ch == '"':
            in_string = True
            result_chars.append(ch)
            idx += 1
            continue
        if ch == '/' and idx + 1 < length:
            nxt = text[idx + 1]
            if nxt == '/':
                idx += 2
                while idx < length and text[idx] not in ('\n', '\r'):
                    idx += 1
                continue
            if nxt == '*':
                idx += 2
                while idx + 1 < length and not (text[idx] == '*' and text[idx + 1] == '/'):
                    idx += 1
                idx += 2
                continue
        result_chars.append(ch)
        idx += 1
    return ''.join(result_chars)


def _load_json_relaxed(text: str) -> dict:
    cleaned = _strip_json_comments(text)
    return json.loads(cleaned)


_SIMPRO_FIELD_MAP_DEFAULT: dict[str, str] = {
    'codigo': 'codigo',
    'codigo_simpro': 'codigo',
    'codigo_usuario': 'codigo_interno',
    'codigo_fracao': 'codigo_alt',
    'codigo_alt': 'codigo_alt',
    'codigo_interno': 'codigo_interno',
    'codigo_alternativo': 'codigo_alt',
    'descricao': 'descricao',
    'descricao_completa': 'descricao',
    'data_ref': 'data_ref',
    'data_vigencia': 'data_ref',
    'tipo_reg': 'tipo_reg',
    'tipo_registro': 'tipo_reg',
    'preco1': 'preco1',
    'preco_pf': 'preco1',
    'preco2': 'preco2',
    'preco_pmc': 'preco2',
    'preco3': 'preco3',
    'preco_ph': 'preco3',
    'preco4': 'preco4',
    'preco_outro': 'preco4',
    'unidade': 'unidade',
    'unidade_comercial': 'unidade',
    'qtd_unidade': 'qtd_unidade',
    'fabricante': 'fabricante',
    'referencia': 'referencia',
    'registro_anvisa': 'anvisa',
    'anvisa': 'anvisa',
    'validade_anvisa': 'validade_anvisa',
    'ean': 'ean',
    'situacao': 'situacao',
    'fracionavel': 'fracionavel',
    'versao': 'versao',
    'tuss_prefix': 'tuss_prefix',
    'tuss_numero': 'tuss_numero',
    'classificacao': 'status_final',
    'classificacao_produto': 'status_final',
    'status_final': 'status_final',
}

_SIMPRO_ALLOWED_COLUMNS: set[str] = {column.name for column in SimproItemNormalized.__table__.columns}
_TERNARY_CONCAT_RE = re.compile(
    r'^\s*([A-Za-z_][A-Za-z0-9_]*)\s*&&\s*([A-Za-z_][A-Za-z0-9_]*)\s*\?\s*([A-Za-z_][A-Za-z0-9_]*)\s*\+\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(null|[A-Za-z_][A-Za-z0-9_]*)\s*$',
    re.IGNORECASE,
)
_TUSS_INLINE_RE = re.compile(r'(?:[#\-\+\s]?)([NS][A-Z])\s*([0-9]{6,12})', re.IGNORECASE)
_TUSS_SUFFIX_MARKER_RE = re.compile(r'([#\-\+\s]?)([NS][A-Z])\s*$', re.IGNORECASE)
_ANVISA_DIGITS_RE = re.compile(r'(\d{13,})')
_ANVISA_INLINE_RE = re.compile(r'[A-Z]{3,}\s{2,}(\d{13,})')


def _extract_tuss_parts(text: str | None) -> tuple[str, str, int, int] | None:
    if not text:
        return None
    for match in _TUSS_INLINE_RE.finditer(str(text)):
        prefix = match.group(1).upper()
        if prefix[0] not in {'N', 'S'}:
            continue
        numero = match.group(2)
        return prefix, numero, match.start(1), match.end(2)
    return None


def _resolve_simpro_field_map(map_config: dict | None) -> dict[str, str]:
    field_map = dict(_SIMPRO_FIELD_MAP_DEFAULT)
    overrides = (map_config or {}).get('field_map')
    if isinstance(overrides, dict):
        for source, target in overrides.items():
            if not isinstance(source, str) or not isinstance(target, str):
                continue
            source_key = source.strip()
            target_key = target.strip()
            if not source_key or not target_key:
                continue
            if target_key not in _SIMPRO_ALLOWED_COLUMNS:
                continue
            field_map[source_key] = target_key
    return field_map


def _evaluate_postprocess_expr(expr: str, record: dict[str, object | None]) -> object | None:
    expr = (expr or '').strip()
    if not expr or expr.lower() == 'null':
        return None
    match = _TERNARY_CONCAT_RE.match(expr)
    if match:
        left_key, right_key, true_left_key, true_right_key, false_branch = match.groups()
        left_val = record.get(left_key)
        right_val = record.get(right_key)
        if left_val and right_val:
            left_true = record.get(true_left_key)
            right_true = record.get(true_right_key)
            left_text = (str(left_true).strip() if left_true is not None else '')
            right_text = (str(right_true).strip() if right_true is not None else '')
            combined = left_text + right_text
            return combined or None
        if false_branch.lower() == 'null':
            return None
        return record.get(false_branch)
    return record.get(expr)


def _apply_simpro_postprocess(record: dict[str, object | None], postprocess_cfg: dict | None) -> None:
    if not isinstance(postprocess_cfg, dict):
        return

    extracts = postprocess_cfg.get('extract') or []
    for spec in extracts:
        if not isinstance(spec, dict):
            continue
        source_field = spec.get('from')
        regex_pattern = spec.get('regex')
        fields_map = spec.get('fields') or {}
        if not source_field or not regex_pattern or not isinstance(fields_map, dict):
            continue
        value = record.get(source_field)
        if not isinstance(value, str):
            for field_name in fields_map:
                record.setdefault(field_name, None)
            continue
        try:
            pattern = re.compile(regex_pattern)
        except re.error:
            continue
        match = pattern.search(value)
        for field_name, group_index in fields_map.items():
            record.setdefault(field_name, None)
            try:
                idx = int(group_index)
            except (TypeError, ValueError):
                continue
            if not match:
                continue
            try:
                captured = match.group(idx)
            except IndexError:
                captured = None
            if isinstance(captured, str):
                captured = captured.strip() or None
            record[field_name] = captured

    derives = postprocess_cfg.get('derive') or []
    for spec in derives:
        if not isinstance(spec, dict):
            continue
        target_field = spec.get('name')
        expr = spec.get('expr')
        if not target_field or not isinstance(expr, str):
            continue
        record[target_field] = _evaluate_postprocess_expr(expr, record)

    cleanup_fields = postprocess_cfg.get('cleanup') or []
    for field_name in cleanup_fields:
        if isinstance(field_name, str):
            record.pop(field_name, None)


def _enrich_tuss_from_ean(record: dict[str, object | None]) -> None:
    raw = record.get('ean')
    if raw is None:
        return
    text = str(raw).strip()
    if not text:
        record['ean'] = None
        return

    extracted = _extract_tuss_parts(text)
    if not extracted:
        marker_match = _TUSS_SUFFIX_MARKER_RE.search(text)
        if marker_match:
            base = text[:marker_match.start(1)].rstrip('#-+ ').strip()
            record['ean'] = base or None
            if not record.get('tuss_prefix'):
                record['tuss_prefix'] = marker_match.group(2).upper()
            return
        record['ean'] = text
        return

    prefix, numero, prefix_start, _ = extracted
    base = text[:prefix_start].rstrip('#-+ ').strip()
    if not base:
        for sep in ('#', '+'):
            if sep in text:
                base = text.split(sep, 1)[0].strip()
                if base:
                    break

    record['ean'] = base or None
    record['tuss_prefix'] = prefix
    record['tuss_numero'] = numero


def _ensure_tuss_from_line(record: dict[str, object | None], line: str) -> None:
    if record.get('tuss_numero'):
        return
    extracted = _extract_tuss_parts(line)
    if not extracted:
        return

    prefix, numero, _, _ = extracted
    if not record.get('tuss_prefix'):
        record['tuss_prefix'] = prefix
    record['tuss_numero'] = numero


def _ensure_tuss_field(record: dict[str, object | None]) -> None:
    if record.get('tuss'):
        return
    prefix_raw = record.get('tuss_prefix')
    numero_raw = record.get('tuss_numero')
    if not prefix_raw or not numero_raw:
        return
    prefix = str(prefix_raw).strip().upper()
    numero = ''.join(ch for ch in str(numero_raw).strip() if ch.isdigit())
    if prefix and numero:
        record['tuss_prefix'] = prefix
        record['tuss_numero'] = numero
        record['tuss'] = f'{prefix}{numero}'


def _normalize_anvisa_field(record: dict[str, object | None]) -> None:
    source_key = 'anvisa'
    raw = record.get(source_key)
    if raw is None:
        source_key = 'registro_anvisa'
        raw = record.get(source_key)
        if raw is None:
            return

    text = str(raw).strip()
    if not text:
        record[source_key] = None
        record['anvisa'] = None
        return

    stripped = text.lstrip()
    if stripped and stripped[0].isalpha():
        record[source_key] = text
        record['anvisa'] = text
        return

    matches = _ANVISA_DIGITS_RE.findall(text)
    if matches:
        candidate = max(matches, key=len)
        normalized = (candidate[-13:] if len(candidate) > 13 else candidate) or None
        record[source_key] = normalized
        record['anvisa'] = normalized
        return

    digit_tokens = [token for token in text.split() if token.isdigit() and len(token) >= 13]
    if digit_tokens:
        candidate = max(digit_tokens, key=len)
        normalized = candidate[-13:] if len(candidate) > 13 else candidate
        record[source_key] = normalized
        record['anvisa'] = normalized
        return

    if ' ' not in text and not re.search(r'[A-Za-z]', text):
        compact = ''.join(ch for ch in text if ch.isdigit())
        if compact:
            normalized = compact[-13:] if len(compact) > 13 else compact
            record[source_key] = normalized
            record['anvisa'] = normalized
            return

    compact = ''.join(ch for ch in text if ch.isdigit())
    if len(compact) >= 13:
        normalized = compact[-13:]
        record[source_key] = normalized
        record['anvisa'] = normalized
        return

    record[source_key] = text
    record['anvisa'] = text


def _ensure_anvisa_from_line(record: dict[str, object | None], line: str) -> None:
    current_raw = record.get('anvisa') or record.get('registro_anvisa')
    current_text = str(current_raw).strip() if current_raw is not None else ''
    if current_text and any(ch.isalpha() for ch in current_text):
        return

    current_digits = ''.join(ch for ch in current_text if ch.isdigit()) if current_text else ''
    if 11 <= len(current_digits) <= 13:
        # Já temos um registro plausível (11-13 dígitos); evita sobrescrever
        # com concatenações longas encontradas no meio da linha fixa.
        return

    preferred: str | None = None
    inline_match = _ANVISA_INLINE_RE.search(line)
    if inline_match:
        digits = ''.join(ch for ch in inline_match.group(1) if ch.isdigit())
        if 13 <= len(digits) <= 40:
            preferred = digits

    candidates: list[tuple[int, str]] = []
    if preferred is None:
        for match in _ANVISA_DIGITS_RE.finditer(line):
            digits = ''.join(ch for ch in match.group(0) if ch.isdigit())
            length = len(digits)
            if length < 13 or length > 40:
                continue
            following = line[match.end():match.end() + 1]
            if following in {'-', '#', '+'}:
                continue
            candidates.append((match.start(), digits))

    if preferred is None and candidates:
        if current_digits:
            for _, digits in candidates:
                if len(digits) > len(current_digits) and digits.endswith(current_digits):
                    preferred = digits
                    break
        if preferred is None:
            candidates.sort(key=lambda item: item[0])
            preferred = candidates[-1][1]

    if preferred:
        normalized = preferred[-13:] if len(preferred) > 13 else preferred
        if not normalized:
            return
        if normalized != current_digits or not record.get('anvisa'):
            record['registro_anvisa'] = normalized
            record['anvisa'] = normalized
def _format_tuss_display(value: str | None, numero: str | None = None) -> str | None:
    numero_text = ''.join(ch for ch in str(numero).strip() if ch.isdigit()) if numero is not None else ''
    if numero_text:
        return numero_text
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if '#' in text:
        candidate = text.split('#')[-1]
        digits = ''.join(ch for ch in candidate if ch.isdigit())
        if digits:
            return digits
    digits = ''.join(ch for ch in text if ch.isdigit())
    return digits or text


def _build_simpro_payload(record: dict[str, object | None], field_map: dict[str, str]) -> dict[str, object | None]:
    payload: dict[str, object | None] = {}
    for source, target in field_map.items():
        if target not in _SIMPRO_ALLOWED_COLUMNS:
            continue
        if source not in record:
            continue
        value = record[source]
        if target == 'tuss_numero':
            value = _format_tuss_display(None, value)
        elif target == 'unidade' and isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                value = cleaned.split()[0]
            else:
                value = None
        elif isinstance(value, str):
            value = value.strip() or None
        payload[target] = value

    if payload.get('codigo') in (None, '') and record.get('codigo_interno'):
        payload['codigo'] = record.get('codigo_interno')
    return payload


class InsumoContextoClinico(db.Model):
    __tablename__ = 'insumo_contexto_clinico'

    id = db.Column(db.BigInteger, primary_key=True)
    origem = db.Column(db.Enum('BRAS', 'SIMPRO', name='insumo_origem'), nullable=False)
    item_id = db.Column(db.BigInteger, nullable=False)
    drg = db.Column(db.String(50), nullable=True)
    procedimento_codigo = db.Column(db.String(50), nullable=True)
    procedimento_descricao = db.Column(db.String(255), nullable=True)
    frequencia_relativa = db.Column(db.Numeric(8, 6), nullable=True)
    custo_procedimento = db.Column(db.Numeric(14, 2), nullable=True)
    substitutos_raw = db.Column(db.Text, nullable=True)
    narrativa = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, server_default=text('CURRENT_TIMESTAMP'))
    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
        server_onupdate=text('CURRENT_TIMESTAMP'),
    )

    __table_args__ = (
        db.Index('idx_icc_origem_item', 'origem', 'item_id'),
    )


class CatalogoBrasindice(db.Model):
    __tablename__ = 'mv_catalogo_vigente_brasindice'
    __table_args__ = {'extend_existing': True}

    uf = db.Column(db.String(2), primary_key=True)
    aliquota_bp = db.Column(db.Integer, nullable=False)
    periodo = db.Column(db.String(6), nullable=True)
    sequencia = db.Column(db.SmallInteger, nullable=True)
    etag_versao = db.Column(db.String(128), nullable=True)
    item_id = db.Column(db.BigInteger, primary_key=True)
    produto_codigo = db.Column(db.String(50), nullable=True)
    apresentacao_codigo = db.Column(db.String(50), nullable=True)
    produto_nome = db.Column(db.String(255), nullable=True)
    apresentacao_descricao = db.Column(db.String(255), nullable=True)
    ean = db.Column(db.String(20), nullable=True)
    registro_anvisa = db.Column(db.String(50), nullable=True)
    preco_pmc_unit = db.Column(db.Numeric(15, 4))
    preco_pfb_unit = db.Column(db.Numeric(15, 4))
    preco_pmc_pacote = db.Column(db.Numeric(15, 4))
    preco_pfb_pacote = db.Column(db.Numeric(15, 4))
    laboratorio_nome = db.Column(db.String(255), nullable=True)
    edicao = db.Column(db.String(50), nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True)
    etag_catalogo = db.Column(db.String(255), nullable=True)


class CatalogoSimpro(db.Model):
    __tablename__ = 'mv_catalogo_vigente_simpro'
    __table_args__ = {'extend_existing': True}

    uf = db.Column(db.String(2), primary_key=True)
    aliquota_bp = db.Column(db.Integer, nullable=False)
    periodo = db.Column(db.String(6), nullable=True)
    sequencia = db.Column(db.SmallInteger, nullable=True)
    etag_versao = db.Column(db.String(128), nullable=True)
    item_id = db.Column(db.BigInteger, primary_key=True)
    codigo_interno = db.Column(db.String(20), nullable=True)
    codigo = db.Column(db.String(20), nullable=True)
    codigo_alt = db.Column(db.String(20), nullable=True)
    tuss_numero = db.Column(db.String(16), nullable=True)
    descricao = db.Column(db.String(255), nullable=True)
    data_ref = db.Column(db.Date, nullable=True)
    preco1 = db.Column(db.Numeric(15, 4))
    preco2 = db.Column(db.Numeric(15, 4))
    preco3 = db.Column(db.Numeric(15, 4))
    preco4 = db.Column(db.Numeric(15, 4))
    qtd_unidade = db.Column(db.Integer, nullable=True)
    fabricante = db.Column(db.String(80), nullable=True)
    referencia = db.Column(db.String(120), nullable=True)
    anvisa = db.Column(db.String(20), nullable=True)
    validade_anvisa = db.Column(db.Date, nullable=True)
    ean = db.Column(db.String(32), nullable=True)
    situacao = db.Column(db.String(40), nullable=True)
    fracionavel = db.Column(db.String(1), nullable=True)
    status_final = db.Column(db.String(8), nullable=True)
    imported_at = db.Column(db.DateTime, nullable=True)
    etag_catalogo = db.Column(db.String(255), nullable=True)


class InsumoIndex(db.Model):
    __tablename__ = 'insumos_index'

    origem = db.Column(db.Enum('BRAS', 'SIMPRO', name='insumo_origem'), primary_key=True)
    item_id = db.Column(db.Integer, primary_key=True)
    tuss = db.Column(db.String(50), index=True, nullable=True)
    tiss = db.Column(db.String(50), index=True, nullable=True)
    descricao = db.Column(db.String(500), index=True, nullable=True)
    preco = db.Column(db.Numeric(12, 4), nullable=True)
    aliquota = db.Column(
        db.Numeric(12, 4),
        primary_key=True,
        nullable=False,
        server_default=text('0'),
    )
    fabricante = db.Column(db.String(255), nullable=True)
    anvisa = db.Column(db.String(50), index=True, nullable=True)
    versao_tabela = db.Column(db.String(100), nullable=True)
    data_atualizacao = db.Column(db.Date, nullable=True)
    uf_referencia = db.Column(db.String(64), nullable=True)
    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        server_default=text('CURRENT_TIMESTAMP'),
        server_onupdate=text('CURRENT_TIMESTAMP'),
    )


BRAS_DEFAULT_COLUMNS = ['tuss', 'tiss', 'anvisa', 'descricao', 'preco', 'fabricante', 'aliquota']
SIMPRO_DEFAULT_COLUMNS = [
    'codigo_interno', 'codigo', 'codigo_alt', 'tuss_numero', 'descricao', 'data_ref', 'tipo_reg',
    'preco1', 'preco2', 'preco3', 'preco4', 'unidade', 'qtd_unidade',
    'fabricante', 'referencia', 'anvisa', 'validade_anvisa', 'ean', 'situacao', 'status_final'
]
DECIMAL_FIELDS = {'preco', 'aliquota'}
DATE_FIELDS = {'data_atualizacao'}
DEFAULT_IMPORT_ENCODINGS = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
TETO_PREVIEW_DIR = Path(tempfile.gettempdir()) / 'cbhpm_teto_previews'
INSUMO_IMPORT_ASYNC_DIR = Path(tempfile.gettempdir()) / 'insumo_async_imports'
INSUMO_IMPORT_ASYNC_DIR.mkdir(parents=True, exist_ok=True)

REEMBOLSO_STORAGE_DIR = Path(os.getenv('REEMBOLSO_STORAGE_DIR', Path(__file__).parent / 'data' / 'reembolsos'))
REEMBOLSO_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
REEMBOLSO_MAX_FILE_MB = _safe_int_env('REEMBOLSO_MAX_FILE_MB', 12)
REEMBOLSO_OCR_MAX_PAGES = _safe_int_env('REEMBOLSO_OCR_MAX_PAGES', 2)
REEMBOLSO_OCR_DPI = _safe_int_env('REEMBOLSO_OCR_DPI', 220)
REEMBOLSO_PDF_TEXT_MIN_CHARS = _safe_int_env('REEMBOLSO_PDF_TEXT_MIN_CHARS', 120)
REEMBOLSO_TESSERACT_LANG = (os.getenv('REEMBOLSO_TESSERACT_LANG') or 'por+eng').strip()
REEMBOLSO_ALLOWED_EXTENSIONS = {'.pdf', '.png', '.jpg', '.jpeg'}

REEMBOLSO_FIELDS_BY_TYPE: dict[str, list[tuple[str, str]]] = {
    'NOTA_FISCAL': [
        ('cpf_cnpj', 'CPF/CNPJ'),
        ('razao_social', 'Razão social'),
        ('valor', 'Valor'),
        ('data', 'Data'),
        ('numero_documento', 'Número'),
        ('serie', 'Série'),
        ('chave_acesso', 'Chave de acesso'),
        ('descricao', 'Descrição'),
    ],
    'RECIBO': [
        ('cpf_cnpj', 'CPF/CNPJ'),
        ('nome_razao', 'Nome/Razão social'),
        ('valor', 'Valor'),
        ('data', 'Data'),
        ('descricao', 'Descrição'),
    ],
    'COMPROVANTE': [
        ('cpf_cnpj', 'CPF/CNPJ'),
        ('pagador', 'Pagador'),
        ('recebedor', 'Recebedor'),
        ('valor', 'Valor'),
        ('data', 'Data'),
        ('meio_pagamento', 'Meio de pagamento'),
        ('status_pagamento', 'Status'),
    ],
    'DESCONHECIDO': [
        ('cpf_cnpj', 'CPF/CNPJ'),
        ('valor', 'Valor'),
        ('data', 'Data'),
        ('descricao', 'Descrição'),
    ],
}


def _reembolso_allowed_extension(filename: str) -> bool:
    if not filename:
        return False
    return Path(filename).suffix.lower() in REEMBOLSO_ALLOWED_EXTENSIONS


def _reembolso_fields_for_type(tipo_documento: str | None) -> list[tuple[str, str]]:
    tipo = (tipo_documento or 'DESCONHECIDO').upper()
    return REEMBOLSO_FIELDS_BY_TYPE.get(tipo, REEMBOLSO_FIELDS_BY_TYPE['DESCONHECIDO'])


def _reembolso_infer_tipo(texto: str) -> str:
    if not texto:
        return 'DESCONHECIDO'
    raw = texto.lower()
    if 'nota fiscal' in raw or 'nfe' in raw or 'nf-e' in raw:
        return 'NOTA_FISCAL'
    if 'recibo' in raw:
        return 'RECIBO'
    if 'comprovante' in raw or 'pagamento' in raw or 'transa' in raw or 'pix' in raw:
        return 'COMPROVANTE'
    return 'DESCONHECIDO'


def _reembolso_extract_pdf_text(path: Path, max_pages: int) -> str:
    if PdfReader is None:
        return ''
    try:
        reader = PdfReader(str(path))
    except Exception:
        return ''
    chunks: list[str] = []
    for idx, page in enumerate(reader.pages):
        if idx >= max_pages:
            break
        try:
            text = page.extract_text() or ''
        except Exception:
            text = ''
        if text:
            chunks.append(text)
    return '\n'.join(chunks).strip()


def _reembolso_is_pdf_native(texto: str) -> bool:
    if not texto:
        return False
    compact = re.sub(r'\s+', '', texto)
    return len(compact) >= REEMBOLSO_PDF_TEXT_MIN_CHARS


def _reembolso_ocr_image(image: "PILImage") -> str:
    if pytesseract is None or PILImage is None or ImageOps is None:
        return ''
    prepared = ImageOps.grayscale(image)
    prepared = ImageOps.autocontrast(prepared)
    return pytesseract.image_to_string(prepared, lang=REEMBOLSO_TESSERACT_LANG) or ''


def _reembolso_extract_text(path: Path) -> tuple[str, bool, str | None, str | None]:
    suffix = path.suffix.lower()
    if suffix == '.pdf':
        text = _reembolso_extract_pdf_text(path, REEMBOLSO_OCR_MAX_PAGES)
        if _reembolso_is_pdf_native(text):
            return text, True, 'PDF_NATIVE', None
        if convert_from_path is None or pytesseract is None or PILImage is None:
            return text, False, 'OCR_NAO_DISPONIVEL', 'OCR indisponível (dependências ausentes).'
        try:
            images = convert_from_path(
                str(path),
                first_page=1,
                last_page=REEMBOLSO_OCR_MAX_PAGES,
                dpi=REEMBOLSO_OCR_DPI,
            )
        except Exception as exc:  # noqa: BLE001
            return text, False, 'OCR_ERRO', f'Falha ao converter PDF: {exc}'
        ocr_chunks = []
        for image in images:
            ocr_chunks.append(_reembolso_ocr_image(image))
        ocr_text = '\n'.join(chunk for chunk in ocr_chunks if chunk).strip()
        return ocr_text, False, 'OCR_OK', None
    if pytesseract is None or PILImage is None:
        return '', False, 'OCR_NAO_DISPONIVEL', 'OCR indisponível (dependências ausentes).'
    try:
        with PILImage.open(path) as img:
            text = _reembolso_ocr_image(img)
    except Exception as exc:  # noqa: BLE001
        return '', False, 'OCR_ERRO', f'Falha ao ler imagem: {exc}'
    return text.strip(), False, 'OCR_OK', None


def _reembolso_extract_fields(texto: str) -> dict[str, str]:
    payload: dict[str, str] = {}
    if not texto:
        return payload
    raw = texto
    cpf_match = re.search(r'\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b', raw)
    cnpj_match = re.search(r'\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b', raw)
    if cnpj_match:
        payload['cpf_cnpj'] = cnpj_match.group(0)
    elif cpf_match:
        payload['cpf_cnpj'] = cpf_match.group(0)

    date_match = re.search(r'\b\d{2}[/-]\d{2}[/-]\d{4}\b', raw)
    if date_match:
        payload['data'] = date_match.group(0)

    value_match = re.search(r'(R\$)?\s*\d{1,3}(?:\.\d{3})*,\d{2}', raw)
    if value_match:
        payload['valor'] = value_match.group(0).strip()

    access_match = re.search(r'\b\d{44}\b', raw)
    if access_match:
        payload['chave_acesso'] = access_match.group(0)

    numero_match = re.search(r'(?i)\b(n[ºo°]|numero)\s*[:#]?\s*(\d{3,})', raw)
    if numero_match:
        payload['numero_documento'] = numero_match.group(2)

    serie_match = re.search(r'(?i)\bs[eé]rie\s*[:#]?\s*([A-Z0-9]+)', raw)
    if serie_match:
        payload['serie'] = serie_match.group(1)

    if 'pix' in raw.lower():
        payload['meio_pagamento'] = 'PIX'
    elif 'boleto' in raw.lower():
        payload['meio_pagamento'] = 'Boleto'
    elif 'cart' in raw.lower():
        payload['meio_pagamento'] = 'Cartão'
    elif 'transfer' in raw.lower() or 'ted' in raw.lower():
        payload['meio_pagamento'] = 'Transferência'

    return payload


def _reembolso_extract_tuss_codes(texto: str) -> list[str]:
    if not texto:
        return []
    candidates: list[str] = []
    seen: set[str] = set()

    for match in re.finditer(r'(?i)\bTUSS\s*[:#-]?\s*([0-9]{6,12})\b', texto):
        code = match.group(1)
        if code and code not in seen:
            seen.add(code)
            candidates.append(code)

    for line in texto.splitlines():
        if 'tuss' not in line.lower():
            continue
        for match in re.finditer(r'\b([0-9]{6,12})\b', line):
            code = match.group(1)
            if code and code not in seen:
                seen.add(code)
                candidates.append(code)

    if not candidates:
        for match in re.finditer(r'\b([0-9]{8,12})\b', texto):
            code = match.group(1)
            if code and code not in seen:
                seen.add(code)
                candidates.append(code)
            if len(candidates) >= 10:
                break

    return candidates


def _reembolso_lookup_tuss_values(codigos: list[str], operadora_id: int | None) -> dict[str, list[dict[str, object]]]:
    if not codigos:
        return {}
    result: dict[str, list[dict[str, object]]] = {code: [] for code in codigos}

    query = (
        db.session.query(
            Procedimento.codigo,
            Procedimento.descricao,
            Procedimento.valor,
            Procedimento.prestador,
            Tabela.nome,
            Tabela.tipo_tabela,
            Tabela.uf,
            Tabela.data_vigencia,
        )
        .join(Tabela, Procedimento.id_tabela == Tabela.id)
        .filter(Procedimento.codigo.in_(codigos))
    )
    if operadora_id:
        query = query.filter(Procedimento.operadora_id == operadora_id)
    query = query.order_by(
        (Tabela.data_vigencia.is_(None)).asc(),
        Tabela.data_vigencia.desc(),
        Tabela.nome.asc(),
    )

    for row in query.all():
        bucket = result.get(row.codigo)
        if bucket is None or len(bucket) >= 6:
            continue
        bucket.append({
            'codigo': row.codigo,
            'descricao': row.descricao,
            'valor': _stringify_for_output(row.valor),
            'tabela': row.nome,
            'tipo_tabela': row.tipo_tabela,
            'prestador': row.prestador,
            'uf': row.uf,
            'vigencia': row.data_vigencia.isoformat() if row.data_vigencia else None,
        })

    missing = [code for code, items in result.items() if not items]
    if missing:
        cbhpm_query = (
            db.session.query(
                CBHPMItem.codigo,
                CBHPMItem.procedimento,
                CBHPMItem.subtotal,
                CBHPMItem.total_porte,
                CBHPMItem.valor_porte,
                CBHPMItem.total_uco,
                CBHPMItem.uco,
                CBHPMItem.total_filme,
                CBHPMItem.filme,
                Tabela.nome,
                Tabela.uf,
                Tabela.data_vigencia,
            )
            .join(Tabela, CBHPMItem.id_tabela == Tabela.id)
            .filter(Tabela.tipo_tabela == 'cbhpm', CBHPMItem.codigo.in_(missing))
        )
        if operadora_id:
            cbhpm_query = cbhpm_query.filter(Tabela.id_operadora == operadora_id)
        cbhpm_query = cbhpm_query.order_by(
            (Tabela.data_vigencia.is_(None)).asc(),
            Tabela.data_vigencia.desc(),
            Tabela.nome.asc(),
        )

        for row in cbhpm_query.all():
            bucket = result.get(row.codigo)
            if bucket is None or len(bucket) >= 6:
                continue
            valor = row.subtotal or row.total_porte or row.valor_porte or row.total_uco or row.uco or row.total_filme or row.filme
            bucket.append({
                'codigo': row.codigo,
                'descricao': row.procedimento,
                'valor': _stringify_for_output(valor),
                'tabela': row.nome,
                'tipo_tabela': 'cbhpm',
                'prestador': None,
                'uf': row.uf,
                'vigencia': row.data_vigencia.isoformat() if row.data_vigencia else None,
            })

    return result

_TETO_IMPORT_JOBS_LOCK = threading.Lock()
_TETO_IMPORT_JOBS: dict[str, dict] = {}

BRAS_RAW_DEFAULT_COLUMNS = [
    'col01', 'col02', 'col03', 'col04', 'col05', 'col06', 'col07', 'col08', 'col09', 'col10',
    'col11', 'col12', 'col13', 'col14', 'col15', 'col16', 'col17', 'col18', 'col19', 'col20',
    'col21', 'col22', 'col23'
]


def _clean_decimal_expression(column: str) -> str:
    sanitized = f"REPLACE(REPLACE(REPLACE({column}, '.', ''), ' ', ''), ',', '.')"
    integer_part = f"SUBSTRING_INDEX({sanitized}, '.', 1)"
    scale_expr = f"GREATEST(CHAR_LENGTH({integer_part}) - 8, 2)"
    return (
        "CAST(\n"
        "    CASE\n"
        f"        WHEN {column} IS NULL THEN NULL\n"
        f"        WHEN {sanitized} = '' THEN NULL\n"
        f"        WHEN {sanitized} NOT REGEXP '^[0-9]+(\\.[0-9]+)?$' THEN NULL\n"
        f"        WHEN CHAR_LENGTH({sanitized}) > 32 THEN NULL\n"
        f"        ELSE (CAST({sanitized} AS DECIMAL(38,6)) / POW(10, {scale_expr}))\n"
        "    END AS DECIMAL(15,4)\n"
        ")"
    )


def _build_bras_item_view_sql() -> str:
    preco_pmc = _clean_decimal_expression('r.col07')
    preco_pfb = _clean_decimal_expression('r.col13')

    return (
        "CREATE OR REPLACE VIEW bras_item_v AS\n"
        "SELECT\n"
        "    r.id,\n"
        "    r.arquivo,\n"
        "    r.linha_num,\n"
        "    r.col01 AS laboratorio_codigo,\n"
        "    r.col02 AS laboratorio_nome,\n"
        "    r.col20 AS produto_codigo,\n"
        "    r.col04 AS produto_nome,\n"
        "    r.col18 AS apresentacao_codigo,\n"
        "    r.col06 AS apresentacao_descricao,\n"
        "    r.col17 AS ean,\n"
        "    r.col22 AS registro_anvisa,\n"
        "    r.col14 AS edicao,\n"
        f"    {preco_pmc} AS preco_pmc_unit,\n"
        f"    {preco_pfb} AS preco_pfb_unit,\n"
        f"    {preco_pmc} AS preco_pmc_pacote,\n"
        f"    {preco_pfb} AS preco_pfb_pacote,\n"
        "    NULL AS aliquota_ou_ipi,\n"
        "    NULL AS quantidade_embalagem,\n"
        "    r.imported_at\n"
        "FROM bras_raw r\n"
    )


BRAS_ITEM_VIEW_SQL = _build_bras_item_view_sql()


def _normalize_column_token(name: str | None) -> str:
    if name is None:
        return ''
    token = str(name).strip().strip('"').strip("'")
    if not token:
        return ''
    normalized = unicodedata.normalize('NFKD', token)
    normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.replace('-', '_').replace(' ', '_')
    normalized = re.sub(r'[^0-9a-zA-Z_]', '', normalized)
    return normalized.lower()


def _columns_valid_for_model(model_cls, columns: list[str]) -> bool:
    if not columns:
        return False
    valid = {col.name for col in model_cls.__table__.columns}
    for col in columns:
        if not col or col not in valid:
            return False
    return True


def _build_encoding_list(primary: str | None) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()

    def _add(candidate: str | None) -> None:
        if not candidate:
            return
        normalized = candidate.strip()
        if not normalized:
            return
        key = normalized.lower()
        if key in seen:
            return
        seen.add(key)
        ordered.append(normalized)

    _add(primary)
    for fallback in DEFAULT_IMPORT_ENCODINGS:
        _add(fallback)
    return ordered


MYSQL_CHARSET_MAP = {
    'utf-8-sig': 'utf8mb4',
    'utf8-sig': 'utf8mb4',
    'utf-8': 'utf8mb4',
    'utf8': 'utf8mb4',
    'utf8mb4': 'utf8mb4',
    'latin-1': 'latin1',
    'latin1': 'latin1',
    'iso-8859-1': 'latin1',
    'cp1252': 'cp1252',
}


def _encoding_to_mysql_charset(encoding: str | None) -> str:
    if not encoding:
        return 'utf8mb4'
    return MYSQL_CHARSET_MAP.get(encoding.lower(), 'utf8mb4')


def _sql_escape_literal(value: str) -> str:
    escaped = value.replace('\\', r'\\').replace("'", r"\'")
    return f"'{escaped}'"


def _encode_line_terminator(value: str | None) -> str:
    if not value:
        return '\n'
    return value.encode('unicode_escape').decode('ascii')


def _delete_in_batches(sql: str, params: dict, batch_size: int = 2000, max_retries: int = 5) -> int:
    """
    Executa DELETE em batches para evitar lock timeout.
    Adiciona LIMIT ao SQL e repete até não haver mais registros.
    """
    import time
    
    total_deleted = 0
    if 'LIMIT' not in sql.upper():
        sql_with_limit = f"{sql} LIMIT {batch_size}"
    else:
        sql_with_limit = sql
    
    while True:
        deleted_this_batch = 0
        for attempt in range(max_retries):
            try:
                result = db.session.execute(text(sql_with_limit), params)
                db.session.commit()
                deleted_this_batch = result.rowcount or 0
                break
            except Exception as exc:
                db.session.rollback()
                error_str = str(exc)
                if ('1205' in error_str or '1213' in error_str) and attempt < max_retries - 1:
                    time.sleep((attempt + 1) * 2)
                    continue
                raise
        
        total_deleted += deleted_this_batch
        
        if deleted_this_batch < batch_size:
            break
        
        time.sleep(0.1)
    
    return total_deleted


def _delete_with_retry(sql: str, params: dict, max_retries: int = 5) -> None:
    """Executa DELETE em batches para evitar lock timeout."""
    _delete_in_batches(sql, params, batch_size=2000, max_retries=max_retries)


def _delete_insumos_by_arquivo(origem: str, item_table: str, arquivo_label: str, batch_size: int = 2000) -> int:
    """
    Deleta insumos_index baseado em item_id de outra tabela, em batches.
    Evita subquery com LIMIT que o MySQL não suporta.
    """
    import time
    
    total_deleted = 0
    last_id = 0

    while True:
        # Primeiro, busca os IDs a deletar usando cursor por ID para evitar repetir lote.
        ids_result = db.session.execute(
            text(
                f"""
                SELECT id
                FROM {item_table}
                WHERE arquivo = :arquivo
                  AND id > :last_id
                ORDER BY id ASC
                LIMIT :batch_size
                """
            ),
            {'arquivo': arquivo_label, 'last_id': last_id, 'batch_size': batch_size}
        ).fetchall()
        
        if not ids_result:
            break
        
        ids = [row[0] for row in ids_result]
        last_id = ids[-1]
        
        # Deleta do insumos_index usando os IDs
        if ids:
            placeholders = ','.join([':id' + str(i) for i in range(len(ids))])
            params = {f'id{i}': id_val for i, id_val in enumerate(ids)}
            params['origem'] = origem
            
            try:
                result = db.session.execute(
                    text(f"DELETE FROM insumos_index WHERE origem = :origem AND item_id IN ({placeholders})"),
                    params
                )
                db.session.commit()
                total_deleted += result.rowcount or 0
            except Exception as exc:
                db.session.rollback()
                app.logger.warning('Delete insumos batch error: %s', str(exc)[:100])
                time.sleep(1)
                continue
        
        # Pequena pausa entre batches
        time.sleep(0.05)
        
        # Se buscou menos que o batch, chegou ao fim da paginação.
        if len(ids) < batch_size:
            break
    
    return total_deleted


def _delete_existing_bras_records(
    arquivo_label: str | None,
    truncate: bool,
    max_retries: int = 10,
    *,
    aliquota_filter: Decimal | None = None,
    uf_filter: str | None = None,
) -> None:
    """
    Limpa registros da Brasíndice.
    
    Modos de operação:
    - truncate=True, aliquota_filter=None: Limpa TUDO da Brasíndice
    - truncate=True, aliquota_filter=X: Limpa índice da alíquota X + dados do arquivo atual
    - truncate=False, arquivo_label=X: Limpa apenas registros do arquivo X
    """
    
    # Modo 1: Limpar por alíquota específica no índice + arquivo atual nas tabelas raw
    if truncate and aliquota_filter is not None:
        app.logger.info('Limpando Brasíndice: alíquota %.2f / UF %s / arquivo %s', 
                       aliquota_filter, uf_filter or 'todas', arquivo_label or 'N/A')
        
        # Construir filtro de UF
        uf_condition = ""
        params: dict = {'aliquota': float(aliquota_filter)}
        if uf_filter:
            uf_condition = " AND (uf_referencia LIKE :uf_pattern OR uf_referencia = :uf_exact)"
            params['uf_pattern'] = f'%{uf_filter}%'
            params['uf_exact'] = uf_filter
        
        # Deletar do índice de insumos pela alíquota (com retry)
        _delete_with_retry(
            f"DELETE FROM insumos_index WHERE origem = 'BRAS' AND aliquota = :aliquota{uf_condition}",
            params,
        )
        
        # TAMBÉM deletar dados do arquivo específico das tabelas raw/normalized
        if arquivo_label:
            arquivo_params = {'arquivo': arquivo_label}
            _delete_with_retry('DELETE FROM bras_item_n WHERE arquivo = :arquivo', arquivo_params)
            _delete_with_retry('DELETE FROM bras_raw WHERE arquivo = :arquivo', arquivo_params)
            _delete_with_retry('DELETE FROM bras_fixed_stage WHERE arquivo = :arquivo', arquivo_params)
        
        return
    
    # Modo 2: Limpar TUDO (truncate sem filtro)
    if truncate:
        _delete_with_retry("DELETE FROM insumos_index WHERE origem = 'BRAS'", {})
        # TRUNCATE é mais rápido e não precisa de lock por muito tempo
        db.session.execute(text('TRUNCATE TABLE bras_item_n'))
        db.session.execute(text('TRUNCATE TABLE bras_raw'))
        db.session.execute(text('TRUNCATE TABLE bras_fixed_stage'))
        try:
            db.session.execute(text('TRUNCATE TABLE bras_catalog_snapshot'))
        except Exception as exc:  # noqa: BLE001
            app.logger.warning('Falha ao TRUNCATE bras_catalog_snapshot: %s', exc)
        try:
            db.session.execute(text('TRUNCATE TABLE bras_item_preco'))
        except Exception as exc:  # noqa: BLE001
            app.logger.warning('Falha ao TRUNCATE bras_item_preco: %s', exc)
        try:
            db.session.execute(text('TRUNCATE TABLE bras_item_cadastro'))
        except Exception as exc:  # noqa: BLE001
            app.logger.warning('Falha ao TRUNCATE bras_item_cadastro: %s', exc)
        db.session.commit()
        return

    # Modo 3: Limpar por arquivo específico
    if not arquivo_label:
        return

    params = {'arquivo': arquivo_label}
    
    # Deletar insumos_index primeiro (usando a função de batch por IDs)
    _delete_insumos_by_arquivo('BRAS', 'bras_item_n', arquivo_label)
    _delete_with_retry('DELETE FROM bras_item_n WHERE arquivo = :arquivo', params)
    _delete_with_retry('DELETE FROM bras_raw WHERE arquivo = :arquivo', params)
    _delete_with_retry('DELETE FROM bras_fixed_stage WHERE arquivo = :arquivo', params)


def _delete_existing_simpro_records(
    arquivo_label: str | None,
    truncate: bool,
    *,
    aliquota_filter: Decimal | None = None,
    uf_filter: str | None = None,
) -> None:
    """
    Limpa registros do SIMPRO.
    
    Modos de operação:
    - truncate=True, aliquota_filter=None: Limpa TUDO do SIMPRO
    - truncate=True, aliquota_filter=X: Limpa apenas dados da alíquota X (mantém outras)
    - truncate=False, arquivo_label=X: Limpa apenas registros do arquivo X
    """
    
    # Modo 1: Limpar por alíquota específica (mantém outras alíquotas)
    if truncate and aliquota_filter is not None:
        app.logger.info('Limpando SIMPRO apenas para alíquota %.2f / UF %s', aliquota_filter, uf_filter or 'todas')
        
        # Construir filtro de UF
        uf_condition = ""
        params: dict = {'aliquota': float(aliquota_filter)}
        if uf_filter:
            uf_condition = " AND (uf_referencia LIKE :uf_pattern OR uf_referencia = :uf_exact)"
            params['uf_pattern'] = f'%{uf_filter}%'
            params['uf_exact'] = uf_filter
        
        # Deletar do índice de insumos pela alíquota (com retry)
        _delete_with_retry(
            f"DELETE FROM insumos_index WHERE origem = 'SIMPRO' AND aliquota = :aliquota{uf_condition}",
            params,
        )
        return
    
    # Modo 2: Limpar TUDO (truncate sem filtro)
    if truncate:
        _delete_with_retry("DELETE FROM insumos_index WHERE origem = 'SIMPRO'", {})
        db.session.execute(text('TRUNCATE TABLE simpro_item_preco'))
        db.session.execute(text('TRUNCATE TABLE simpro_item_cadastro'))
        db.session.execute(text('TRUNCATE TABLE simpro_item_norm'))
        db.session.execute(text('TRUNCATE TABLE simpro_fixed_stage'))
        db.session.commit()
        return

    # Modo 3: Limpar por arquivo específico
    if not arquivo_label:
        return

    params = {'arquivo': arquivo_label}

    # Novo split SIMPRO: limpar preços por arquivo + índice vinculado aos cadastros tocados.
    cadastro_ids = [
        int(row.cadastro_id)
        for row in db.session.query(SimproItemPreco.cadastro_id)
        .filter(SimproItemPreco.arquivo_fonte == arquivo_label)
        .all()
        if row.cadastro_id is not None
    ]
    if cadastro_ids:
        (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == 'SIMPRO',
                InsumoIndex.item_id.in_(cadastro_ids),
            )
            .delete(synchronize_session=False)
        )
    _delete_with_retry('DELETE FROM simpro_item_preco WHERE arquivo_fonte = :arquivo', params)
    _cleanup_simpro_cadastro_orphans(cadastro_ids)

    # Legacy pipeline SIMPRO (mantido para retrocompatibilidade durante transição).
    _delete_with_retry('DELETE FROM simpro_item_norm WHERE arquivo = :arquivo', params)
    _delete_with_retry('DELETE FROM simpro_fixed_stage WHERE arquivo = :arquivo', params)
    db.session.commit()


def _bras_load_data_delimited(
    *,
    file_path: Path,
    delimiter: str,
    quotechar: str | None,
    line_terminator: str,
    skip_header: bool,
    encoding: str | None,
    arquivo_label: str,
) -> int:
    charset = _encoding_to_mysql_charset(encoding)
    delimiter_lit = _sql_escape_literal(delimiter)
    line_term_lit = _sql_escape_literal(_encode_line_terminator(line_terminator))
    file_literal = _sql_escape_literal(str(file_path))
    arquivo_literal = _sql_escape_literal(arquivo_label)
    quote_clause = ''
    if quotechar:
        quote_clause = f"OPTIONALLY ENCLOSED BY {_sql_escape_literal(quotechar)}\n"

    ignore_clause = 'IGNORE 1 LINES\n' if skip_header else ''

    bindings = [f"@col{idx:02d}" for idx in range(1, 24)]
    set_lines = [
        f"col{idx:02d} = NULLIF(@col{idx:02d}, '')"
        for idx in range(1, 24)
    ]
    set_lines.append(f"arquivo = {arquivo_literal}")
    set_lines.append("linha_num = (@row := @row + 1)")
    set_clause = ',\n        '.join(set_lines)

    load_stmt = (
        f"LOAD DATA LOCAL INFILE {file_literal}\n"
        "INTO TABLE bras_raw\n"
        f"CHARACTER SET {charset}\n"
        f"FIELDS TERMINATED BY {delimiter_lit}\n"
        f"{quote_clause}"
        f"LINES TERMINATED BY {line_term_lit}\n"
        f"{ignore_clause}"
        f"({', '.join(bindings)})\n"
        f"SET {set_clause}"
    )

    with db.engine.begin() as conn:
        conn.exec_driver_sql('SET @row := 0')
        result = conn.exec_driver_sql(load_stmt)
        return result.rowcount or 0


def _bras_csv_fallback(
    *,
    file_path: Path,
    delimiter: str,
    quotechar: str | None,
    skip_header: bool,
    encodings: list[str],
    arquivo_label: str,
    batch_size: int = BRAS_RAW_CSV_BATCH,
) -> int:
    for enc in encodings:
        try:
            with file_path.open('r', encoding=enc, newline='') as handle:
                reader = csv.reader(handle, delimiter=delimiter, quotechar=quotechar or '"')
                rows: list[dict] = []
                total_inserted = 0
                linha_num = 0
                
                for idx, raw in enumerate(reader, start=1):
                    _raise_if_import_paused()
                    if skip_header and idx == 1:
                        continue
                    linha_num += 1
                    values = (raw or [])[:23]
                    values += [''] * (23 - len(values))
                    mapping = {
                        'arquivo': arquivo_label,
                        'linha_num': linha_num,
                        **{f'col{pos:02d}': (val.strip() or None) if isinstance(val, str) else None for pos, val in enumerate(values, start=1)}
                    }
                    rows.append(mapping)
                    
                    # Commit em batches para evitar lock timeout
                    if len(rows) >= batch_size:
                        db.session.bulk_insert_mappings(BrasRaw, rows)
                        db.session.commit()
                        total_inserted += len(rows)
                        rows = []
                
                # Inserir linhas restantes
                if rows:
                    db.session.bulk_insert_mappings(BrasRaw, rows)
                    db.session.commit()
                    total_inserted += len(rows)
                
                return total_inserted
        except UnicodeDecodeError:
            db.session.rollback()
            continue
    raise click.ClickException('Não foi possível decodificar o arquivo com as codificações testadas.')


def _stage_simpro_fixed(
    *,
    file_path: Path,
    map_config: dict,
    encoding: str | None,
    arquivo_label: str,
    batch_size: int = 2000,
) -> tuple[int, str]:
    encodings = _build_encoding_list(encoding)
    skip_header = bool(map_config.get('skip_header'))
    inserted = 0
    for enc in encodings:
        try:
            rows: list[dict] = []
            total_inserted = 0
            with file_path.open('r', encoding=enc, newline='') as handle:
                logical_idx = 0
                for raw_idx, raw_line in enumerate(handle, start=1):
                    _raise_if_import_paused()
                    if skip_header and raw_idx == 1:
                        continue
                    line = raw_line.rstrip('\r\n')
                    logical_idx += 1
                    rows.append({
                        'arquivo': arquivo_label,
                        'linha_num': logical_idx,
                        'linha': line,
                    })
                    
                    # Commit em batches para evitar lock timeout
                    if len(rows) >= batch_size:
                        db.session.bulk_insert_mappings(SimproFixedStage, rows)
                        db.session.commit()
                        total_inserted += len(rows)
                        rows = []
            
            # Inserir linhas restantes
            if rows:
                db.session.bulk_insert_mappings(SimproFixedStage, rows)
                db.session.commit()
                total_inserted += len(rows)
            
            if total_inserted > 0:
                inserted = total_inserted
                break
        except UnicodeDecodeError:
            db.session.rollback()
            continue
    if not inserted:
        raise click.ClickException('Não foi possível decodificar o arquivo de largura fixa do SIMPRO.')
    return inserted, 'python_fixed'


def _parse_fixed_date(value: str | None, fmt: str | None) -> date | None:
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    fmt = (fmt or 'DDMMYYYY').upper()
    python_fmt = fmt.replace('YYYY', '%Y').replace('YY', '%y').replace('MM', '%m').replace('DD', '%d')
    try:
        return datetime.strptime(value, python_fmt).date()
    except ValueError:
        return None


def _parse_simpro_json_date(value: str | None) -> date | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%d%m%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _sanitize_numeric(value: str) -> str:
    return ''.join(ch for ch in value if ch.isdigit() or ch in ',.-')


def _auto_scale_decimal(value: Decimal | None, *, min_fraction_digits: int = 2) -> Decimal | None:
    if value is None:
        return None
    if not isinstance(value, Decimal):
        try:
            value = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    quantize_pattern = '0.' + ('0' * max(min_fraction_digits, 4))
    try:
        return value.quantize(Decimal(quantize_pattern))
    except (InvalidOperation, ValueError):
        return value


def _materialize_simpro_items(
    *,
    arquivo_label: str,
    map_config: dict,
    versao: str,
    uf_default: str | None,
) -> int:
    columns_cfg = map_config.get('columns') or []
    if not columns_cfg:
        raise click.ClickException('Mapa SIMPRO precisa definir "columns".')

    try:
        decimal_divisor = Decimal(str(map_config.get('decimal_divisor') or '1'))
    except (InvalidOperation, ValueError):
        decimal_divisor = Decimal('1')
    if not decimal_divisor:
        decimal_divisor = Decimal('1')

    postprocess_cfg = map_config.get('postprocess') if isinstance(map_config.get('postprocess'), dict) else None
    field_map = _resolve_simpro_field_map(map_config)

    read_batch_size = 2000
    write_batch_size = 500
    last_stage_id = 0
    total_inserted = 0
    while True:
        stage_rows = (
            SimproFixedStage.query
            .filter(
                SimproFixedStage.arquivo == arquivo_label,
                SimproFixedStage.id > last_stage_id,
            )
            .order_by(SimproFixedStage.id.asc())
            .limit(read_batch_size)
            .all()
        )
        if not stage_rows:
            break

        parsed_batch: list[dict] = []
        for stage in stage_rows:
            _raise_if_import_paused()
            line = stage.linha or ''
            record: dict[str, object | None] = {
                'id': stage.id,
                'arquivo': stage.arquivo,
                'linha_num': stage.linha_num,
                'versao': versao,
                'uf_referencia': uf_default,
                'imported_at': stage.imported_at,
            }
            for cfg in columns_cfg:
                if not isinstance(cfg, dict):
                    continue
                name = (cfg.get('name') or '').strip()
                if not name:
                    continue
                start = max(int(cfg.get('start', 1)) - 1, 0)
                length = max(int(cfg.get('length', 0)), 0)
                if length <= 0:
                    record[name] = None
                    continue
                raw_value = line[start:start + length]
                if cfg.get('strip') and isinstance(cfg['strip'], (list, tuple)):
                    for ch in cfg['strip']:
                        raw_value = raw_value.replace(str(ch), '')
                if cfg.get('rtrim'):
                    raw_value = raw_value.rstrip()
                value = raw_value.strip()
                if not value:
                    record[name] = None
                    continue

                value_type = (cfg.get('type') or '').strip().lower()
                if value_type == 'decimal':
                    coerced = _coerce_decimal(_sanitize_numeric(value))
                    if coerced is None:
                        record[name] = None
                    else:
                        divisor_raw = cfg.get('divide_by', decimal_divisor)
                        try:
                            divisor = Decimal(str(divisor_raw or '1'))
                        except (InvalidOperation, ValueError):
                            divisor = Decimal('1')
                        if not divisor:
                            divisor = Decimal('1')
                        scaled = Decimal(coerced) / divisor
                        if scaled >= Decimal('10000000'):
                            adjusted = scaled / Decimal('1000000')
                            if adjusted >= Decimal('0.01'):
                                scaled = adjusted
                        record[name] = _auto_scale_decimal(scaled)
                elif value_type == 'date':
                    record[name] = _parse_fixed_date(value, cfg.get('date_fmt'))
                elif value_type == 'int':
                    digits = ''.join(ch for ch in value if ch.isdigit() or ch == '-')
                    try:
                        record[name] = int(digits) if digits else None
                    except ValueError:
                        record[name] = None
                else:
                    record[name] = value

            _apply_simpro_postprocess(record, postprocess_cfg)
            _enrich_tuss_from_ean(record)
            _ensure_tuss_from_line(record, line)
            _ensure_tuss_field(record)
            _normalize_anvisa_field(record)
            _ensure_anvisa_from_line(record, line)

            payload = {
                'id': stage.id,
                'arquivo': stage.arquivo,
                'linha_num': stage.linha_num,
                'versao': versao,
                'uf_referencia': uf_default,
                'imported_at': stage.imported_at,
            }
            payload.update(_build_simpro_payload(record, field_map))
            parsed_batch.append(payload)

            if len(parsed_batch) >= write_batch_size:
                db.session.bulk_insert_mappings(SimproItemNormalized, parsed_batch)
                db.session.commit()
                total_inserted += len(parsed_batch)
                parsed_batch = []

        if parsed_batch:
            db.session.bulk_insert_mappings(SimproItemNormalized, parsed_batch)
            db.session.commit()
            total_inserted += len(parsed_batch)

        last_stage_id = int(stage_rows[-1].id)
    
    return total_inserted


def _load_simpro_json_payload(file_path: Path, encoding: str | None = None) -> list[dict]:
    encodings = _build_encoding_list(encoding)
    last_error: Exception | None = None
    for enc in encodings:
        try:
            with file_path.open('r', encoding=enc, newline='') as fp:
                payload = json.load(fp)
            if isinstance(payload, dict):
                items = payload.get('produtos')
            elif isinstance(payload, list):
                items = payload
            else:
                items = None
            if not isinstance(items, list):
                raise click.ClickException('JSON SIMPRO deve conter uma lista em "produtos" ou na raiz.')
            normalized_items = [item for item in items if isinstance(item, dict)]
            if not normalized_items:
                raise click.ClickException('JSON SIMPRO não possui registros válidos.')
            return normalized_items
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except json.JSONDecodeError as exc:
            raise click.ClickException(f'Não foi possível interpretar o JSON do SIMPRO: {exc}') from exc
    raise click.ClickException('Não foi possível decodificar o arquivo JSON do SIMPRO.') from last_error


def _materialize_simpro_json_items(
    *,
    arquivo_label: str,
    records: Sequence[dict],
    versao: str,
    uf_default: str | None,
    job_id: str | None = None,
) -> int:
    if not records:
        return 0

    max_id = db.session.query(func.max(SimproItemNormalized.id)).scalar() or 0
    next_id = int(max_id) + 1
    raw_total = len(records)
    batch_size = SIMPRO_JSON_MATERIALIZE_BATCH
    batch: list[dict[str, object | None]] = []
    total_inserted = 0

    def _flush_batch() -> None:
        nonlocal batch, total_inserted
        if not batch:
            return
        _raise_if_import_paused()
        db.session.bulk_insert_mappings(SimproItemNormalized, batch)
        db.session.commit()
        total_inserted += len(batch)
        if job_id:
            _touch_import_job_progress(
                job_id,
                message=(
                    f'SIMPRO JSON: {total_inserted} linhas materializadas '
                    f'({raw_total} registros no arquivo; lote {batch_size})…'
                ),
                total_linhas=raw_total,
                linhas_materializadas=total_inserted,
            )
        batch = []

    for idx, source in enumerate(records, start=1):
        _raise_if_import_paused()
        codigo_simpro = (source.get('codigoSimpro') or source.get('codigoUsuario') or '').strip()
        descricao = (source.get('descricao') or '').strip()
        if not codigo_simpro or not descricao:
            continue

        tuss_raw = source.get('codigoTUSS')
        tuss_digits = ''.join(ch for ch in str(tuss_raw).strip() if ch.isdigit()) if tuss_raw is not None else ''
        anvisa_record = {
            'anvisa': source.get('anvisa'),
        }
        _normalize_anvisa_field(anvisa_record)

        preco_fabrica = _coerce_decimal(source.get('precoFabrica'))
        preco_usuario = _coerce_decimal(source.get('precoUsuario'))
        preco_fabrica_fracao = _coerce_decimal(source.get('precoFabricaFracao'))
        preco_usuario_fracao = _coerce_decimal(source.get('precoUsuarioFracao'))

        batch.append({
            'id': next_id + idx - 1,
            'arquivo': arquivo_label,
            'linha_num': idx,
            'codigo_interno': (source.get('codigoUsuario') or '').strip() or None,
            'codigo': codigo_simpro,
            'codigo_alt': (source.get('codigoFracao') or '').strip() or None,
            'descricao': descricao,
            'data_ref': _parse_simpro_json_date(source.get('vigencia')),
            'tipo_reg': (source.get('identificacao') or '').strip() or None,
            'preco1': Decimal(preco_fabrica) if preco_fabrica is not None else None,
            'preco2': Decimal(preco_usuario) if preco_usuario is not None else None,
            'preco3': Decimal(preco_fabrica_fracao) if preco_fabrica_fracao is not None else None,
            'preco4': Decimal(preco_usuario_fracao) if preco_usuario_fracao is not None else None,
            'unidade': ((source.get('embalagem') or '').strip() or (source.get('fracao') or '').strip() or None),
            'qtd_unidade': int(Decimal(str(source.get('quantidadeEmbalagem')))) if source.get('quantidadeEmbalagem') not in (None, '') else None,
            'fabricante': (source.get('fabricante') or '').strip() or None,
            'referencia': (source.get('referencia') or '').strip() or None,
            'anvisa': anvisa_record.get('anvisa'),
            'validade_anvisa': _parse_simpro_json_date(source.get('validadeAnvisa')),
            'ean': (source.get('codigoEAN') or '').strip() or None,
            'situacao': (source.get('tipoAlteracao') or '').strip() or None,
            'fracionavel': ((source.get('fracionavel') or '').strip()[:1].upper() or None),
            'versao': versao,
            'uf_referencia': uf_default,
            'tuss_prefix': None,
            'tuss_numero': tuss_digits or None,
            'status_final': (source.get('classificacao') or '').strip() or None,
        })
        if len(batch) >= batch_size:
            _flush_batch()

    _flush_batch()
    return total_inserted


def _stage_bras_delimited(
    *,
    file_path: Path,
    delimiter: str,
    quotechar: str | None,
    line_terminator: str,
    skip_header: bool,
    encoding: str | None,
    arquivo_label: str,
    use_load_data: bool,
) -> tuple[int, str]:
    encodings = _build_encoding_list(encoding)
    inserted = 0
    strategy = 'load_data'
    if use_load_data:
        try:
            inserted = _bras_load_data_delimited(
                file_path=file_path,
                delimiter=delimiter,
                quotechar=quotechar,
                line_terminator=line_terminator,
                skip_header=skip_header,
                encoding=encodings[0],
                arquivo_label=arquivo_label,
            )
        except Exception as exc:
            db.session.rollback()
            app.logger.warning('LOAD DATA falhou (%s); usando fallback Python.', exc)
            inserted = 0
            strategy = 'python'

    if not inserted:
        inserted = _bras_csv_fallback(
            file_path=file_path,
            delimiter=delimiter,
            quotechar=quotechar,
            skip_header=skip_header,
            encodings=encodings,
            arquivo_label=arquivo_label,
        )
        strategy = 'python'
    return inserted, strategy


def _stage_bras_fixed(
    *,
    file_path: Path,
    map_config: dict,
    encoding: str | None,
    line_terminator: str,
    arquivo_label: str,
    batch_size: int = 2000,
) -> tuple[int, str]:
    columns_cfg = map_config.get('columns') or []
    if not columns_cfg:
        raise click.ClickException('Arquivo de mapeamento precisa definir "columns".')

    encodings = _build_encoding_list(encoding)
    inserted = 0
    for enc in encodings:
        try:
            rows_stage: list[dict] = []
            rows_raw: list[dict] = []
            total_inserted = 0
            
            with file_path.open('r', encoding=enc, newline='') as handle:
                for idx, raw_line in enumerate(handle, start=1):
                    _raise_if_import_paused()
                    line = raw_line.rstrip('\r\n')
                    rows_stage.append({'arquivo': arquivo_label, 'linha_num': idx, 'linha': line})
                    mapping = {'arquivo': arquivo_label, 'linha_num': idx}
                    for col in columns_cfg:
                        name = col.get('name')
                        start = int(col.get('start', 1)) - 1
                        length = int(col.get('length', 0))
                        if not name or length <= 0:
                            continue
                        snippet = line[start:start + length]
                        mapping[name] = snippet.strip() or None
                    rows_raw.append(mapping)
                    
                    # Commit em batches para evitar lock timeout
                    if len(rows_raw) >= batch_size:
                        if rows_stage:
                            db.session.bulk_insert_mappings(BrasFixedStage, rows_stage)
                        if rows_raw:
                            db.session.bulk_insert_mappings(BrasRaw, rows_raw)
                        db.session.commit()
                        total_inserted += len(rows_raw)
                        rows_stage = []
                        rows_raw = []
            
            # Inserir linhas restantes
            if rows_stage:
                db.session.bulk_insert_mappings(BrasFixedStage, rows_stage)
            if rows_raw:
                db.session.bulk_insert_mappings(BrasRaw, rows_raw)
            db.session.commit()
            total_inserted += len(rows_raw)
            
            inserted = total_inserted
            break
        except UnicodeDecodeError:
            db.session.rollback()
            continue
    if not inserted:
        raise click.ClickException('Não foi possível decodificar o arquivo de largura fixa.')
    return inserted, 'python_fixed'


_bras_view_created = False

def _ensure_bras_item_view_exists() -> None:
    global _bras_view_created
    if _bras_view_created:
        return
    with db.engine.begin() as conn:
        conn.exec_driver_sql(BRAS_ITEM_VIEW_SQL)
    _bras_view_created = True


def _execute_with_retry(sql, params: dict, max_retries: int = 10) -> int:
    """Executa SQL com retry automático para deadlocks e lock timeouts."""
    import time
    
    for attempt in range(max_retries):
        try:
            result = db.session.execute(sql, params)
            db.session.commit()
            return result.rowcount or 0
        except Exception as exc:
            db.session.rollback()
            error_str = str(exc)
            # Erro 1213: Deadlock, Erro 1205: Lock wait timeout
            if ('1213' in error_str or '1205' in error_str) and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2  # Backoff progressivo: 2s, 4s, 6s, 8s...
                app.logger.warning('DB lock error (attempt %d/%d), retrying in %.1fs: %s', 
                                  attempt + 1, max_retries, wait_time, error_str[:100])
                time.sleep(wait_time)
                continue
            raise
    return 0


def _materialize_bras_items(
    arquivo_label: str | None, batch_size: int | None = None,
) -> int:
    """Materializa itens da view bras_item_v para bras_item_n em batches (keyset id; evita OFFSET lento)."""
    _ensure_bras_item_view_exists()
    if batch_size is not None:
        bs = max(200, min(4000, int(batch_size)))
    else:
        bs = BRAS_MATERIALIZE_BATCH

    params: dict[str, object] = {}
    where_w = 'w.arquivo = :arquivo'
    if arquivo_label:
        params['arquivo'] = arquivo_label
    else:
        where_w = '1=1'

    count_sql = text(
        f"SELECT COUNT(*) FROM bras_item_v w WHERE {where_w}" if arquivo_label
        else 'SELECT COUNT(*) FROM bras_item_v w',
    )
    if arquivo_label is None:
        count_params: dict = {}
    else:
        count_params = params
    total_count = db.session.execute(count_sql, count_params).scalar() or 0

    if total_count == 0:
        return 0

    # Keyset: lotes consecutivos por `id` (não depende de OFFSET; escala muito melhor)
    last_id: int = 0
    batch_idx = 0
    while True:
        _raise_if_import_paused()
        if arquivo_label is not None:
            id_list_sql = text(
                f"""
                SELECT w.id FROM bras_item_v w
                WHERE {where_w} AND w.id > :last_id
                ORDER BY w.id
                LIMIT :bs
                """
            )
        else:
            id_list_sql = text(
                """
                SELECT w.id FROM bras_item_v w
                WHERE w.id > :last_id
                ORDER BY w.id
                LIMIT :bs
                """
            )
        batch_params = {**params, 'last_id': last_id, 'bs': bs}
        rows = db.session.execute(id_list_sql, batch_params).fetchall()
        if not rows:
            break
        ids: list[int] = [int(r[0]) for r in rows]
        last_id = ids[-1]
        in_clause = f"({ids[0]})" if len(ids) == 1 else str(tuple(ids))

        insert_sql = text(
            f"""
            INSERT INTO bras_item_n (
                id, arquivo, linha_num,
                laboratorio_codigo, laboratorio_nome,
                produto_codigo, produto_nome,
                apresentacao_codigo, apresentacao_descricao,
                ean, registro_anvisa, edicao,
                preco_pmc_pacote, preco_pfb_pacote, preco_pmc_unit, preco_pfb_unit,
                aliquota_ou_ipi, quantidade_embalagem, imported_at
            )
            SELECT
                v.id, v.arquivo, v.linha_num,
                v.laboratorio_codigo, v.laboratorio_nome,
                v.produto_codigo, v.produto_nome,
                v.apresentacao_codigo, v.apresentacao_descricao,
                v.ean, v.registro_anvisa, v.edicao,
                v.preco_pmc_pacote, v.preco_pfb_pacote, v.preco_pmc_unit, v.preco_pfb_unit,
                v.aliquota_ou_ipi, v.quantidade_embalagem, v.imported_at
            FROM bras_item_v v
            WHERE v.id IN {in_clause}
            ON DUPLICATE KEY UPDATE
                arquivo = VALUES(arquivo),
                linha_num = VALUES(linha_num),
                laboratorio_codigo = VALUES(laboratorio_codigo),
                laboratorio_nome = VALUES(laboratorio_nome),
                produto_codigo = VALUES(produto_codigo),
                produto_nome = VALUES(produto_nome),
                apresentacao_codigo = VALUES(apresentacao_codigo),
                apresentacao_descricao = VALUES(apresentacao_descricao),
                ean = VALUES(ean),
                registro_anvisa = VALUES(registro_anvisa),
                edicao = VALUES(edicao),
                preco_pmc_pacote = VALUES(preco_pmc_pacote),
                preco_pfb_pacote = VALUES(preco_pfb_pacote),
                preco_pmc_unit = VALUES(preco_pmc_unit),
                preco_pfb_unit = VALUES(preco_pfb_unit),
                aliquota_ou_ipi = VALUES(aliquota_ou_ipi),
                quantidade_embalagem = VALUES(quantidade_embalagem),
                imported_at = VALUES(imported_at)
            """,
        )

        _execute_with_retry(insert_sql, {})
        batch_idx += 1
        _bras_throttle_between_batches()
        app.logger.debug(
            'Materialize BRAS keyset: batch=%d, last_id=%d, tamanho_lote=%d, total_espalhado~=%d',
            batch_idx,
            last_id,
            len(ids),
            total_count,
        )

    return total_count


def _normalize_uf_codes(
    uf_values: Sequence[str] | None,
    *,
    uf_default: str | None = None,
) -> list[str]:
    codes: list[str] = []
    pool = list(uf_values or [])
    if uf_default is not None:
        pool.append(uf_default)
    for raw in pool:
        raw_str = (raw or '').strip().upper()
        if not raw_str:
            continue
        if raw_str not in codes:
            codes.append(raw_str)
    return codes


def _encode_uf_codes(codes: Sequence[str] | None) -> str | None:
    if not codes:
        return None
    filtered = [code.strip().upper() for code in codes if code and code.strip()]
    if not filtered:
        return None
    return '|' + '|'.join(filtered) + '|'


_SIMPRO_ALIQUOTA_UF_ROWS: tuple[tuple[str, list[str]], ...] = tuple(
    (s, ufs)
    for s, ufs in [
        ('17', ['DF', 'ES', 'MT', 'MS', 'RS', 'SC']),
        ('18', ['AP', 'MG', 'SP']),
        ('19', ['AC', 'AL', 'GO', 'PA', 'SE']),
        ('19.5', ['PR', 'RO']),
        ('20', ['AM', 'CE', 'PB', 'RN', 'RR', 'TO']),
        ('20.5', ['BA', 'PE']),
        ('22', ['RJ']),
        ('22.5', ['PI']),
        ('23', ['MA']),
    ]
)


def _simpro_piso_aliquota_for_uf(uf: str | None) -> Decimal | None:
    """Alíquota (piso) SIMPRO associada à UF no mapa ``_SIMPRO_ALIQUOTA_UF_ROWS``, ou None."""
    u = (uf or '').strip().upper()
    if not u:
        return None
    for rate_str, ufs in _SIMPRO_ALIQUOTA_UF_ROWS:
        if u in {x.upper() for x in ufs}:
            try:
                return Decimal(str(rate_str))
            except (InvalidOperation, ValueError, TypeError):
                return None
    return None


def _simpro_aliquota_includes_uf(aliquota: Decimal | None, uf: str | None) -> bool:
    """True se a UF pertence ao piso SIMPRO da alíquota (mapa ``_SIMPRO_ALIQUOTA_UF_ROWS``)."""
    if not (uf or '').strip():
        return True
    if aliquota is None:
        return False
    u = uf.strip().upper()
    aq = _br_norm_aliquota(aliquota) or aliquota
    for rate_str, ufs in _SIMPRO_ALIQUOTA_UF_ROWS:
        try:
            rq = Decimal(str(rate_str))
        except Exception:
            continue
        if abs(aq - rq) < Decimal('0.02'):
            return u in {x.upper() for x in ufs}
    return True


def _sql_insumo_uf_referencia_from_aliquota_column(col_sql: str) -> str:
    """
    Fragmento SQL: retorna literal `|UF|…|` alinhado a `_ufs_pertencentes_a_aliquota_piso`,
    com tolerância como `_br_aliquota_certeiro`: ABS(a−b) < 0.02.
    """
    parts = ['CASE']
    for rate_str, ufs in _SIMPRO_ALIQUOTA_UF_ROWS:
        enc = _encode_uf_codes(ufs)
        if not enc:
            continue
        esc = enc.replace("'", "''")
        rq = Decimal(str(rate_str))
        rate_lit = format(rq, 'f').replace("'", "''")
        parts.append(
            '    WHEN '
            f'{col_sql} IS NOT NULL AND ABS(CAST({col_sql} AS DECIMAL(14, 4)) - CAST(\'{rate_lit}\' AS DECIMAL(14, 4)))'
            ' < CAST(0.02 AS DECIMAL(14, 4)) '
            f"THEN '{esc}'"
        )
    parts.append('    ELSE NULL')
    parts.append('END')
    return '\n'.join(parts)


UF_SPLIT_RE = re.compile(r'[|,\s]+')


def _decode_uf_codes(value: str | None) -> list[str]:
    if not value:
        return []
    text = value.strip()
    if not text:
        return []
    if text.startswith('|') and text.endswith('|') and len(text) >= 2:
        text = text.strip('|')
    parts = [segment.strip().upper() for segment in UF_SPLIT_RE.split(text) if segment and segment.strip()]
    if not parts and value:
        fallback = value.strip().upper()
        if fallback:
            parts = [fallback]
    deduped: list[str] = []
    for code in parts:
        if code and code not in deduped:
            deduped.append(code)
    return deduped


def _combine_uf_codes(*values: str | None) -> list[str]:
    combined: list[str] = []
    for raw in values:
        for code in _decode_uf_codes(raw):
            if code not in combined:
                combined.append(code)
    return combined


def _clean_bras_catalog_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _build_bras_catalog_key(
    *,
    ean: str | None,
    codigo_composto: str | None,
    laboratorio_codigo: str | None,
    produto_codigo: str | None,
    apresentacao_codigo: str | None,
) -> tuple[str, str]:
    ean_clean = ''.join(ch for ch in str(ean or '').strip() if ch.isdigit())
    if ean_clean:
        return f'ean:{ean_clean}', 'ean'

    composed = _clean_bras_catalog_text(codigo_composto)
    if composed:
        return f'comp:{composed.upper()}', 'codigo_composto'

    parts = [
        (_clean_bras_catalog_text(laboratorio_codigo) or '').upper(),
        (_clean_bras_catalog_text(produto_codigo) or '').upper(),
        (_clean_bras_catalog_text(apresentacao_codigo) or '').upper(),
    ]
    return f'fallback:{"|".join(parts)}', 'fallback'


def _build_bras_catalog_row_hash(payload: dict[str, object | None]) -> str:
    fingerprint = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=_json_default)
    return hashlib.sha256(fingerprint.encode('utf-8')).hexdigest()


def _parse_bras_catalog_row(row: Sequence[str]) -> dict[str, str | None] | None:
    if len(row) < 10:
        return None

    laboratorio_codigo = _clean_bras_catalog_text(row[0])
    laboratorio_nome = _clean_bras_catalog_text(row[1])
    produto_codigo = _clean_bras_catalog_text(row[2])
    produto_nome = _clean_bras_catalog_text(row[3])
    apresentacao_codigo = _clean_bras_catalog_text(row[4])
    apresentacao_descricao = _clean_bras_catalog_text(row[5])
    codigo_composto = _clean_bras_catalog_text(row[6])
    ean = _clean_bras_catalog_text(row[7])
    codigo_interno = _clean_bras_catalog_text(row[8])
    tuss = _clean_bras_catalog_text(row[9])

    item_key, key_kind = _build_bras_catalog_key(
        ean=ean,
        codigo_composto=codigo_composto,
        laboratorio_codigo=laboratorio_codigo,
        produto_codigo=produto_codigo,
        apresentacao_codigo=apresentacao_codigo,
    )

    payload = {
        'item_key': item_key,
        'key_kind': key_kind,
        'laboratorio_codigo': laboratorio_codigo,
        'laboratorio_nome': laboratorio_nome,
        'produto_codigo': produto_codigo,
        'produto_nome': produto_nome,
        'apresentacao_codigo': apresentacao_codigo,
        'apresentacao_descricao': apresentacao_descricao,
        'codigo_composto': codigo_composto,
        'ean': ean,
        'codigo_interno': codigo_interno,
        'tuss': tuss,
    }
    payload['row_hash'] = _build_bras_catalog_row_hash(payload)
    return payload


def _load_bras_catalog_file(
    file_path: Path,
    *,
    delimiter: str = ';',
    quotechar: str = '"',
    encoding: str | None = 'latin-1',
) -> list[dict[str, str | None]]:
    import csv

    rows: list[dict[str, str | None]] = []
    encodings = _build_encoding_list(encoding)
    last_error: Exception | None = None

    for enc in encodings:
        try:
            with file_path.open('r', encoding=enc, errors='strict', newline='') as handle:
                reader = csv.reader(handle, delimiter=delimiter, quotechar=quotechar)
                for row in reader:
                    parsed = _parse_bras_catalog_row(row)
                    if parsed:
                        rows.append(parsed)
            return rows
        except UnicodeDecodeError as exc:
            rows.clear()
            last_error = exc
            continue

    raise click.ClickException('Não foi possível decodificar o arquivo de catálogo Brasíndice.') from last_error


def _sync_bras_catalog_snapshot(
    *,
    file_path: Path,
    versao: str,
    delimiter: str = ';',
    quotechar: str = '"',
    encoding: str | None = 'latin-1',
) -> dict[str, int | str]:
    max_id = db.session.query(func.max(BrasCatalogSnapshot.id)).scalar() or 0
    records = _load_bras_catalog_file(
        file_path,
        delimiter=delimiter,
        quotechar=quotechar,
        encoding=encoding,
    )

    db.session.query(BrasCatalogSnapshot).filter_by(versao=versao).delete(synchronize_session=False)
    db.session.commit()

    if not records:
        return {'versao': versao, 'rows': 0}

    source_name = file_path.name
    batch: list[dict[str, object | None]] = []
    for idx, record in enumerate(records, start=1):
        batch.append({
            'id': int(max_id) + idx,
            'versao': versao,
            'source_file': source_name,
            **record,
        })

    db.session.bulk_insert_mappings(BrasCatalogSnapshot, batch)
    db.session.commit()
    return {'versao': versao, 'rows': len(batch)}


def _resolve_previous_bras_catalog_version(current_version: str) -> str | None:
    rows = (
        db.session.query(BrasCatalogSnapshot.versao)
        .filter(BrasCatalogSnapshot.versao != current_version)
        .distinct()
        .order_by(BrasCatalogSnapshot.versao.desc())
        .all()
    )
    return rows[0][0] if rows else None


def _analyze_bras_catalog_delta(
    *,
    current_version: str,
    previous_version: str | None = None,
) -> dict[str, object]:
    previous_version = previous_version or _resolve_previous_bras_catalog_version(current_version)

    current_rows = (
        BrasCatalogSnapshot.query
        .with_entities(
            BrasCatalogSnapshot.item_key,
            BrasCatalogSnapshot.row_hash,
            BrasCatalogSnapshot.ean,
            BrasCatalogSnapshot.produto_nome,
            BrasCatalogSnapshot.apresentacao_descricao,
            BrasCatalogSnapshot.tuss,
        )
        .filter_by(versao=current_version)
        .all()
    )
    current_map = {
        row.item_key: {
            'row_hash': row.row_hash,
            'ean': row.ean,
            'produto_nome': row.produto_nome,
            'apresentacao_descricao': row.apresentacao_descricao,
            'tuss': row.tuss,
        }
        for row in current_rows
    }

    previous_map: dict[str, dict[str, object | None]] = {}
    if previous_version:
        previous_rows = (
            BrasCatalogSnapshot.query
            .with_entities(
                BrasCatalogSnapshot.item_key,
                BrasCatalogSnapshot.row_hash,
                BrasCatalogSnapshot.ean,
                BrasCatalogSnapshot.produto_nome,
                BrasCatalogSnapshot.apresentacao_descricao,
                BrasCatalogSnapshot.tuss,
            )
            .filter_by(versao=previous_version)
            .all()
        )
        previous_map = {
            row.item_key: {
                'row_hash': row.row_hash,
                'ean': row.ean,
                'produto_nome': row.produto_nome,
                'apresentacao_descricao': row.apresentacao_descricao,
                'tuss': row.tuss,
            }
            for row in previous_rows
        }

    new_keys = sorted(key for key in current_map.keys() if key not in previous_map)
    changed_keys = sorted(
        key for key, current in current_map.items()
        if key in previous_map and current['row_hash'] != previous_map[key]['row_hash']
    )
    removed_keys = sorted(key for key in previous_map.keys() if key not in current_map)

    def _sample(keys: list[str], source_map: dict[str, dict[str, object | None]]) -> list[dict[str, object | None]]:
        sample: list[dict[str, object | None]] = []
        for key in keys[:20]:
            payload = dict(source_map.get(key) or {})
            payload['item_key'] = key
            sample.append(payload)
        return sample

    return {
        'current_version': current_version,
        'previous_version': previous_version,
        'current_total': len(current_map),
        'previous_total': len(previous_map),
        'new_keys': set(new_keys),
        'changed_keys': set(changed_keys),
        'removed_keys': set(removed_keys),
        'new_count': len(new_keys),
        'changed_count': len(changed_keys),
        'removed_count': len(removed_keys),
        'sample_new': _sample(new_keys, current_map),
        'sample_changed': _sample(changed_keys, current_map),
        'sample_removed': _sample(removed_keys, previous_map),
    }


def _build_bras_main_row_key(row: Sequence[str]) -> tuple[str | None, str | None]:
    if len(row) < 20:
        return None, None

    ean = row[16].strip() if len(row) > 16 and row[16] else None
    laboratorio_codigo = row[0].strip() if len(row) > 0 and row[0] else None
    produto_codigo = row[19].strip() if len(row) > 19 and row[19] else None
    apresentacao_codigo = row[17].strip() if len(row) > 17 and row[17] else None
    codigo_composto = None
    if laboratorio_codigo and produto_codigo and apresentacao_codigo:
        codigo_composto = f'{laboratorio_codigo}.{produto_codigo}.{apresentacao_codigo}'

    return _build_bras_catalog_key(
        ean=ean,
        codigo_composto=codigo_composto,
        laboratorio_codigo=laboratorio_codigo,
        produto_codigo=produto_codigo,
        apresentacao_codigo=apresentacao_codigo,
    )


def _sql_clamp_decimal(expr: str, *, integer_digits: int = 8, scale: int = 4) -> str:
    max_value = f"{'9' * integer_digits}.{ '9' * scale}"
    return (
        "CASE\n"
        f"    WHEN {expr} IS NULL THEN NULL\n"
        f"    WHEN {expr} > {max_value} THEN {max_value}\n"
        f"    WHEN {expr} < -{max_value} THEN -{max_value}\n"
        f"    ELSE {expr}\n"
        "END"
    )


def _sync_bras_insumo_index(
    arquivo_label: str | None,
    *,
    uf_default: str | None = None,
    uf_values: Sequence[str] | None = None,
    aliquota_default: Decimal | None = None,
    versao_label: str | None = None,
    batch_size: int | None = None,
) -> None:
    """Sincroniza itens BRAS para insumos_index em lotes (keyset por n.id; evita OFFSET lento)."""
    target_ufs = list(dict.fromkeys([*(uf_values or []), *( [uf_default] if uf_default else [] )]))
    uf_codes = _normalize_uf_codes(target_ufs, uf_default=uf_default)
    uf_storage = _encode_uf_codes(uf_codes)

    if batch_size is not None:
        bs = max(200, min(4000, int(batch_size)))
    else:
        bs = BRAS_INDEX_SYNC_BATCH

    params_base: dict[str, object] = {
        'aliquota_default': aliquota_default,
        'uf_default': uf_default,
        'uf_storage': uf_storage,
        'versao_label': (versao_label or '').strip() or None,
    }
    if arquivo_label:
        params_base['arquivo'] = arquivo_label

    where_count = 'WHERE n.arquivo = :arquivo' if arquivo_label else ''
    count_sql = text(f'SELECT COUNT(*) FROM bras_item_n n {where_count}')
    total_count = db.session.execute(count_sql, params_base if arquivo_label else {}).scalar() or 0
    if total_count == 0:
        return

    preco_expr = "COALESCE(n.preco_pmc_unit, n.preco_pmc_pacote, n.preco_pfb_unit, n.preco_pfb_pacote)"
    preco_sql = _sql_clamp_decimal(preco_expr)
    aliquota_expr = "COALESCE(n.aliquota_ou_ipi, :aliquota_default)"
    aliquota_sql = _sql_clamp_decimal(aliquota_expr, integer_digits=4, scale=4)

    last_n_id: int = 0
    batch_n = 0
    while True:
        if arquivo_label:
            id_list_sql = text(
                """
                SELECT n.id FROM bras_item_n n
                WHERE n.arquivo = :arquivo AND n.id > :last_n_id
                ORDER BY n.id
                LIMIT :bs
                """,
            )
        else:
            id_list_sql = text(
                """
                SELECT n.id FROM bras_item_n n
                WHERE n.id > :last_n_id
                ORDER BY n.id
                LIMIT :bs
                """,
            )
        list_params = {**params_base, 'last_n_id': last_n_id, 'bs': bs}
        rows = db.session.execute(id_list_sql, list_params).fetchall()
        if not rows:
            break
        ids: list[int] = [int(r[0]) for r in rows]
        last_n_id = ids[-1]
        in_clause = f"({ids[0]})" if len(ids) == 1 else str(tuple(ids))
        batch_n += 1

        upsert_template = text(
            f"""
            INSERT INTO insumos_index (
                origem, item_id, tuss, tiss, descricao, preco, aliquota,
                fabricante, anvisa, versao_tabela, data_atualizacao,
                uf_referencia, updated_at
            )
            SELECT
                'BRAS' AS origem,
                n.id AS item_id,
                n.produto_codigo AS tuss,
                n.apresentacao_codigo AS tiss,
                TRIM(CONCAT_WS(' • ', NULLIF(n.produto_nome, ''), NULLIF(n.apresentacao_descricao, ''))) AS descricao,
                {preco_sql} AS preco,
                {aliquota_sql} AS aliquota,
                n.laboratorio_nome AS fabricante,
                n.registro_anvisa AS anvisa,
                COALESCE(:versao_label, n.edicao, n.arquivo) AS versao_tabela,
                NULL AS data_atualizacao,
                COALESCE(:uf_storage, :uf_default, idx.uf_referencia) AS uf_referencia,
                NOW() AS updated_at
            FROM bras_item_n n
            LEFT JOIN insumos_index idx
              ON idx.origem = 'BRAS'
             AND idx.item_id = n.id
             AND idx.aliquota = ({aliquota_sql})
            WHERE n.id IN {in_clause}
            ON DUPLICATE KEY UPDATE
                tuss = VALUES(tuss),
                tiss = VALUES(tiss),
                descricao = VALUES(descricao),
                preco = VALUES(preco),
                aliquota = VALUES(aliquota),
                fabricante = VALUES(fabricante),
                anvisa = VALUES(anvisa),
                versao_tabela = VALUES(versao_tabela),
                data_atualizacao = VALUES(data_atualizacao),
                uf_referencia = COALESCE(VALUES(uf_referencia), insumos_index.uf_referencia),
                updated_at = VALUES(updated_at)
            """
        )
        _execute_with_retry(upsert_template, params_base)
        _bras_throttle_between_batches()
        app.logger.debug('Sync BRAS index keyset: batch=%d, last_n_id=%d, lote=%d, total~%d', batch_n, last_n_id, len(ids), total_count)


def _br_norm_aliquota(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _simpro_item_identity_key(tuss_numero: str | None, codigo: str | None) -> str | None:
    code = str(codigo or '').strip()
    if code:
        return f'SIMPRO:{code}'
    tuss = ''.join(ch for ch in str(tuss_numero or '').strip() if ch.isdigit())
    if tuss:
        return f'TUSS:{tuss}'
    return None


def _sync_simpro_split_from_norm_fast(
    *,
    arquivo_label: str,
    versao: str,
    aliquota_override: Decimal | None,
) -> dict:
    stat = {
        'linhas_simpro_n': 0,
        'cadastros_criados': 0,
        'cadastros_atualizados': 0,
        'precos_criados': 0,
        'precos_atualizados': 0,
        'puladas_sem_identidade': 0,
        'puladas_sem_aliquota': 0,
    }

    aliquota = _br_norm_aliquota(aliquota_override) if aliquota_override is not None else None
    if aliquota is None:
        count_rows = (
            db.session.query(func.count(SimproItemNormalized.id))
            .filter(SimproItemNormalized.arquivo == arquivo_label)
            .scalar()
        ) or 0
        stat['linhas_simpro_n'] = int(count_rows)
        stat['puladas_sem_aliquota'] = int(count_rows)
        return stat

    version_label = (versao or '').strip()
    now_ts = _now_utc()
    next_cadastro_id = int(db.session.query(func.max(SimproItemCadastro.id)).scalar() or 0) + 1
    next_preco_id = int(db.session.query(func.max(SimproItemPreco.id)).scalar() or 0) + 1

    prepared_by_key: dict[str, dict[str, object | None]] = {}
    last_norm_id = 0
    read_batch_size = 2000
    while True:
        norm_rows = (
            db.session.query(
                SimproItemNormalized.id,
                SimproItemNormalized.codigo_interno,
                SimproItemNormalized.codigo,
                SimproItemNormalized.codigo_alt,
                SimproItemNormalized.descricao,
                SimproItemNormalized.data_ref,
                SimproItemNormalized.preco1,
                SimproItemNormalized.preco2,
                SimproItemNormalized.preco3,
                SimproItemNormalized.preco4,
                SimproItemNormalized.unidade,
                SimproItemNormalized.qtd_unidade,
                SimproItemNormalized.fabricante,
                SimproItemNormalized.referencia,
                SimproItemNormalized.anvisa,
                SimproItemNormalized.ean,
                SimproItemNormalized.fracionavel,
                SimproItemNormalized.status_final,
                SimproItemNormalized.linha_num,
                SimproItemNormalized.tuss_numero,
            )
            .filter(
                SimproItemNormalized.arquivo == arquivo_label,
                SimproItemNormalized.id > last_norm_id,
            )
            .order_by(SimproItemNormalized.id.asc())
            .limit(read_batch_size)
            .all()
        )
        if not norm_rows:
            break
        stat['linhas_simpro_n'] += len(norm_rows)
        for row in norm_rows:
            item_key = _simpro_item_identity_key(row.tuss_numero, row.codigo)
            if not item_key:
                stat['puladas_sem_identidade'] += 1
                continue
            # Mantém a última ocorrência da chave no arquivo para evitar duplicatas
            # no mesmo lote e refletir o valor mais recente da importação.
            prepared_by_key[item_key] = {
                'codigo_interno': row.codigo_interno,
                'codigo': row.codigo,
                'codigo_alt': row.codigo_alt,
                'descricao': row.descricao,
                'data_ref': row.data_ref,
                'preco1': row.preco1,
                'preco2': row.preco2,
                'preco3': row.preco3,
                'preco4': row.preco4,
                'unidade': row.unidade,
                'qtd_unidade': row.qtd_unidade,
                'fabricante': row.fabricante,
                'referencia': row.referencia,
                'anvisa': row.anvisa,
                'ean': row.ean,
                'fracionavel': row.fracionavel,
                'status_final': row.status_final,
                'linha_num': row.linha_num,
                'tuss_numero': row.tuss_numero,
            }
        last_norm_id = int(norm_rows[-1].id)
    if not prepared_by_key:
        return stat
    identity_keys = list(prepared_by_key.keys())

    existing_cadastros: list[SimproItemCadastro] = []
    for i in range(0, len(identity_keys), SIMPRO_SPLIT_IN_CHUNK):
        chunk = identity_keys[i : i + SIMPRO_SPLIT_IN_CHUNK]
        existing_cadastros.extend(
            SimproItemCadastro.query.filter(
                SimproItemCadastro.versao == version_label,
                SimproItemCadastro.item_key.in_(chunk),
            ).all()
        )
    cadastro_map = {(row.versao, row.item_key): row for row in existing_cadastros}

    new_cadastros: list[dict[str, object | None]] = []
    for item_key, row in prepared_by_key.items():
        map_key = (version_label, item_key)
        cad = cadastro_map.get(map_key)
        if cad is None:
            new_cadastros.append({
                'id': next_cadastro_id,
                'versao': version_label,
                'item_key': item_key,
                'tuss_numero': row.get('tuss_numero'),
                'codigo': row.get('codigo'),
                'codigo_interno': row.get('codigo_interno'),
                'codigo_alt': row.get('codigo_alt'),
                'descricao': row.get('descricao'),
                'fabricante': row.get('fabricante'),
                'referencia': row.get('referencia'),
                'anvisa': row.get('anvisa'),
                'ean': row.get('ean'),
                'unidade': row.get('unidade'),
                'qtd_unidade': row.get('qtd_unidade'),
                'fracionavel': row.get('fracionavel'),
                'status_final': row.get('status_final'),
                'data_ref': row.get('data_ref'),
                'linha_num': row.get('linha_num'),
                'imported_at': now_ts,
            })
            next_cadastro_id += 1
            stat['cadastros_criados'] += 1
            continue

        incoming_tuss = row.get('tuss_numero')
        current_tuss = getattr(cad, 'tuss_numero', None)
        if incoming_tuss in (None, '') and current_tuss not in (None, ''):
            effective_tuss = current_tuss
        else:
            effective_tuss = incoming_tuss

        changed = False
        for field_name, field_value in (
            ('tuss_numero', effective_tuss),
            ('codigo', row.get('codigo')),
            ('codigo_interno', row.get('codigo_interno')),
            ('codigo_alt', row.get('codigo_alt')),
            ('descricao', row.get('descricao')),
            ('fabricante', row.get('fabricante')),
            ('referencia', row.get('referencia')),
            ('anvisa', row.get('anvisa')),
            ('ean', row.get('ean')),
            ('unidade', row.get('unidade')),
            ('qtd_unidade', row.get('qtd_unidade')),
            ('fracionavel', row.get('fracionavel')),
            ('status_final', row.get('status_final')),
            ('data_ref', row.get('data_ref')),
            ('linha_num', row.get('linha_num')),
        ):
            if getattr(cad, field_name) != field_value:
                setattr(cad, field_name, field_value)
                changed = True
        if changed:
            cad.imported_at = now_ts
            stat['cadastros_atualizados'] += 1

    if new_cadastros:
        db.session.bulk_insert_mappings(SimproItemCadastro, new_cadastros)

    cadastro_rows: list = []
    for i in range(0, len(identity_keys), SIMPRO_SPLIT_IN_CHUNK):
        chunk = identity_keys[i : i + SIMPRO_SPLIT_IN_CHUNK]
        cadastro_rows.extend(
            SimproItemCadastro.query.with_entities(SimproItemCadastro.id, SimproItemCadastro.item_key)
            .filter(
                SimproItemCadastro.versao == version_label,
                SimproItemCadastro.item_key.in_(chunk),
            )
            .all()
        )
    cadastro_id_by_key = {row.item_key: int(row.id) for row in cadastro_rows if row.item_key and row.id is not None}
    if not cadastro_id_by_key:
        db.session.commit()
        return stat

    cadastro_ids = list(cadastro_id_by_key.values())
    existing_precos: list[SimproItemPreco] = []
    for i in range(0, len(cadastro_ids), SIMPRO_SPLIT_IN_CHUNK):
        chunk = cadastro_ids[i : i + SIMPRO_SPLIT_IN_CHUNK]
        existing_precos.extend(
            SimproItemPreco.query.filter(
                SimproItemPreco.cadastro_id.in_(chunk),
                SimproItemPreco.aliquota == aliquota,
            ).all()
        )
    preco_by_cadastro = {int(row.cadastro_id): row for row in existing_precos if row.cadastro_id is not None}

    new_precos: list[dict[str, object | None]] = []
    for item_key, row in prepared_by_key.items():
        cadastro_id = cadastro_id_by_key.get(item_key)
        if cadastro_id is None:
            continue
        p_row = preco_by_cadastro.get(cadastro_id)
        if p_row is None:
            new_precos.append({
                'id': next_preco_id,
                'cadastro_id': cadastro_id,
                'aliquota': aliquota,
                'preco1': row.get('preco1'),
                'preco2': row.get('preco2'),
                'preco3': row.get('preco3'),
                'preco4': row.get('preco4'),
                'arquivo_fonte': arquivo_label,
                'imported_at': now_ts,
            })
            next_preco_id += 1
            stat['precos_criados'] += 1
            continue
        p_row.preco1 = row.get('preco1')
        p_row.preco2 = row.get('preco2')
        p_row.preco3 = row.get('preco3')
        p_row.preco4 = row.get('preco4')
        p_row.arquivo_fonte = arquivo_label
        p_row.imported_at = now_ts
        stat['precos_atualizados'] += 1

    if new_precos:
        db.session.bulk_insert_mappings(SimproItemPreco, new_precos)

    db.session.commit()
    return stat


def _br_aliquota_certeiro(a: Decimal | None, b: Decimal | None) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) < Decimal('0.02')


def _candidatas_edicao_bras(rotulo: str) -> list[str]:
    """
    Gera rótulos alternativos para bater `edicao` no cadastro (vem da **col. 14** do TXT),
    que muitas vezes é só `1091`, enquanto o usuário informa `2026-1091` na tela.
    """
    import re

    t = (rotulo or '').strip()
    if not t:
        return []
    out: list[str] = []
    seen: set[str] = set()

    def add(x: str) -> None:
        v = (x or '').strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)

    add(t)
    if '-' in t:
        tail = t.rsplit('-', 1)[-1].strip()
        if tail and tail != t and re.match(r'^\d{3,5}$', tail):
            add(tail)
    return out


def _align_bras_n_import_version(
    versao: str | None,
    *,
    arquivo_label: str | None = None,
) -> int:
    """
    Faz a versão importada prevalecer em `bras_item_n.edicao`.

    A Brasíndice pode trazer a col. 14 com edições antigas dentro do mesmo arquivo publicado.
    Para o fluxo operacional do sistema, a chave do split deve seguir a versão importada
    (ex.: `1091`), mantendo o valor original apenas implicitamente no arquivo fonte.
    """
    versao_norm = (versao or '').strip()
    if not versao_norm:
        return 0

    query = BrasItemNormalized.query
    if arquivo_label:
        query = query.filter(BrasItemNormalized.arquivo == arquivo_label)
    else:
        version_tokens = _candidatas_edicao_bras(versao_norm) or [versao_norm]
        like_clauses = [
            func.coalesce(BrasItemNormalized.arquivo, '').like(f'%{token}%')
            for token in version_tokens
            if token
        ]
        if like_clauses:
            query = query.filter(or_(*like_clauses))
        else:
            return 0

    rows = query.with_entities(BrasItemNormalized.id, BrasItemNormalized.edicao).all()
    touched_ids = [int(row.id) for row in rows if (row.edicao or '').strip() != versao_norm]
    if not touched_ids:
        return 0

    (
        BrasItemNormalized.query
        .filter(BrasItemNormalized.id.in_(touched_ids))
        .update({'edicao': versao_norm}, synchronize_session=False)
    )
    db.session.commit()
    return len(touched_ids)


def _patch_insumo_from_bras_item_n(n: BrasItemNormalized) -> None:
    """Atualiza linha de `insumos_index` a partir de um `bras_item_n` (após ajuste de preço leve)."""
    alq_key = _br_norm_aliquota(n.aliquota_ou_ipi) if n.aliquota_ou_ipi is not None else Decimal('0')
    idx = InsumoIndex.query.filter_by(origem='BRAS', item_id=n.id, aliquota=alq_key).first()
    if not idx:
        return
    preco = n.preco_pmc_unit or n.preco_pmc_pacote or n.preco_pfb_unit or n.preco_pfb_pacote
    idx.preco = preco
    idx.aliquota = n.aliquota_ou_ipi
    idx.tuss = n.produto_codigo
    idx.tiss = n.apresentacao_codigo
    idx.anvisa = n.registro_anvisa
    idx.fabricante = n.laboratorio_nome
    idx.versao_tabela = n.edicao or n.arquivo
    p1 = (n.produto_nome or '').strip()
    p2 = (n.apresentacao_descricao or '').strip()
    idx.descricao = ' • '.join([p for p in (p1, p2) if p]) or None
    idx.updated_at = _now_utc()
    db.session.add(idx)


def _ufs_pertencentes_a_aliquota_piso(alq: Decimal) -> list[str]:
    """
    UFs atribuídas à alíquota na tabela piso (ANVISA), inverso do mapa UF→alíquota.
    Usado para preencher `insumos_index.uf_referencia` após import por alíquota.
    """
    aln = _br_norm_aliquota(alq) or alq
    out: set[str] = set()
    for a_str, ufs in _SIMPRO_ALIQUOTA_UF_ROWS:
        a = _br_norm_aliquota(Decimal(a_str)) or Decimal(a_str)
        if _br_aliquota_certeiro(a, aln):
            out.update(ufs)
    return sorted(out)


def _reatribuir_insumo_bras_apos_import_precos(
    n_ids: set[int],
    aliquota: Decimal,
    versao_label: str | None = None,
) -> int:
    """
    Após `import-precos`, alinha `insumos_index` com a alíquota do ficheiro e com as UFs do piso,
    e com preço vindo de `bras_item_preco` (ou `bras_item_n`). Cria linha de índice se ainda não existir.
    """
    if not n_ids:
        return 0
    alq = _br_norm_aliquota(aliquota) or aliquota
    ufs = _ufs_pertencentes_a_aliquota_piso(alq)
    uf_str = _encode_uf_codes(ufs)
    n_atual = 0
    for nid in n_ids:
        n = BrasItemNormalized.query.get(nid)
        if not n:
            continue
        ean = (n.ean or '').strip()
        if not ean:
            continue
        ed = (n.edicao or '').strip() or (n.arquivo or '')[:50] or '—'
        ed = (versao_label or '').strip() or ed
        cands = _candidatas_edicao_bras(ed)
        cad = (
            BrasItemCadastro.query.filter(
                and_(
                    func.trim(BrasItemCadastro.ean) == ean,
                    or_(*[func.trim(BrasItemCadastro.edicao) == c for c in cands]),
                )
            )
            .first()
        )
        p_row: BrasItemPreco | None = None
        if cad is not None:
            p_row = (
                BrasItemPreco.query.filter(BrasItemPreco.cadastro_id == cad.id)
                .filter(BrasItemPreco.aliquota == alq)
                .first()
            )
        preco = None
        if p_row is not None:
            preco = p_row.preco_pmc_unit or p_row.preco_pmc_pacote or p_row.preco_pfb_unit or p_row.preco_pfb_pacote
        if preco is None:
            preco = n.preco_pmc_unit or n.preco_pmc_pacote or n.preco_pfb_unit or n.preco_pfb_pacote
        p1 = (n.produto_nome or '').strip()
        p2 = (n.apresentacao_descricao or '').strip()
        desc = ' • '.join([x for x in (p1, p2) if x]) or None
        vtab_source = (versao_label or '').strip() or n.edicao or n.arquivo
        vtab = vtab_source[:100] if vtab_source else None
        idx = InsumoIndex.query.filter_by(origem='BRAS', item_id=nid, aliquota=alq).first()
        if idx is None:
            idx = InsumoIndex(
                origem='BRAS',
                item_id=int(nid),
                tuss=n.produto_codigo,
                tiss=n.apresentacao_codigo,
                descricao=desc,
                preco=preco,
                aliquota=alq,
                fabricante=n.laboratorio_nome,
                anvisa=n.registro_anvisa,
                versao_tabela=vtab,
                data_atualizacao=None,
                uf_referencia=uf_str,
                updated_at=_now_utc(),
            )
            db.session.add(idx)
        else:
            idx.aliquota = alq
            idx.uf_referencia = uf_str
            if preco is not None:
                idx.preco = preco
            idx.tuss = n.produto_codigo
            idx.tiss = n.apresentacao_codigo
            idx.descricao = desc
            idx.fabricante = n.laboratorio_nome
            idx.anvisa = n.registro_anvisa
            idx.versao_tabela = vtab
            idx.updated_at = _now_utc()
        n_atual += 1
    if n_atual:
        db.session.commit()
    return n_atual


def _autobackfill_bras_cadastro_se_necessario() -> dict | None:
    """
    Se `bras_item_n` tem dados mas `bras_item_cadastro` está vazio, o backfill antigo não criava
    nada (alíquota nula no `n`). Garante uma passada de preenchimento de cadastro antes de `import-precos`.
    """
    try:
        n_n = int(db.session.query(func.count(BrasItemNormalized.id)).scalar() or 0)
        n_c = int(db.session.query(func.count(BrasItemCadastro.id)).scalar() or 0)
    except Exception as exc:  # noqa: BLE001
        app.logger.warning('autobackfill BRAS: contagem: %s', exc)
        return None
    if n_n == 0 or n_c > 0:
        return None
    st = _sync_bras_split_from_bras_n_fast()
    app.logger.info('autobackfill bras_item_cadastro a partir de bras_item_n: %s', st)
    return st


def _autobackfill_bras_cadastro_para_edicao_se_necessario(edicao: str) -> dict | None:
    """
    Garante que o split BRAS tenha o cadastro da ediÃ§Ã£o alvo antes de uma carga "sÃ³ preÃ§os".

    Caso comum: o banco jÃ¡ possui alguns cadastros antigos, entÃ£o o autobackfill global nÃ£o dispara,
    mas a ediÃ§Ã£o atual ainda nÃ£o foi sincronizada para `bras_item_cadastro`.
    """
    edicao_cands = _candidatas_edicao_bras((edicao or '').strip())
    if not edicao_cands:
        return None

    n_rows = (
        BrasItemNormalized.query.with_entities(BrasItemNormalized.id, BrasItemNormalized.ean)
        .filter(
            BrasItemNormalized.ean.isnot(None),
            func.trim(BrasItemNormalized.ean) != '',
            or_(*[func.coalesce(BrasItemNormalized.edicao, '') == c for c in edicao_cands]),
        )
        .all()
    )
    if not n_rows:
        return None

    n_ids = [int(row.id) for row in n_rows if row.id is not None]
    n_eans = {(row.ean or '').strip() for row in n_rows if (row.ean or '').strip()}
    if not n_ids or not n_eans:
        return None

    cad_rows = (
        BrasItemCadastro.query.with_entities(BrasItemCadastro.ean)
        .filter(
            BrasItemCadastro.ean.isnot(None),
            func.trim(BrasItemCadastro.ean) != '',
            or_(*[func.trim(BrasItemCadastro.edicao) == c for c in edicao_cands]),
        )
        .all()
    )
    cad_eans = {(row.ean or '').strip() for row in cad_rows if (row.ean or '').strip()}
    missing_eans = n_eans - cad_eans
    if not missing_eans:
        return None

    st = _sync_bras_split_from_bras_n_fast(n_ids=n_ids)
    st['edicao_candidatas'] = edicao_cands
    st['cadastros_faltantes_antes'] = len(missing_eans)
    app.logger.info('autobackfill bras_item_cadastro para edição %s: %s', edicao, st)
    return st


def _import_bras_somente_precos(
    *,
    file_path: Path,
    edicao: str,
    aliquota: Decimal,
    delimiter: str = ',',
    quotechar: str = '"',
    encoding: str | None = 'latin-1',
    skip_header: bool = False,
    arquivo_fonte: str = 'import-precos',
    update_legacy: bool = True,
    commit_cada: int = 500,
) -> dict:
    """
    Atualiza só preços (mesmo layout TXT D delimitado) para itens **já** presentes em `BrasItemCadastro`.
    Também replica em `bras_item_n` (linhas cujo EAN+edição+alíquota batem) e em `insumos_index` quando possível.
    """
    import csv

    aliquota = _br_norm_aliquota(aliquota) or aliquota
    edicao = (edicao or '').strip()
    if not edicao:
        raise ValueError('edição/versão é obrigatória.')
    if aliquota is None:
        raise ValueError('alíquota é obrigatória.')
    aligned_rows = _align_bras_n_import_version(edicao)
    edicao_cands = _candidatas_edicao_bras(edicao)

    autobf = _autobackfill_bras_cadastro_se_necessario()
    autobf_edicao = _autobackfill_bras_cadastro_para_edicao_se_necessario(edicao)
    enc = encoding or 'latin-1'
    stat = {
        'linhas_lidas': 0,
        'atualizados_preco': 0,
        'sem_cadastro': 0,
        'erros': 0,
        'legacy_atualizados': 0,
        'edicao_candidatas': edicao_cands,
        'bras_n_aligned': aligned_rows,
        'autobackfill_cadastro': autobf,
        'autobackfill_edicao': autobf_edicao,
    }
    next_preco_id = int(db.session.query(func.max(BrasItemPreco.id)).scalar() or 0) + 1
    n_ids_tocados: set[int] = set()
    n_ids_para_insumo_index: set[int] = set()
    price_by_ean: dict[str, dict[str, Decimal | None]] = {}
    eans_in_file: set[str] = set()

    with file_path.open('r', encoding=enc, errors='replace', newline='') as f:
        reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar or '"')
        for linha_num, row in enumerate(reader, start=1):
            _raise_if_import_paused()
            if skip_header and linha_num == 1:
                continue
            if len(row) < 17:
                continue
            try:
                ean = (row[16] or '').strip()
                if not ean:
                    continue
                stat['linhas_lidas'] += 1
                eans_in_file.add(ean)
                price_by_ean[ean] = {
                    'pmc_p': Decimal((row[6] or '0').replace(',', '.')) if (row[6] or '').strip() else None,
                    'pfb_p': Decimal((row[7] or '0').replace(',', '.')) if (row[7] or '').strip() else None,
                    'pmc_u': Decimal((row[10] or '0').replace(',', '.')) if (row[10] or '').strip() else None,
                    'pfb_u': Decimal((row[12] or '0').replace(',', '.')) if (row[12] or '').strip() else None,
                }
            except (InvalidOperation, ValueError, TypeError) as exc:
                stat['erros'] += 1
                app.logger.debug('import precos leve linha %s: %s', linha_num, exc)
                continue

    if eans_in_file:
        cad_rows = (
            BrasItemCadastro.query.filter(
                BrasItemCadastro.ean.in_(list(eans_in_file)),
                BrasItemCadastro.edicao.in_(edicao_cands),
            )
            .all()
        )
        cadastro_by_ean = {(row.ean or '').strip(): row for row in cad_rows if (row.ean or '').strip()}
        cadastro_ids = [int(row.id) for row in cad_rows if row.id is not None]
        preco_rows = (
            BrasItemPreco.query.filter(
                BrasItemPreco.cadastro_id.in_(cadastro_ids),
                BrasItemPreco.aliquota == aliquota,
            )
            .all()
            if cadastro_ids else []
        )
        preco_by_cadastro = {int(row.cadastro_id): row for row in preco_rows if row.cadastro_id is not None}
        normalized_rows = (
            BrasItemNormalized.query.filter(
                BrasItemNormalized.ean.in_(list(eans_in_file)),
                BrasItemNormalized.edicao.in_(edicao_cands),
            )
            .all()
        )
        normalized_by_ean: dict[str, list[BrasItemNormalized]] = {}
        for row in normalized_rows:
            ean = (row.ean or '').strip()
            if ean:
                normalized_by_ean.setdefault(ean, []).append(row)

        now_ts = _now_utc()
        new_precos: list[BrasItemPreco] = []
        for ean, payload in price_by_ean.items():
            _raise_if_import_paused()
            cad = cadastro_by_ean.get(ean)
            if cad is None:
                stat['sem_cadastro'] += 1
                continue

            p_row = preco_by_cadastro.get(int(cad.id))
            if p_row is None:
                p_row = BrasItemPreco(
                    id=next_preco_id,
                    cadastro_id=cad.id,
                    aliquota=aliquota,
                    arquivo_fonte=arquivo_fonte,
                    imported_at=now_ts,
                )
                next_preco_id += 1
                preco_by_cadastro[int(cad.id)] = p_row
                new_precos.append(p_row)

            p_row.preco_pmc_pacote = payload['pmc_p']
            p_row.preco_pfb_pacote = payload['pfb_p']
            p_row.preco_pmc_unit = payload['pmc_u']
            p_row.preco_pfb_unit = payload['pfb_u']
            p_row.arquivo_fonte = arquivo_fonte
            p_row.imported_at = now_ts
            stat['atualizados_preco'] += 1

            n_match = normalized_by_ean.get(ean, [])
            for n_row in n_match:
                n_ids_para_insumo_index.add(int(n_row.id))

            if update_legacy:
                for n in n_match:
                    if n is None or (n.aliquota_ou_ipi is not None and not _br_aliquota_certeiro(n.aliquota_ou_ipi, aliquota)):
                        continue
                    n.preco_pmc_pacote = payload['pmc_p']
                    n.preco_pfb_pacote = payload['pfb_p']
                    n.preco_pmc_unit = payload['pmc_u']
                    n.preco_pfb_unit = payload['pfb_u']
                    n.aliquota_ou_ipi = aliquota
                    stat['legacy_atualizados'] += 1
                    n_ids_tocados.add(int(n.id))

        if new_precos:
            db.session.add_all(new_precos)

    db.session.commit()
    n_idx = _reatribuir_insumo_bras_apos_import_precos(
        n_ids_para_insumo_index,
        aliquota,
        versao_label=edicao,
    )
    stat['insumos_index_vinculados'] = n_idx
    if n_ids_tocados or n_ids_para_insumo_index:
        _clear_insumo_cache()
    return stat


def _backfill_bras_cadastro_preco_from_bras_n(*, commit_cada: int = 2000, dry_run: bool = False) -> dict:
    """
    A partir de `bras_item_n`, cria/ajusta `BrasItemCadastro` (dedup por edição+EAN) e `BrasItemPreco` por alíquota.
    A view `bras_item_v` e o pipeline raw deixam `aliquota_ou_ipi` nulo; mesmo assim o **cadastro** é criado,
    e `BrasItemPreco` só recebe linha quando há alíquota (senão preços vêm de `import-precos` / delta).
    """
    stat = {
        'linhas_bras_n': 0,
        'cadastros_unicos': 0,
        'preco_linhas': 0,
        'puladas_sem_ean': 0,
        'n_sem_aliquota_so_cadastro': 0,
    }
    next_cadastro_id = int(db.session.query(func.max(BrasItemCadastro.id)).scalar() or 0) + 1
    next_preco_id = int(db.session.query(func.max(BrasItemPreco.id)).scalar() or 0) + 1
    n_desde = 0
    q = BrasItemNormalized.query.order_by(BrasItemNormalized.id.asc())
    for row in q.yield_per(2000):
        stat['linhas_bras_n'] += 1
        ean = (row.ean or '').strip()
        if not ean:
            stat['puladas_sem_ean'] += 1
            continue
        ed = (row.edicao or '').strip() or (row.arquivo or '')[:50]
        if not ed:
            ed = (row.arquivo or '')[:50] or '—'

        cad: BrasItemCadastro | None = None
        if not dry_run:
            cad = (
                BrasItemCadastro.query.filter(
                    and_(
                        func.trim(BrasItemCadastro.edicao) == ed,
                        func.trim(BrasItemCadastro.ean) == ean,
                    )
                )
                .first()
            )
            if cad is None:
                cad = BrasItemCadastro(
                    edicao=ed,
                    ean=ean,
                    laboratorio_codigo=row.laboratorio_codigo,
                    laboratorio_nome=row.laboratorio_nome,
                    produto_codigo=row.produto_codigo,
                    produto_nome=row.produto_nome,
                    apresentacao_codigo=row.apresentacao_codigo,
                    apresentacao_descricao=row.apresentacao_descricao,
                    registro_anvisa=row.registro_anvisa,
                    quantidade_embalagem=row.quantidade_embalagem,
                    linha_num=row.linha_num,
                    imported_at=_now_utc(),
                )
                db.session.add(cad)
                db.session.flush()
                stat['cadastros_unicos'] += 1
            n_desde += 1
            if n_desde >= commit_cada:
                db.session.commit()
                n_desde = 0

        if row.aliquota_ou_ipi is None:
            stat['n_sem_aliquota_so_cadastro'] += 1
            continue

        if dry_run:
            continue

        if not cad:
            continue
        ali = _br_norm_aliquota(row.aliquota_ou_ipi)
        p_row = (
            BrasItemPreco.query.filter(BrasItemPreco.cadastro_id == cad.id)
            .filter(BrasItemPreco.aliquota == ali)
            .first()
        )
        if p_row is None:
            p_row = BrasItemPreco(
                id=next_preco_id,
                cadastro_id=cad.id,
                aliquota=ali,
                arquivo_fonte=row.arquivo,
                imported_at=row.imported_at or _now_utc(),
            )
            next_preco_id += 1
            db.session.add(p_row)
        p_row.preco_pmc_pacote = row.preco_pmc_pacote
        p_row.preco_pfb_pacote = row.preco_pfb_pacote
        p_row.preco_pmc_unit = row.preco_pmc_unit
        p_row.preco_pfb_unit = row.preco_pfb_unit
        p_row.arquivo_fonte = row.arquivo
        stat['preco_linhas'] += 1
        n_desde += 1
        if n_desde >= commit_cada:
            db.session.commit()
            n_desde = 0

    if not dry_run:
        db.session.commit()
    return stat


def _sync_bras_split_from_bras_n(
    *,
    arquivo_label: str | None = None,
    n_ids: Sequence[int] | None = None,
    aliquota_override: Decimal | None = None,
    commit_cada: int = 1000,
) -> dict:
    """
    Sincroniza o split BRAS (`bras_item_cadastro` + `bras_item_preco`) a partir de linhas jÃ¡ materializadas
    em `bras_item_n`, limitado por arquivo ou por IDs tocados.

    Fecha o ciclo operacional da arquitetura nova sem depender de backfill manual apÃ³s a carga base/delta.
    """
    stat = {
        'linhas_bras_n': 0,
        'cadastros_criados': 0,
        'cadastros_atualizados': 0,
        'precos_criados': 0,
        'precos_atualizados': 0,
        'puladas_sem_ean': 0,
        'puladas_sem_aliquota': 0,
    }
    if n_ids is not None and not list(n_ids):
        return stat

    normalized_override = None
    if aliquota_override is not None:
        normalized_override = _br_norm_aliquota(aliquota_override) or aliquota_override

    query = BrasItemNormalized.query
    if arquivo_label:
        query = query.filter(BrasItemNormalized.arquivo == arquivo_label)
    if n_ids is not None:
        query = query.filter(BrasItemNormalized.id.in_(list(n_ids)))
    query = query.order_by(BrasItemNormalized.id.asc())

    next_cadastro_id = int(db.session.query(func.max(BrasItemCadastro.id)).scalar() or 0) + 1
    next_preco_id = int(db.session.query(func.max(BrasItemPreco.id)).scalar() or 0) + 1
    cadastro_cache: dict[tuple[str, str], BrasItemCadastro] = {}
    preco_cache: dict[tuple[int, str], BrasItemPreco] = {}
    n_desde = 0

    rows = query.all()
    for row in rows:
        stat['linhas_bras_n'] += 1
        ean = (row.ean or '').strip()
        if not ean:
            stat['puladas_sem_ean'] += 1
            continue

        ed = (row.edicao or '').strip() or (row.arquivo or '')[:50] or 'â€”'
        cad_key = (ed, ean)
        cad = cadastro_cache.get(cad_key)
        if cad is None:
            cad = (
                BrasItemCadastro.query.filter(
                    and_(
                        func.trim(BrasItemCadastro.edicao) == ed,
                        func.trim(BrasItemCadastro.ean) == ean,
                    )
                )
                .first()
            )
            if cad is None:
                cad = BrasItemCadastro(
                    id=next_cadastro_id,
                    edicao=ed,
                    ean=ean,
                    laboratorio_codigo=row.laboratorio_codigo,
                    laboratorio_nome=row.laboratorio_nome,
                    produto_codigo=row.produto_codigo,
                    produto_nome=row.produto_nome,
                    apresentacao_codigo=row.apresentacao_codigo,
                    apresentacao_descricao=row.apresentacao_descricao,
                    registro_anvisa=row.registro_anvisa,
                    quantidade_embalagem=row.quantidade_embalagem,
                    linha_num=row.linha_num,
                    imported_at=_now_utc(),
                )
                next_cadastro_id += 1
                db.session.add(cad)
                db.session.flush()
                stat['cadastros_criados'] += 1
            cadastro_cache[cad_key] = cad

        cadastro_changed = False
        cadastro_fields = {
            'laboratorio_codigo': row.laboratorio_codigo,
            'laboratorio_nome': row.laboratorio_nome,
            'produto_codigo': row.produto_codigo,
            'produto_nome': row.produto_nome,
            'apresentacao_codigo': row.apresentacao_codigo,
            'apresentacao_descricao': row.apresentacao_descricao,
            'registro_anvisa': row.registro_anvisa,
            'quantidade_embalagem': row.quantidade_embalagem,
            'linha_num': row.linha_num,
        }
        for field_name, field_value in cadastro_fields.items():
            if getattr(cad, field_name) != field_value:
                setattr(cad, field_name, field_value)
                cadastro_changed = True
        if cadastro_changed:
            cad.imported_at = _now_utc()
            stat['cadastros_atualizados'] += 1

        alq = normalized_override
        if alq is None and row.aliquota_ou_ipi is not None:
            alq = _br_norm_aliquota(row.aliquota_ou_ipi) or row.aliquota_ou_ipi
        if alq is None:
            stat['puladas_sem_aliquota'] += 1
            n_desde += 1
            if n_desde >= commit_cada:
                db.session.commit()
                n_desde = 0
            continue

        preco_key = (int(cad.id), format(alq, 'f'))
        preco_row = preco_cache.get(preco_key)
        if preco_row is None:
            preco_row = (
                BrasItemPreco.query.filter(BrasItemPreco.cadastro_id == cad.id)
                .filter(BrasItemPreco.aliquota == alq)
                .first()
            )
            if preco_row is None:
                preco_row = BrasItemPreco(
                    id=next_preco_id,
                    cadastro_id=cad.id,
                    aliquota=alq,
                    arquivo_fonte=row.arquivo,
                    imported_at=_now_utc(),
                )
                next_preco_id += 1
                db.session.add(preco_row)
                stat['precos_criados'] += 1
            else:
                stat['precos_atualizados'] += 1
            preco_cache[preco_key] = preco_row
        else:
            stat['precos_atualizados'] += 1

        preco_row.preco_pmc_pacote = row.preco_pmc_pacote
        preco_row.preco_pfb_pacote = row.preco_pfb_pacote
        preco_row.preco_pmc_unit = row.preco_pmc_unit
        preco_row.preco_pfb_unit = row.preco_pfb_unit
        preco_row.arquivo_fonte = row.arquivo
        preco_row.imported_at = _now_utc()

        n_desde += 1
        if n_desde >= commit_cada:
            db.session.commit()
            n_desde = 0

    db.session.commit()
    return stat


def _sync_bras_split_from_bras_n_fast(
    *,
    arquivo_label: str | None = None,
    n_ids: Sequence[int] | None = None,
    aliquota_override: Decimal | None = None,
    versao_override: str | None = None,
) -> dict:
    """
    Variante otimizada do sync BRAS split.
    Evita SELECT por linha em `bras_item_cadastro`/`bras_item_preco` e trabalha em lote.
    """
    stat = {
        'linhas_bras_n': 0,
        'cadastros_criados': 0,
        'cadastros_atualizados': 0,
        'precos_criados': 0,
        'precos_atualizados': 0,
        'puladas_sem_ean': 0,
        'puladas_sem_aliquota': 0,
    }
    if n_ids is not None and not list(n_ids):
        return stat

    normalized_override = None
    if aliquota_override is not None:
        normalized_override = _br_norm_aliquota(aliquota_override) or aliquota_override

    query = BrasItemNormalized.query
    if arquivo_label:
        query = query.filter(BrasItemNormalized.arquivo == arquivo_label)
    if n_ids is not None:
        query = query.filter(BrasItemNormalized.id.in_(list(n_ids)))
    rows = query.order_by(BrasItemNormalized.id.asc()).all()
    if not rows:
        return stat

    stat['linhas_bras_n'] = len(rows)
    now_ts = _now_utc()
    next_cadastro_id = int(db.session.query(func.max(BrasItemCadastro.id)).scalar() or 0) + 1
    next_preco_id = int(db.session.query(func.max(BrasItemPreco.id)).scalar() or 0) + 1

    prepared_rows: list[tuple[BrasItemNormalized, str, str, Decimal | None]] = []
    eans: set[str] = set()
    edicoes: set[str] = set()
    for row in rows:
        ean = (row.ean or '').strip()
        if not ean:
            stat['puladas_sem_ean'] += 1
            continue
        ed = (versao_override or '').strip() or (row.edicao or '').strip() or (row.arquivo or '')[:50] or '-'
        alq = normalized_override
        if alq is None and row.aliquota_ou_ipi is not None:
            alq = _br_norm_aliquota(row.aliquota_ou_ipi) or row.aliquota_ou_ipi
        if alq is None:
            stat['puladas_sem_aliquota'] += 1
        prepared_rows.append((row, ed, ean, alq))
        eans.add(ean)
        edicoes.add(ed)

    existing_cadastros = (
        BrasItemCadastro.query.filter(
            BrasItemCadastro.ean.in_(list(eans)),
            BrasItemCadastro.edicao.in_(list(edicoes)),
        ).all()
        if eans and edicoes else []
    )
    cadastro_map = {
        ((cad.edicao or '').strip(), (cad.ean or '').strip()): cad
        for cad in existing_cadastros
        if (cad.edicao or '').strip() and (cad.ean or '').strip()
    }

    new_cadastros: list[dict[str, object]] = []
    pending_cadastro_keys: set[tuple[str, str]] = set()
    for row, ed, ean, _alq in prepared_rows:
        cadastro_key = (ed, ean)
        cad = cadastro_map.get(cadastro_key)
        if cad is None:
            if cadastro_key in pending_cadastro_keys:
                continue
            mapping = {
                'id': next_cadastro_id,
                'edicao': ed,
                'ean': ean,
                'laboratorio_codigo': row.laboratorio_codigo,
                'laboratorio_nome': row.laboratorio_nome,
                'produto_codigo': row.produto_codigo,
                'produto_nome': row.produto_nome,
                'apresentacao_codigo': row.apresentacao_codigo,
                'apresentacao_descricao': row.apresentacao_descricao,
                'registro_anvisa': row.registro_anvisa,
                'quantidade_embalagem': row.quantidade_embalagem,
                'linha_num': row.linha_num,
                'imported_at': now_ts,
            }
            new_cadastros.append(mapping)
            pending_cadastro_keys.add(cadastro_key)
            next_cadastro_id += 1
            stat['cadastros_criados'] += 1
            continue

        changed = False
        for field_name, field_value in (
            ('laboratorio_codigo', row.laboratorio_codigo),
            ('laboratorio_nome', row.laboratorio_nome),
            ('produto_codigo', row.produto_codigo),
            ('produto_nome', row.produto_nome),
            ('apresentacao_codigo', row.apresentacao_codigo),
            ('apresentacao_descricao', row.apresentacao_descricao),
            ('registro_anvisa', row.registro_anvisa),
            ('quantidade_embalagem', row.quantidade_embalagem),
            ('linha_num', row.linha_num),
        ):
            if getattr(cad, field_name) != field_value:
                setattr(cad, field_name, field_value)
                changed = True
        if changed:
            cad.imported_at = now_ts
            stat['cadastros_atualizados'] += 1

    if new_cadastros:
        db.session.bulk_insert_mappings(BrasItemCadastro, new_cadastros)
        db.session.commit()
        existing_cadastros = BrasItemCadastro.query.filter(
            BrasItemCadastro.ean.in_(list(eans)),
            BrasItemCadastro.edicao.in_(list(edicoes)),
        ).all()
        cadastro_map = {
            ((cad.edicao or '').strip(), (cad.ean or '').strip()): cad
            for cad in existing_cadastros
            if (cad.edicao or '').strip() and (cad.ean or '').strip()
        }
    else:
        db.session.flush()

    cadastro_ids = [int(cad.id) for cad in cadastro_map.values() if cad.id is not None]
    existing_precos = (
        BrasItemPreco.query.filter(BrasItemPreco.cadastro_id.in_(cadastro_ids)).all()
        if cadastro_ids else []
    )
    preco_map = {
        (int(preco.cadastro_id), format((_br_norm_aliquota(preco.aliquota) or preco.aliquota), 'f')): preco
        for preco in existing_precos
        if preco.cadastro_id is not None and preco.aliquota is not None
    }

    new_precos: list[dict[str, object]] = []
    pending_preco_keys: set[tuple[int, str]] = set()
    for row, ed, ean, alq in prepared_rows:
        if alq is None:
            continue
        cad = cadastro_map.get((ed, ean))
        if cad is None or cad.id is None:
            continue
        preco_key = (int(cad.id), format(alq, 'f'))
        preco_row = preco_map.get(preco_key)
        if preco_row is None:
            if preco_key in pending_preco_keys:
                continue
            new_precos.append({
                'id': next_preco_id,
                'cadastro_id': cad.id,
                'aliquota': alq,
                'preco_pmc_pacote': row.preco_pmc_pacote,
                'preco_pfb_pacote': row.preco_pfb_pacote,
                'preco_pmc_unit': row.preco_pmc_unit,
                'preco_pfb_unit': row.preco_pfb_unit,
                'arquivo_fonte': row.arquivo,
                'imported_at': now_ts,
            })
            pending_preco_keys.add(preco_key)
            next_preco_id += 1
            stat['precos_criados'] += 1
            continue

        preco_row.preco_pmc_pacote = row.preco_pmc_pacote
        preco_row.preco_pfb_pacote = row.preco_pfb_pacote
        preco_row.preco_pmc_unit = row.preco_pmc_unit
        preco_row.preco_pfb_unit = row.preco_pfb_unit
        preco_row.arquivo_fonte = row.arquivo
        preco_row.imported_at = now_ts
        stat['precos_atualizados'] += 1

    if new_precos:
        db.session.bulk_insert_mappings(BrasItemPreco, new_precos)
    db.session.commit()
    return stat


def _simpro_insumo_index_upsert_sql_template() -> str:
    """SQL INSERT…SELECT para popular ``insumos_index`` a partir de ``simpro_item_preco`` (inclui ``{where_clause}``)."""
    preco_expr = (
        "COALESCE("
        "NULLIF(p.preco2, 0), NULLIF(p.preco1, 0), NULLIF(p.preco3, 0), NULLIF(p.preco4, 0), "
        "p.preco2, p.preco1, p.preco3, p.preco4)"
    )
    preco_sql = _sql_clamp_decimal(preco_expr)
    aliquota_sql = _sql_clamp_decimal("p.aliquota", integer_digits=4, scale=4)
    descricao_expr = "TRIM(COALESCE(c.descricao, ''))"
    uf_from_preco_aliquota = _sql_insumo_uf_referencia_from_aliquota_column('p.aliquota')

    return (
        """
        INSERT INTO insumos_index (
            origem, item_id, tuss, tiss, descricao, preco, aliquota,
            fabricante, anvisa, versao_tabela, data_atualizacao,
            uf_referencia, updated_at
        )
        SELECT
            'SIMPRO' AS origem,
            c.id AS item_id,
            COALESCE(
                NULLIF(c.tuss_numero, ''),
                NULLIF(c.codigo, ''),
                c.item_key
            ) AS tuss,
            c.codigo_alt AS tiss,
            {descricao_expr} AS descricao,
            {preco_sql} AS preco,
            {aliquota_sql} AS aliquota,
            c.fabricante AS fabricante,
            c.anvisa AS anvisa,
            c.versao AS versao_tabela,
            c.data_ref AS data_atualizacao,
            COALESCE(:uf_storage, :uf_default,
            {uf_from_preco_aliquota}
            ) AS uf_referencia,
            NOW() AS updated_at
        FROM simpro_item_preco p
        INNER JOIN simpro_item_cadastro c ON c.id = p.cadastro_id
        {where_clause}
        ON DUPLICATE KEY UPDATE
            tuss = VALUES(tuss),
            tiss = VALUES(tiss),
            descricao = VALUES(descricao),
            preco = VALUES(preco),
            aliquota = VALUES(aliquota),
            fabricante = VALUES(fabricante),
            anvisa = VALUES(anvisa),
            versao_tabela = VALUES(versao_tabela),
            data_atualizacao = VALUES(data_atualizacao),
            uf_referencia = VALUES(uf_referencia),
            updated_at = VALUES(updated_at)
        """
        .replace('{preco_sql}', preco_sql)
        .replace('{aliquota_sql}', aliquota_sql)
        .replace('{descricao_expr}', descricao_expr)
        .replace('{uf_from_preco_aliquota}', uf_from_preco_aliquota)
    )


def _sync_simpro_insumo_index_for_preco_ids(preco_ids: Sequence[int]) -> None:
    """Atualiza ``insumos_index`` só para ``simpro_item_preco.id`` informados (após preço manual ou patch)."""
    ids = sorted({int(x) for x in preco_ids if x is not None})
    if not ids:
        return
    params_base: dict[str, object] = {'uf_default': None, 'uf_storage': None}
    tpl = _simpro_insumo_index_upsert_sql_template()
    chunk = 800
    for i in range(0, len(ids), chunk):
        part = ids[i : i + chunk]
        wc = 'WHERE p.id IN (' + ','.join(str(x) for x in part) + ')'
        db.session.execute(text(tpl.replace('{where_clause}', wc)), params_base)
        db.session.commit()


def _sync_simpro_insumo_index(
    arquivo_label: str | None,
    *,
    uf_default: str | None = None,
    uf_values: Sequence[str] | None = None,
    aliquota_default: Decimal | None = None,
) -> None:
    """Replica linhas de ``simpro_item_preco`` em ``insumos_index`` em lotes (evita INSERT único enorme)."""
    target_ufs = list(dict.fromkeys([*(uf_values or []), *( [uf_default] if uf_default else [] )]))
    uf_codes = _normalize_uf_codes(target_ufs, uf_default=uf_default)
    if not uf_codes and aliquota_default is not None:
        uf_codes = _ufs_pertencentes_a_aliquota_piso(aliquota_default)
        if uf_codes and not uf_default:
            uf_default = uf_codes[0]
    uf_storage = _encode_uf_codes(uf_codes)

    params_base: dict[str, object] = {
        'uf_default': uf_default,
        'uf_storage': uf_storage,
    }

    upsert_sql = _simpro_insumo_index_upsert_sql_template()

    arquivo_strip = (arquivo_label or '').strip() or None
    last_id = 0
    batch_num = 0

    while True:
        id_parts = ['SELECT p.id FROM simpro_item_preco p WHERE ']
        id_params: dict[str, object] = {'last_id': last_id, 'bs': SIMPRO_INDEX_SYNC_BATCH}
        if arquivo_strip:
            id_parts.append('p.arquivo_fonte = :arquivo AND ')
            id_params['arquivo'] = arquivo_strip
        id_parts.append('p.id > :last_id ORDER BY p.id ASC LIMIT :bs')
        id_rows = db.session.execute(text(''.join(id_parts)), id_params).fetchall()
        if not id_rows:
            break

        ids = [int(r[0]) for r in id_rows]
        last_id = ids[-1]
        batch_num += 1

        ids_sql = ','.join(str(i) for i in ids)
        where_clause = f'WHERE p.id IN ({ids_sql})'
        upsert_template = text(upsert_sql.replace('{where_clause}', where_clause))

        db.session.execute(upsert_template, params_base)
        db.session.commit()

        app.logger.info(
            'SIMPRO insumos_index lote %s (%s linhas preço, último id=%s)',
            batch_num,
            len(ids),
            last_id,
        )
        _bras_throttle_between_batches()

    if batch_num == 0:
        app.logger.info('SIMPRO insumos_index: nenhuma linha em simpro_item_preco para sincronizar.')


def _backfill_catalogo_simpro_identifiers() -> None:
    try:
        db.session.execute(text("""
            UPDATE mv_catalogo_vigente_simpro c
            JOIN simpro_item_norm n ON n.id = c.item_id
            SET
                c.codigo_interno = n.codigo_interno,
                c.tuss_numero = n.tuss_numero,
                c.referencia = n.referencia,
                c.fracionavel = n.fracionavel,
                c.status_final = n.status_final
            WHERE
                (c.codigo_interno IS NULL OR c.codigo_interno = '')
                OR (c.tuss_numero IS NULL OR c.tuss_numero = '')
                OR (c.referencia IS NULL OR c.referencia = '')
                OR (c.fracionavel IS NULL OR c.fracionavel = '')
                OR (c.status_final IS NULL OR c.status_final = '')
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _purge_bras_versions_except(keep_version: str) -> dict[str, int]:
    """Remove dados BRAS de versões antigas, mantendo apenas a versão informada."""
    keep = (keep_version or '').strip()
    if not keep:
        return {}

    summary = {
        'insumos_index': 0,
        'bras_item_preco': 0,
        'bras_item_cadastro': 0,
        'bras_item_n': 0,
        'bras_catalog_snapshot': 0,
    }

    summary['insumos_index'] = (
        db.session.query(InsumoIndex)
        .filter(
            InsumoIndex.origem == 'BRAS',
            func.coalesce(InsumoIndex.versao_tabela, '') != keep,
        )
        .delete(synchronize_session=False)
    ) or 0

    dialect = (db.session.bind.dialect.name if db.session.bind is not None else '').lower()
    if dialect in {'mysql', 'mariadb'}:
        summary['bras_item_preco'] = db.session.execute(
            text(
                """
                DELETE p
                FROM bras_item_preco p
                INNER JOIN bras_item_cadastro c ON c.id = p.cadastro_id
                WHERE COALESCE(c.edicao, '') <> :keep_version
                """
            ),
            {'keep_version': keep},
        ).rowcount or 0
    else:
        cadastro_ids = [
            row[0]
            for row in db.session.query(BrasItemCadastro.id)
            .filter(func.coalesce(BrasItemCadastro.edicao, '') != keep)
            .all()
        ]
        if cadastro_ids:
            summary['bras_item_preco'] = (
                db.session.query(BrasItemPreco)
                .filter(BrasItemPreco.cadastro_id.in_(cadastro_ids))
                .delete(synchronize_session=False)
            ) or 0

    summary['bras_item_cadastro'] = (
        db.session.query(BrasItemCadastro)
        .filter(func.coalesce(BrasItemCadastro.edicao, '') != keep)
        .delete(synchronize_session=False)
    ) or 0

    summary['bras_item_n'] = (
        db.session.query(BrasItemNormalized)
        .filter(func.coalesce(BrasItemNormalized.edicao, '') != keep)
        .delete(synchronize_session=False)
    ) or 0

    summary['bras_catalog_snapshot'] = (
        db.session.query(BrasCatalogSnapshot)
        .filter(func.coalesce(BrasCatalogSnapshot.versao, '') != keep)
        .delete(synchronize_session=False)
    ) or 0

    db.session.commit()
    app.logger.info('BRAS purge versões antigas (mantida=%s): %s', keep, summary)
    return summary


def _import_bras(
    *,
    file_path: Path,
    versao: str,
    data_ref: str | None,
    fmt: str,
    delimiter: str,
    quotechar: str | None,
    line_terminator: str,
    skip_header: bool,
    encoding: str | None,
    map_config: dict,
    truncate: bool,
    uf_default: str | None = None,
    uf_values: Sequence[str] | None = None,
    aliquota_default: Decimal | None = None,
    arquivo_label_override: str | None = None,
    keep_only_latest_version: bool = False,
) -> dict:
    del data_ref
    arquivo_label_base = arquivo_label_override or map_config.get('arquivo') or versao or file_path.name
    arquivo_label = arquivo_label_base
    if uf_default:
        arquivo_label = f"{arquivo_label}_{uf_default.upper()}"

    # Se truncate E alíquota informada, limpa apenas dados daquela alíquota
    # Caso contrário, comportamento padrão (limpar tudo ou por arquivo)
    _delete_existing_bras_records(
        arquivo_label,
        truncate,
        aliquota_filter=aliquota_default if truncate else None,
        uf_filter=uf_default if truncate else None,
    )

    inserted = 0
    stage_strategy = 'unknown'
    if fmt == 'delimited':
        inserted, stage_strategy = _stage_bras_delimited(
            file_path=file_path,
            delimiter=delimiter,
            quotechar=quotechar,
            line_terminator=line_terminator,
            skip_header=skip_header,
            encoding=encoding,
            arquivo_label=arquivo_label,
            use_load_data=not map_config.get('disable_load_data', False),
        )
    else:
        inserted, stage_strategy = _stage_bras_fixed(
            file_path=file_path,
            map_config=map_config,
            encoding=encoding,
            line_terminator=line_terminator,
            arquivo_label=arquivo_label,
        )

    # Sempre materializa apenas o arquivo atual
    _materialize_bras_items(arquivo_label)
    aligned_rows = _align_bras_n_import_version(versao, arquivo_label=arquivo_label)
    
    # Conta as linhas reais materializadas para este arquivo
    count_result = db.session.execute(
        text("SELECT COUNT(*) FROM bras_item_n WHERE arquivo = :arquivo"),
        {'arquivo': arquivo_label}
    ).scalar() or 0
    split_sync = _sync_bras_split_from_bras_n_fast(
        arquivo_label=arquivo_label,
        aliquota_override=aliquota_default,
        versao_override=versao,
    )
    sync_ufs: list[str] = list(uf_values) if uf_values else []
    if aliquota_default is not None and not sync_ufs and not uf_default:
        sync_ufs = _ufs_pertencentes_a_aliquota_piso(aliquota_default)
    _sync_bras_insumo_index(
        arquivo_label,
        uf_default=uf_default,
        uf_values=sync_ufs or None,
        aliquota_default=aliquota_default,
        versao_label=versao,
    )

    purge_summary: dict[str, int] | None = None
    if keep_only_latest_version:
        purge_summary = _purge_bras_versions_except(versao)

    return {
        'arquivo': arquivo_label,
        'linhas_raw': inserted,
        'linhas_materializadas': count_result,
        'load_strategy': stage_strategy,
        'aligned_version_rows': aligned_rows,
        'split_sync': split_sync,
        'purge_summary': purge_summary,
    }


def _analyze_bras_delta(
    *,
    file_path: Path,
    delimiter: str = ',',
    quotechar: str = '"',
    encoding: str | None = 'latin-1',
) -> dict:
    """
    Analisa um arquivo Brasíndice e compara com dados existentes.
    Retorna estatísticas de novos, alterados e inalterados.
    """
    import csv
    
    # Carregar dados existentes indexados por EAN
    existing_items: dict[str, dict] = {}
    items = BrasItemNormalized.query.with_entities(
        BrasItemNormalized.ean,
        BrasItemNormalized.preco_pmc_pacote,
        BrasItemNormalized.preco_pfb_pacote,
        BrasItemNormalized.preco_pmc_unit,
        BrasItemNormalized.preco_pfb_unit,
        BrasItemNormalized.produto_nome,
        BrasItemNormalized.apresentacao_descricao,
    ).filter(BrasItemNormalized.ean.isnot(None), BrasItemNormalized.ean != '').all()
    
    for item in items:
        if item.ean:
            existing_items[item.ean.strip()] = {
                'pmc_pacote': item.preco_pmc_pacote,
                'pfb_pacote': item.preco_pfb_pacote,
                'pmc_unit': item.preco_pmc_unit,
                'pfb_unit': item.preco_pfb_unit,
                'produto': item.produto_nome,
                'apresentacao': item.apresentacao_descricao,
            }
    
    # Analisar arquivo novo
    novos: list[dict] = []
    alterados: list[dict] = []
    inalterados: list[dict] = []
    eans_no_arquivo: set[str] = set()
    
    enc = encoding or 'latin-1'
    with open(file_path, 'r', encoding=enc, errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar)
        for linha_num, row in enumerate(reader, start=1):
            if len(row) < 17:
                continue
            
            try:
                ean = row[16].strip() if row[16] else ''
                if not ean:
                    continue
                
                eans_no_arquivo.add(ean)
                
                # Parsear preços do arquivo
                pmc_pacote = Decimal(row[6].replace(',', '.')) if row[6] else None
                pfb_pacote = Decimal(row[7].replace(',', '.')) if row[7] else None
                pmc_unit = Decimal(row[10].replace(',', '.')) if row[10] else None
                pfb_unit = Decimal(row[12].replace(',', '.')) if row[12] else None
                
                item_info = {
                    'linha': linha_num,
                    'ean': ean,
                    'laboratorio': row[1].strip() if len(row) > 1 else '',
                    'produto': row[3].strip() if len(row) > 3 else '',
                    'apresentacao': row[5].strip() if len(row) > 5 else '',
                    'pmc_pacote': pmc_pacote,
                    'pfb_pacote': pfb_pacote,
                    'pmc_unit': pmc_unit,
                    'pfb_unit': pfb_unit,
                }
                
                if ean not in existing_items:
                    novos.append(item_info)
                else:
                    existente = existing_items[ean]
                    # Comparar preços (com tolerância para diferenças de arredondamento)
                    preco_mudou = False
                    diferencas = []
                    
                    if pmc_pacote and existente['pmc_pacote']:
                        if abs(pmc_pacote - existente['pmc_pacote']) > Decimal('0.01'):
                            preco_mudou = True
                            diferencas.append(f"PMC: {existente['pmc_pacote']} → {pmc_pacote}")
                    
                    if pfb_pacote and existente['pfb_pacote']:
                        if abs(pfb_pacote - existente['pfb_pacote']) > Decimal('0.01'):
                            preco_mudou = True
                            diferencas.append(f"PFB: {existente['pfb_pacote']} → {pfb_pacote}")
                    
                    if preco_mudou:
                        item_info['diferencas'] = diferencas
                        item_info['precos_antigos'] = {
                            'pmc_pacote': existente['pmc_pacote'],
                            'pfb_pacote': existente['pfb_pacote'],
                        }
                        alterados.append(item_info)
                    else:
                        inalterados.append(item_info)
                        
            except (ValueError, IndexError) as e:
                continue
    
    # Itens removidos (existem no banco mas não no arquivo)
    eans_existentes = set(existing_items.keys())
    eans_removidos = eans_existentes - eans_no_arquivo
    removidos = [{'ean': ean, **existing_items[ean]} for ean in list(eans_removidos)[:100]]
    
    return {
        'total_arquivo': len(eans_no_arquivo),
        'total_existente': len(existing_items),
        'novos': len(novos),
        'alterados': len(alterados),
        'inalterados': len(inalterados),
        'removidos': len(eans_removidos),
        'detalhes_novos': novos[:50],  # Limitar para preview
        'detalhes_alterados': alterados[:50],
        'detalhes_removidos': removidos[:50],
    }


def _import_bras_delta(
    *,
    file_path: Path,
    versao: str,
    delimiter: str = ',',
    quotechar: str = '"',
    encoding: str | None = 'latin-1',
    skip_header: bool = False,
    data_ref: str | None = None,
    uf_default: str | None = None,
    uf_values: Sequence[str] | None = None,
    aliquota_default: Decimal | None = None,
    catalog_file: Path | None = None,
    catalog_encoding: str | None = 'latin-1',
    catalog_delimiter: str = ';',
    previous_catalog_version: str | None = None,
) -> dict:
    """
    Importa apenas itens novos ou com preços alterados da Brasíndice.
    Mais eficiente que reimportar o arquivo inteiro.
    Suporta alíquota, UFs e data de referência para vincular corretamente os dados.
    """
    import csv

    aligned_rows = _align_bras_n_import_version(versao)
    catalog_delta: dict[str, object] | None = None
    catalog_candidate_keys: set[str] = set()
    if catalog_file is not None:
        _sync_bras_catalog_snapshot(
            file_path=catalog_file,
            versao=versao,
            delimiter=catalog_delimiter,
            encoding=catalog_encoding,
        )
        catalog_delta = _analyze_bras_catalog_delta(
            current_version=versao,
            previous_version=previous_catalog_version,
        )
        catalog_candidate_keys = set(catalog_delta['new_keys']) | set(catalog_delta['changed_keys'])
    
    # Carregar dados existentes indexados por EAN
    existing_items: dict[str, dict] = {}
    items = BrasItemNormalized.query.with_entities(
        BrasItemNormalized.id,
        BrasItemNormalized.ean,
        BrasItemNormalized.preco_pmc_pacote,
        BrasItemNormalized.preco_pfb_pacote,
    ).filter(BrasItemNormalized.ean.isnot(None), BrasItemNormalized.ean != '').all()
    
    for item in items:
        if item.ean:
            existing_items[item.ean.strip()] = {
                'id': item.id,
                'pmc_pacote': item.preco_pmc_pacote,
                'pfb_pacote': item.preco_pfb_pacote,
            }
    
    # Processar arquivo e coletar alterações
    linhas_para_importar: list[int] = []
    atualizacoes: list[dict] = []
    novos_count = 0
    alterados_count = 0
    
    enc = encoding or 'latin-1'
    with open(file_path, 'r', encoding=enc, errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar)
        if skip_header:
            next(reader, None)
        
        for linha_num, row in enumerate(reader, start=1):
            _raise_if_import_paused()
            if len(row) < 17:
                continue
            
            try:
                ean = row[16].strip() if row[16] else ''
                row_key, _key_kind = _build_bras_main_row_key(row)
                pmc_pacote = Decimal(row[6].replace(',', '.')) if row[6] else None
                pfb_pacote = Decimal(row[7].replace(',', '.')) if row[7] else None
                force_catalog_import = bool(row_key and row_key in catalog_candidate_keys)
                if not ean and not force_catalog_import:
                    continue
                
                if force_catalog_import or ean not in existing_items:
                    # Novo item - marcar linha para importação
                    linhas_para_importar.append(linha_num)
                    novos_count += 1
                else:
                    existente = existing_items[ean]
                    preco_mudou = False
                    
                    if pmc_pacote and existente['pmc_pacote']:
                        if abs(pmc_pacote - existente['pmc_pacote']) > Decimal('0.01'):
                            preco_mudou = True
                    
                    if pfb_pacote and existente['pfb_pacote']:
                        if abs(pfb_pacote - existente['pfb_pacote']) > Decimal('0.01'):
                            preco_mudou = True
                    
                    if preco_mudou:
                        # Atualizar registro existente diretamente
                        atualizacoes.append({
                            'id': existente['id'],
                            'pmc_pacote': pmc_pacote,
                            'pfb_pacote': pfb_pacote,
                            'pmc_unit': Decimal(row[10].replace(',', '.')) if row[10] else None,
                            'pfb_unit': Decimal(row[12].replace(',', '.')) if row[12] else None,
                        })
                        alterados_count += 1
                        
            except (ValueError, IndexError):
                continue
    
    # Aplicar atualizações em batch
    updated_ids: set[int] = set()
    for upd in atualizacoes:
        _raise_if_import_paused()
        BrasItemNormalized.query.filter_by(id=upd['id']).update({
            'preco_pmc_pacote': upd['pmc_pacote'],
            'preco_pfb_pacote': upd['pfb_pacote'],
            'preco_pmc_unit': upd['pmc_unit'],
            'preco_pfb_unit': upd['pfb_unit'],
            'edicao': versao,
        })
        updated_ids.add(int(upd['id']))
    
    db.session.commit()
    split_sync_updates = None
    if updated_ids:
        split_sync_updates = _sync_bras_split_from_bras_n_fast(
            n_ids=sorted(updated_ids),
            aliquota_override=aliquota_default,
            versao_override=versao,
        )
        if aliquota_default is not None:
            _reatribuir_insumo_bras_apos_import_precos(
                updated_ids,
                aliquota_default,
                versao_label=versao,
            )
    
    # Se houver itens novos, importar via pipeline normal (somente linhas específicas)
    novos_importados = 0
    if linhas_para_importar:
        # Se muitos itens novos e banco quase vazio, usar importação normal
        # é mais eficiente que criar arquivo temporário
        if novos_count > 5000 and len(existing_items) < 1000:
            app.logger.info(
                'Delta Brasíndice: %d novos, banco com %d itens. Recomenda-se usar importação normal.',
                novos_count, len(existing_items)
            )
        
        # Converter para set para busca O(1) em vez de O(n)
        linhas_set = set(linhas_para_importar)
        
        # Criar arquivo temporário só com as linhas novas
        # Usar newline='' para csv.writer controlar os line endings
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding=enc, newline='') as tmp:
            tmp_path = Path(tmp.name)
            with open(file_path, 'r', encoding=enc, errors='replace', newline='') as f:
                reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar)
                writer = csv.writer(tmp, delimiter=delimiter, quotechar=quotechar, lineterminator='\n')
                linhas_escritas = 0
                for linha_num, row in enumerate(reader, start=1):
                    _raise_if_import_paused()
                    if linha_num in linhas_set:
                        writer.writerow(row)
                        linhas_escritas += 1
        
        app.logger.info('Delta Brasíndice: arquivo temporário criado com %d linhas', linhas_escritas)
        
        # Importar arquivo temporário
        try:
            result = _import_bras(
                file_path=tmp_path,
                versao=versao,
                data_ref=data_ref,
                fmt='delimited',
                delimiter=delimiter,
                quotechar=quotechar,
                line_terminator='\n',
                skip_header=False,
                encoding=enc,
                map_config={'disable_load_data': True},  # Forçar fallback Python que é mais robusto
                truncate=False,
                uf_default=uf_default,
                uf_values=uf_values,
                aliquota_default=aliquota_default,
            )
            novos_importados = result.get('linhas_materializadas', 0)
            app.logger.info('Delta Brasíndice: %d linhas materializadas', novos_importados)
        except Exception as exc:
            app.logger.exception('Erro ao importar novos itens Brasíndice: %s', exc)
            raise
        finally:
            tmp_path.unlink(missing_ok=True)
    
    return {
        'versao': versao,
        'novos': novos_count,
        'novos_importados': novos_importados,
        'alterados': alterados_count,
        'total_processado': novos_count + alterados_count,
        'catalog_current_version': catalog_delta['current_version'] if catalog_delta else None,
        'catalog_previous_version': catalog_delta['previous_version'] if catalog_delta else None,
        'catalog_new': catalog_delta['new_count'] if catalog_delta else 0,
        'catalog_changed': catalog_delta['changed_count'] if catalog_delta else 0,
        'catalog_removed': catalog_delta['removed_count'] if catalog_delta else 0,
        'aligned_version_rows': aligned_rows,
        'split_sync_updates': split_sync_updates,
    }


def _import_simpro(
    *,
    file_path: Path,
    versao: str,
    fmt: str,
    map_config: dict,
    encoding: str | None,
    truncate: bool,
    uf_default: str | None,
    uf_values: Sequence[str] | None = None,
    aliquota_default: Decimal | None,
    arquivo_label_override: str | None = None,
    job_id: str | None = None,
) -> dict:
    arquivo_label = _build_simpro_arquivo_label(
        arquivo_label_override=arquivo_label_override,
        map_config=map_config,
        versao=versao,
        fallback_name=file_path.name,
        uf_default=uf_default,
        aliquota_default=aliquota_default,
    )

    # Se truncate E alíquota informada, limpa apenas dados daquela alíquota
    _delete_existing_simpro_records(
        arquivo_label,
        truncate,
        aliquota_filter=aliquota_default if truncate else None,
        uf_filter=uf_default if truncate else None,
    )

    if fmt == 'json':
        if job_id:
            _touch_import_job_progress(
                job_id,
                message='SIMPRO JSON: lendo e interpretando arquivo (pode demorar em arquivos grandes)…',
            )
        payload = _load_simpro_json_payload(file_path, encoding)
        inserted = len(payload)
        if job_id:
            _touch_import_job_progress(
                job_id,
                message=f'SIMPRO JSON: arquivo com {inserted} registros; iniciando materialização em lotes…',
                total_linhas=inserted,
                linhas_materializadas=0,
            )
        materialized = _materialize_simpro_json_items(
            arquivo_label=arquivo_label,
            records=payload,
            versao=versao,
            uf_default=uf_default,
            job_id=job_id,
        )
        stage_strategy = 'json_native'
    elif fmt == 'fixed':
        inserted, stage_strategy = _stage_simpro_fixed(
            file_path=file_path,
            map_config=map_config,
            encoding=encoding,
            arquivo_label=arquivo_label,
        )

        materialized = _materialize_simpro_items(
            arquivo_label=arquivo_label,
            map_config=map_config,
            versao=versao,
            uf_default=uf_default,
        )
    else:
        raise click.ClickException('Importação SIMPRO suporta apenas formatos JSON e largura fixa.')

    if job_id:
        _touch_import_job_progress(
            job_id,
            message='SIMPRO: sincronizando cadastros e preços (split)…',
        )
    split_sync = _sync_simpro_split_from_norm_fast(
        arquivo_label=arquivo_label,
        versao=versao,
        aliquota_override=aliquota_default,
    )
    if job_id:
        _touch_import_job_progress(
            job_id,
            message='SIMPRO: atualizando índice de insumos em lotes…',
        )
    _sync_simpro_insumo_index(
        arquivo_label,
        uf_default=uf_default,
        uf_values=uf_values,
        aliquota_default=aliquota_default,
    )
    _backfill_catalogo_simpro_identifiers()

    return {
        'arquivo': arquivo_label,
        'linhas_raw': inserted,
        'linhas_materializadas': materialized,
        'load_strategy': stage_strategy,
        'split_sync': split_sync,
    }


def _simpro_aliquota_label_fragment(aliquota: Decimal | None) -> str | None:
    if aliquota is None:
        return None
    try:
        decimal_value = Decimal(str(aliquota))
    except (InvalidOperation, ValueError, TypeError):
        return None
    normalized = format(decimal_value, 'f')
    if '.' in normalized:
        normalized = normalized.rstrip('0').rstrip('.')
    if not normalized:
        normalized = '0'
    safe = normalized.replace('-', 'N').replace('.', '_')
    return f'ALQ{safe}'


def _build_simpro_arquivo_label(
    *,
    arquivo_label_override: str | None,
    map_config: dict,
    versao: str,
    fallback_name: str,
    uf_default: str | None,
    aliquota_default: Decimal | None,
) -> str:
    arquivo_label = arquivo_label_override or map_config.get('arquivo') or versao or fallback_name
    if uf_default:
        arquivo_label = f"{arquivo_label}_{uf_default.upper()}"
    aliquota_fragment = _simpro_aliquota_label_fragment(aliquota_default)
    if aliquota_fragment:
        arquivo_label = f"{arquivo_label}_{aliquota_fragment}"
    return arquivo_label


DECIMAL_SANITIZE_RE = re.compile(r'[^0-9,\.-]')


def _coerce_decimal(value: str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, 'f')
    if isinstance(value, (int, float)):
        try:
            decimal_value = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
        return format(decimal_value, 'f')

    raw = str(value).strip()
    if not raw:
        return None

    raw = raw.replace('\xa0', '').replace(' ', '')
    if raw.endswith('-') and raw.count('-') == 1:
        raw = '-' + raw[:-1]
    raw = DECIMAL_SANITIZE_RE.sub('', raw)

    if not raw or raw in {'-', '.', ',', '-.', '-,'}:
        return None
    if raw.count('-') > 1 or (raw[0] == '-' and '-' in raw[1:]):
        return None

    negative = raw.startswith('-')
    if negative:
        raw = raw[1:]

    if not raw:
        return None

    decimal_sep = None
    if ',' in raw and '.' in raw:
        last_comma = raw.rfind(',')
        last_dot = raw.rfind('.')
        sep_index = max(last_comma, last_dot)
        decimal_sep = raw[sep_index]
        integer_part = raw[:sep_index].replace(',', '').replace('.', '')
        fractional_part = raw[sep_index + 1:]
    else:
        sep = None
        if ',' in raw:
            sep = ','
        elif '.' in raw:
            sep = '.'

        if sep is not None:
            occurrences = [idx for idx, ch in enumerate(raw) if ch == sep]
            last_idx = occurrences[-1]
            decimals_len = len(raw) - last_idx - 1
            if 0 < decimals_len <= 6:
                decimal_sep = sep
                integer_part = raw[:last_idx].replace(',', '').replace('.', '')
                fractional_part = raw[last_idx + 1:]
            else:
                integer_part = raw.replace(',', '').replace('.', '')
                fractional_part = ''
        else:
            integer_part = raw
            fractional_part = ''

    if decimal_sep is None:
        normalized = integer_part
    else:
        normalized = f"{integer_part}.{fractional_part}" if fractional_part else integer_part

    if negative:
        normalized = f"-{normalized}"

    if not normalized or normalized in {'-', '.', '-.'}:
        return None

    try:
        decimal_value = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None

    return format(decimal_value, 'f')


def _coerce_date(value: str | None) -> date | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_delimiter(delimiter: str) -> str:
    if delimiter.lower() in {'\t', 'tab'}:
        return '\t'
    return delimiter


def _resolve_columns(config_columns: list[str] | None, default_columns: list[str], header: list[str] | None) -> list[str]:
    if config_columns:
        sanitized = [_normalize_column_token(col) for col in config_columns]
        cols = [col for col in sanitized if col]
        return cols or default_columns
    if header:
        sanitized = [_normalize_column_token(h) for h in header]
        cols = [col for col in sanitized if col]
        return cols or default_columns
    return default_columns


def _parse_positive_int(value: str | None, default: int, *, minimum: int = 1, maximum: int | None = None) -> int:
    try:
        parsed = int(str(value))
    except (TypeError, ValueError):
        return default
    if parsed < minimum:
        return minimum
    if maximum is not None and parsed > maximum:
        return maximum
    return parsed


def _decimal_to_string(value: Decimal | None, precision: int = 4) -> str | None:
    if value is None:
        return None
    quantize_target = Decimal('1').scaleb(-precision)
    try:
        normalized = value.quantize(quantize_target)
    except (InvalidOperation, ValueError):
        normalized = value
    normalized = normalized.normalize()
    as_str = format(normalized, 'f')
    if '.' in as_str:
        as_str = as_str.rstrip('0').rstrip('.')
    return as_str


def _decimal_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError, InvalidOperation):
        try:
            return float(Decimal(str(value)))
        except Exception:
            return None


def _format_money_decimal(value: Decimal | None) -> float | None:
    if value is None:
        return None
    try:
        quantized = Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return float(quantized)
    except (InvalidOperation, ValueError):
        return _decimal_to_float(value)


def _ensure_teto_preview_dir() -> Path:
    directory = TETO_PREVIEW_DIR
    try:
        directory.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return directory


def _format_brl(value: Decimal | None) -> str:
    if value is None:
        return ''
    try:
        quantized = Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return ''
    formatted = format(quantized, ',.2f')
    return 'R$ ' + formatted.replace(',', 'X').replace('.', ',').replace('X', '.')


def _store_teto_preview(payload: dict) -> str:
    directory = _ensure_teto_preview_dir()
    token = uuid4().hex
    rows = []
    for row in payload.get('rows', []):
        value = row.get('valor_total')
        if value is not None and not isinstance(value, str):
            try:
                value = format(Decimal(value), 'f')
            except (InvalidOperation, ValueError):
                value = None
        rows.append({
            'codigo': row.get('codigo'),
            'descricao': row.get('descricao'),
            'valor_total': value,
            'row_number': row.get('row_number'),
        })
    data = {
        'rows': rows,
        'meta': payload.get('meta', {}),
        'errors': payload.get('errors', []),
    }
    file_path = directory / f'{token}.json'
    file_path.write_text(json.dumps(data, ensure_ascii=False))
    return token


def _load_teto_preview(token: str) -> dict | None:
    if not token:
        return None
    file_path = _ensure_teto_preview_dir() / f'{token}.json'
    if not file_path.exists():
        return None
    try:
        raw = json.loads(file_path.read_text(encoding='utf-8'))
    except (json.JSONDecodeError, OSError):
        return None
    rows: list[dict] = []
    for row in raw.get('rows', []):
        valor = row.get('valor_total')
        try:
            valor_decimal = Decimal(str(valor)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if valor is not None else None
        except (InvalidOperation, ValueError):
            valor_decimal = None
        rows.append({
            'codigo': row.get('codigo'),
            'descricao': row.get('descricao'),
            'valor_total': valor_decimal,
            'row_number': row.get('row_number'),
        })
    return {
        'rows': rows,
        'meta': raw.get('meta', {}),
        'errors': raw.get('errors', []),
        'token': token,
    }


def _discard_teto_preview(token: str) -> None:
    if not token:
        return
    file_path = _ensure_teto_preview_dir() / f'{token}.json'
    try:
        file_path.unlink(missing_ok=True)
    except Exception:
        pass


def _register_teto_job(job: dict) -> None:
    with _TETO_IMPORT_JOBS_LOCK:
        _TETO_IMPORT_JOBS[job['id']] = job
        if len(_TETO_IMPORT_JOBS) > 25:
            oldest_id = min(_TETO_IMPORT_JOBS.items(), key=lambda item: item[1]['created_at'])[0]
            _TETO_IMPORT_JOBS.pop(oldest_id, None)


def _update_teto_job(job_id: str, **fields) -> None:
    with _TETO_IMPORT_JOBS_LOCK:
        job = _TETO_IMPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(fields)


def _snapshot_teto_jobs() -> list[SimpleNamespace]:
    with _TETO_IMPORT_JOBS_LOCK:
        jobs = [job.copy() for job in _TETO_IMPORT_JOBS.values()]
    jobs.sort(key=lambda item: item['created_at'], reverse=True)
    snapshot: list[SimpleNamespace] = []
    for job in jobs:
        created_fmt = job['created_at'].strftime('%d/%m/%Y %H:%M:%S') if job.get('created_at') else None
        started_fmt = job.get('started_at').strftime('%d/%m/%Y %H:%M:%S') if job.get('started_at') else None
        finished_fmt = job.get('finished_at').strftime('%d/%m/%Y %H:%M:%S') if job.get('finished_at') else None
        job_ns = SimpleNamespace(**job)
        job_ns.created_at_fmt = created_fmt
        job_ns.started_at_fmt = started_fmt
        job_ns.finished_at_fmt = finished_fmt
        snapshot.append(job_ns)
    return snapshot


def _run_teto_import_job(job_id: str, preview_payload: dict, confirm_token: str) -> None:
    _update_teto_job(job_id, status='running', started_at=datetime.utcnow(), message='Processando registros…')
    try:
        # Multi-operadora: obter operadora_id do payload
        operadora_id = preview_payload.get('operadora_id', 1)

        rows_raw = preview_payload.get('rows') or []
        rows = []
        for row in rows_raw:
            cod = (row.get('codigo') or '').strip()
            val = row.get('valor_total')
            if not cod or val is None:
                continue
            rows.append({
                'codigo': cod,
                'descricao': (row.get('descricao') or '').strip()[:255],
                'valor_total': val,
                'operadora_id': operadora_id,  # Multi-operadora
            })
        codes = [row['codigo'] for row in rows]
        with app.app_context():
            existing: set[tuple[str, int]] = set()
            if codes:
                # Multi-operadora: verificar existentes por (codigo, operadora_id)
                existing = {(c, o) for (c, o) in db.session.query(CbhpmTeto.codigo, CbhpmTeto.operadora_id)
                           .filter(CbhpmTeto.codigo.in_(codes), CbhpmTeto.operadora_id == operadora_id).all()}
            if rows:
                chunk_size = 800
                for start in range(0, len(rows), chunk_size):
                    chunk = rows[start:start + chunk_size]
                    insert_rows = [
                        row
                        for row in chunk
                    ]
                    stmt = mysql_insert(CbhpmTeto.__table__).values(insert_rows)
                    upsert_stmt = stmt.on_duplicate_key_update(
                        descricao=stmt.inserted.descricao,
                        valor_total=stmt.inserted.valor_total,
                        updated_at=text('CURRENT_TIMESTAMP'),
                    )
                    db.session.execute(upsert_stmt)
                    db.session.commit()
            db.session.remove()
            # Multi-operadora: contar inseridos/atualizados por (codigo, operadora_id)
            inserted = len([code for code in codes if (code, operadora_id) not in existing])
            updated = len(codes) - inserted
            error_count = len(preview_payload.get('errors', []))
            _update_teto_job(
                job_id,
                status='success',
                finished_at=datetime.utcnow(),
                inserted=inserted,
                updated=updated,
                total=len(codes),
                errors=error_count,
                message=f'Importação concluída: {len(codes)} linha(s), {inserted} inserida(s), {updated} atualizada(s).',
            )
            app.logger.info(
                'Importação CBHPM teto concluída (job %s): total=%s, inseridos=%s, atualizados=%s, erros=%s',
                job_id, len(codes), inserted, updated, error_count
            )
    except Exception as exc:
        with app.app_context():
            db.session.rollback()
            db.session.remove()
        app.logger.exception('Erro ao processar importação de tetos (job %s)', job_id)
        _update_teto_job(
            job_id,
            status='error',
            finished_at=datetime.utcnow(),
            message=str(exc),
        )
    finally:
        _discard_teto_preview(confirm_token)


def _start_teto_import_job(preview_payload: dict, confirm_token: str) -> str:
    codes = [row['codigo'] for row in preview_payload.get('rows') or [] if row.get('codigo')]
    job_id = uuid4().hex[:8].upper()
    job = {
        'id': job_id,
        'status': 'queued',
        'created_at': datetime.utcnow(),
        'started_at': None,
        'finished_at': None,
        'total': len(codes),
        'inserted': 0,
        'updated': 0,
        'errors': len(preview_payload.get('errors', []) or []),
        'message': 'Aguardando processamento…',
    }
    _register_teto_job(job)
    thread = threading.Thread(target=_run_teto_import_job, args=(job_id, preview_payload, confirm_token), daemon=True)
    thread.start()
    return job_id


def _read_teto_rows_from_csv(file_path: Path) -> list[tuple[int, dict[str, object]]]:
    encodings = DEFAULT_IMPORT_ENCODINGS + ['utf-8']
    for enc in encodings:
        try:
            with file_path.open('r', encoding=enc, newline='') as handle:
                first_line = handle.readline()
                if not first_line:
                    return []
                delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','
                handle.seek(0)
                reader = csv.reader(handle, delimiter=delimiter)
                try:
                    header = next(reader)
                except StopIteration:
                    return []
                headers_norm = [_norm_header(h) for h in header]
                rows: list[tuple[int, dict[str, object]]] = []
                for row_idx, raw_row in enumerate(reader, start=2):
                    values: dict[str, object] = {}
                    for col_idx, key in enumerate(headers_norm):
                        if not key:
                            continue
                        value = raw_row[col_idx] if col_idx < len(raw_row) else ''
                        if isinstance(value, str):
                            value = value.strip()
                        values[key] = value
                    rows.append((row_idx, values))
                return rows
        except UnicodeDecodeError:
            continue
    raise click.ClickException('Não foi possível decodificar o arquivo CSV (UTF-8/Latin-1).')


def _read_teto_rows_from_xlsx(file_path: Path) -> list[tuple[int, dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise click.ClickException('Dependência openpyxl não disponível para ler arquivos XLSX.') from exc
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    headers_norm = [_norm_header(h) for h in headers]
    parsed: list[tuple[int, dict[str, object]]] = []
    for idx, line in enumerate(rows[1:], start=2):
        values: dict[str, object] = {}
        for col_idx, key in enumerate(headers_norm):
            if not key:
                continue
            value = line[col_idx] if col_idx < len(line) else None
            if isinstance(value, str):
                value = value.strip()
            values[key] = value
        parsed.append((idx, values))
    return parsed


def _parse_teto_import_file(file_path: Path) -> dict:
    ext = file_path.suffix.lower()
    if ext in {'.csv', '.txt'}:
        raw_rows = _read_teto_rows_from_csv(file_path)
    elif ext in {'.xlsx'}:
        raw_rows = _read_teto_rows_from_xlsx(file_path)
    else:
        raise click.ClickException('Formato não suportado. Envie um arquivo CSV ou XLSX.')

    records: dict[str, dict[str, object]] = {}
    order: list[str] = []
    errors: list[str] = []
    total_input = 0
    duplicate_count = 0

    for row_number, row in raw_rows:
        total_input += 1
        codigo_raw = row.get('codigo') or row.get('codigoprocedimento')
        codigo = str(codigo_raw or '').strip().upper()
        if not codigo:
            errors.append(f"Linha {row_number}: campo 'codigo' é obrigatório.")
            continue

        descricao_raw = row.get('descricao') or row.get('procedimento')
        descricao = str(descricao_raw or '').strip()
        if not descricao:
            errors.append(f"Linha {row_number} ({codigo}): campo 'descricao' é obrigatório.")
            continue

        valor_raw = row.get('valor_total') or row.get('valortotal') or row.get('valor')
        valor_str = _coerce_decimal(valor_raw if valor_raw is not None else '')
        if valor_str is None:
            errors.append(f"Linha {row_number} ({codigo}): valor_total inválido.")
            continue
        try:
            valor_decimal = Decimal(valor_str)
        except (InvalidOperation, ValueError):
            errors.append(f"Linha {row_number} ({codigo}): valor_total inválido.")
            continue
        if valor_decimal <= Decimal('0'):
            errors.append(f"Linha {row_number} ({codigo}): valor_total deve ser maior que zero.")
            continue
        valor_decimal = valor_decimal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        record = {
            'codigo': codigo,
            'descricao': descricao,
            'valor_total': valor_decimal,
            'row_number': row_number,
        }
        if codigo in records:
            duplicate_count += 1
            try:
                order.remove(codigo)
            except ValueError:
                pass
        order.append(codigo)
        records[codigo] = record

    final_rows = [records[cod] for cod in order]
    return {
        'rows': final_rows,
        'errors': errors,
        'total_input': total_input,
        'valid_count': len(final_rows),
        'duplicate_count': duplicate_count,
    }



@dataclass(frozen=True)
class SupplierConfig:
    fornecedor_key: str
    origem: str
    model: type
    hash_fields: Sequence[str]
    item_key_fields: Sequence[str]


_SUPPLIER_CONFIGS: dict[str, SupplierConfig] = {
    'BRASINDICE': SupplierConfig(
        fornecedor_key='BRASINDICE',
        origem='BRAS',
        model=BrasItemNormalized,
        hash_fields=(
            'produto_codigo', 'apresentacao_codigo', 'ean', 'registro_anvisa',
            'preco_pmc_unit', 'preco_pfb_unit', 'preco_pmc_pacote', 'preco_pfb_pacote',
            'laboratorio_nome', 'edicao'
        ),
        item_key_fields=('produto_codigo', 'apresentacao_codigo', 'ean'),
    ),
    'SIMPRO': SupplierConfig(
        fornecedor_key='SIMPRO',
        origem='SIMPRO',
        model=SimproItemNormalized,
        hash_fields=(
            'codigo_interno', 'codigo', 'codigo_alt', 'tuss_numero', 'descricao', 'data_ref', 'tipo_reg',
            'preco1', 'preco2', 'preco3', 'preco4', 'fabricante', 'anvisa',
            'validade_anvisa', 'ean', 'situacao', 'fracionavel'
        ),
        item_key_fields=('codigo', 'tuss_numero', 'ean'),
    ),
}


class AliquotaIngestionError(RuntimeError):
    """Erro de negócio durante ingestão/publicação de lotes por alíquota."""


def _normalize_fornecedor(value: str) -> str:
    if not value:
        raise AliquotaIngestionError('Fornecedor é obrigatório.')
    normalized = ''.join(
        c for c in unicodedata.normalize('NFKD', value.upper()) if not unicodedata.combining(c)
    )
    return normalized.strip()


def _resolve_supplier_config(fornecedor: str, origem: str | None = None) -> SupplierConfig:
    key = _normalize_fornecedor(fornecedor)
    config = _SUPPLIER_CONFIGS.get(key)
    if not config:
        raise AliquotaIngestionError(f'Fornecedor "{fornecedor}" não suportado.')
    if origem and config.origem != origem.upper():
        raise AliquotaIngestionError(
            f'Origem {origem} não corresponde ao fornecedor {fornecedor} (esperado {config.origem}).'
        )
    return config


def _normalize_periodo(periodo: str) -> str:
    if not periodo:
        raise AliquotaIngestionError('Período é obrigatório.')
    periodo = str(periodo).strip()
    if len(periodo) != 6 or not periodo.isdigit():
        raise AliquotaIngestionError('Período deve estar no formato YYYYMM.')
    return periodo


def _normalize_sequencia(sequencia: int | str) -> int:
    try:
        seq = int(sequencia)
    except (TypeError, ValueError) as exc:
        raise AliquotaIngestionError('Sequência deve ser 1 ou 2.') from exc
    if seq not in {1, 2}:
        raise AliquotaIngestionError('Sequência deve ser 1 ou 2.')
    return seq


def _normalize_aliquota_bp(value: int | str | Decimal) -> int:
    if value is None:
        raise AliquotaIngestionError('Informe a alíquota em basis points ou percentual.')
    if isinstance(value, int):
        if value < 0:
            raise AliquotaIngestionError('Alíquota não pode ser negativa.')
        return value
    if isinstance(value, Decimal):
        if value < 0:
            raise AliquotaIngestionError('Alíquota não pode ser negativa.')
        return int((value * Decimal('100')).to_integral_value(rounding=ROUND_HALF_UP))
    s = str(value).strip().replace('%', '').replace(' ', '')
    if not s:
        raise AliquotaIngestionError('Informe a alíquota.')
    s = s.replace(',', '.')
    try:
        decimal_value = Decimal(s)
    except InvalidOperation as exc:
        raise AliquotaIngestionError('Alíquota inválida.') from exc
    if decimal_value < 0:
        raise AliquotaIngestionError('Alíquota não pode ser negativa.')
    return int((decimal_value * Decimal('100')).to_integral_value(rounding=ROUND_HALF_UP))


def _compute_file_hash(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open('rb') as handle:
        for chunk in iter(lambda: handle.read(65536), b''):
            hasher.update(chunk)
    return hasher.hexdigest()


def _json_default(value):
    if isinstance(value, Decimal):
        return format(value, 'f')
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _serialize_row(row, fields: Sequence[str]) -> dict[str, object]:
    payload: dict[str, object] = {}
    for field in fields:
        payload[field] = _json_default(getattr(row, field, None))
    return payload


def _build_item_key(row, fields: Sequence[str]) -> str:
    parts: list[str] = []
    for field in fields:
        value = getattr(row, field, None)
        if value in (None, ''):
            continue
        if isinstance(value, Decimal):
            parts.append(format(value, 'f'))
        else:
            parts.append(str(value).strip())
    if not parts:
        parts.append(str(getattr(row, 'id', '')))
    return '|'.join(parts)


def ingestir_arquivo(
    fornecedor: str,
    origem: str,
    aliquota_bp: int | str | Decimal,
    periodo: str,
    sequencia: int | str,
    arquivo_label: str,
    *,
    arquivo_path: Path | None = None,
    arquivo_bytes: bytes | None = None,
    commit: bool = True,
    session=None,
) -> Lote:
    session = session or db.session
    config = _resolve_supplier_config(fornecedor, origem)
    periodo_norm = _normalize_periodo(periodo)
    sequencia_norm = _normalize_sequencia(sequencia)
    aliquota_bp_norm = _normalize_aliquota_bp(aliquota_bp)

    if not arquivo_label:
        raise AliquotaIngestionError('arquivo_label é obrigatório para correlacionar os itens normalizados.')

    rows = session.query(config.model).filter(config.model.arquivo == arquivo_label).all()
    if not rows:
        raise AliquotaIngestionError(f'Nenhum item carregado encontrado para arquivo "{arquivo_label}".')

    if arquivo_bytes is not None:
        hash_arquivo = hashlib.sha256(arquivo_bytes).hexdigest()
    elif arquivo_path is not None:
        hash_arquivo = _compute_file_hash(arquivo_path)
    else:
        hash_arquivo = None

    aggregator = hashlib.sha256()
    line_entries: list[tuple[str, str, str]] = []
    for row in rows:
        payload = _serialize_row(row, config.hash_fields)
        payload['linha_num'] = getattr(row, 'linha_num', None)
        payload['arquivo'] = getattr(row, 'arquivo', None)
        item_key = _build_item_key(row, config.item_key_fields)
        payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=_json_default)
        line_hash = hashlib.sha256(payload_json.encode('utf-8')).hexdigest()
        aggregator.update(item_key.encode('utf-8'))
        aggregator.update(line_hash.encode('utf-8'))
        line_entries.append((item_key, line_hash, payload_json))

    if hash_arquivo is None:
        hash_arquivo = aggregator.hexdigest()

    fornecedor_norm = _normalize_fornecedor(fornecedor)
    lote = (
        session.query(Lote)
        .filter_by(
            fornecedor=fornecedor_norm,
            aliquota_bp=aliquota_bp_norm,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
        )
        .one_or_none()
    )
    if lote is None:
        lote = Lote(
            fornecedor=fornecedor_norm,
            aliquota_bp=aliquota_bp_norm,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
            arquivo_label=arquivo_label,
            status=LoteStatus.PENDENTE,
        )
        session.add(lote)
        session.flush()
    else:
        lote.arquivo_label = arquivo_label

    lote.hash_arquivo = hash_arquivo
    lote.total_itens = len(line_entries)
    lote.status = LoteStatus.VALIDADO
    lote.validado_em = datetime.utcnow()

    _upsert_linha_hashes(lote, line_entries, session=session)

    if commit:
        session.commit()
    else:
        session.flush()

    app.logger.info(
        'Lote %s/%s período %s sequência %s validado (%s itens)',
        fornecedor_norm,
        aliquota_bp_norm,
        periodo_norm,
        sequencia_norm,
        len(line_entries),
    )
    return lote


def _upsert_linha_hashes(lote: Lote, entries: Sequence[tuple[str, str, str]], *, session=None) -> None:
    session = session or db.session
    if lote.id is None:
        session.flush()
    existing = {
        row.item_chave: row for row in session.query(LinhaHash).filter_by(lote_id=lote.id).all()
    }
    
    # Deduplica entries - mantém a última ocorrência de cada item_key
    unique_entries: dict[str, tuple[str, str, str]] = {}
    for item_key, line_hash, payload_json in entries:
        unique_entries[item_key] = (item_key, line_hash, payload_json)
    
    new_keys = set()
    added_in_session: set[str] = set()  # Rastreia o que já foi adicionado nesta sessão
    
    for item_key, line_hash, payload_json in unique_entries.values():
        new_keys.add(item_key)
        row = existing.get(item_key)
        if row:
            if row.hash_linha != line_hash or row.payload_snapshot != payload_json:
                row.hash_linha = line_hash
                row.payload_snapshot = payload_json
        elif item_key not in added_in_session:
            # Só adiciona se não adicionamos nesta sessão ainda
            session.add(
                LinhaHash(
                    lote_id=lote.id,
                    item_chave=item_key,
                    hash_linha=line_hash,
                    payload_snapshot=payload_json,
                )
            )
            added_in_session.add(item_key)
    
    for key, row in existing.items():
        if key not in new_keys:
            session.delete(row)


def publicar_lote(
    fornecedor: str,
    aliquota_bp: int | str | Decimal,
    periodo: str,
    sequencia: int | str,
    *,
    session=None,
    commit: bool = True,
) -> Publicacao:
    session = session or db.session
    fornecedor_norm = _normalize_fornecedor(fornecedor)
    aliquota_bp_norm = _normalize_aliquota_bp(aliquota_bp)
    periodo_norm = _normalize_periodo(periodo)
    sequencia_norm = _normalize_sequencia(sequencia)

    # Verifica se já existe uma publicação com esses dados (evita duplicatas)
    existing_pub = (
        session.query(Publicacao)
        .filter_by(
            fornecedor=fornecedor_norm,
            aliquota_bp=aliquota_bp_norm,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
        )
        .order_by(Publicacao.publicado_em.desc())
        .first()
    )
    if existing_pub:
        return existing_pub

    lote = (
        session.query(Lote)
        .filter_by(
            fornecedor=fornecedor_norm,
            aliquota_bp=aliquota_bp_norm,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
        )
        .one_or_none()
    )
    if not lote:
        raise AliquotaIngestionError('Lote não encontrado para publicação.')
    if lote.status not in {LoteStatus.VALIDADO, LoteStatus.PUBLICADO}:
        raise AliquotaIngestionError(f'Lote em status {lote.status.value} não pode ser publicado.')

    lote.status = LoteStatus.PUBLICADO
    lote.publicado_em = datetime.utcnow()

    etag = f"{fornecedor_norm}:{periodo_norm}:{sequencia_norm}"
    publication = Publicacao(
        fornecedor=fornecedor_norm,
        aliquota_bp=aliquota_bp_norm,
        periodo=periodo_norm,
        sequencia=sequencia_norm,
        lote_id=lote.id,
        etag_versao=etag,
    )
    session.add(publication)

    if commit:
        session.commit()
    else:
        session.flush()

    _refresh_materialized_catalogs(fornecedor_norm)
    return publication


def _refresh_materialized_catalogs(fornecedor: str) -> None:
    bind = db.session.get_bind()
    dialect = getattr(bind, 'dialect', None) if bind is not None else None
    dialect_name = getattr(dialect, 'name', '').lower() if dialect is not None else ''

    view_map = {
        'BRASINDICE': ['mv_catalogo_vigente_brasindice'],
        'SIMPRO': ['mv_catalogo_vigente_simpro'],
    }
    targets = view_map.get(fornecedor, [])
    if not targets:
        return

    if dialect_name != 'postgresql':
        app.logger.debug(
            'Ignorando refresh de materialized view para %s (dialeto=%s).',
            fornecedor,
            dialect_name or 'desconhecido',
        )
        return

    for view_name in targets:
        try:
            db.session.execute(text(f'REFRESH MATERIALIZED VIEW IF EXISTS {view_name}'))
            db.session.commit()
            app.logger.info('Materialized view %s atualizada.', view_name)
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            app.logger.warning('Falha ao atualizar view %s (%s).', view_name, exc)


def _normalize_periodo_from_label(label: str | None) -> str | None:
    if not label:
        return None
    digits = ''.join(ch for ch in str(label) if ch.isdigit())
    if len(digits) >= 6:
        return digits[:6]
    return None


def _assign_uf_aliquota(ufs: Sequence[str], aliquota_bp: int) -> None:
    if not ufs:
        return
    now_ts = datetime.utcnow()
    today = date.today()
    for uf in ufs:
        record = (UfAliquota.query
                  .filter_by(uf=uf)
                  .order_by(UfAliquota.valid_from.desc())
                  .first())
        if record:
            record.aliquota_bp = aliquota_bp
            record.is_current = True
            record.updated_at = now_ts
        else:
            db.session.add(UfAliquota(
                uf=uf,
                valid_from=today,
                valid_to=None,
                aliquota_bp=aliquota_bp,
                is_current=True,
                created_at=now_ts,
                updated_at=now_ts,
            ))
    db.session.commit()


def _post_catalog_ingest(*, origem: str, arquivo_label: str, versao: str, sequencia_input: str | None, aliquota_value: Decimal | None, uf_values: Sequence[str]) -> None:
    if aliquota_value is None:
        return
    fornecedor = 'BRASINDICE' if origem == 'BRAS' else 'SIMPRO'
    try:
        aliquota_bp = _normalize_aliquota_bp(aliquota_value)
    except AliquotaIngestionError as exc:
        _safe_flash(f'Falha ao validar alíquota: {exc}', 'warning')
        return

    periodo_norm = _normalize_periodo_from_label(versao)
    if not periodo_norm:
        periodo_norm = datetime.utcnow().strftime('%Y%m')

    try:
        sequencia_norm = _normalize_sequencia(sequencia_input or 1)
    except AliquotaIngestionError as exc:
        _safe_flash(f'Sequência inválida: {exc}', 'warning')
        sequencia_norm = 1

    try:
        lote = ingestir_arquivo(
            fornecedor=fornecedor,
            origem=origem,
            aliquota_bp=aliquota_bp,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
            arquivo_label=arquivo_label,
        )
        publicar_lote(
            fornecedor=fornecedor,
            aliquota_bp=aliquota_bp,
            periodo=periodo_norm,
            sequencia=sequencia_norm,
        )
        if uf_values:
            _assign_uf_aliquota(uf_values, aliquota_bp)
        app.logger.info('Lote %s consolidado para %s/%s (seq %s)', lote.id, fornecedor, periodo_norm, sequencia_norm)
    except AliquotaIngestionError as exc:
        _safe_flash(f'Falha ao consolidar o catálogo por alíquota: {exc}', 'warning')
    except Exception as exc:  # noqa: BLE001
        app.logger.exception('Falha ao consolidar catálogo por alíquota', exc_info=exc)
        _safe_flash('Falha ao consolidar catálogo por alíquota. Verifique os logs.', 'warning')


def _catalogo_filter_bras(query, filters: dict):
    """
    Aplica filtros na query de Brasíndice.
    
    Otimizações:
    - Filtros exatos (=) primeiro para usar índices
    - LIKE com % só no final quando possível
    - Evita func.lower() quando não necessário
    """
    # Filtros exatos primeiro (usam índices)
    if filters.get('uf_referencia'):
        query = query.filter(CatalogoBrasindice.uf == filters['uf_referencia'])
    if filters.get('aliquota') is not None:
        target_bp = int((filters['aliquota'] * Decimal('100')).to_integral_value(rounding=ROUND_HALF_UP))
        query = query.filter(CatalogoBrasindice.aliquota_bp == target_bp)
    if filters.get('tuss'):
        query = query.filter(CatalogoBrasindice.produto_codigo == filters['tuss'])
    if filters.get('tiss'):
        query = query.filter(CatalogoBrasindice.apresentacao_codigo == filters['tiss'])
    if filters.get('anvisa'):
        query = query.filter(CatalogoBrasindice.registro_anvisa == filters['anvisa'])
    if filters.get('versao_tabela'):
        query = query.filter(CatalogoBrasindice.periodo == filters['versao_tabela'])
    
    # Filtros de texto (mais lentos)
    if filters.get('fabricante'):
        fabricante = filters['fabricante'].lower()
        query = query.filter(func.lower(CatalogoBrasindice.laboratorio_nome).like(f"%{fabricante}%"))
    
    tokens = filters.get('tokens') or []
    for token in tokens:
        pattern = f"%{token}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(CatalogoBrasindice.produto_nome, '')).like(pattern),
                func.lower(func.coalesce(CatalogoBrasindice.apresentacao_descricao, '')).like(pattern),
                func.lower(func.coalesce(CatalogoBrasindice.ean, '')).like(pattern),
                func.lower(func.coalesce(CatalogoBrasindice.registro_anvisa, '')).like(pattern),
            )
        )
    return query


def _catalogo_filter_simpro(query, filters: dict):
    """
    Aplica filtros na query de SIMPRO.
    
    Otimizações:
    - Filtros exatos (=) primeiro para usar índices
    - LIKE com % só no final quando possível
    """
    # Filtros exatos primeiro (usam índices)
    if filters.get('uf_referencia'):
        query = query.filter(CatalogoSimpro.uf == filters['uf_referencia'])
    if filters.get('aliquota') is not None:
        target_bp = int((filters['aliquota'] * Decimal('100')).to_integral_value(rounding=ROUND_HALF_UP))
        query = query.filter(CatalogoSimpro.aliquota_bp == target_bp)
    if filters.get('tuss'):
        query = query.filter(CatalogoSimpro.tuss_numero == filters['tuss'])
    if filters.get('tiss'):
        query = query.filter(CatalogoSimpro.codigo_alt == filters['tiss'])
    if filters.get('anvisa'):
        query = query.filter(CatalogoSimpro.anvisa == filters['anvisa'])
    if filters.get('versao_tabela'):
        query = query.filter(CatalogoSimpro.periodo == filters['versao_tabela'])
    
    # Filtros de texto (mais lentos)
    if filters.get('fabricante'):
        fabricante = filters['fabricante'].lower()
        query = query.filter(func.lower(CatalogoSimpro.fabricante).like(f"%{fabricante}%"))
    
    tokens = filters.get('tokens') or []
    for token in tokens:
        pattern = f"%{token}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(CatalogoSimpro.descricao, '')).like(pattern),
                func.lower(func.coalesce(CatalogoSimpro.referencia, '')).like(pattern),
                func.lower(func.coalesce(CatalogoSimpro.codigo, '')).like(pattern),
                func.lower(func.coalesce(CatalogoSimpro.tuss_numero, '')).like(pattern),
                func.lower(func.coalesce(CatalogoSimpro.ean, '')).like(pattern),
            )
        )
    return query


def _serialize_catalogo_bras(row: CatalogoBrasindice) -> dict:
    descricao = ' • '.join([part for part in [row.produto_nome, row.apresentacao_descricao] if part])
    preco_pmc = row.preco_pmc_unit or row.preco_pmc_pacote
    preco_pfb = row.preco_pfb_unit or row.preco_pfb_pacote
    base_preco = preco_pmc or preco_pfb
    aliquota_decimal = _aliquota_bp_to_decimal(row.aliquota_bp)
    return {
        'origem': 'BRAS',
        'item_id': row.item_id,
        'tuss': row.produto_codigo,
        'tiss': row.apresentacao_codigo,
        'descricao': descricao or None,
        'preco': _decimal_to_float(base_preco),
        'preco_pmc': _decimal_to_float(preco_pmc),
        'preco_pfb': _decimal_to_float(preco_pfb),
        'aliquota': _decimal_to_float(aliquota_decimal),
        'fabricante': row.laboratorio_nome,
        'anvisa': row.registro_anvisa,
        'versao_tabela': row.periodo or row.etag_versao,
        'data_atualizacao': None,
        'updated_at': row.imported_at.isoformat() if isinstance(row.imported_at, datetime) else None,
        'uf_referencia': row.uf,
    }


def _serialize_catalogo_simpro(row: CatalogoSimpro) -> dict:
    preco_fav, preco_pmc, preco_pfb = _split_simpro_prices(
        row.preco1,
        row.preco2,
        row.preco3,
        row.preco4,
    )
    aliquota_decimal = _aliquota_bp_to_decimal(row.aliquota_bp)
    tuss_digits = _format_tuss_display(None, getattr(row, 'tuss_numero', None))
    descricao = (row.descricao or '').strip() or None
    return {
        'origem': 'SIMPRO',
        'item_id': row.item_id,
        'codigo_simpro': row.codigo,
        'codigo_usuario': getattr(row, 'codigo_interno', None),
        'codigo_fracao': row.codigo_alt,
        'tuss': tuss_digits,
        'tuss_numero': tuss_digits,
        'tuss_raw': getattr(row, 'tuss_numero', None),
        'tiss': row.codigo_alt,
        'descricao': descricao,
        'embalagem': row.unidade,
        'qtd_unidade': row.qtd_unidade,
        'preco': _decimal_to_float(preco_fav),
        'preco_pmc': _decimal_to_float(preco_pmc),
        'preco_pfb': _decimal_to_float(preco_pfb),
        'aliquota': _decimal_to_float(aliquota_decimal),
        'fabricante': row.fabricante,
        'anvisa': row.anvisa,
        'versao_tabela': row.periodo or row.etag_versao,
        'data_atualizacao': row.data_ref.isoformat() if isinstance(row.data_ref, date) else None,
        'updated_at': row.imported_at.isoformat() if isinstance(row.imported_at, datetime) else None,
        'uf_referencia': row.uf,
    }


def _build_uf_piso_aliquota_bras() -> dict[str, Decimal]:
    """
    Tabela piso (ANVISA / fórmula) por UF, mesma usada no front da importação.
    Usada na busca quando a UF não tem linha no índice, mas o preço da alíquota está em `bras_item_preco`.
    """
    pairs: list[tuple[str, list[str]]] = [
        ('17', ['DF', 'ES', 'MT', 'MS', 'RS', 'SC']),
        ('18', ['AP', 'MG', 'SP']),
        ('19', ['AC', 'AL', 'GO', 'PA', 'SE']),
        ('19.5', ['PR', 'RO']),
        ('20', ['AM', 'CE', 'PB', 'RN', 'RR', 'TO']),
        ('20.5', ['BA', 'PE']),
        ('22', ['RJ']),
        ('22.5', ['PI']),
        ('23', ['MA']),
    ]
    out: dict[str, Decimal] = {}
    for a_str, ufs in pairs:
        al = _br_norm_aliquota(Decimal(a_str)) or Decimal(a_str)
        for u in ufs:
            out[u] = al
    return out


_UF_PISO_ALIQUOTA_BRAS: dict[str, Decimal] = _build_uf_piso_aliquota_bras()


def _serialize_bras_split_preco(
    n: 'BrasItemNormalized',
    cad: 'BrasItemCadastro',
    p: 'BrasItemPreco',
    uf: str,
    alq_display: Decimal,
) -> dict:
    pmc = p.preco_pmc_unit or p.preco_pmc_pacote
    pfb = p.preco_pfb_unit or p.preco_pfb_pacote
    preco_disp = pmc or pfb
    tuss_display = _format_tuss_display(n.produto_codigo)
    p1 = (n.produto_nome or '').strip()
    p2 = (n.apresentacao_descricao or '').strip()
    desc = ' • '.join([x for x in (p1, p2) if x]) or None
    return {
        'origem': 'BRAS',
        'item_id': n.id,
        'codigo_simpro': None,
        'codigo_usuario': None,
        'codigo_fracao': None,
        'embalagem': None,
        'qtd_unidade': n.quantidade_embalagem,
        'tuss': tuss_display,
        'tuss_numero': tuss_display,
        'tuss_raw': n.produto_codigo,
        'tiss': n.apresentacao_codigo,
        'descricao': desc,
        'preco': _decimal_to_string(preco_disp),
        'preco_pmc': _decimal_to_string(pmc),
        'preco_pfb': _decimal_to_string(pfb),
        'aliquota': _decimal_to_string(_br_norm_aliquota(alq_display) or alq_display),
        'fabricante': n.laboratorio_nome,
        'anvisa': n.registro_anvisa,
        'versao_tabela': cad.edicao,
        'data_atualizacao': None,
        'updated_at': p.imported_at.isoformat() if isinstance(p.imported_at, datetime) else None,
        'uf_referencia': uf,
        'uf_referencia_codes': [uf],
    }


def _serialize_bras_split_somente_n(
    n: 'BrasItemNormalized',
    cad: 'BrasItemCadastro',
    uf: str,
    alq_referencia: Decimal,
) -> dict:
    """Quando ainda não há `bras_item_preco`, mas `bras_item_n` traz preços (import base)."""
    pmc = n.preco_pmc_unit or n.preco_pmc_pacote
    pfb = n.preco_pfb_unit or n.preco_pfb_pacote
    preco_disp = pmc or pfb
    alq_show: Decimal | None = n.aliquota_ou_ipi
    if alq_show is None:
        alq_show = alq_referencia
    alq_show = _br_norm_aliquota(alq_show) or alq_show
    tuss_display = _format_tuss_display(n.produto_codigo)
    p1 = (n.produto_nome or '').strip()
    p2 = (n.apresentacao_descricao or '').strip()
    desc = ' • '.join([x for x in (p1, p2) if x]) or None
    return {
        'origem': 'BRAS',
        'item_id': n.id,
        'codigo_simpro': None,
        'codigo_usuario': None,
        'codigo_fracao': None,
        'embalagem': None,
        'qtd_unidade': n.quantidade_embalagem,
        'tuss': tuss_display,
        'tuss_numero': tuss_display,
        'tuss_raw': n.produto_codigo,
        'tiss': n.apresentacao_codigo,
        'descricao': desc,
        'preco': _decimal_to_string(preco_disp),
        'preco_pmc': _decimal_to_string(pmc),
        'preco_pfb': _decimal_to_string(pfb),
        'aliquota': _decimal_to_string(alq_show),
        'fabricante': n.laboratorio_nome,
        'anvisa': n.registro_anvisa,
        'versao_tabela': cad.edicao,
        'data_atualizacao': None,
        'updated_at': n.imported_at.isoformat() if isinstance(n.imported_at, datetime) else None,
        'uf_referencia': uf,
        'uf_referencia_codes': [uf],
    }


def _escolher_preco_cadastro(
    cad: 'BrasItemCadastro',
    alq_ideal: Decimal,
) -> 'BrasItemPreco | None':
    """
    1) linha com alíquota exata; 2) a mais próxima em valor; 3) qualquer linha.
    Útil quando a UF piso exige 20% mas o sistema só importou 20,5% (arquivo de outro estado).
    """
    q = (
        BrasItemPreco.query.filter(BrasItemPreco.cadastro_id == cad.id)
        .order_by(BrasItemPreco.id.asc())
        .all()
    )
    if not q:
        return None
    for pr in q:
        if _br_aliquota_certeiro(pr.aliquota, alq_ideal):
            return pr
    def dist(pr: 'BrasItemPreco') -> float:
        try:
            a = pr.aliquota
            if a is None:
                return 999.0
            return float(abs(a - alq_ideal))
        except (InvalidOperation, TypeError, ValueError):
            return 999.0
    return min(q, key=dist)


def _catalogo_search_bras_cadastro_preco_fallback(
    filters: dict,
    page: int,
    per_page: int,
) -> dict | None:
    """
    Se catálogo MV e `insumos_index` estão vazios para a combinação UF+TUSS, resolve:
    `bras_item_n` (identidade) + `bras_item_cadastro` + `bras_item_preco` (preço da alíquota: filtro
    ou tabela piso ANVISA por UF).
    """
    if (filters.get('origem') or '').upper() == 'SIMPRO':
        return None
    if page > 1:
        return None
    tuss, tiss, anvisa = filters.get('tuss'), filters.get('tiss'), filters.get('anvisa')
    if not tuss and not tiss and not anvisa:
        return None
    uf = (filters.get('uf_referencia') or '').strip().upper()
    if not uf:
        return None
    alq_f = filters.get('aliquota')
    if alq_f is not None:
        alq: Decimal | None = _br_norm_aliquota(aliq_f) or alq_f
    else:
        alq = _UF_PISO_ALIQUOTA_BRAS.get(uf)
    if alq is None:
        return None
    nq = BrasItemNormalized.query
    if tuss:
        nq = nq.filter(BrasItemNormalized.produto_codigo == tuss)
    if tiss:
        nq = nq.filter(BrasItemNormalized.apresentacao_codigo == tiss)
    if anvisa:
        nq = nq.filter(BrasItemNormalized.registro_anvisa == anvisa)
    nq = nq.filter(BrasItemNormalized.ean.isnot(None), func.trim(BrasItemNormalized.ean) != '')
    if filters.get('fabricante'):
        fab = filters['fabricante'].lower()
        nq = nq.filter(func.lower(BrasItemNormalized.laboratorio_nome).like(f'%{fab}%'))
    vtab = (filters.get('versao_tabela') or '').strip()
    if vtab and vtab.lower() not in ('todas', 'all', '*'):
        vc = _candidatas_edicao_bras(vtab)
        if vc:
            nq = nq.filter(
                or_(
                    *[
                        func.coalesce(BrasItemNormalized.edicao, '') == c
                        for c in vc
                    ]
                )
            )
    n = nq.order_by(BrasItemNormalized.id.desc()).first()
    if not n:
        return None
    ean = (n.ean or '').strip()
    ed = (n.edicao or '').strip() or (n.arquivo or '')[:50] or '—'
    vtab_filtro = vtab if (vtab and vtab.lower() not in ('todas', 'all', '*')) else ''
    cands: list[str] = list(
        dict.fromkeys(
            _candidatas_edicao_bras(ed) + (_candidatas_edicao_bras(vtab_filtro) if vtab_filtro else []),
        )
    ) or [ed]
    cad = (
        BrasItemCadastro.query.filter(
            and_(
                func.trim(BrasItemCadastro.ean) == ean,
                or_(*[func.trim(BrasItemCadastro.edicao) == c for c in cands]),
            )
        )
        .first()
    )
    if not cad:
        return None
    p_row = (
        BrasItemPreco.query.filter(BrasItemPreco.cadastro_id == cad.id)
        .filter(BrasItemPreco.aliquota == alq)
        .first()
    )
    if p_row:
        alq_mostrada = _br_norm_aliquota(p_row.aliquota) or p_row.aliquota
        item = _serialize_bras_split_preco(n, cad, p_row, uf, alq_mostrada)
        return {
            'items': [item],
            'empty_hint': None,
            'preco_fonte': 'bras_cadastro_preco',
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': 1,
                'pages': 1,
            },
        }
    return None


# Cache simples para contagens na busca de insumos (TTL configurável; default 120s — ver INSUMOS_COUNT_CACHE_TTL).
_CATALOGO_COUNT_CACHE: dict[str, tuple[int, float]] = {}


def _get_cached_count(cache_key: str, query_fn) -> int:
    """Retorna contagem do cache ou executa query se expirado."""
    import time
    now = time.time()
    ttl = _catalog_count_cache_ttl_seconds()
    cached = _CATALOGO_COUNT_CACHE.get(cache_key)
    if cached and (now - cached[1]) < ttl:
        return cached[0]
    count = query_fn()
    _CATALOGO_COUNT_CACHE[cache_key] = (count, now)
    return count


def _decimal_lookup_key(value: Decimal | None) -> str | None:
    if value is None:
        return None
    try:
        return format(Decimal(str(value)).normalize(), 'f')
    except (InvalidOperation, ValueError):
        return str(value)


def _prefetch_insumo_related(rows: list['InsumoIndex']) -> dict[str, dict]:
    simpro_ids = [int(row.item_id) for row in rows if row.origem == 'SIMPRO' and row.item_id is not None]
    if not simpro_ids:
        return {
            'simpro_cadastro_by_id': {},
            'simpro_preco_by_key': {},
        }

    simpro_cadastros = (
        SimproItemCadastro.query
        .filter(SimproItemCadastro.id.in_(simpro_ids))
        .all()
    )
    cadastro_by_id = {int(row.id): row for row in simpro_cadastros if row.id is not None}
    if not cadastro_by_id:
        return {
            'simpro_cadastro_by_id': {},
            'simpro_preco_by_key': {},
        }

    aliquotas = list({
        row.aliquota
        for row in rows
        if row.origem == 'SIMPRO' and row.aliquota is not None
    })
    if not aliquotas:
        return {
            'simpro_cadastro_by_id': cadastro_by_id,
            'simpro_preco_by_key': {},
        }

    precos = (
        SimproItemPreco.query
        .filter(
            SimproItemPreco.cadastro_id.in_(list(cadastro_by_id.keys())),
            SimproItemPreco.aliquota.in_(aliquotas),
        )
        .all()
    )
    preco_by_key = {
        (int(row.cadastro_id), _decimal_lookup_key(row.aliquota)): row
        for row in precos
        if row.cadastro_id is not None
    }
    return {
        'simpro_cadastro_by_id': cadastro_by_id,
        'simpro_preco_by_key': preco_by_key,
    }


def _wrap_insumos_index_latest_version(filters: dict):
    """
    Deduplica `insumos_index` quando há vários cadastros (versões) para o mesmo
    código comercial (TISS+tuss…) + alíquota — mantém o mais recente por
    `data_ref` / `imported_at` do cadastro e preço, depois por `updated_at` e `item_id`.
    """
    base_q = _apply_insumo_filters(InsumoIndex.query, filters)
    rn = (
        func.row_number()
        .over(
            partition_by=[
                case(
                    (
                        InsumoIndex.origem == literal('SIMPRO'),
                        func.concat_ws(
                            literal('|'),
                            func.coalesce(InsumoIndex.tiss, literal('')),
                            cast(InsumoIndex.aliquota, Numeric(14, 4)),
                        ),
                    ),
                    else_=func.concat_ws(
                        literal('|'),
                        cast(InsumoIndex.origem, Unicode(12)),
                        cast(InsumoIndex.item_id, Unicode(32)),
                        cast(InsumoIndex.aliquota, Numeric(14, 4)),
                    ),
                ),
            ],
            order_by=(
                func.coalesce(func.unix_timestamp(SimproItemCadastro.data_ref), 0).desc(),
                func.coalesce(func.unix_timestamp(SimproItemCadastro.imported_at), 0).desc(),
                func.coalesce(func.unix_timestamp(SimproItemPreco.imported_at), 0).desc(),
                func.coalesce(func.unix_timestamp(InsumoIndex.updated_at), 0).desc(),
                InsumoIndex.item_id.desc(),
            ),
        )
        .label('_rn_versao')
    )
    inner = (
        base_q.outerjoin(
            SimproItemCadastro,
            and_(
                InsumoIndex.origem == literal('SIMPRO'),
                SimproItemCadastro.id == InsumoIndex.item_id,
            ),
        ).outerjoin(
            SimproItemPreco,
            and_(
                SimproItemPreco.cadastro_id == InsumoIndex.item_id,
                SimproItemPreco.aliquota == InsumoIndex.aliquota,
            ),
        ).add_columns(rn)
    )
    ranked = inner.subquery('rn_ix')
    rc = ranked.c
    return (
        db.session.query(InsumoIndex)
        .select_from(ranked)
        .join(
            InsumoIndex,
            and_(
                InsumoIndex.origem == rc.origem,
                InsumoIndex.item_id == rc.item_id,
                InsumoIndex.aliquota == rc.aliquota,
            ),
        )
        .filter(rc._rn_versao == 1)
    )


def _simpro_grupo_versao_chave_serializado(it: dict) -> tuple[str, Decimal]:
    """Alinha ao PARTITION do dedupe SQL: mesmo TISS (texto índice) + alíquota normalizada."""
    tiss_txt = str(it.get('tiss') or '').strip()
    aq_str = _coerce_decimal(str(it.get('aliquota') or '0'))
    al_dec = Decimal(aq_str) if aq_str is not None else Decimal('0')
    al_fin = _br_norm_aliquota(al_dec) or al_dec
    return (tiss_txt, al_fin)


def _marcar_destaque_versao_mais_recente(items: list[dict]) -> None:
    """
    Para SIMPRO com mais de uma linha no resultado (mesmo TISS + alíquota),
    marca a cadastro/preço mais recente com ``destaque_versao_mais_recente``.
    Critérios alinhados a ``_wrap_insumos_index_latest_version`` / ROW_NUMBER().
    """
    grupos: dict[tuple[str, Decimal], list[dict]] = {}
    for it in items:
        if str(it.get('origem')).upper() != 'SIMPRO':
            continue
        grupos.setdefault(_simpro_grupo_versao_chave_serializado(it), []).append(it)

    candidatos = [g for g in grupos.values() if len(g) >= 2]
    if not candidatos:
        return

    ids_flat: set[int] = set()
    for grp in candidatos:
        for it in grp:
            try:
                ids_flat.add(int(it['item_id']))
            except (TypeError, ValueError, KeyError):
                continue
    if not ids_flat:
        return

    cadastros = (
        SimproItemCadastro.query.filter(SimproItemCadastro.id.in_(ids_flat)).all()
    )
    cad_por_id = {int(c.id): c for c in cadastros if c.id is not None}

    precos = SimproItemPreco.query.filter(SimproItemPreco.cadastro_id.in_(list(ids_flat))).all()
    prec_por_ca: dict[tuple[int, str | None], SimproItemPreco] = {}
    for p in precos:
        if p.cadastro_id is None:
            continue
        key = (int(p.cadastro_id), _decimal_lookup_key(p.aliquota))
        prec_por_ca[key] = p

    def _parse_upd_iso(s: object | None) -> datetime:
        if not s:
            return datetime.min.replace(microsecond=0)
        try:
            t = str(s).strip()
            if '+' not in t and t.endswith('Z'):
                t = t[:-1] + '+00:00'
            dt = datetime.fromisoformat(t)
            if getattr(dt, 'tzinfo', None):
                dt = dt.replace(tzinfo=None)
            return dt
        except Exception:
            return datetime.min.replace(microsecond=0)

    def _chave_ordem_recente(rec: dict) -> tuple:
        cid = int(rec['item_id'])
        aq_raw = Decimal(str(_coerce_decimal(str(rec.get('aliquota'))) or '0'))
        aq_key = _decimal_lookup_key(_br_norm_aliquota(aq_raw) or aq_raw)
        cad = cad_por_id.get(cid)
        pr = prec_por_ca.get((cid, aq_key))
        dt_ref_ord = cad.data_ref.toordinal() if cad and cad.data_ref else float('-inf')
        ts_c_imp = cad.imported_at.timestamp() if cad and cad.imported_at else float('-inf')
        ts_p_imp = pr.imported_at.timestamp() if pr and pr.imported_at else float('-inf')
        ts_ix = _parse_upd_iso(rec.get('updated_at')).timestamp()
        iid = int(rec['item_id'])
        # Maior tuple = mais recente (consistente com ROW_NUMBER no SQL).
        return (dt_ref_ord, ts_c_imp, ts_p_imp, ts_ix, iid)

    for grp in candidatos:
        best = max(grp, key=_chave_ordem_recente)
        best['destaque_versao_mais_recente'] = True


def _catalogo_search(filters: dict, page: int, per_page: int) -> dict:
    """
    Busca da tela de insumos priorizando `insumos_index`, que ? muito mais barato
    do que consultar views de cat?logo em bases grandes.
    """
    import hashlib

    filter_key = hashlib.md5(str(sorted(filters.items())).encode()).hexdigest()[:16]
    versao_tabela_f = (filters.get('versao_tabela') or '').strip()
    dedupe_latest = bool(filters.get('versao_unica')) and not versao_tabela_f
    base_query = (
        _wrap_insumos_index_latest_version(filters)
        if dedupe_latest
        else _apply_insumo_filters(InsumoIndex.query, filters)
    )
    query = base_query.order_by(
        InsumoIndex.origem.asc(),
        InsumoIndex.item_id.asc(),
        InsumoIndex.aliquota.asc(),
    )
    cache_key = f'insumos_index_{filter_key}'
    total = _get_cached_count(cache_key, lambda: base_query.order_by(None).count())

    if total <= 0:
        sp = _catalogo_search_bras_cadastro_preco_fallback(filters, page, per_page)
        if sp is not None:
            return sp
        if page == 1:
            simpro_xuf = _simpro_cross_uf_preview_search(filters, per_page)
            if simpro_xuf is not None:
                return simpro_xuf
        return {
            'items': [],
            'empty_hint': _build_empty_catalog_hint(filters),
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': 0,
                'pages': 0,
            }
        }

    rows = query.offset(max(page - 1, 0) * per_page).limit(per_page).all()
    related_context = _prefetch_insumo_related(rows)
    serialized = [
        _serialize_insumo_index(row, include_related=False, related_context=related_context)
        for row in rows
    ]
    _marcar_destaque_versao_mais_recente(serialized)

    payload = {
        'items': serialized,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': math.ceil(total / per_page) if per_page else 0,
        }
    }
    if (
        not serialized
        and page == 1
        and (filters.get('origem') or '').upper() != 'SIMPRO'
    ):
        sp = _catalogo_search_bras_cadastro_preco_fallback(filters, page, per_page)
        if sp is not None:
            return sp
    return payload

def _build_empty_catalog_hint(filters: dict) -> str | None:
    selected_uf = (filters.get('uf_referencia') or '').strip().upper()
    if not selected_uf:
        return None

    relaxed_filters = dict(filters)
    relaxed_filters.pop('uf_referencia', None)

    query = _apply_insumo_filters(InsumoIndex.query, relaxed_filters).limit(50)
    rows = query.all()
    if not rows:
        return None

    origin_labels: list[str] = []
    uf_codes: list[str] = []
    for row in rows:
        origem = (getattr(row, 'origem', None) or '').strip().upper()
        if origem and origem not in origin_labels:
            origin_labels.append(origem)
        for code in _decode_uf_codes(getattr(row, 'uf_referencia', None)):
            if code not in uf_codes:
                uf_codes.append(code)

    if selected_uf in uf_codes:
        return None

    origem_text = ', '.join(origin_labels) if origin_labels else 'os itens'
    if uf_codes:
        uf_text = ', '.join(uf_codes)
        return f'Nenhum item encontrado para a UF {selected_uf}. Com os demais filtros atuais, há resultados em: {uf_text} ({origem_text}).'

    return f'Nenhum item encontrado para a UF {selected_uf} com os filtros atuais.'


def _simpro_cross_uf_preview_has_narrow_filters(filters: dict) -> bool:
    """
    Preview entre UFs só faz sentido com algum critério além de Origem+UF (senão a consulta
    sem UF seria arbitrária e pesada). Aceita termo (tokens), códigos, fabricante, versão, alíquota.
    """
    if any(filters.get(k) for k in ('tuss', 'tiss', 'anvisa')):
        return True
    tokens = filters.get('tokens') or []
    if isinstance(tokens, list) and len(tokens) > 0:
        return True
    if (filters.get('fabricante') or '').strip():
        return True
    vt = (filters.get('versao_tabela') or '').strip()
    if vt and vt.lower() not in ('todas', 'all', '*'):
        return True
    if filters.get('aliquota') is not None:
        return True
    return False


def _simpro_cross_uf_preview_search(filters: dict, per_page: int) -> dict | None:
    """
    Se não há linha no índice para a UF filtrada, mas o mesmo critério (TISS, termo ``q``, etc.)
    existe em outra UF, devolve linhas SIMPRO para o usuário abrir Detalhes e registrar preço
    manual na UF dos filtros.
    """
    origem = (filters.get('origem') or '').upper()
    if origem not in ('SIMPRO', ''):
        return None
    uf_sel = (filters.get('uf_referencia') or '').strip().upper()
    if not uf_sel:
        return None
    if not _simpro_cross_uf_preview_has_narrow_filters(filters):
        return None

    relaxed = dict(filters)
    relaxed.pop('uf_referencia', None)
    relaxed['origem'] = 'SIMPRO'
    query = _apply_insumo_filters(InsumoIndex.query, relaxed).filter(InsumoIndex.origem == 'SIMPRO')
    query = query.order_by(InsumoIndex.item_id.desc())
    limit_n = min(max(int(per_page or 50), 1), 50)
    rows = query.limit(limit_n).all()
    if not rows:
        return None

    related_context = _prefetch_insumo_related(rows)
    serialized = [
        _serialize_insumo_index(row, include_related=False, related_context=related_context)
        for row in rows
    ]
    for s in serialized:
        s['cross_uf_preview'] = True
    _marcar_destaque_versao_mais_recente(serialized)
    total = len(serialized)
    return {
        'items': serialized,
        'pagination': {
            'page': 1,
            'per_page': per_page,
            'total': total,
            'pages': 1,
        },
        'cross_uf_preview': True,
        'empty_hint': _build_empty_catalog_hint(filters),
    }


def _catalogo_fetch_all(filters: dict, limit: int | None = None) -> list[dict]:
    include_bras = filters.get('origem') in (None, 'BRAS')
    include_simpro = filters.get('origem') in (None, 'SIMPRO')
    items: list[dict] = []
    remaining = limit

    if include_bras and (remaining is None or remaining > 0):
        query = _catalogo_filter_bras(CatalogoBrasindice.query, filters)
        query = query.order_by(CatalogoBrasindice.produto_nome.asc(), CatalogoBrasindice.item_id.asc())
        if remaining is not None and remaining > 0:
            query = query.limit(remaining)
        bras_rows = query.all()
        items.extend(_serialize_catalogo_bras(row) for row in bras_rows)
        if remaining is not None:
            remaining = max(remaining - len(bras_rows), 0)

    if include_simpro and (remaining is None or remaining > 0):
        query = _catalogo_filter_simpro(CatalogoSimpro.query, filters)
        query = query.order_by(CatalogoSimpro.descricao.asc(), CatalogoSimpro.item_id.asc())
        if remaining is not None and remaining > 0:
            query = query.limit(remaining)
        simpro_rows = query.all()
        items.extend(_serialize_catalogo_simpro(row) for row in simpro_rows)
        if remaining is not None:
            remaining = max(remaining - len(simpro_rows), 0)

    return items[:limit] if limit is not None else items


def _pick_simpro_display_price(*values: Decimal | None) -> Decimal | None:
    fallback: Decimal | None = None
    for value in values:
        if value is None:
            continue
        decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
        if fallback is None:
            fallback = decimal_value
        if decimal_value > 0:
            return decimal_value
    return fallback


def _split_simpro_prices(
    preco_pfb_pacote: Decimal | None,
    preco_pmc_pacote: Decimal | None,
    preco_pfb_fracao: Decimal | None,
    preco_pmc_fracao: Decimal | None,
) -> tuple[Decimal | None, Decimal | None, Decimal | None]:
    preco_pfb = _pick_simpro_display_price(preco_pfb_pacote, preco_pfb_fracao)
    preco_pmc = next((value for value in (preco_pmc_pacote, preco_pmc_fracao) if value is not None), None)
    preco_base = _pick_simpro_display_price(preco_pfb, preco_pmc)
    return preco_base, preco_pmc, preco_pfb


def _compose_simpro_description(
    descricao: str | None,
    referencia: str | None = None,
    classificacao: str | None = None,
) -> str | None:
    base = (descricao or '').strip()
    extras: list[str] = []
    referencia_value = (referencia or '').strip()
    classificacao_value = (classificacao or '').strip()
    if referencia_value and 'Ref:' not in base:
        extras.append(f"Ref: {referencia_value}")
    if classificacao_value and 'Classificação:' not in base:
        extras.append(f"Classificação: {classificacao_value}")
    if base and extras:
        return f"{base} • {' • '.join(extras)}"
    if base:
        return base
    if extras:
        return ' • '.join(extras)
    return None



def _serialize_insumo_index(
    item: 'InsumoIndex',
    *,
    preco_pmc: Decimal | None = None,
    preco_pfb: Decimal | None = None,
    include_related: bool = True,
    related_context: dict[str, dict] | None = None,
) -> dict:
    uf_codes = _decode_uf_codes(item.uf_referencia)
    uf_display = ', '.join(uf_codes) if uf_codes else item.uf_referencia
    preco_pmc_value: Decimal | None = preco_pmc if preco_pmc is not None else item.preco
    preco_pfb_value: Decimal | None = preco_pfb if preco_pfb is not None else item.preco
    preco_display_value: Decimal | None = item.preco
    codigo_simpro: str | None = None
    codigo_usuario: str | None = None
    codigo_fracao: str | None = None
    embalagem: str | None = None
    qtd_unidade: int | None = None
    simpro_ref: str | None = None
    simpro_status: str | None = None
    simpro_descricao: str | None = None
    simpro_split: SimproItemCadastro | None = None

    if include_related and item.origem == 'BRAS':
        bras_row = BrasItemNormalized.query.get(item.item_id)
        if bras_row:
            split_preco = None
            ean = (bras_row.ean or '').strip()
            produto_codigo = (bras_row.produto_codigo or '').strip()
            apresentacao_codigo = (bras_row.apresentacao_codigo or '').strip()
            if ean and produto_codigo and apresentacao_codigo and item.aliquota is not None:
                cad = (
                    BrasItemCadastro.query.filter_by(
                        edicao=(item.versao_tabela or bras_row.edicao or '').strip() or None,
                        ean=ean,
                        produto_codigo=produto_codigo,
                        apresentacao_codigo=apresentacao_codigo,
                    )
                    .first()
                )
                if cad is not None:
                    split_preco = (
                        BrasItemPreco.query.filter(
                            BrasItemPreco.cadastro_id == cad.id,
                            BrasItemPreco.aliquota == item.aliquota,
                        )
                        .first()
                    )
            if split_preco is not None:
                if split_preco.preco_pmc_unit is not None or split_preco.preco_pmc_pacote is not None:
                    preco_pmc_value = split_preco.preco_pmc_unit or split_preco.preco_pmc_pacote
                if split_preco.preco_pfb_unit is not None or split_preco.preco_pfb_pacote is not None:
                    preco_pfb_value = split_preco.preco_pfb_unit or split_preco.preco_pfb_pacote
            elif bras_row.arquivo and bras_row.linha_num is not None:
                raw_row = (
                    BrasRaw.query
                    .with_entities(BrasRaw.col07, BrasRaw.col08)
                    .filter_by(arquivo=bras_row.arquivo, linha_num=bras_row.linha_num)
                    .first()
                )
                if raw_row is not None:
                    raw_pmc = _coerce_decimal(raw_row.col07)
                    raw_pfb = _coerce_decimal(raw_row.col08)
                    if raw_pmc is not None:
                        try:
                            preco_pmc_value = Decimal(raw_pmc)
                        except (InvalidOperation, ValueError):
                            pass
                    if raw_pfb is not None:
                        try:
                            preco_pfb_value = Decimal(raw_pfb)
                        except (InvalidOperation, ValueError):
                            pass

    elif item.origem == 'SIMPRO':
        simpro_split = None
        split_preco = None
        if related_context:
            simpro_split = (related_context.get('simpro_cadastro_by_id') or {}).get(item.item_id)
            if simpro_split is not None and item.aliquota is not None:
                split_preco = (related_context.get('simpro_preco_by_key') or {}).get(
                    (int(simpro_split.id), _decimal_lookup_key(item.aliquota))
                )
        if simpro_split is None and include_related:
            simpro_split = SimproItemCadastro.query.get(item.item_id)
        if simpro_split is not None:
            if split_preco is None and include_related and item.aliquota is not None:
                split_preco = (
                    SimproItemPreco.query
                    .filter(
                        SimproItemPreco.cadastro_id == simpro_split.id,
                        SimproItemPreco.aliquota == item.aliquota,
                    )
                    .first()
                )
            preco_effective, preco_pmc_simpro, preco_pfb_simpro = _split_simpro_prices(
                split_preco.preco1 if split_preco is not None else None,
                split_preco.preco2 if split_preco is not None else None,
                split_preco.preco3 if split_preco is not None else None,
                split_preco.preco4 if split_preco is not None else None,
            )
            if preco_effective is not None:
                preco_display_value = preco_effective
            # Com linha em ``simpro_item_preco``, não usar ``item.preco`` como fallback só para um
            # dos dois — senão PMC repetia o PFB quando só um valor foi informado.
            if split_preco is not None:
                preco_pmc_value = preco_pmc_simpro
                preco_pfb_value = preco_pfb_simpro
            else:
                if preco_pmc_simpro is not None:
                    preco_pmc_value = preco_pmc_simpro
                if preco_pfb_simpro is not None:
                    preco_pfb_value = preco_pfb_simpro
            codigo_simpro = simpro_split.codigo
            codigo_usuario = simpro_split.codigo_interno
            codigo_fracao = simpro_split.codigo_alt
            embalagem = simpro_split.unidade
            qtd_unidade = simpro_split.qtd_unidade
            simpro_ref = simpro_split.referencia
            simpro_status = simpro_split.status_final
            simpro_descricao = simpro_split.descricao
        elif include_related:
            simpro_row = SimproItemNormalized.query.get(item.item_id)
            if simpro_row is not None:
                preco_effective, preco_pmc_simpro, preco_pfb_simpro = _split_simpro_prices(
                    simpro_row.preco1,
                    simpro_row.preco2,
                    simpro_row.preco3,
                    simpro_row.preco4,
                )
                if preco_effective is not None:
                    preco_display_value = preco_effective
                preco_pmc_value = preco_pmc_simpro
                preco_pfb_value = preco_pfb_simpro
                codigo_simpro = simpro_row.codigo
                codigo_usuario = simpro_row.codigo_interno
                codigo_fracao = simpro_row.codigo_alt
                embalagem = simpro_row.unidade
                qtd_unidade = simpro_row.qtd_unidade
                simpro_ref = simpro_row.referencia
                simpro_status = simpro_row.status_final
                simpro_descricao = simpro_row.descricao

    tuss_display = _format_tuss_display(item.tuss)
    return {
        'origem': item.origem,
        'item_id': item.item_id,
        'codigo_simpro': codigo_simpro,
        'codigo_usuario': codigo_usuario,
        'codigo_fracao': codigo_fracao,
        'referencia': simpro_ref if item.origem == 'SIMPRO' else None,
        'classificacao': simpro_status if item.origem == 'SIMPRO' else None,
        'embalagem': embalagem,
        'qtd_unidade': qtd_unidade,
        'tuss': tuss_display,
        'tuss_numero': tuss_display,
        'tuss_raw': item.tuss,
        'tiss': item.tiss,
        'descricao': ((simpro_descricao if item.origem == 'SIMPRO' else item.descricao) or '').strip() or None,
        'preco': _decimal_to_string(preco_display_value),
        'preco_pmc': _decimal_to_string(preco_pmc_value),
        'preco_pfb': _decimal_to_string(preco_pfb_value),
        'aliquota': _decimal_to_string(item.aliquota),
        'fabricante': (simpro_split.fabricante if item.origem == 'SIMPRO' and simpro_split is not None and simpro_split.fabricante else item.fabricante),
        'anvisa': item.anvisa,
        'versao_tabela': item.versao_tabela,
        'data_atualizacao': item.data_atualizacao.isoformat() if isinstance(item.data_atualizacao, date) else None,
        'updated_at': item.updated_at.isoformat() if isinstance(item.updated_at, datetime) else None,
        'uf_referencia': uf_display,
        'uf_referencia_codes': uf_codes,
    }


def _serialize_insumo_detail(
    origem: str,
    item: BrasItemNormalized | SimproItemNormalized | SimproItemCadastro | SimproItem,
    index_entry: InsumoIndex | None = None,
    *,
    catalog_entry: CatalogoBrasindice | CatalogoSimpro | None = None,
    selected_uf: str | None = None,
) -> dict:
    index_aliquota = _decimal_to_float(index_entry.aliquota) if index_entry else None
    index_uf = index_entry.uf_referencia if index_entry else None
    index_data = index_entry.data_atualizacao.isoformat() if isinstance(getattr(index_entry, 'data_atualizacao', None), date) else None
    index_created = index_entry.updated_at.isoformat() if isinstance(getattr(index_entry, 'updated_at', None), datetime) else None

    catalog_aliquota: str | None = None
    catalog_uf: str | None = None
    catalog_periodo: str | None = None
    catalog_data_ref: str | None = None
    catalog_updated: str | None = None
    catalog_preco_pmc: Decimal | None = None
    catalog_preco_pfb: Decimal | None = None

    if catalog_entry is not None:
        aliquota_bp = getattr(catalog_entry, 'aliquota_bp', None)
        if aliquota_bp is not None:
            catalog_aliquota = _decimal_to_float(_aliquota_bp_to_decimal(aliquota_bp))
        catalog_uf = getattr(catalog_entry, 'uf', None)
        catalog_periodo = getattr(catalog_entry, 'periodo', None) or getattr(catalog_entry, 'etag_versao', None)
        imported_at = getattr(catalog_entry, 'imported_at', None)
        if isinstance(imported_at, datetime):
            catalog_updated = imported_at.isoformat()
        data_ref = getattr(catalog_entry, 'data_ref', None)
        if isinstance(data_ref, date):
            catalog_data_ref = data_ref.isoformat()

        if isinstance(catalog_entry, CatalogoBrasindice):
            catalog_preco_pmc = catalog_entry.preco_pmc_unit or catalog_entry.preco_pmc_pacote
            catalog_preco_pfb = catalog_entry.preco_pfb_unit or catalog_entry.preco_pfb_pacote
        elif isinstance(catalog_entry, CatalogoSimpro):
            _catalog_preco_base, catalog_preco_pmc, catalog_preco_pfb = _split_simpro_prices(
                catalog_entry.preco1,
                catalog_entry.preco2,
                catalog_entry.preco3,
                catalog_entry.preco4,
            )

    def _first_defined(*values):
        return next((value for value in values if value not in (None, '')), None)

    if isinstance(item, BrasItemNormalized):
        descricao = item.produto_nome or ''
        if item.apresentacao_descricao:
            descricao = f"{descricao} • {item.apresentacao_descricao}" if descricao else item.apresentacao_descricao
        split_preco_pmc: Decimal | None = None
        split_preco_pfb: Decimal | None = None
        ean = (item.ean or '').strip()
        produto_codigo = (item.produto_codigo or '').strip()
        apresentacao_codigo = (item.apresentacao_codigo or '').strip()
        versao_lookup = (index_entry.versao_tabela if index_entry else None) or catalog_periodo or item.edicao or item.arquivo
        aliquota_lookup = None
        if index_entry is not None and index_entry.aliquota is not None:
            aliquota_lookup = index_entry.aliquota
        elif catalog_aliquota is not None:
            try:
                aliquota_lookup = Decimal(str(catalog_aliquota))
            except (InvalidOperation, ValueError, TypeError):
                aliquota_lookup = None
        if ean and produto_codigo and apresentacao_codigo and versao_lookup and aliquota_lookup is not None:
            cad = (
                BrasItemCadastro.query.filter_by(
                    edicao=str(versao_lookup).strip(),
                    ean=ean,
                    produto_codigo=produto_codigo,
                    apresentacao_codigo=apresentacao_codigo,
                )
                .first()
            )
            if cad is not None:
                split_preco = (
                    BrasItemPreco.query.filter(
                        BrasItemPreco.cadastro_id == cad.id,
                        BrasItemPreco.aliquota == aliquota_lookup,
                    )
                    .first()
                )
                if split_preco is not None:
                    split_preco_pmc = split_preco.preco_pmc_unit or split_preco.preco_pmc_pacote
                    split_preco_pfb = split_preco.preco_pfb_unit or split_preco.preco_pfb_pacote
        uf_codes = _combine_uf_codes(catalog_uf, index_uf)
        uf_display = ', '.join(uf_codes) if uf_codes else _first_defined(catalog_uf, index_uf, selected_uf)
        return {
            'origem': 'BRAS',
            'item_id': item.id,
            'tuss': item.produto_codigo,
            'tiss': item.apresentacao_codigo,
            'anvisa': item.registro_anvisa,
            'descricao': descricao,
            'preco': _decimal_to_float(_first_defined(split_preco_pmc, catalog_preco_pmc, item.preco_pmc_unit, item.preco_pmc_pacote)),
            'preco_pmc': _decimal_to_float(_first_defined(split_preco_pmc, catalog_preco_pmc, item.preco_pmc_unit, item.preco_pmc_pacote)),
            'preco_pfb': _decimal_to_float(_first_defined(split_preco_pfb, catalog_preco_pfb, item.preco_pfb_unit, item.preco_pfb_pacote)),
            'aliquota': _first_defined(catalog_aliquota, _decimal_to_float(item.aliquota_ou_ipi), index_aliquota),
            'fabricante': item.laboratorio_nome,
            'versao_tabela': _first_defined(index_entry.versao_tabela if index_entry else None, catalog_periodo, item.edicao, item.arquivo),
            'data_atualizacao': _first_defined(catalog_data_ref, index_data),
            'updated_at': _first_defined(catalog_updated, item.imported_at.isoformat() if isinstance(item.imported_at, datetime) else None, index_created),
            'created_at': None,
            'uf_referencia': uf_display,
            'uf_referencia_codes': uf_codes,
            'arquivo': item.arquivo,
            'linha_num': item.linha_num,
            'preco_pmc_pacote': _decimal_to_float(item.preco_pmc_pacote),
            'preco_pfb_pacote': _decimal_to_float(item.preco_pfb_pacote),
            'preco_pfb_unit': _decimal_to_float(item.preco_pfb_unit),
            'quantidade_embalagem': item.quantidade_embalagem,
        }

    if isinstance(item, SimproItemCadastro):
        split_preco = None
        if index_entry is not None and index_entry.aliquota is not None:
            split_preco = (
                SimproItemPreco.query
                .filter(
                    SimproItemPreco.cadastro_id == item.id,
                    SimproItemPreco.aliquota == index_entry.aliquota,
                )
                .first()
            )
        preco_effective, simpro_preco_pmc, simpro_preco_pfb = _split_simpro_prices(
            split_preco.preco1 if split_preco is not None else None,
            split_preco.preco2 if split_preco is not None else None,
            split_preco.preco3 if split_preco is not None else None,
            split_preco.preco4 if split_preco is not None else None,
        )
        uf_codes = _combine_uf_codes(catalog_uf, index_uf)
        uf_display = ', '.join(uf_codes) if uf_codes else _first_defined(catalog_uf, index_uf, selected_uf)
        tuss_digits = _format_tuss_display(None, getattr(item, 'tuss_numero', None))
        return {
            'origem': 'SIMPRO',
            'item_id': item.id,
            'codigo_simpro': item.codigo,
            'codigo_usuario': item.codigo_interno,
            'codigo_fracao': item.codigo_alt,
            'fracionavel': (item.fracionavel or '').strip().upper() or None,
            'referencia': item.referencia,
            'classificacao': item.status_final,
            'embalagem': item.unidade,
            'qtd_unidade': item.qtd_unidade,
            'tuss': tuss_digits,
            'tuss_numero': tuss_digits,
            'tuss_raw': item.tuss_numero,
            'tiss': item.codigo_alt,
            'anvisa': item.anvisa,
            'descricao': (item.descricao or '').strip() or None,
            'preco': _decimal_to_float(_first_defined(catalog_preco_pfb, simpro_preco_pfb, preco_effective)),
            'preco_pmc': _decimal_to_float(_first_defined(catalog_preco_pmc, simpro_preco_pmc)),
            'preco_pfb': _decimal_to_float(_first_defined(catalog_preco_pfb, simpro_preco_pfb, preco_effective)),
            'aliquota': _first_defined(catalog_aliquota, index_aliquota),
            'fabricante': item.fabricante,
            'versao_tabela': _first_defined(catalog_periodo, item.versao),
            'data_atualizacao': _first_defined(catalog_data_ref, item.data_ref.isoformat() if isinstance(item.data_ref, date) else None, index_data),
            'updated_at': _first_defined(catalog_updated, item.imported_at.isoformat() if isinstance(item.imported_at, datetime) else None, index_created),
            'created_at': None,
            'uf_referencia': uf_display,
            'uf_referencia_codes': uf_codes,
            'situacao': None,
            'validade_anvisa': None,
            'ean': item.ean,
        }

    if isinstance(item, SimproItemNormalized):
        preco_effective, simpro_preco_pmc, simpro_preco_pfb = _split_simpro_prices(
            item.preco1,
            item.preco2,
            item.preco3,
            item.preco4,
        )
        uf_codes = _combine_uf_codes(catalog_uf, item.uf_referencia, index_uf)
        uf_display = ', '.join(uf_codes) if uf_codes else _first_defined(catalog_uf, item.uf_referencia, index_uf, selected_uf)
        tuss_digits = _format_tuss_display(None, getattr(item, 'tuss_numero', None))
        return {
            'origem': 'SIMPRO',
            'item_id': item.id,
            'codigo_simpro': item.codigo,
            'codigo_usuario': item.codigo_interno,
            'codigo_fracao': item.codigo_alt,
            'fracionavel': (item.fracionavel or '').strip().upper() or None,
            'referencia': item.referencia,
            'classificacao': item.status_final,
            'embalagem': item.unidade,
            'qtd_unidade': item.qtd_unidade,
            'tuss': tuss_digits,
            'tuss_numero': tuss_digits,
            'tuss_raw': item.tuss_numero,
            'tiss': item.codigo_alt,
            'anvisa': item.anvisa,
            'descricao': (item.descricao or '').strip() or None,
            'preco': _decimal_to_float(_first_defined(catalog_preco_pfb, simpro_preco_pfb, preco_effective)),
            'preco_pmc': _decimal_to_float(_first_defined(catalog_preco_pmc, simpro_preco_pmc)),
            'preco_pfb': _decimal_to_float(_first_defined(catalog_preco_pfb, simpro_preco_pfb, preco_effective)),
            'aliquota': _first_defined(catalog_aliquota, index_aliquota),
            'fabricante': item.fabricante,
            'versao_tabela': _first_defined(catalog_periodo, item.versao, item.arquivo),
            'data_atualizacao': _first_defined(catalog_data_ref, item.data_ref.isoformat() if isinstance(item.data_ref, date) else None, index_data),
            'updated_at': _first_defined(catalog_updated, item.imported_at.isoformat() if isinstance(item.imported_at, datetime) else None, index_created),
            'created_at': None,
            'uf_referencia': uf_display,
            'uf_referencia_codes': uf_codes,
            'situacao': item.situacao,
            'validade_anvisa': item.validade_anvisa.isoformat() if isinstance(item.validade_anvisa, date) else None,
            'ean': item.ean,
        }

    if isinstance(item, SimproItem):  # fallback legacy
        uf_codes = _combine_uf_codes(catalog_uf, item.uf_referencia, index_uf)
        uf_display = ', '.join(uf_codes) if uf_codes else _first_defined(catalog_uf, item.uf_referencia, index_uf, selected_uf)
        tuss_digits = _format_tuss_display(item.tuss)
        return {
            'origem': origem,
            'item_id': item.id,
            'tuss': tuss_digits,
            'tuss_numero': tuss_digits,
            'tuss_raw': item.tuss,
            'tiss': item.tiss,
            'anvisa': item.anvisa,
            'descricao': item.descricao,
            'preco': _decimal_to_float(item.preco),
            'preco_pmc': _decimal_to_float(item.preco),
            'preco_pfb': _decimal_to_float(item.preco),
            'aliquota': _decimal_to_float(item.aliquota),
            'fabricante': item.fabricante,
            'versao_tabela': item.versao_tabela,
            'data_atualizacao': item.data_atualizacao.isoformat() if isinstance(item.data_atualizacao, date) else None,
            'updated_at': item.updated_at.isoformat() if isinstance(item.updated_at, datetime) else None,
            'created_at': item.created_at.isoformat() if isinstance(item.created_at, datetime) else None,
            'uf_referencia': uf_display,
            'uf_referencia_codes': uf_codes,
        }

    uf_codes = _combine_uf_codes(catalog_uf, index_uf)
    uf_display = ', '.join(uf_codes) if uf_codes else _first_defined(catalog_uf, index_uf, selected_uf)
    return {
        'origem': origem,
        'item_id': item.id,
        'tuss': None,
        'tiss': None,
        'anvisa': None,
        'descricao': '',
        'preco': None,
        'preco_pmc': None,
        'preco_pfb': None,
        'aliquota': None,
        'fabricante': None,
        'versao_tabela': None,
        'data_atualizacao': None,
        'updated_at': None,
        'created_at': None,
        'uf_referencia': uf_display,
        'uf_referencia_codes': uf_codes,
    }


def _extract_insumo_filters(args) -> dict:
    origem = (args.get('origem') or '').strip().upper()
    uf_ref = (args.get('uf_referencia') or args.get('uf') or '').strip().upper()
    aliquota_raw = (args.get('aliquota') or '').strip()
    aliquota_filter = _coerce_decimal(aliquota_raw) if aliquota_raw else None
    aliquota_value = Decimal(aliquota_filter) if aliquota_filter is not None else None
    if aliquota_value is not None:
        aliquota_value = _br_norm_aliquota(aliquota_value) or aliquota_value
    filters = {
        'origem': origem or None,
        'tuss': (args.get('tuss') or '').strip() or None,
        'tiss': (args.get('tiss') or '').strip() or None,
        'anvisa': (args.get('anvisa') or '').strip() or None,
        'fabricante': (args.get('fabricante') or '').strip() or None,
        'versao_tabela': (args.get('versao_tabela') or '').strip() or None,
        'uf_referencia': uf_ref or None,
        'aliquota': aliquota_value,
    }
    q = (args.get('q') or '').strip()
    tokens = [token.lower() for token in re.split(r'\s+', q) if token]
    filters['tokens'] = tokens[:6]
    filters['raw_q'] = q
    _vu_raw = str(args.get('versao_unica', '') or '').strip().lower()
    filters['versao_unica'] = _vu_raw in ('1', 'true', 'yes', 'y', 'on') if _vu_raw else False
    return filters


def _insumo_fulltext_search_enabled() -> bool:
    """MySQL/MariaDB: exige índice FULLTEXT — ver ``scripts/sql/insumos_index_fulltext_mysql.sql``."""
    return os.getenv('INSUMOS_FULLTEXT_SEARCH', '').strip().lower() in {'1', 'true', 'yes', 'on'}


def _insumo_token_ref_lookup_enabled() -> bool:
    """
    Por padrão **desligado**: por token, duas subconsultas IN em ``referencia`` deixam a busca por texto muito lenta.
    Ative só se precisar pesquisar código de referência SIMPRO pelo campo «Buscar termo»:
    INSUMOS_TOKEN_SEARCH_INCLUDE_REF=1
    """
    return os.getenv('INSUMOS_TOKEN_SEARCH_INCLUDE_REF', '').strip().lower() in {'1', 'true', 'yes', 'on'}


def _apply_insumo_token_filters(query, filters: dict):
    """Filtros por palavras do campo ``q`` (tokens): FULLTEXT no MySQL se configurado; senão LIKE."""
    tokens = [token for token in (filters.get('tokens') or []) if token]
    if not tokens:
        return query

    sess = getattr(query, 'session', None) or db.session
    bind = sess.bind if sess is not None else None
    dialect = (bind.dialect.name if bind is not None else '').lower()

    if dialect in {'mysql', 'mariadb'} and _insumo_fulltext_search_enabled():
        parts: list[str] = []
        for t in tokens[:6]:
            clean = re.sub(r'[^\w]', '', t, flags=re.UNICODE)
            if len(clean) >= 2:
                parts.append(f'+{clean}*')
        if parts:
            ft_str = ' '.join(parts)
            try:
                return query.filter(
                    text(
                        'MATCH(insumos_index.descricao, insumos_index.fabricante) '
                        'AGAINST (:ft IN BOOLEAN MODE)'
                    ).bindparams(ft=ft_str)
                )
            except Exception as exc:  # noqa: BLE001
                app.logger.warning(
                    'Busca FULLTEXT insumos falhou (%s); usando LIKE. Índice criado? Desative INSUMOS_FULLTEXT_SEARCH se necessário.',
                    exc,
                )

    allow_simpro_ref_lookup = (
        (filters.get('origem') or '').upper() in ('', 'SIMPRO')
        and _insumo_token_ref_lookup_enabled()
    )
    for token in tokens:
        pattern = f"%{token}%"
        simpro_ref_predicate = false()
        if allow_simpro_ref_lookup:
            simpro_ref_subquery_norm = (
                db.session.query(SimproItemNormalized.id)
                .filter(func.lower(func.coalesce(SimproItemNormalized.referencia, '')).like(pattern))
            )
            simpro_ref_subquery_split = (
                db.session.query(SimproItemCadastro.id)
                .filter(func.lower(func.coalesce(SimproItemCadastro.referencia, '')).like(pattern))
            )
            simpro_ref_predicate = and_(
                InsumoIndex.origem == 'SIMPRO',
                or_(
                    InsumoIndex.item_id.in_(simpro_ref_subquery_norm),
                    InsumoIndex.item_id.in_(simpro_ref_subquery_split),
                ),
            )
        query = query.filter(
            or_(
                func.lower(InsumoIndex.descricao).like(pattern),
                func.lower(InsumoIndex.fabricante).like(pattern),
                func.lower(func.coalesce(InsumoIndex.tuss, '')).like(pattern),
                func.lower(func.coalesce(InsumoIndex.tiss, '')).like(pattern),
                func.lower(func.coalesce(InsumoIndex.anvisa, '')).like(pattern),
                simpro_ref_predicate,
            )
        )

    return query


def _apply_insumo_filters(query, filters: dict):
    origem = filters.get('origem')
    if origem:
        query = query.filter(InsumoIndex.origem == origem)

    if filters.get('tuss'):
        query = query.filter(InsumoIndex.tuss == filters['tuss'])
    if filters.get('tiss'):
        query = query.filter(InsumoIndex.tiss == filters['tiss'])
    if filters.get('anvisa'):
        query = query.filter(InsumoIndex.anvisa == filters['anvisa'])
    if filters.get('fabricante'):
        fabricante = filters['fabricante'].lower()
        query = query.filter(func.lower(InsumoIndex.fabricante).like(f"%{fabricante}%"))
    if filters.get('versao_tabela'):
        query = query.filter(InsumoIndex.versao_tabela == filters['versao_tabela'])
    if filters.get('uf_referencia'):
        uf_target = filters['uf_referencia']
        pattern = f"%|{uf_target}|%"
        query = query.filter(
            or_(
                func.upper(InsumoIndex.uf_referencia) == uf_target,
                func.upper(func.coalesce(InsumoIndex.uf_referencia, '')).like(pattern)
            )
        )
    if filters.get('aliquota') is not None:
        aq = filters['aliquota']
        if isinstance(aq, Decimal):
            aq_dec = aq
        else:
            try:
                aq_dec = Decimal(str(aq))
            except (InvalidOperation, ValueError, TypeError):
                aq_dec = None
        if aq_dec is not None:
            tol = Decimal('0.02')
            query = query.filter(
                InsumoIndex.aliquota.isnot(None),
                InsumoIndex.aliquota >= aq_dec - tol,
                InsumoIndex.aliquota <= aq_dec + tol,
            )

    query = _apply_insumo_token_filters(query, filters)

    return query


def _insumo_summary(model_cls) -> dict:
    # O resumo do dashboard precisa refletir limpeza/importação imediatamente.
    cache_key = f"insumo_summary:{getattr(model_cls, '__tablename__', model_cls.__name__)}"
    cached = _insumo_cache.get(cache_key)
    if cached is not None:
        return cached
    updated_column = None
    for candidate in ('updated_at', 'imported_at'):
        updated_column = getattr(model_cls, candidate, None)
        if updated_column is not None:
            break

    data_column = getattr(model_cls, 'data_atualizacao', None)
    if data_column is None:
        data_column = getattr(model_cls, 'data_ref', None)

    version_column = None
    for candidate in ('versao_tabela', 'versao', 'edicao', 'arquivo'):
        version_column = getattr(model_cls, candidate, None)
        if version_column is not None:
            break

    # Uma única query com todas as agregações
    aggregations = [func.count(model_cls.id).label('total')]
    if updated_column is not None:
        aggregations.append(func.max(updated_column).label('last_updated'))
    if data_column is not None:
        aggregations.append(func.max(data_column).label('last_data'))
    if version_column is not None:
        aggregations.append(func.max(version_column).label('latest_version'))

    row = db.session.query(*aggregations).one()

    total = int(row.total) if row.total else 0
    last_updated = getattr(row, 'last_updated', None) if hasattr(row, 'last_updated') else None
    last_data = getattr(row, 'last_data', None) if hasattr(row, 'last_data') else None
    latest_version = getattr(row, 'latest_version', None) if hasattr(row, 'latest_version') else None

    result = {
        'total': int(total),
        'last_updated': last_updated,
        'last_data_ref': last_data,
        'latest_version': latest_version,
    }
    _insumo_cache[cache_key] = result
    return result


def _insumo_distinct_versions(model_cls) -> list[str]:
    cache_key = f"insumo_versions:{getattr(model_cls, '__tablename__', model_cls.__name__)}"
    cached = _insumo_cache.get(cache_key)
    if cached is not None:
        return cached
    version_column = None
    for candidate in ('versao_tabela', 'versao', 'edicao', 'arquivo'):
        version_column = getattr(model_cls, candidate, None)
        if version_column is not None:
            break

    if version_column is None:
        return []

    rows = (
        db.session.query(version_column)
        .filter(version_column.isnot(None))
        .distinct()
        .order_by(version_column)
        .all()
    )
    result = [row[0] for row in rows if row[0]]
    _insumo_cache[cache_key] = result
    return result


def _table_exists(table_name: str) -> bool:
    try:
        with db.engine.connect() as connection:
            return bool(db.engine.dialect.has_table(connection, table_name))
    except Exception:
        return False


def _get_teto_map(codigos: list[str], operadora_id: int | None = None) -> dict[str, 'CbhpmTeto']:
    """
    Retorna mapa de tetos CBHPM por código

    Args:
        codigos: Lista de códigos CBHPM
        operadora_id: ID da operadora (obrigatório para multi-operadora)
    """
    unique_codes = {str(c or '').strip().upper() for c in codigos if str(c or '').strip()}
    if not unique_codes:
        return {}

    query = CbhpmTeto.query.filter(CbhpmTeto.codigo.in_(unique_codes))

    # Se operadora_id fornecida, filtrar por ela
    if operadora_id:
        query = query.filter(CbhpmTeto.operadora_id == operadora_id)

    rows = query.all()
    return {row.codigo.upper(): row for row in rows}


def _get_user_operadoras_list():
    """
    Retorna lista de operadoras ativas filtradas pelo usuário logado.

    Se o usuário tiver operadoras associadas, retorna apenas as operadoras dele.
    Se não tiver operadoras associadas (admin geral), retorna todas as operadoras ativas.

    Returns:
        List[Operadora]: Lista de operadoras ativas que o usuário pode acessar
    """
    # Buscar todas as operadoras ativas
    query = Operadora.query.filter_by(status='Ativa').order_by(Operadora.nome)

    # Se o usuário tem operadoras específicas associadas, filtrar por elas
    if hasattr(g, 'current_user') and g.current_user:
        user_operadoras = g.current_user.operadoras
        if user_operadoras:
            # Filtrar apenas pelas operadoras do usuário
            operadora_ids = [op.id for op in user_operadoras]
            query = query.filter(Operadora.id.in_(operadora_ids))

    return query.all()


def _load_data_local_infile(engine, table_name: str, columns: list[str], file_path: Path, delimiter: str,
                             quotechar: str | None, skip_lines: int, extra_assignments: dict[str, str | None],
                             charset: str | None) -> int:
    bindings = [f"@col{idx}" for idx in range(len(columns))]
    set_clauses: list[str] = []
    params: dict[str, str | None] = {'file_path': str(file_path), 'delimiter': delimiter}
    if quotechar:
        params['enclosed'] = quotechar

    for name, binding in zip(columns, bindings):
        if name in DECIMAL_FIELDS:
            normalized = f"REPLACE({binding}, ',', '.')"
            set_clauses.append(
                f"{name} = CASE WHEN {normalized} REGEXP '^-?[0-9]+(\\.[0-9]+)?$' "
                f"THEN NULLIF({normalized}, '') ELSE NULL END"
            )
        else:
            set_clauses.append(f"{name} = NULLIF({binding}, '')")

    for key, value in extra_assignments.items():
        param_key = f"extra_{key}"
        set_clauses.append(f"{key} = :{param_key}")
        params[param_key] = value

    sql_parts = [
        "LOAD DATA LOCAL INFILE :file_path",
        f"INTO TABLE {table_name}",
        "FIELDS TERMINATED BY :delimiter",
    ]
    if charset:
        sql_parts.insert(2, f"CHARACTER SET {charset}")
    if quotechar:
        sql_parts.append("OPTIONALLY ENCLOSED BY :enclosed")
    sql_parts.append("LINES TERMINATED BY '\n'")
    if skip_lines:
        sql_parts.append(f"IGNORE {skip_lines} LINES")
    sql_parts.append(f"({', '.join(bindings)})")
    sql_parts.append(f"SET {', '.join(set_clauses)}")
    sql = "\n".join(sql_parts)

    with engine.begin() as conn:
        result = conn.exec_driver_sql(sql, params)
        return result.rowcount or 0


def _fallback_delimited(model_cls, columns: list[str], file_path: Path, delimiter: str, quotechar: str | None,
                        skip_header: bool, extra_assignments: dict[str, object | None],
                        encodings: list[str]) -> int:
    delimiter = delimiter or ';'
    quotechar = quotechar or '"'
    tried: list[str] = []

    for encoding in encodings:
        try:
            created = 0
            with file_path.open('r', encoding=encoding, newline='') as fh:
                reader = csv.reader(fh, delimiter=delimiter, quotechar=quotechar)
                if skip_header:
                    next(reader, None)
                rows = []
                for raw_row in reader:
                    record: dict[str, object | None] = {}
                    for idx, col in enumerate(columns):
                        value = raw_row[idx] if idx < len(raw_row) else ''
                        value = value.strip() if isinstance(value, str) else value
                        if not value:
                            record[col] = None
                        elif col in DECIMAL_FIELDS:
                            coerced = _coerce_decimal(value)
                            record[col] = Decimal(coerced) if coerced is not None else None
                        elif col in DATE_FIELDS:
                            record[col] = _coerce_date(value)
                        else:
                            record[col] = value
                    record.update(extra_assignments)
                    rows.append(model_cls(**record))
                if rows:
                    db.session.bulk_save_objects(rows)
                    db.session.commit()
                    created = len(rows)
            return created
        except UnicodeDecodeError:
            tried.append(encoding)
            db.session.rollback()
            continue

    tried_display = ', '.join(tried) if tried else 'utf-8'
    raise click.ClickException(
        f'Não foi possível decodificar o arquivo com as codificações testadas ({tried_display}). '
        'Informe a codificação correta ou converta o arquivo para UTF-8.'
    )


def _handle_delimited_import(*, model_cls, table_name: str, file_path: Path, versao: str,
                             data_ref: date | None, delimiter: str, quotechar: str | None,
                             columns_cfg: list[str] | None, skip_header: bool, use_load_data: bool,
                             truncate: bool, encoding: str | None,
                             extra_assignments: dict[str, object | None]) -> int:
    if truncate:
        db.session.query(model_cls).delete(synchronize_session=False)
        db.session.commit()

    encodings = _build_encoding_list(encoding)
    chosen_encoding = encodings[0] if encodings else 'utf-8-sig'
    header: list[str] | None = None
    if skip_header:
        effective_delimiter = delimiter or ';'
        effective_quotechar = (quotechar or '"') if quotechar is not None else '"'
        raw_header: list[str] = []
        try:
            for encoding_option in encodings:
                try:
                    with file_path.open('r', encoding=encoding_option, newline='') as fh:
                        reader = csv.reader(fh, delimiter=effective_delimiter, quotechar=effective_quotechar)
                        raw_header = next(reader, [])
                        chosen_encoding = encoding_option
                        break
                except UnicodeDecodeError:
                    db.session.rollback()
                    raw_header = []
                    continue
        except Exception:
            raw_header = []

        header = [_normalize_column_token(h) for h in raw_header]
        if not _columns_valid_for_model(model_cls, header):
            header = None
            skip_header = False

    default_cols = BRAS_DEFAULT_COLUMNS if table_name == 'bras_item' else SIMPRO_DEFAULT_COLUMNS
    columns = _resolve_columns(columns_cfg, default_cols, header)

    has_decimal_columns = any(col in DECIMAL_FIELDS for col in columns)
    if has_decimal_columns:
        use_load_data = False

    inserted = 0
    if use_load_data:
        try:
            inserted = _load_data_local_infile(
                db.engine,
                table_name,
                columns,
                file_path,
                delimiter,
                quotechar,
                1 if skip_header else 0,
                {k: (_decimal_to_string(v) if isinstance(v, (Decimal, float)) else (v.isoformat() if isinstance(v, date) else v)) for k, v in extra_assignments.items()},
                _encoding_to_mysql_charset(chosen_encoding),
            )
        except Exception:
            db.session.rollback()
            inserted = 0

    if not inserted:
        inserted = _fallback_delimited(
            model_cls,
            columns,
            file_path,
            delimiter,
            quotechar,
            skip_header,
            extra_assignments,
            encodings,
        )
    return inserted


def _handle_fixed_import(*, model_cls, file_path: Path, versao: str, data_ref: date | None,
                         map_config: dict, truncate: bool,
                         extra_assignments: dict[str, object | None]) -> int:
    columns_cfg = map_config.get('columns') or []
    if not columns_cfg:
        raise click.ClickException('Arquivo de mapeamento precisa definir "columns".')
    if truncate:
        db.session.query(model_cls).delete(synchronize_session=False)
        db.session.commit()

    extra_assignments_local = dict(extra_assignments)

    rows = []
    encoding = map_config.get('encoding', 'utf-8-sig')
    multiplier = Decimal(str(map_config.get('decimal_divisor', '100')))
    with file_path.open('r', encoding=encoding) as fh:
        for raw_line in fh:
            record: dict[str, object | None] = {}
            for cfg in columns_cfg:
                name = cfg.get('name')
                if not name:
                    continue
                start = int(cfg.get('start', 1)) - 1
                length = int(cfg.get('length', 0))
                value = raw_line[start:start + length].strip()
                divides_by = cfg.get('divide_by')
                if divides_by:
                    try:
                        divisor = Decimal(str(divides_by))
                    except Exception:
                        divisor = multiplier
                else:
                    divisor = multiplier if name in DECIMAL_FIELDS or cfg.get('type') == 'decimal' else Decimal('1')

                if not value:
                    record[name] = None
                elif name in DECIMAL_FIELDS or cfg.get('type') == 'decimal':
                    coerced = _coerce_decimal(value)
                    if coerced is None:
                        record[name] = None
                    else:
                        record[name] = (Decimal(coerced) / divisor) if divisor else Decimal(coerced)
                elif name in DATE_FIELDS or cfg.get('type') == 'date':
                    record[name] = _coerce_date(value)
                else:
                    record[name] = value
            record.update(extra_assignments_local)
            rows.append(model_cls(**record))
    if rows:
        db.session.bulk_save_objects(rows)
        db.session.commit()
    return len(rows)


def _run_insumo_import(resource: str, model_cls, table_name: str, file_path: Path, versao: str,
                       data_str: str | None, fmt: str, delimiter: str, quotechar: str | None,
                       map_path: Path | None, no_header: bool, truncate: bool, encoding: str | None,
                       uf_referencia: str | None, aliquota: Decimal | None) -> None:
    file_path = file_path.resolve()
    if not file_path.exists():
        raise click.ClickException(f'Arquivo não encontrado: {file_path}')

    data_ref = _coerce_date(data_str)
    map_config: dict = {}
    if map_path:
        try:
            map_config = json.loads(map_path.read_text(encoding='utf-8'))
        except json.JSONDecodeError as exc:
            raise click.ClickException(f'Não foi possível ler o arquivo de mapeamento: {exc}') from exc
        if not isinstance(map_config, dict):
            raise click.ClickException('Arquivo de mapeamento deve conter um objeto JSON na raiz.')

    delimiter = _normalize_delimiter(map_config.get('delimiter', delimiter)) if fmt == 'delimited' else delimiter
    quotechar_cfg = map_config.get('quotechar') if fmt == 'delimited' else None
    if quotechar_cfg is not None:
        quotechar = quotechar_cfg
    if quotechar is not None and not str(quotechar).strip():
        quotechar = None

    encoding_cfg = map_config.get('encoding') if fmt == 'delimited' else map_config.get('encoding')
    if isinstance(encoding_cfg, str) and encoding_cfg.strip():
        encoding = encoding_cfg.strip()
    elif isinstance(encoding_cfg, list) and encoding_cfg:
        encoding = str(encoding_cfg[0]).strip() or encoding

    skip_header = map_config.get('skip_header') if fmt == 'delimited' and 'skip_header' in map_config else (not no_header)
    columns_cfg = map_config.get('columns') if fmt == 'delimited' else map_config.get('columns')

    base_assignments: dict[str, object | None] = {
        'versao_tabela': versao,
        'data_atualizacao': data_ref,
    }
    if uf_referencia:
        base_assignments['uf_referencia'] = uf_referencia
    if aliquota is not None:
        base_assignments['aliquota'] = aliquota

    extra_from_map = map_config.get('extra') if isinstance(map_config.get('extra'), dict) else {}

    merged_assignments = dict(extra_from_map)
    merged_assignments.update(base_assignments)

    for key, value in list(merged_assignments.items()):
        if key in DECIMAL_FIELDS and value is not None:
            if isinstance(value, Decimal):
                continue
            if isinstance(value, (int, float)):
                merged_assignments[key] = Decimal(str(value))
            else:
                coerced = _coerce_decimal(str(value))
                merged_assignments[key] = Decimal(coerced) if coerced is not None else None

    if fmt == 'delimited':
        _handle_delimited_import(
            model_cls=model_cls,
            table_name=table_name,
            file_path=file_path,
            versao=versao,
            data_ref=data_ref,
            delimiter=_normalize_delimiter(delimiter or ';'),
            quotechar=quotechar,
            columns_cfg=columns_cfg,
            skip_header=bool(skip_header),
            use_load_data=not map_config.get('disable_load_data', False),
            truncate=truncate,
            encoding=encoding,
            extra_assignments=merged_assignments,
        )
    else:
        if not map_path:
            raise click.ClickException('Formato fixed requer arquivo de mapeamento (--map).')
        _handle_fixed_import(
            model_cls=model_cls,
            file_path=file_path,
            versao=versao,
            data_ref=data_ref,
            map_config=map_config,
            truncate=truncate,
            extra_assignments=merged_assignments,
        )


def _common_import_options(func):
    func = click.option('--truncate', is_flag=True, default=False, help='Limpa a tabela antes de importar.')(func)
    func = click.option('--no-header', is_flag=True, default=False, help='Arquivo sem cabeçalho (delimited).')(func)
    func = click.option('--map', 'map_path', type=click.Path(exists=True, dir_okay=False, path_type=Path), help='Arquivo JSON com configuração.')(func)
    func = click.option('--quotechar', default='"', show_default=True, help='Delimitador de texto (apenas delimited).')(func)
    func = click.option('--delimiter', default=';', show_default=True, help='Delimitador (apenas delimited).')(func)
    func = click.option('--lines-terminated', 'lines_terminated', default='\n', show_default=True, help='Terminador de linha do arquivo.')(func)
    func = click.option('--encoding', default=None, help='Codificação do arquivo (tenta auto se omitido).')(func)
    func = click.option('--uf', 'uf_referencia', default=None, help='UF de referência da tabela importada.')(func)
    func = click.option('--aliquota', default=None, help='Alíquota associada à tabela (percentual).')(func)
    func = click.option('--format', 'fmt', type=click.Choice(['delimited', 'fixed', 'json']), default='delimited', show_default=True)(func)
    func = click.option('--data', 'data_str', required=False, help='Data de atualização (YYYY-MM-DD).')(func)
    func = click.option('--versao', required=True, help='Versão de referência da tabela.')(func)
    func = click.option('--file', 'file_path', type=click.Path(exists=True, dir_okay=False, path_type=Path), required=True)(func)
    return func


@app.cli.command('bras:import')
@_common_import_options
def bras_import(file_path: Path, versao: str, data_str: str | None, fmt: str, delimiter: str,
                quotechar: str, map_path: Path | None, no_header: bool, truncate: bool,
                encoding: str | None, uf_referencia: str | None, aliquota: str | None,
                lines_terminated: str) -> None:
    """Importa arquivo da Brasíndice (pipeline staging + materialização)."""
    uf_value = (uf_referencia or '').strip().upper() or None
    aliquota_value: Decimal | None = None
    if aliquota:
        aliquota_str = _coerce_decimal(aliquota)
        if aliquota_str is None:
            raise click.ClickException('Valor de alíquota inválido.')
        aliquota_value = Decimal(aliquota_str)

    file_path = file_path.resolve()
    if not file_path.exists():
        raise click.ClickException(f'Arquivo não encontrado: {file_path}')

    map_config: dict = {}
    if map_path:
        try:
            map_config = json.loads(map_path.read_text(encoding='utf-8'))
        except json.JSONDecodeError as exc:
            raise click.ClickException(f'Não foi possível ler o arquivo de mapeamento: {exc}') from exc
        if not isinstance(map_config, dict):
            raise click.ClickException('Arquivo de mapeamento deve conter um objeto JSON na raiz.')

    delimiter = map_config.get('delimiter', delimiter) if fmt == 'delimited' else delimiter
    quote_cfg = map_config.get('quotechar') if fmt == 'delimited' else None
    if quote_cfg is not None:
        quotechar = quote_cfg
    if quotechar is not None and not str(quotechar).strip():
        quotechar = None

    encoding_cfg = map_config.get('encoding')
    if isinstance(encoding_cfg, str) and encoding_cfg.strip():
        encoding = encoding_cfg.strip()

    line_cfg = map_config.get('lines_terminated') or map_config.get('line_terminator')
    if line_cfg:
        lines_terminated = line_cfg

    skip_header_cfg = map_config.get('skip_header') if 'skip_header' in map_config else None
    skip_header = bool(skip_header_cfg) if skip_header_cfg is not None else (not no_header)

    result = _import_bras(
        file_path=file_path,
        versao=versao,
        data_ref=data_str,
        fmt=fmt,
        delimiter=_normalize_delimiter(delimiter) if fmt == 'delimited' else delimiter,
        quotechar=quotechar,
        line_terminator=lines_terminated or '\n',
        skip_header=skip_header,
        encoding=encoding,
        map_config=map_config,
        truncate=truncate,
        uf_default=uf_value,
        uf_values=[uf_value] if uf_value else None,
        aliquota_default=aliquota_value,
    )

    click.echo(f"Brasíndice importado: arquivo={result['arquivo']} linhas_raw={result['linhas_raw']} materializadas={result['linhas_materializadas']}")


@app.cli.command('bras:analyze-delta')
@click.argument('file_path', type=click.Path(exists=True, path_type=Path))
@click.option('--delimiter', '-d', default=',', help='Delimitador de colunas')
@click.option('--encoding', '-e', default='latin-1', help='Encoding do arquivo')
def bras_analyze_delta(file_path: Path, delimiter: str, encoding: str) -> None:
    """Analisa diferenças entre arquivo Brasíndice e dados existentes."""
    file_path = file_path.resolve()
    
    click.echo(f"Analisando arquivo: {file_path.name}")
    click.echo("Comparando com dados existentes no banco...")
    
    result = _analyze_bras_delta(
        file_path=file_path,
        delimiter=_normalize_delimiter(delimiter),
        encoding=encoding,
    )
    
    click.echo("\n" + "=" * 50)
    click.echo("RESULTADO DA ANÁLISE")
    click.echo("=" * 50)
    click.echo(f"Total de itens no arquivo:    {result['total_arquivo']:,}")
    click.echo(f"Total de itens existentes:    {result['total_existente']:,}")
    click.echo(f"Itens NOVOS:                  {result['novos']:,}")
    click.echo(f"Itens ALTERADOS:              {result['alterados']:,}")
    click.echo(f"Itens inalterados:            {result['inalterados']:,}")
    click.echo(f"Itens removidos (no banco):   {result['removidos']:,}")
    
    if result['detalhes_alterados']:
        click.echo("\n--- Amostra de itens ALTERADOS ---")
        for item in result['detalhes_alterados'][:10]:
            click.echo(f"  EAN: {item['ean']} | {item['produto'][:30]}")
            for diff in item.get('diferencas', []):
                click.echo(f"       {diff}")
    
    if result['detalhes_novos']:
        click.echo(f"\n--- Amostra de itens NOVOS ({len(result['detalhes_novos'])} mostrados) ---")
        for item in result['detalhes_novos'][:5]:
            click.echo(f"  EAN: {item['ean']} | {item['produto'][:30]} | PMC: {item['pmc_pacote']}")


@app.cli.command('bras:import-delta')
@click.argument('file_path', type=click.Path(exists=True, path_type=Path))
@click.argument('versao')
@click.option('--delimiter', '-d', default=',', help='Delimitador de colunas')
@click.option('--encoding', '-e', default='latin-1', help='Encoding do arquivo')
@click.option('--skip-header', is_flag=True, help='Pular primeira linha (cabeçalho)')
@click.option('--catalog-file', type=click.Path(exists=True, path_type=Path), default=None, help='Arquivo catálogo Brasíndice (TXT D INCL).')
@click.option('--catalog-encoding', default='latin-1', help='Encoding do arquivo catálogo.')
@click.option('--catalog-delimiter', default=';', help='Delimitador do arquivo catálogo.')
@click.option('--catalog-prev-versao', default=None, help='Versão anterior do snapshot de catálogo a comparar.')
def bras_import_delta(
    file_path: Path,
    versao: str,
    delimiter: str,
    encoding: str,
    skip_header: bool,
    catalog_file: Path | None,
    catalog_encoding: str,
    catalog_delimiter: str,
    catalog_prev_versao: str | None,
) -> None:
    """Importa apenas itens novos ou alterados da Brasíndice (incremental)."""
    file_path = file_path.resolve()
    
    click.echo(f"Importação incremental: {file_path.name}")
    click.echo(f"Versão: {versao}")
    click.echo("Processando...")
    
    result = _import_bras_delta(
        file_path=file_path,
        versao=versao,
        delimiter=_normalize_delimiter(delimiter),
        encoding=encoding,
        skip_header=skip_header,
        catalog_file=catalog_file.resolve() if catalog_file else None,
        catalog_encoding=catalog_encoding,
        catalog_delimiter=_normalize_delimiter(catalog_delimiter),
        previous_catalog_version=catalog_prev_versao,
    )
    
    click.echo("\n" + "=" * 50)
    click.echo("IMPORTAÇÃO INCREMENTAL CONCLUÍDA")
    click.echo("=" * 50)
    click.echo(f"Versão:                {result['versao']}")
    click.echo(f"Itens novos:           {result['novos']:,}")
    click.echo(f"  - Importados:        {result['novos_importados']:,}")
    click.echo(f"Itens atualizados:     {result['alterados']:,}")
    click.echo(f"Total processado:      {result['total_processado']:,}")
    if catalog_file:
        click.echo(f"Catálogo atual:        {result['catalog_current_version']}")
        click.echo(f"Catálogo anterior:     {result['catalog_previous_version'] or '—'}")
        click.echo(f"Catálogo novos:        {result['catalog_new']:,}")
        click.echo(f"Catálogo alterados:    {result['catalog_changed']:,}")
        click.echo(f"Catálogo removidos:    {result['catalog_removed']:,}")


@app.cli.command('bras:truncate')
@click.option('--yes', is_flag=True, help='Obrigatório para executar a exclusão.')
def bras_truncate_all(yes: bool) -> None:
    """
    Apaga **todos** os dados de Brasíndice no banco: insumos_index (origem BRAS),
    bras_item_n, bras_raw, bras_fixed_stage e bras_catalog_snapshot.
    Não apaga SIMPRO. Depois, rode a consolidação de catálogo se o sistema não atualizar a MV sozinho.
    """
    if not yes:
        click.echo('Isso apaga toda a Brasíndice. Para confirmar:')
        click.echo('  flask --app app bras:truncate --yes')
        return
    with app.app_context():
        _delete_existing_bras_records(None, True)
        _clear_insumo_cache()
    click.echo('Concluído: dados da Brasíndice removidos (índice + tabelas de staging, catálogo, cadastro+preço split).')


@app.cli.command('bras:backfill-cadastro-preco')
@click.option('--dry-run', is_flag=True, help='Só exibe contadores, não grava.')
def cli_bras_backfill_cadastro_preco(dry_run: bool) -> None:
    """Cria/ajusta tabelas `bras_item_cadastro` e `bras_item_preco` a partir de `bras_item_n`."""
    with app.app_context():
        st = _backfill_bras_cadastro_preco_from_bras_n(dry_run=dry_run)
    for k, v in st.items():
        click.echo(f'  {k}: {v}')


@app.cli.command('bras:import-precos')
@click.argument('file_path', type=click.Path(exists=True, path_type=Path))
@click.argument('edicao', type=str)
@click.option('--aliquota', required=True, help='Ex.: 20.5 ou 18')
@click.option('--delimiter', default=',')
@click.option('--encoding', default='latin-1')
@click.option('--no-legacy', is_flag=True, help='Não atualizar bras_item_n/insumos (só tabelas split).')
def cli_bras_import_precos(
    file_path: Path, edicao: str, aliquota: str, delimiter: str, encoding: str, no_legacy: bool,
) -> None:
    """Carga leve: só preços, para itens que já existem no cadastro (após `bras:backfill` ou carga canônica)."""
    a_str = _coerce_decimal(aliquota)
    if a_str is None:
        raise click.ClickException('Alíquota inválida.')
    alq = _br_norm_aliquota(Decimal(a_str))
    if alq is None:
        raise click.ClickException('Alíquota inválida.')
    with app.app_context():
        st = _import_bras_somente_precos(
            file_path=file_path,
            edicao=edicao.strip(),
            aliquota=alq,
            delimiter=delimiter,
            quotechar='"',
            encoding=encoding,
            skip_header=False,
            arquivo_fonte=file_path.name,
            update_legacy=not no_legacy,
        )
    for k, v in st.items():
        click.echo(f'  {k}: {v}')


@app.cli.command('insumos:create-indexes')
def insumos_create_indexes() -> None:
    """Cria índices otimizados para busca de insumos."""
    click.echo("Criando índices para otimização de consultas de insumos...")
    
    # Lista de índices: (nome, tabela, colunas)
    indexes = [
        # Índices para mv_catalogo_vigente_brasindice
        ("idx_bras_uf_aliquota", "mv_catalogo_vigente_brasindice", "uf, aliquota_bp"),
        ("idx_bras_produto_codigo", "mv_catalogo_vigente_brasindice", "produto_codigo"),
        ("idx_bras_apresentacao_codigo", "mv_catalogo_vigente_brasindice", "apresentacao_codigo"),
        ("idx_bras_anvisa", "mv_catalogo_vigente_brasindice", "registro_anvisa"),
        ("idx_bras_ean", "mv_catalogo_vigente_brasindice", "ean"),
        
        # Índices para mv_catalogo_vigente_simpro
        ("idx_simpro_uf_aliquota", "mv_catalogo_vigente_simpro", "uf, aliquota_bp"),
        ("idx_simpro_codigo", "mv_catalogo_vigente_simpro", "codigo"),
        ("idx_simpro_codigo_alt", "mv_catalogo_vigente_simpro", "codigo_alt"),
        ("idx_simpro_anvisa", "mv_catalogo_vigente_simpro", "anvisa"),
        ("idx_simpro_ean", "mv_catalogo_vigente_simpro", "ean"),
        
        # Índices para insumos_index (fallback)
        ("idx_insumos_origem_uf", "insumos_index", "origem, uf_referencia"),
        ("idx_insumos_origem_aliquota", "insumos_index", "origem, aliquota"),
        ("idx_insumos_fabricante", "insumos_index", "fabricante(100)"),
    ]
    
    created = 0
    skipped = 0
    errors = []
    
    for name, table, columns in indexes:
        try:
            # Verifica se tabela existe
            table_check = db.session.execute(text(f"SHOW TABLES LIKE '{table}'")).fetchone()
            if not table_check:
                click.echo(f"  ○ {name} (tabela {table} não existe)")
                skipped += 1
                continue
            
            # Verifica se índice já existe
            check_sql = text(f"SHOW INDEX FROM {table} WHERE Key_name = :name")
            existing = db.session.execute(check_sql, {'name': name}).fetchone()
            
            if existing:
                click.echo(f"  ○ {name} (já existe)")
                skipped += 1
                continue
            
            # Cria índice
            create_sql = f"CREATE INDEX {name} ON {table} ({columns})"
            db.session.execute(text(create_sql))
            db.session.commit()
            click.echo(f"  ✓ {name}")
            created += 1
        except Exception as e:
            db.session.rollback()
            error_msg = str(e)
            if 'Duplicate' in error_msg or 'already exists' in error_msg.lower():
                click.echo(f"  ○ {name} (já existe)")
                skipped += 1
            else:
                click.echo(f"  ✗ {name}: {error_msg[:80]}")
                errors.append(name)
    
    click.echo(f"\nResultado: {created} criados, {skipped} ignorados, {len(errors)} erros")
    
    if errors:
        click.echo(f"Índices com erro: {', '.join(errors)}")


@app.cli.command('insumos:optimize')
def insumos_optimize() -> None:
    """Otimiza tabelas de insumos (ANALYZE + OPTIMIZE)."""
    click.echo("Otimizando tabelas de insumos...")
    
    tables = [
        'mv_catalogo_vigente_brasindice',
        'mv_catalogo_vigente_simpro',
        'insumos_index',
        'bras_raw',
        'simpro_item_normalized',
    ]
    
    for table in tables:
        try:
            # Verifica se tabela existe
            result = db.session.execute(text(f"SHOW TABLES LIKE '{table}'")).fetchone()
            if not result:
                click.echo(f"  ○ {table} (não existe)")
                continue
            
            # ANALYZE para atualizar estatísticas
            db.session.execute(text(f"ANALYZE TABLE {table}"))
            click.echo(f"  ✓ {table} analisada")
        except Exception as e:
            click.echo(f"  ✗ {table}: {str(e)[:60]}")
    
    db.session.commit()
    click.echo("Otimização concluída!")


@app.cli.command('insumos:clear-cache')
def insumos_clear_cache() -> None:
    """Limpa caches de insumos (resumos na tela e contagens de busca por filtro)."""
    _clear_insumo_cache()
    click.echo("Cache de insumos e contagens limpo!")


@app.cli.command('insumos:sync-simpro-index')
@click.option('--purge-only', is_flag=True, help='Apenas remove linhas SIMPRO do insumos_index.')
def insumos_sync_simpro_index(purge_only: bool) -> None:
    """
    Recria entradas de busca SIMPRO em `insumos_index` a partir de `simpro_item_cadastro`
    + `simpro_item_preco`. Use após mysqldump/restore só das tabelas SIMPRO, quando a lista
    de insumos fica incompleta por UF/alíquota.
    """
    with app.app_context():
        db.session.execute(text("DELETE FROM insumos_index WHERE origem = 'SIMPRO'"))
        db.session.commit()
        click.echo('Removido SIMPRO de insumos_index.')
        if purge_only:
            _clear_insumo_cache()
            return
        _sync_simpro_insumo_index(None)
        _backfill_catalogo_simpro_identifiers()
        _clear_insumo_cache()
        total_ix = db.session.execute(
            text("SELECT COUNT(*) FROM insumos_index WHERE origem = 'SIMPRO'")
        ).scalar()
        distinct_aq = db.session.execute(
            text(
                "SELECT ROUND(aliquota, 2), COUNT(*) FROM insumos_index "
                "WHERE origem = 'SIMPRO' GROUP BY ROUND(aliquota, 2) ORDER BY ROUND(aliquota, 2)"
            )
        ).fetchall()
        click.echo(f'Reindexação SIMPRO concluída. Linhas SIMPRO em insumos_index: {total_ix}. Por alíquota (arred.):')
        for col_aq, ct in distinct_aq[:24]:
            click.echo(f'  {col_aq} -> {ct}')
        if distinct_aq and len(distinct_aq) > 24:
            click.echo(f'  … (+{len(distinct_aq) - 24} faixas)')


@app.cli.command('unlock-user')
@click.argument('email')
def unlock_user(email: str) -> None:
    """Desbloqueia um usuário por e-mail (zera tentativas e locked_until)."""
    u = Usuario.query.filter_by(email=email.strip()).first()
    if not u:
        click.echo(f"Usuário com e-mail '{email}' não encontrado.", err=True)
        raise SystemExit(1)
    u.failed_login_attempts = 0
    u.locked_until = None
    db.session.commit()
    click.echo(f"Usuário {u.email} desbloqueado.")


@app.cli.command('set-password')
@click.argument('email')
@click.argument('nova_senha', required=False)
def set_password(email: str, nova_senha: str | None) -> None:
    """Define a senha de um usuário por e-mail. Uso: flask set-password email@exemplo.com NovaSenha"""
    email = email.strip()
    u = Usuario.query.filter_by(email=email).first()
    if not u:
        click.echo(f"Usuário com e-mail '{email}' não encontrado.", err=True)
        raise SystemExit(1)
    if not nova_senha or not nova_senha.strip():
        click.echo("Informe a nova senha: flask set-password email@exemplo.com 'SuaSenha'", err=True)
        raise SystemExit(1)
    u.senha = _hash_password(nova_senha)
    u.failed_login_attempts = 0
    u.locked_until = None
    db.session.commit()
    click.echo(f"Senha de {u.email} atualizada. Faça login com a nova senha.")


@app.cli.command('simpro:import')
@_common_import_options
def simpro_import(file_path: Path, versao: str, data_str: str | None, fmt: str, delimiter: str,
                  quotechar: str, map_path: Path | None, no_header: bool, truncate: bool,
                  encoding: str | None, uf_referencia: str | None, aliquota: str | None,
                  lines_terminated: str) -> None:
    """Importa arquivo do SIMPRO."""
    del lines_terminated, delimiter, quotechar, no_header, data_str
    uf_value = (uf_referencia or '').strip().upper() or None
    aliquota_value: Decimal | None = None
    if aliquota:
        aliquota_str = _coerce_decimal(aliquota)
        if aliquota_str is None:
            raise click.ClickException('Valor de alíquota inválido.')
        aliquota_value = Decimal(aliquota_str)

    if fmt not in {'fixed', 'json'}:
        raise click.ClickException('Importação SIMPRO suporta apenas JSON ou largura fixa.')

    file_path = file_path.resolve()
    if not file_path.exists():
        raise click.ClickException(f'Arquivo não encontrado: {file_path}')

    map_config: dict = {}
    if map_path:
        try:
            map_config = json.loads(map_path.read_text(encoding='utf-8'))
        except json.JSONDecodeError as exc:
            raise click.ClickException(f'Não foi possível ler o arquivo de mapeamento: {exc}') from exc
        if not isinstance(map_config, dict):
            raise click.ClickException('Arquivo de mapeamento deve conter um objeto JSON na raiz.')
    if fmt == 'fixed' and not map_config:
        raise click.ClickException('Informe um mapa JSON contendo as posições do arquivo SIMPRO.')

    encoding_cfg = map_config.get('encoding')
    if isinstance(encoding_cfg, str) and encoding_cfg.strip():
        encoding = encoding_cfg.strip()

    result = _import_simpro(
        file_path=file_path,
        versao=versao,
        fmt=fmt,
        map_config=map_config,
        encoding=encoding,
        truncate=truncate,
        uf_default=uf_value,
        uf_values=[uf_value] if uf_value else None,
        aliquota_default=aliquota_value,
    )

    click.echo(
        f"Importação SIMPRO concluída: arquivo={result['arquivo']} linhas_raw={result['linhas_raw']} "
        f"materializadas={result['linhas_materializadas']}"
    )


@app.cli.command('aliquota:ingest')
@click.option('--fornecedor', required=True)
@click.option('--origem', type=click.Choice(['BRAS', 'SIMPRO']), required=True)
@click.option('--aliquota', 'aliquota_bp', required=True, help='Alíquota em basis points ou percentual (ex.: 1700 ou 17.0).')
@click.option('--periodo', required=True, help='Formato YYYYMM')
@click.option('--sequencia', required=True, type=int)
@click.option('--arquivo-label', required=True, help='Rótulo utilizado nas tabelas normalizadas (campo arquivo).')
@click.option('--arquivo', 'arquivo_path', type=click.Path(exists=True), default=None, help='Arquivo bruto para cálculo de hash (opcional).')
def cli_aliquota_ingest(fornecedor, origem, aliquota_bp, periodo, sequencia, arquivo_label, arquivo_path):
    lote = ingestir_arquivo(
        fornecedor=fornecedor,
        origem=origem,
        aliquota_bp=aliquota_bp,
        periodo=periodo,
        sequencia=sequencia,
        arquivo_label=arquivo_label,
        arquivo_path=Path(arquivo_path) if arquivo_path else None,
    )
    click.echo(
        f"Lote {lote.id} validado: fornecedor={lote.fornecedor}, período={lote.periodo}, sequência={lote.sequencia}, itens={lote.total_itens}"
    )


@app.cli.command('aliquota:publicar')
@click.option('--fornecedor', required=True)
@click.option('--aliquota', 'aliquota_bp', required=True)
@click.option('--periodo', required=True)
@click.option('--sequencia', required=True, type=int)
def cli_aliquota_publicar(fornecedor, aliquota_bp, periodo, sequencia):
    publicacao = publicar_lote(
        fornecedor=fornecedor,
        aliquota_bp=aliquota_bp,
        periodo=periodo,
        sequencia=sequencia,
    )
    click.echo(
        f"Lote {publicacao.lote_id} publicado em {publicacao.publicado_em:%Y-%m-%d %H:%M:%S} (etag={publicacao.etag_versao})."
    )


@app.cli.command('insumos-import-worker')
@click.option('--poll-interval', default=5, type=int, show_default=True, help='Tempo em segundos entre cada verificação de jobs pendentes.')
@click.option('--run-once', is_flag=True, help='Processa apenas um job e encerra.')
def cli_insumos_import_worker(poll_interval: int, run_once: bool) -> None:
    """Worker simples para processar importações de insumos."""

    poll_interval = max(1, poll_interval)
    _run_import_worker_loop(poll_interval=poll_interval, run_once=run_once)


class CBHPMRuleSet(db.Model):
    __tablename__ = 'cbhpm_rulesets'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), nullable=False)
    versao = db.Column(db.String(50), nullable=True)
    descricao = db.Column(db.Text, nullable=True)
    ativo = db.Column(db.Boolean, nullable=False, default=False)
    regras = db.Column(db.JSON, nullable=False, default=dict)
    criado_em = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    atualizado_em = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)


DEFAULT_CBHPM_RULES = {
    "descricao": "Regras base CBHPM",
    "porte": {
        "reducoes_simultaneos": [1.0, 0.5, 0.3, 0.2]
    },
    "auxiliares": {
        "percentuais": [0.3, 0.2, 0.1, 0.1],
        "max_por_porte": {
            "0": 0,
            "1": 0,
            "2": 1,
            "3": 2,
            "4": 2,
            "5": 3,
            "6": 3,
            "default": 2
        }
    },
    "uco": {"multiplicador": 1.0},
    "filme": {"multiplicador": 1.0}
}

class PorteValorItem(db.Model):
    __tablename__ = 'porte_valores'
    id = db.Column(db.Integer, primary_key=True)
    porte = db.Column(db.String(50), nullable=False)
    valor = db.Column(db.Numeric(12, 2), nullable=False)
    uf = db.Column(db.String(2), nullable=True)
    id_tabela = db.Column(db.Integer, db.ForeignKey('tabelas.id'), nullable=False)


class PorteAnestesicoValorItem(db.Model):
    __tablename__ = 'porte_anestesico_valores'
    id = db.Column(db.Integer, primary_key=True)
    porte_an = db.Column(db.String(50), nullable=False)
    valor = db.Column(db.Numeric(12, 2), nullable=False)
    uf = db.Column(db.String(2), nullable=True)
    id_tabela = db.Column(db.Integer, db.ForeignKey('tabelas.id'), nullable=False)


# --- 3. ROTAS (PÁGINAS) ---

@app.route('/')
@login_required
def dashboard():
    return render_template('index.html')



@app.route('/price-changes')
@login_required
def price_changes():
    """Dashboard de Alterações de Preços Recentes"""

    # Parâmetros de filtro
    days = request.args.get('days', 7, type=int)  # Últimos N dias
    origem = request.args.get('origem', '')  # BRAS ou SIMPRO
    uf = request.args.get('uf', '')  # Estado
    min_change = request.args.get('min_change', 0, type=float)  # % mínima de mudança

    price_changes = {
        'summary': {},
        'bras_changes': [],
        'simpro_changes': [],
        'top_increases': [],
        'top_decreases': []
    }

    try:
        # 1. BRAS - Comparar as 2 últimas versões de cada produto
        bras_changes = db.session.query(
            BrasItemNormalized.produto_codigo,
            BrasItemNormalized.produto_nome,
            BrasItemNormalized.apresentacao_descricao,
            BrasItemNormalized.preco_pmc_unit,
            BrasItemNormalized.imported_at,
            BrasItemNormalized.edicao
        ).filter(
            BrasItemNormalized.imported_at >= datetime.utcnow() - timedelta(days=days)
        ).order_by(
            BrasItemNormalized.produto_codigo,
            BrasItemNormalized.imported_at.desc()
        ).all()

        # Processa mudanças de BRAS
        bras_dict = {}
        for item in bras_changes:
            codigo = item[0]
            if codigo not in bras_dict:
                bras_dict[codigo] = []
            bras_dict[codigo].append(item)

        bras_result = []
        for codigo, versions in bras_dict.items():
            if len(versions) >= 2:
                current = versions[0]  # Mais recente
                previous = versions[1]  # Anterior

                preco_atual = float(current[3]) if current[3] else 0
                preco_anterior = float(previous[3]) if previous[3] else 0

                if preco_anterior > 0:
                    percentual = ((preco_atual - preco_anterior) / preco_anterior) * 100

                    if abs(percentual) >= min_change:
                        bras_result.append({
                            'origem': 'BRAS',
                            'codigo': codigo,
                            'nome': current[1],
                            'apresentacao': current[2],
                            'preco_anterior': round(preco_anterior, 2),
                            'preco_novo': round(preco_atual, 2),
                            'diferenca': round(preco_atual - preco_anterior, 2),
                            'percentual': round(percentual, 2),
                            'data_mudanca': current[4],
                            'edicao_nova': current[5],
                            'edicao_anterior': previous[5]
                        })

        price_changes['bras_changes'] = sorted(bras_result, key=lambda x: x['data_mudanca'], reverse=True)[:50]
        print(f'✅ BRAS: {len(price_changes["bras_changes"])} produtos com mudança de preço')

    except Exception as e:
        print(f'❌ Erro ao processar BRAS: {e}')
        price_changes['bras_changes'] = []

    try:
        # 2. SIMPRO - Comparar as 2 últimas versões de cada produto
        simpro_changes = db.session.query(
            SimproItemNormalized.codigo,
            SimproItemNormalized.descricao,
            SimproItemNormalized.preco1,
            SimproItemNormalized.imported_at,
            SimproItemNormalized.versao,
            SimproItemNormalized.data_ref
        ).filter(
            SimproItemNormalized.imported_at >= datetime.utcnow() - timedelta(days=days)
        ).order_by(
            SimproItemNormalized.codigo,
            SimproItemNormalized.imported_at.desc()
        ).all()

        # Processa mudanças de SIMPRO
        simpro_dict = {}
        for item in simpro_changes:
            codigo = item[0]
            if codigo not in simpro_dict:
                simpro_dict[codigo] = []
            simpro_dict[codigo].append(item)

        simpro_result = []
        for codigo, versions in simpro_dict.items():
            if len(versions) >= 2:
                current = versions[0]  # Mais recente
                previous = versions[1]  # Anterior

                preco_atual = float(current[2]) if current[2] else 0
                preco_anterior = float(previous[2]) if previous[2] else 0

                if preco_anterior > 0:
                    percentual = ((preco_atual - preco_anterior) / preco_anterior) * 100

                    if abs(percentual) >= min_change:
                        simpro_result.append({
                            'origem': 'SIMPRO',
                            'codigo': codigo,
                            'nome': current[1],
                            'apresentacao': '',
                            'preco_anterior': round(preco_anterior, 2),
                            'preco_novo': round(preco_atual, 2),
                            'diferenca': round(preco_atual - preco_anterior, 2),
                            'percentual': round(percentual, 2),
                            'data_mudanca': current[3],
                            'edicao_nova': current[4],
                            'edicao_anterior': previous[4]
                        })

        price_changes['simpro_changes'] = sorted(simpro_result, key=lambda x: x['data_mudanca'], reverse=True)[:50]
        print(f'✅ SIMPRO: {len(price_changes["simpro_changes"])} produtos com mudança de preço')

    except Exception as e:
        print(f'❌ Erro ao processar SIMPRO: {e}')
        price_changes['simpro_changes'] = []

    try:
        # 3. Resumo de mudanças
        total_bras = len(price_changes['bras_changes'])
        total_simpro = len(price_changes['simpro_changes'])

        # Top aumentos
        all_changes = price_changes['bras_changes'] + price_changes['simpro_changes']
        increases = [c for c in all_changes if c['percentual'] > 0]
        decreases = [c for c in all_changes if c['percentual'] < 0]

        price_changes['summary'] = {
            'total_mudancas': len(all_changes),
            'total_bras': total_bras,
            'total_simpro': total_simpro,
            'total_aumentos': len(increases),
            'total_reducoes': len(decreases),
            'media_aumento': round(sum(c['percentual'] for c in increases) / len(increases), 2) if increases else 0,
            'media_reducao': round(sum(c['percentual'] for c in decreases) / len(decreases), 2) if decreases else 0,
        }

        price_changes['top_increases'] = sorted(increases, key=lambda x: x['percentual'], reverse=True)[:10]
        price_changes['top_decreases'] = sorted(decreases, key=lambda x: x['percentual'])[:10]

        print(f"✅ Resumo: {price_changes['summary']}")

    except Exception as e:
        print(f'❌ Erro ao compilar resumo: {e}')

    return render_template('price_changes.html', changes=price_changes, days=days, min_change=min_change)


@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    senha_alterada = request.args.get('senha_alterada') == '1'
    disable_login_lock = os.getenv('DISABLE_LOGIN_LOCK', '0') == '1'
    auto_login = os.getenv('DEV_AUTO_LOGIN', '0') == '1'
    if auto_login and request.method == 'GET':
        usuario = Usuario.query.order_by(Usuario.id.asc()).first()
        if usuario:
            session.clear()
            session.permanent = True
            session['user_id'] = usuario.id
            session['perfil'] = usuario.perfil
            session['nome'] = usuario.nome
            nomes = [op.nome for op in usuario.operadoras]
            ids = [op.id for op in usuario.operadoras]
            session['operadora_ids'] = ids
            session['operadora_id'] = ids[0] if ids else None
            session['operadora_nomes'] = nomes
            session['operadora_nome'] = ', '.join(nomes) if nomes else None
            session['feature_insumos'] = bool(usuario.acesso_insumos) or (usuario.perfil == 'adm')
            session['feature_consulta'] = bool(usuario.acesso_consulta) or (usuario.perfil == 'adm')
            session['feature_contratos'] = bool(getattr(usuario, 'acesso_contratos', True)) or (usuario.perfil in {'adm', 'adm de contrato', 'operadora'})
            session['feature_tuss_rol'] = bool(usuario.acesso_tuss_rol) or (usuario.perfil == 'adm')
            session['login_time'] = _now_utc().isoformat()
            session['session_nonce'] = uuid4().hex
            session['login_ip'] = _get_remote_addr()
            session['password_changed_at'] = usuario.senha_atualizada_em.isoformat() if usuario.senha_atualizada_em else None
            session['must_change_senha'] = False
            session.modified = True
            return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip()
        senha = request.form.get('senha') or ''
        agora = _now_utc()
        print(f"[login] incoming email={repr(email)} senha_len={len(senha)} content_type={request.content_type}", flush=True)
        usuario = Usuario.query.filter_by(email=email).first()
        if usuario:
            ok = _verify_password(usuario.senha, senha)
        else:
            ok = False

        # Dev-only: permite login mesmo se a senha não bater quando auto-login está ligado
        if disable_login_lock and os.getenv('DEV_AUTO_LOGIN', '0') == '1':
            if not usuario:
                usuario = Usuario.query.order_by(Usuario.id.asc()).first()
            ok = bool(usuario)
            if usuario:
                print(
                    f"[login] email={usuario.email} ok={ok} "
                    f"locked_until={usuario.locked_until} failed={usuario.failed_login_attempts} "
                    f"senha_len={len(senha)} senha_repr={repr(senha)} "
                    f"hash_prefix={(usuario.senha or '')[:12]}",
                    flush=True,
                )

        if usuario and not disable_login_lock and usuario.locked_until and usuario.locked_until > agora:
            if ok:
                # Permite desbloqueio imediato com senha correta
                usuario.failed_login_attempts = 0
                usuario.locked_until = None
            else:
                minutos = max(int((usuario.locked_until - agora).total_seconds() // 60) + 1, 1)
                erro = f'Conta temporariamente bloqueada. Tente novamente em aproximadamente {minutos} minuto(s).'
                _register_audit(
                    'login.locked',
                    usuario=usuario,
                    detalhes={'locked_until': usuario.locked_until.isoformat()}
                )
                db.session.commit()
                return render_template('login.html', erro=erro, senha_alterada=senha_alterada, hide_chrome=True)

        if usuario and ok:
            # migra senhas legadas sem hash
            if not _is_password_hashed(usuario.senha):
                novo_hash = _hash_password(senha)
                usuario.senha = novo_hash
                usuario.senha_atualizada_em = usuario.senha_atualizada_em or agora
                try:
                    _append_password_history(usuario, novo_hash)
                except Exception:
                    app.logger.warning('Falha ao salvar histórico de senha para o usuário %s', usuario.email)

            usuario.failed_login_attempts = 0
            usuario.locked_until = None

            session.clear()
            session.permanent = True
            session['user_id'] = usuario.id
            session['perfil'] = usuario.perfil
            session['nome'] = usuario.nome
            nomes = [op.nome for op in usuario.operadoras]
            ids = [op.id for op in usuario.operadoras]
            session['operadora_ids'] = ids
            session['operadora_id'] = ids[0] if ids else None
            session['operadora_nomes'] = nomes
            session['operadora_nome'] = ', '.join(nomes) if nomes else None
            session['feature_insumos'] = bool(usuario.acesso_insumos) or (usuario.perfil == 'adm')
            session['feature_consulta'] = bool(usuario.acesso_consulta) or (usuario.perfil == 'adm')
            session['feature_contratos'] = bool(getattr(usuario, 'acesso_contratos', True)) or (usuario.perfil in {'adm', 'adm de contrato', 'operadora'})
            session['feature_tuss_rol'] = bool(usuario.acesso_tuss_rol) or (usuario.perfil == 'adm')
            session['login_time'] = agora.isoformat()
            session['session_nonce'] = uuid4().hex
            session['login_ip'] = _get_remote_addr()
            session['password_changed_at'] = usuario.senha_atualizada_em.isoformat() if usuario.senha_atualizada_em else None

            must_change = bool(usuario.must_reset_senha)
            last_change = usuario.senha_atualizada_em
            if not must_change:
                if last_change is None:
                    must_change = True
                else:
                    try:
                        delta = _now_utc() - last_change
                        if delta > timedelta(days=PASSWORD_EXPIRATION_DAYS):
                            must_change = True
                    except Exception:
                        must_change = True
            if must_change and not usuario.must_reset_senha:
                usuario.must_reset_senha = True
            session['must_change_senha'] = must_change

            _register_audit('login.success', usuario=usuario)
            try:
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                app.logger.error('Falha ao confirmar login: %s', exc)
                erro = 'Não foi possível concluir o login. Tente novamente em instantes.'
                session.clear()
                return render_template('login.html', erro=erro, senha_alterada=senha_alterada, hide_chrome=True)

            if must_change:
                return redirect(url_for('alterar_senha'))
            return redirect(url_for('dashboard'))

        if usuario:
            app.logger.warning('Login failure: email=%s', usuario.email)
            bloqueado = False
            if not disable_login_lock:
                usuario.failed_login_attempts = (usuario.failed_login_attempts or 0) + 1
                if usuario.failed_login_attempts >= MAX_FAILED_LOGIN_ATTEMPTS:
                    usuario.locked_until = agora + timedelta(minutes=ACCOUNT_LOCK_MINUTES)
                    usuario.failed_login_attempts = 0
                    bloqueado = True
            _register_audit(
                'login.failure',
                usuario=usuario,
                detalhes={
                    'reason': 'senha_incorreta',
                    'blocked': bloqueado,
                    'next_unlock': usuario.locked_until.isoformat() if usuario.locked_until else None,
                },
            )
            try:
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                app.logger.error('Falha ao registrar login invalido: %s', exc)
            if bloqueado:
                erro = f'Conta bloqueada por {ACCOUNT_LOCK_MINUTES} minutos após múltiplas tentativas.'
            else:
                erro = 'Credenciais inválidas.'
            return render_template('login.html', erro=erro, senha_alterada=senha_alterada, hide_chrome=True)

        _register_audit(
            'login.failure',
            email_alvo=email,
            detalhes={'reason': 'usuario_inexistente'}
        )
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        erro = 'Credenciais inválidas.'
        return render_template('login.html', erro=erro, senha_alterada=senha_alterada, hide_chrome=True)

    # GET: layout limpo
    return render_template('login.html', senha_alterada=senha_alterada, hide_chrome=True)



@app.route('/logout')
def logout():
    usuario_id = session.get('user_id')
    if usuario_id:
        try:
            usuario = Usuario.query.get(usuario_id)
        except Exception:
            usuario = None
        if usuario:
            usuario.last_logout_at = _now_utc()
            _register_audit('logout', usuario=usuario)
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
    session.clear()
    session.modified = True
    return redirect(url_for('login'))


@app.route('/health')
def health_check():
    """
    Endpoint de health check / status do sistema COMPLETO
    Mostra: banco, CPU, memória, disco, cache, jobs, usuários, alertas
    Acesso público (sem login) para monitoramento externo
    """
    import psutil
    from datetime import datetime, timedelta

    health_data = {
        'timestamp': datetime.now().isoformat(),
        'status': 'healthy',
        'checks': {},
        'alerts': [],
        'recommendations': []
    }

    # 1. Database Connection & Performance
    try:
        start = time.perf_counter()
        db.session.execute(text('SELECT 1')).scalar()
        query_time = (time.perf_counter() - start) * 1000

        # Database size
        db_size_result = db.session.execute(text(
            "SELECT ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) as size_mb "
            "FROM information_schema.tables WHERE table_schema = DATABASE()"
        )).scalar()

        health_data['checks']['database'] = {
            'status': 'ok' if query_time < 100 else 'warning',
            'message': f'Conexão OK ({query_time:.2f}ms)',
            'query_time_ms': round(query_time, 2),
            'size_mb': float(db_size_result) if db_size_result else 0
        }

        if query_time > 100:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'exclamation-triangle',
                'message': f'Database response time alto: {query_time:.2f}ms'
            })
    except Exception as e:
        health_data['status'] = 'unhealthy'
        health_data['checks']['database'] = {
            'status': 'error',
            'message': f'Erro: {str(e)}'
        }
        health_data['alerts'].append({
            'level': 'error',
            'icon': 'x-circle',
            'message': 'Banco de dados inacessível!'
        })

    # 2. Disk Space
    try:
        disk = psutil.disk_usage('/')
        disk_percent = disk.percent
        health_data['checks']['disk'] = {
            'status': 'ok' if disk_percent < 85 else ('warning' if disk_percent < 95 else 'error'),
            'used_percent': round(disk_percent, 1),
            'free_gb': round(disk.free / (1024**3), 2),
            'total_gb': round(disk.total / (1024**3), 2),
            'used_gb': round(disk.used / (1024**3), 2)
        }

        if disk_percent > 95:
            health_data['alerts'].append({
                'level': 'error',
                'icon': 'hdd-fill',
                'message': f'Disco CRÍTICO: {disk_percent}% usado! Libere espaço urgente.'
            })
            health_data['recommendations'].append('Execute limpeza de arquivos temporários e logs antigos')
        elif disk_percent > 85:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'hdd',
                'message': f'Disco com {disk_percent}% usado.'
            })
            health_data['recommendations'].append('Planeje limpeza de arquivos antigos em breve')
    except Exception as e:
        health_data['checks']['disk'] = {'status': 'error', 'message': str(e)}

    # 3. Memory
    try:
        mem = psutil.virtual_memory()
        swap = psutil.swap_memory()
        health_data['checks']['memory'] = {
            'status': 'ok' if mem.percent < 85 else ('warning' if mem.percent < 95 else 'error'),
            'used_percent': round(mem.percent, 1),
            'available_gb': round(mem.available / (1024**3), 2),
            'total_gb': round(mem.total / (1024**3), 2),
            'used_gb': round(mem.used / (1024**3), 2),
            'swap_used_percent': round(swap.percent, 1),
            'swap_used_gb': round(swap.used / (1024**3), 2)
        }

        if mem.percent > 95:
            health_data['alerts'].append({
                'level': 'error',
                'icon': 'memory',
                'message': f'Memória CRÍTICA: {mem.percent:.1f}%!'
            })
        elif mem.percent > 85:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'memory',
                'message': f'Memória alta: {mem.percent:.1f}%'
            })
    except Exception as e:
        health_data['checks']['memory'] = {'status': 'error', 'message': str(e)}

    # 4. CPU Usage
    try:
        cpu_percent = psutil.cpu_percent(interval=1)
        cpu_count = psutil.cpu_count()
        try:
            load_avg = psutil.getloadavg()
        except:
            load_avg = (0, 0, 0)

        health_data['checks']['cpu'] = {
            'status': 'ok' if cpu_percent < 80 else ('warning' if cpu_percent < 95 else 'error'),
            'percent': round(cpu_percent, 1),
            'cores': cpu_count,
            'load_avg_1min': round(load_avg[0], 2),
            'load_avg_5min': round(load_avg[1], 2),
            'load_avg_15min': round(load_avg[2], 2)
        }

        if cpu_percent > 95:
            health_data['alerts'].append({
                'level': 'error',
                'icon': 'cpu',
                'message': f'CPU CRÍTICA: {cpu_percent}%!'
            })
    except Exception as e:
        health_data['checks']['cpu'] = {'status': 'error', 'message': str(e)}

    # 5. Table Counts & Last Updates
    try:
        counts = {}
        counts['usuarios'] = db.session.query(func.count(Usuario.id)).scalar() or 0
        counts['operadoras'] = db.session.query(func.count(Operadora.id)).scalar() or 0
        counts['tabelas'] = db.session.query(func.count(Tabela.id)).scalar() or 0
        counts['procedimentos'] = db.session.query(func.count(Procedimento.id)).scalar() or 0
        counts['simpro'] = db.session.query(func.count(SimproItemNormalized.id)).scalar() or 0
        counts['brasindice'] = db.session.query(func.count(BrasItemNormalized.id)).scalar() or 0
        counts['insumos_index'] = InsumoIndex.query.count()

        # Últimas atualizações
        last_simpro = db.session.query(func.max(SimproItemNormalized.imported_at)).scalar()
        last_bras = db.session.query(func.max(BrasItemNormalized.imported_at)).scalar()

        health_data['checks']['tables'] = {
            'status': 'ok',
            'counts': counts,
            'last_updates': {
                'simpro': last_simpro.isoformat() if last_simpro else None,
                'brasindice': last_bras.isoformat() if last_bras else None
            }
        }

        # Alertas
        if counts['simpro'] == 0:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'database',
                'message': 'Nenhum registro SIMPRO importado'
            })
        if counts['brasindice'] == 0:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'database',
                'message': 'Nenhum registro Brasíndice importado'
            })

        # Verificar dados desatualizados
        if last_simpro:
            days_old = (datetime.utcnow() - last_simpro.replace(tzinfo=None)).days
            if days_old > 60:
                health_data['recommendations'].append(f'SIMPRO está {days_old} dias desatualizado')
        if last_bras:
            days_old = (datetime.utcnow() - last_bras.replace(tzinfo=None)).days
            if days_old > 60:
                health_data['recommendations'].append(f'Brasíndice está {days_old} dias desatualizado')

    except Exception as e:
        health_data['checks']['tables'] = {'status': 'error', 'message': str(e)}

    # 6. Cache Status
    cache_size = len(_insumo_cache)
    health_data['checks']['cache'] = {
        'status': 'ok',
        'insumo_cache_size': cache_size,
        'ttl_seconds': _insumo_cache_ttl,
        'hit_potential': 'Alto' if cache_size > 0 else 'Baixo',
        'efficiency': f'{min(100, cache_size * 10)}%'
    }

    # 7. Import Jobs Analysis
    try:
        recent_jobs = ImportJob.query.order_by(ImportJob.created_at.desc()).limit(10).all()
        running_jobs = ImportJob.query.filter_by(status='RUNNING').count()

        yesterday = datetime.utcnow() - timedelta(hours=24)
        failed_24h = ImportJob.query.filter(
            ImportJob.status == 'FAILED',
            ImportJob.created_at >= yesterday
        ).count()

        total_24h = ImportJob.query.filter(ImportJob.created_at >= yesterday).count()
        success_rate = 0
        if total_24h > 0:
            success_24h = ImportJob.query.filter(
                ImportJob.status == 'SUCCESS',
                ImportJob.created_at >= yesterday
            ).count()
            success_rate = (success_24h / total_24h) * 100

        jobs_data = []
        for job in recent_jobs:
            duration = None
            if job.started_at and job.finished_at:
                duration = round((job.finished_at - job.started_at).total_seconds(), 1)

            jobs_data.append({
                'id': job.id,
                'origem': job.origem,
                'status': job.status,
                'versao': job.versao,
                'linhas': job.total_linhas,
                'created_at': job.created_at.isoformat()[:19] if job.created_at else None,
                'duration_seconds': duration
            })

        health_data['checks']['import_jobs'] = {
            'status': 'ok' if failed_24h == 0 else ('warning' if failed_24h < 5 else 'error'),
            'running_now': running_jobs,
            'failed_24h': failed_24h,
            'success_rate_24h': round(success_rate, 1),
            'total_24h': total_24h,
            'recent_jobs': jobs_data
        }

        if running_jobs > 3:
            health_data['alerts'].append({
                'level': 'warning',
                'icon': 'cloud-upload',
                'message': f'{running_jobs} importações simultâneas'
            })

        if failed_24h > 5:
            health_data['alerts'].append({
                'level': 'error',
                'icon': 'exclamation-circle',
                'message': f'{failed_24h} importações falharam (24h)'
            })

        if success_rate < 80 and total_24h > 0:
            health_data['recommendations'].append(f'Taxa de sucesso baixa: {success_rate:.1f}% - Investigue erros')

    except Exception as e:
        health_data['checks']['import_jobs'] = {'status': 'error', 'message': str(e)}

    # 8. Active Users
    try:
        yesterday = datetime.utcnow() - timedelta(hours=24)
        active_24h = db.session.query(func.count(func.distinct(AuditLog.usuario_id))).filter(
            AuditLog.timestamp >= yesterday
        ).scalar() or 0

        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        logins_today = db.session.query(func.count(AuditLog.id)).filter(
            AuditLog.acao == 'login',
            AuditLog.timestamp >= today
        ).scalar() or 0

        total_users = db.session.query(func.count(Usuario.id)).scalar() or 0

        health_data['checks']['users'] = {
            'status': 'ok',
            'active_24h': active_24h,
            'logins_today': logins_today,
            'total_users': total_users,
            'activity_rate': f'{round((active_24h/total_users)*100, 1)}%' if total_users > 0 else '0%'
        }
    except Exception as e:
        health_data['checks']['users'] = {'status': 'error', 'message': str(e)}

    # 9. System Uptime
    try:
        with open('/proc/uptime', 'r') as f:
            uptime_seconds = float(f.readline().split()[0])
        uptime_days = uptime_seconds / 86400

        health_data['checks']['uptime'] = {
            'status': 'ok',
            'seconds': round(uptime_seconds, 0),
            'days': round(uptime_days, 2),
            'formatted': f'{int(uptime_days)}d {int((uptime_seconds % 86400) / 3600)}h {int((uptime_seconds % 3600) / 60)}m'
        }

        if uptime_days > 90:
            health_data['recommendations'].append('Sistema há mais de 90 dias sem restart - Considere manutenção')
    except:
        health_data['checks']['uptime'] = {
            'status': 'ok',
            'message': 'N/A'
        }

    # Status geral
    has_errors = any(c.get('status') == 'error' for c in health_data['checks'].values() if isinstance(c, dict))
    has_warnings = any(c.get('status') == 'warning' for c in health_data['checks'].values() if isinstance(c, dict))

    if has_errors:
        health_data['status'] = 'unhealthy'
    elif has_warnings:
        health_data['status'] = 'degraded'

    # JSON ou HTML
    if request.args.get('format') == 'json':
        return jsonify(health_data)

    return render_template('health.html', health=health_data)


@app.route('/debug/db')
def debug_db():
    if os.getenv('DEV_AUTO_LOGIN', '0') != '1':
        abort(404)
    return jsonify({
        'database_url': app.config.get('SQLALCHEMY_DATABASE_URI'),
        'disable_login_lock': os.getenv('DISABLE_LOGIN_LOCK', '0'),
        'dev_auto_login': os.getenv('DEV_AUTO_LOGIN', '0'),
    })


@app.route('/minha-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    usuario = Usuario.query.get_or_404(session.get('user_id'))
    erro = None
    aviso = None
    if request.method == 'POST':
        atual = (request.form.get('senha_atual') or '').strip()
        nova = (request.form.get('senha_nova') or '').strip()
        confirma = (request.form.get('senha_confirmacao') or '').strip()

        if not _verify_password(usuario.senha, atual):
            erro = 'Senha atual incorreta.'
        elif not nova:
            erro = 'Informe a nova senha.'
        elif nova != confirma:
            erro = 'Confirmação da senha não confere.'
        elif nova == atual:
            erro = 'A nova senha deve ser diferente da atual.'
        else:
            politica = _password_policy_error(nova)
            if politica:
                erro = politica
            elif _password_was_used_recently(usuario, nova):
                erro = f'Não reutilize as últimas {PASSWORD_HISTORY_SIZE} senhas.'

        if not erro:
            try:
                senha_hash = _hash_password(nova)
                agora = _now_utc()
                usuario.senha = senha_hash
                usuario.must_reset_senha = False
                usuario.senha_atualizada_em = agora
                usuario.failed_login_attempts = 0
                usuario.locked_until = None
                _append_password_history(usuario, senha_hash)
                _register_audit('password.change', usuario=usuario)
                db.session.commit()
                session.clear()
                session.modified = True
                return redirect(url_for('login', senha_alterada=1))
            except Exception as exc:
                db.session.rollback()
                app.logger.error('Erro ao atualizar senha: %s', exc)
                erro = 'Erro ao atualizar a senha. Tente novamente em instantes.'

    must_change = session.get('must_change_senha')
    return render_template('minha-senha.html', erro=erro, must_change=must_change, aviso=aviso)



@app.route('/consulta-comparar')
@login_required
@feature_required('consulta')
def consulta_comparar():
    # Multi-operadora: buscar ID da operadora atual (deve estar no início!)
    current_operadora_id = session.get('operadora_id')

    restore_cbhpm_payload = None
    history_id = request.args.get('sim_hist')
    if history_id:
        history = session.get('sim_history') or []
        entry = next((item for item in history if str(item.get('id')) == str(history_id)), None)
        if entry:
            if entry.get('type') == 'cbhpm' and entry.get('payload'):
                restore_cbhpm_payload = entry.get('payload')
            elif entry.get('type') == 'compare' and entry.get('url_fragment'):
                target = entry.get('url_fragment') or ''
                if target:
                    return redirect(f"{url_for('consulta_comparar')}?{target}")

    q = request.args.get('q', '').strip()
    search_code = None
    search_text = None
    if ' - ' in q:
        parts = q.split(' - ', 1)
        search_code = parts[0].strip()
        search_text = parts[1].strip() if len(parts) > 1 else ''

    tabela_nome = request.args.get('tabela_nome')
    selected_uf = request.args.get('uf') or ''
    selected_prestadores = request.args.getlist('prestadores')
    selected_versoes = request.args.getlist('versoes')
    show_results = (request.args.get('run') == '1')

    procs_raw = request.args.getlist('procedimentos')  # ex.: ['10101012', '10101020 - ...']
    codigos = []
    for s in procs_raw:
        s = (s or '').strip()
        if not s:
            continue
        codigos.append(s.split(' - ', 1)[0].strip())
    if q.isdigit():
        codigos.append(q)
    # Permite códigos duplicados - não remove duplicatas

    nomes = [r[0] for r in db.session.query(Tabela.nome).distinct().order_by(Tabela.nome).all()]
    if not tabela_nome and len(nomes) == 1:
        tabela_nome = nomes[0]

    is_cbhpm = False
    if tabela_nome:
        is_cbhpm = db.session.query(CBHPMItem.id)            .join(Tabela, CBHPMItem.id_tabela == Tabela.id)            .filter(Tabela.nome == tabela_nome).first() is not None

    prestadores_disp, versoes_disp = [], []
    if tabela_nome:
        if is_cbhpm:
            versoes_disp = [r[0] for r in db.session.query(Tabela.nome)
                              .filter(Tabela.tipo_tabela == 'cbhpm')
                              .distinct().order_by(Tabela.nome).all()]
        else:
            q_prest = db.session.query(Procedimento.prestador)                .join(Tabela, Procedimento.id_tabela == Tabela.id)                .filter(Tabela.nome == tabela_nome)

            # Multi-operadora: filtrar prestadores por operadora_id
            if current_operadora_id:
                q_prest = q_prest.filter(Procedimento.operadora_id == current_operadora_id)

            if selected_uf:
                q_prest = q_prest.filter(or_(Tabela.uf == selected_uf, Procedimento.uf == selected_uf))
            prestadores_disp = [r[0] for r in q_prest
                .filter((Procedimento.prestador.isnot(None)) & (Procedimento.prestador != ''))
                .distinct().order_by(Procedimento.prestador).all()]

    columns = (selected_versoes or versoes_disp) if is_cbhpm else (selected_prestadores or prestadores_disp)

    rows = []
    if show_results and tabela_nome:
        data = {}
        if is_cbhpm:
            targets = columns or []
            for ver in targets:
                qv = db.session.query(
                        CBHPMItem.codigo, CBHPMItem.procedimento,
                        CBHPMItem.subtotal, CBHPMItem.total_porte, CBHPMItem.valor_porte,
                        CBHPMItem.total_uco, CBHPMItem.uco, CBHPMItem.total_filme, CBHPMItem.filme
                    )                    .join(Tabela, CBHPMItem.id_tabela == Tabela.id)                    .filter(Tabela.nome == ver, Tabela.tipo_tabela == 'cbhpm')
                if selected_uf:
                    qv = qv.filter(or_(Tabela.uf == selected_uf, CBHPMItem.uf == selected_uf))

                if codigos:
                    qv = qv.filter(CBHPMItem.codigo.in_(codigos))
                elif q:
                    if search_code:
                        qv = qv.filter(or_(
                            CBHPMItem.codigo == search_code,
                            CBHPMItem.codigo.ilike(f"{search_code}%"),
                            (CBHPMItem.procedimento.ilike(f"%{search_text}%") if search_text else False)
                        ))
                    else:
                        like = f"%{q}%"
                        qv = qv.filter(or_(CBHPMItem.codigo.ilike(like), CBHPMItem.procedimento.ilike(like)))

                for cod, desc, sub, tp, vp, tu, u, tf, f in qv.all():
                    v = sub or tp or vp or tu or u or tf or f
                    entry = data.setdefault(cod, {"descricao": desc, "values": {}})
                    entry["values"][ver] = v
        else:
            query = db.session.query(Procedimento, Procedimento.prestador)                .join(Tabela, Procedimento.id_tabela == Tabela.id)                .filter(Tabela.nome == tabela_nome)

            # Multi-operadora: filtrar procedimentos por operadora_id
            if current_operadora_id:
                query = query.filter(Procedimento.operadora_id == current_operadora_id)

            if selected_uf:
                query = query.filter(or_(Tabela.uf == selected_uf, Procedimento.uf == selected_uf))
            if selected_prestadores:
                query = query.filter(Procedimento.prestador.in_(selected_prestadores))

            if codigos:
                query = query.filter(Procedimento.codigo.in_(codigos))
            elif q:
                if search_code:
                    query = query.filter(or_(
                        Procedimento.codigo == search_code,
                        Procedimento.codigo.ilike(f"{search_code}%"),
                        (Procedimento.descricao.ilike(f"%{search_text}%") if search_text else False)
                    ))
                else:
                    like = f"%{q}%"
                    query = query.filter(or_(Procedimento.codigo.ilike(like), Procedimento.descricao.ilike(like)))

            prestadores_usados = set()
            for proc, prest in query.all():
                prest = prest or '-'
                prestadores_usados.add(prest)
                entry = data.setdefault(proc.codigo, {"descricao": proc.descricao, "values": {}})
                entry["values"][prest] = proc.valor
            if not selected_prestadores and prestadores_usados:
                columns = sorted(list(prestadores_usados))

        rol_map = _fetch_tuss_rol_map(list(data.keys()))
        for codigo in sorted(data.keys()):
            item = data[codigo]
            values = [item["values"].get(p) for p in columns]
            numeric = [v for v in values if v is not None]
            min_v = min(numeric) if numeric else None
            max_v = max(numeric) if numeric else None
            avg_v = (sum(numeric) / len(numeric)) if numeric else None
            rol_info = rol_map.get(codigo)
            rows.append({
                "codigo": codigo,
                "descricao": item["descricao"],
                "values": values,
                "min": min_v,
                "max": max_v,
                "avg": avg_v,
                "count": len(numeric),
                "rol": rol_info,
            })

    porte_list = [t.nome for t in Tabela.query.filter_by(tipo_tabela='porte').order_by(Tabela.nome).all()]
    porte_an_list = [t.nome for t in Tabela.query.filter_by(tipo_tabela='porte_anestesico').order_by(Tabela.nome).all()]
    dtp_list = [t.nome for t in Tabela.query.filter_by(tipo_tabela='diarias_taxas_pacotes').order_by(Tabela.nome).all()]
    cbhpm_list_all = [r[0] for r in db.session.query(Tabela.nome).filter(Tabela.tipo_tabela=='cbhpm').distinct().order_by(Tabela.nome).all()]

    # Multi-operadora: buscar lista de operadoras ativas (filtrada pelo usuário)
    operadoras_list = _get_user_operadoras_list()

    ruleset_dict, ruleset_model = _get_active_cbhpm_ruleset(return_model=True)
    rules_meta = {
        'nome': ruleset_model.nome if ruleset_model else 'Padrão',
        'versao': ruleset_model.versao if ruleset_model else None,
        'descricao': ruleset_model.descricao if ruleset_model else None,
        'id': ruleset_model.id if ruleset_model else None,
    }

    if show_results and tabela_nome:
        query_bytes = request.query_string or b''
        if query_bytes:
            try:
                query_str = query_bytes.decode('utf-8', 'ignore')
            except Exception:
                query_str = query_bytes.decode('latin-1', 'ignore')
            label_parts = []
            clean_table = unicodedata.normalize('NFKD', tabela_nome).encode('ascii', 'ignore').decode() if tabela_nome else ''
            if clean_table:
                label_parts.append(clean_table)
            if codigos:
                snippet = ', '.join(codigos[:3])
                if len(codigos) > 3:
                    snippet += ', ...'
                snippet = unicodedata.normalize('NFKD', snippet).encode('ascii', 'ignore').decode()
                label_parts.append(f'codigos {snippet}')
            elif q:
                clean_q = unicodedata.normalize('NFKD', q).encode('ascii', 'ignore').decode()
                label_parts.append(f'busca {clean_q}')
            label = ' | '.join(filter(None, label_parts)) or (clean_table or 'Consulta recente')
            entry_id = hashlib.md5(query_str.encode('utf-8')).hexdigest()[:10] if query_str else uuid4().hex[:8]
            _store_history_entry({
                'type': 'compare',
                'id': entry_id,
                'signature': f'compare:{query_str}',
                'url_fragment': query_str,
                'label': label[:80],
                'timestamp': datetime.now().strftime('%d/%m %H:%M'),
            })

    return render_template(
        'consulta-comparar.html',
        nomes=nomes, tabela_nome=tabela_nome,
        prestadores_disp=prestadores_disp, selected_prestadores=selected_prestadores,
        versoes_disp=versoes_disp, selected_versoes=selected_versoes,
        UFS=BR_UFS, selected_uf=selected_uf, q=q,
        columns=columns, rows=rows, show_results=show_results, is_cbhpm=is_cbhpm,
        porte_list=porte_list, porte_an_list=porte_an_list,
        cbhpm_list_all=cbhpm_list_all, dtp_list=dtp_list,
        restore_cbhpm_payload=restore_cbhpm_payload,
        cbhpm_rules_info=rules_meta,
        operadoras_list=operadoras_list,
        current_operadora_id=current_operadora_id
    )


def _compute_simulacao_cbhpm(data):
    data = data or {}

    ruleset_dict, ruleset_model = _get_active_cbhpm_ruleset(return_model=True)
    rules_meta = {
        'nome': ruleset_model.nome if ruleset_model else 'Padrao',
        'versao': ruleset_model.versao if ruleset_model else None,
        'descricao': ruleset_model.descricao if ruleset_model else None,
        'id': ruleset_model.id if ruleset_model else None,
    }

    quantize_money = Decimal('0.01')
    quantize_pct = Decimal('0.01')

    # Suporta arrays de ajustes por índice (para códigos repetidos com ajustes independentes)
    via_pct_array = data.get('via_entrada_pcts_array') or []
    acomodacao_array = data.get('acomodacao_array') or []

    via_pct_map_raw = data.get('via_entrada_pcts')
    via_pct_map: dict[str, Decimal] = {}
    if isinstance(via_pct_map_raw, dict):
        for key, value in via_pct_map_raw.items():
            key_norm = str(key or '').strip()
            if not key_norm:
                continue
            pct = _as_decimal(value)
            if pct is None:
                continue
            try:
                pct = max(Decimal('0'), min(Decimal('100'), pct))
            except Exception:
                continue
            via_pct_map[key_norm] = pct
    if '__default__' not in via_pct_map:
        via_pct_map['__default__'] = Decimal('100')
    applied_via_map: dict[str, Decimal] = {}

    acomodacao_map_raw = data.get('acomodacao_map')
    acomodacao_map: dict[str, str] = {}
    if isinstance(acomodacao_map_raw, dict):
        for key, value in acomodacao_map_raw.items():
            key_norm = str(key or '').strip().upper()
            if not key_norm:
                continue
            val_norm = str(value or '').strip().lower()
            if val_norm not in {'apartamento', 'enfermaria'}:
                continue
            acomodacao_map[key_norm] = 'apartamento' if val_norm == 'apartamento' else 'enfermaria'

    def _resolve_acomodacao(code_key: str | None, item_index: int | None = None) -> str:
        # Prioriza acomodação por índice se disponível (para códigos repetidos)
        if item_index is not None and item_index < len(acomodacao_array):
            val = str(acomodacao_array[item_index] or '').strip().lower()
            if val in {'apartamento', 'enfermaria'}:
                return val
        # Fallback para mapa por código
        key_norm = str(code_key or '').strip().upper()
        if key_norm and key_norm in acomodacao_map:
            return acomodacao_map[key_norm]
        default_val = acomodacao_map.get('__DEFAULT__')
        if default_val == 'apartamento':
            return 'apartamento'
        return 'enfermaria'

    def apply_via_entrada(breakdown: dict | None, code_key: str | None, item_index: int | None = None):
        if not breakdown:
            return breakdown
        normalized = str(code_key or '').strip().upper() or '__default__'
        # Prioriza percentual por índice se disponível (para códigos repetidos)
        if item_index is not None and item_index < len(via_pct_array):
            pct_raw = via_pct_array[item_index]
            pct = _as_decimal(pct_raw)
            if pct is not None:
                pct = max(Decimal('0'), min(Decimal('100'), pct))
            else:
                pct = Decimal('100')
        else:
            # Fallback para mapa por código
            pct = via_pct_map.get(normalized, via_pct_map.get('__default__', Decimal('100')))
        applied_via_map[normalized] = pct
        factor = (pct / Decimal('100')) if pct is not None else Decimal('1')
        breakdown['via_entrada_pct'] = pct
        breakdown['via_entrada_factor'] = factor

        total_porte_original = _as_decimal(breakdown.get('total_porte'))
        total_filme = _as_decimal(breakdown.get('total_filme'))
        total_uco = _as_decimal(breakdown.get('total_uco'))
        total_an = _as_decimal(breakdown.get('total_porte_an'))
        total_aux = _as_decimal(breakdown.get('total_auxiliares'))
        total_original = _sum_decimals([total_porte_original, total_filme, total_uco, total_an, total_aux])
        breakdown['total_original'] = total_original

        reduced_porte = total_porte_original
        if total_porte_original is not None:
            try:
                reduced_porte = (total_porte_original * factor).quantize(quantize_money, rounding=ROUND_HALF_UP)
            except (InvalidOperation, ValueError):
                reduced_porte = total_porte_original
        breakdown['total_porte'] = reduced_porte

        aux_scaled_total = None
        aux_details = breakdown.get('auxiliares_detalhe')
        if isinstance(aux_details, list) and aux_details:
            new_details: list[dict] = []
            running_total = Decimal('0')
            for entry in aux_details:
                val = _as_decimal(entry.get('valor'))
                if val is None:
                    new_details.append(entry)
                    continue
                try:
                    scaled_val = (val * factor).quantize(quantize_money, rounding=ROUND_HALF_UP)
                except (InvalidOperation, ValueError):
                    scaled_val = val * factor
                running_total += scaled_val
                new_entry = dict(entry)
                new_entry['valor'] = scaled_val
                new_details.append(new_entry)
            breakdown['auxiliares_detalhe'] = new_details
            aux_scaled_total = running_total

        if aux_scaled_total is not None:
            breakdown['total_auxiliares'] = aux_scaled_total
            total_aux = aux_scaled_total
        elif total_aux is not None and factor is not None and factor != Decimal('1'):
            try:
                breakdown['total_auxiliares'] = (total_aux * factor).quantize(quantize_money, rounding=ROUND_HALF_UP)
            except (InvalidOperation, ValueError):
                breakdown['total_auxiliares'] = total_aux * factor
            total_aux = _as_decimal(breakdown.get('total_auxiliares'))

        total_reduzido = _sum_decimals([reduced_porte, total_filme, total_uco, total_an, total_aux])
        breakdown['total_reduzido'] = total_reduzido
        breakdown['total_final'] = total_reduzido
        breakdown['total'] = total_reduzido
        return breakdown

    def apply_acomodacao(breakdown: dict | None, code_key: str | None, item_index: int | None = None):
        if not breakdown:
            return breakdown
        acomodacao_value = _resolve_acomodacao(code_key, item_index)
        breakdown['acomodacao'] = acomodacao_value
        if acomodacao_value != 'apartamento':
            return breakdown
        keys_to_scale = ('total_porte', 'total_porte_an', 'total_auxiliares')
        for key in keys_to_scale:
            current = _as_decimal(breakdown.get(key))
            if current is None:
                continue
            try:
                breakdown[key] = (current * Decimal('2')).quantize(quantize_money, rounding=ROUND_HALF_UP)
            except (InvalidOperation, ValueError):
                breakdown[key] = current * Decimal('2')
        aux_details = breakdown.get('auxiliares_detalhe')
        if isinstance(aux_details, list):
            new_details = []
            for entry in aux_details:
                val = _as_decimal(entry.get('valor'))
                if val is None:
                    new_details.append(entry)
                    continue
                try:
                    new_val = (val * Decimal('2')).quantize(quantize_money, rounding=ROUND_HALF_UP)
                except (InvalidOperation, ValueError):
                    new_val = val * Decimal('2')
                updated = dict(entry)
                updated['valor'] = new_val
                new_details.append(updated)
            breakdown['auxiliares_detalhe'] = new_details
        total_porte = _as_decimal(breakdown.get('total_porte'))
        total_filme = _as_decimal(breakdown.get('total_filme'))
        total_uco = _as_decimal(breakdown.get('total_uco'))
        total_an = _as_decimal(breakdown.get('total_porte_an'))
        total_aux = _as_decimal(breakdown.get('total_auxiliares'))
        total_calculado = _sum_decimals([total_porte, total_filme, total_uco, total_an, total_aux])
        for key in ('total', 'total_final', 'total_reduzido'):
            if key in breakdown:
                breakdown[key] = total_calculado
        if breakdown.get('total_original') is not None:
            breakdown['total_original'] = total_calculado
        return breakdown

    codigo = (data.get('codigo') or '').strip()
    codigos = data.get('codigos') or []
    if isinstance(codigos, list):
        codigos = [str(c or '').split(' - ', 1)[0].strip() for c in codigos if (c or '')]

    dtp_raw = data.get('dtp_items') or []
    dtp_map = {}
    for item in dtp_raw:
        code = str(item.get('codigo') or '').strip()
        if not code:
            continue
        desc = (item.get('descricao') or '').strip()
        valor = _as_decimal(item.get('valor'))
        tabela_nome = (item.get('tabela_nome') or item.get('tabela') or '').strip()
        qtd = max(int(item.get('qtd') or 1), 1)
        uf_item = (item.get('uf') or '').strip()
        dtp_map[code] = {
            'codigo': code,
            'descricao': desc,
            'valor': valor,
            'qtd': qtd,
            'tabela_nome': tabela_nome,
            'uf': uf_item,
        }
    dtp_items = list(dtp_map.values())

    uf = (data.get('uf') or '').strip() or None
    versao = data.get('versao')
    porte_tab_name = data.get('porte_tab')
    porte_an_tab_name = data.get('porte_an_tab')
    uco_valor_in = _as_decimal(data.get('uco_valor'))
    filme_valor_in = _as_decimal(data.get('filme_valor'))
    incid_in = _as_decimal(data.get('incidencias'))
    aj_porte_pct = _as_decimal(data.get('ajuste_porte_pct')) or Decimal('0')
    aj_an_pct = _as_decimal(data.get('ajuste_porte_an_pct')) or Decimal('0')

    # Multi-operadora: obter operadora_id do request ou da sessão
    operadora_id = data.get('operadora_id')
    if not operadora_id:
        operadora_id = session.get('operadora_id')

    if not codigo and not codigos and not dtp_items:
        return {"error": 'Informe "codigo" ou a lista "codigos".'}, 400

    item = None
    t_ref = None
    codigos_list = codigos  # Permite códigos duplicados - usa lista de códigos diretamente
    target_code = codigo or (codigos_list[0] if codigos_list else '')
    if target_code:
        q = (db.session.query(CBHPMItem, Tabela)
             .join(Tabela, CBHPMItem.id_tabela == Tabela.id))
        if versao:
            q = q.filter(Tabela.nome == versao)
        q = q.filter(or_(CBHPMItem.codigo == target_code,
                         CBHPMItem.codigo.ilike(f"{target_code}%")))
        if uf:
            q = q.filter(or_(CBHPMItem.uf == uf, Tabela.uf == uf))
        row = q.first()
        if row:
            item, t_ref = row[0], row[1]

    if not t_ref and versao:
        t_ref = Tabela.query.filter_by(nome=versao).first()

    if not t_ref:
        t_ref = Tabela.query.filter_by(tipo_tabela='cbhpm').first()

    if not t_ref:
        op = Operadora.query.first()
        t_ref = Tabela(nome='SIMULACAO', id_operadora=(op.id if op else 1))

    if uco_valor_in is not None:
        t_ref.uco_valor = uco_valor_in

    fracao_override = _as_decimal(data.get('fracao_porte'))

    base = CBHPMItem(
        codigo=codigo,
        procedimento=(item.procedimento if item else (data.get('descricao') or '')),
        porte=(item.porte if item else data.get('porte')),
        fracao_porte=(fracao_override if fracao_override is not None else (item.fracao_porte if item else None)),
        valor_porte=(item.valor_porte if item else _as_decimal(data.get('valor_porte'))),
        total_porte=(item.total_porte if item else None),
        filme=(item.filme if item else filme_valor_in),
        incidencias=(item.incidencias if item else incid_in),
        total_filme=(item.total_filme if item else None),
        uco=(item.uco if item else _as_decimal(data.get('uco_qtd'))),
        total_uco=(item.total_uco if item else None),
        porte_anestesico=(item.porte_anestesico if item else data.get('porte_an')),
        valor_porte_anestesico=(item.valor_porte_anestesico if item else _as_decimal(data.get('valor_porte_an'))),
        total_porte_anestesico=(item.total_porte_anestesico if item else None),
        numero_auxiliares=(item.numero_auxiliares if item is not None else data.get('numero_auxiliares')),
        total_auxiliares=(item.total_auxiliares if item else None),
        total_1_aux=(item.total_1_aux if item else _as_decimal(data.get('total_1_aux'))),
        total_2_aux=(item.total_2_aux if item else _as_decimal(data.get('total_2_aux'))),
        total_3_aux=(item.total_3_aux if item else _as_decimal(data.get('total_3_aux'))),
        total_4_aux=(item.total_4_aux if item else _as_decimal(data.get('total_4_aux'))),
    )

    base._fracao_input = fracao_override

    if porte_tab_name:
        base.valor_porte = None
        base.total_porte = None
    if porte_an_tab_name:
        base.valor_porte_anestesico = None
        base.total_porte_anestesico = None
    if filme_valor_in is not None:
        base.filme = filme_valor_in
        base.total_filme = None
    if uco_valor_in is not None:
        base.total_uco = None

    porte_hint = porte_tab_name or (t_ref.nome if t_ref else None)
    porte_an_hint = porte_an_tab_name or (t_ref.nome if t_ref else None)

    if codigos or dtp_items:
        itens = []
        cbhpm_results = []
        teto_alerts: list[dict] = []
        d0 = Decimal('0')
        def to_decimal(value, qtd=1):
            val = _as_decimal(value)
            val = val if val is not None else d0
            return val * Decimal(qtd)

        acomodacao_out_map: dict[str, str] = {}
        if codigos:  # Permite códigos duplicados - verifica se há códigos na lista
            for cod_idx, cod in enumerate(codigos):
                it_item = None
                if versao or cod:
                    qit = (db.session.query(CBHPMItem, Tabela)
                           .join(Tabela, CBHPMItem.id_tabela == Tabela.id))
                    if versao:
                        qit = qit.filter(Tabela.nome == versao)
                    if cod:
                        qit = qit.filter(or_(CBHPMItem.codigo == cod, CBHPMItem.codigo.ilike(f"{cod}%")))
                    if uf:
                        qit = qit.filter(or_(CBHPMItem.uf == uf, Tabela.uf == uf))
                    rowi = qit.first()
                    if rowi:
                        it_item = rowi[0]

                base_i = CBHPMItem(
                    codigo=cod,
                    procedimento=(it_item.procedimento if it_item else ''),
                    porte=(it_item.porte if it_item else None),
                    fracao_porte=(fracao_override if fracao_override is not None else (it_item.fracao_porte if it_item else None)),
                    valor_porte=(it_item.valor_porte if it_item else None),
                    total_porte=(it_item.total_porte if it_item else None),
                    filme=(it_item.filme if it_item else filme_valor_in),
                    incidencias=(it_item.incidencias if it_item else incid_in),
                    total_filme=(it_item.total_filme if it_item else None),
                    uco=(it_item.uco if it_item else None),
                    total_uco=(it_item.total_uco if it_item else None),
                    porte_anestesico=(it_item.porte_anestesico if it_item else None),
                    valor_porte_anestesico=(it_item.valor_porte_anestesico if it_item else None),
                    total_porte_anestesico=(it_item.total_porte_anestesico if it_item else None),
                    numero_auxiliares=(it_item.numero_auxiliares if it_item else None),
                    total_auxiliares=(it_item.total_auxiliares if it_item else None),
                    total_1_aux=(it_item.total_1_aux if it_item else None),
                    total_2_aux=(it_item.total_2_aux if it_item else None),
                    total_3_aux=(it_item.total_3_aux if it_item else None),
                    total_4_aux=(it_item.total_4_aux if it_item else None),
                )
                base_i._fracao_input = fracao_override

                if porte_tab_name:
                    base_i.valor_porte = None
                    base_i.total_porte = None
                if porte_an_tab_name:
                    base_i.valor_porte_anestesico = None
                    base_i.total_porte_anestesico = None
                if filme_valor_in is not None:
                    base_i.filme = filme_valor_in
                    base_i.total_filme = None
                if uco_valor_in is not None:
                    base_i.total_uco = None

                br = compute_cbhpm_breakdown(
                    base_i, t_ref,
                    porte_hint=porte_hint, porte_an_hint=porte_an_hint,
                    ajuste_porte_pct=aj_porte_pct, ajuste_porte_an_pct=aj_an_pct,
                    rules=ruleset_dict
                )
                # Passa índice para permitir ajustes individuais em códigos repetidos
                br = apply_via_entrada(br, cod, cod_idx)
                br = apply_acomodacao(br, cod, cod_idx)
                item_out = {k: _stringify_for_output(v) for k, v in br.items()}
                if br.get('applied_rules'):
                    item_out['applied_rules'] = _stringify_for_output(br['applied_rules'])
                if br.get('total_original') is not None:
                    item_out['total_original'] = _stringify_for_output(br.get('total_original'))
                if br.get('total_final') is not None:
                    item_out['total_final'] = _stringify_for_output(br.get('total_final'))
                item_out['percentual_via'] = _stringify_for_output(br.get('via_entrada_pct')) if br.get('via_entrada_pct') is not None else item_out.get('percentual_via')
                item_out.update({'codigo': cod, 'descricao': base_i.procedimento, 'origem': 'cbhpm', 'item_index': cod_idx})
                code_norm = str(cod or '').strip().upper()
                if code_norm:
                    acomodacao_out_map[code_norm] = br.get('acomodacao') or 'enfermaria'
                cbhpm_results.append({
                    'payload': item_out,
                    'totals': {
                        'total_porte': to_decimal(br.get('total_porte')),
                        'total_filme': to_decimal(br.get('total_filme')),
                        'total_uco': to_decimal(br.get('total_uco')),
                        'total_porte_an': to_decimal(br.get('total_porte_an')),
                        'total_auxiliares': to_decimal(br.get('total_auxiliares')),
                        'total': to_decimal(br.get('total')),
                        'total_original': to_decimal(br.get('total_original')),
                        'total_final': to_decimal(br.get('total_final')),
                        'via_pct': to_decimal(br.get('via_entrada_pct')),
                    }
                })

        reducoes = (ruleset_dict.get('porte') or {}).get('reducoes_simultaneos') or []
        if reducoes and len(cbhpm_results) > 1:
            ordered = sorted(
                enumerate(cbhpm_results),
                key=lambda pair: pair[1]['totals']['total_porte'], # A ordenação é feita por valor individual
                reverse=True
            )
            for rank, (idx, entry) in enumerate(ordered):
                original = entry['totals']['total_porte']
                if original <= d0:
                    continue
                factor_raw = reducoes[min(rank, len(reducoes) - 1)]
                try:
                    factor = Decimal(str(factor_raw))
                except (InvalidOperation, ValueError):
                    continue
                if factor > Decimal('5'):
                    factor = factor / Decimal('100')
                if factor > Decimal('1'):
                    factor = Decimal('1')
                if factor < Decimal('0'):
                    factor = Decimal('0')
                adjusted = original * factor
                if adjusted == original:
                    continue
                delta = original - adjusted
                entry['totals']['total_porte'] = adjusted
                # Atualiza o total do item e o total final com a redução
                entry['totals']['total'] = entry['totals']['total'] - delta
                payload_entry = entry['payload']
                payload_entry['total_porte'] = str(adjusted)
                payload_entry['total'] = str(entry['totals']['total'])
                applied = list(payload_entry.get('applied_rules') or [])
                applied.append({
                    'component': 'porte',
                    'rule': 'reducoes_simultaneos',
                    'ordem': rank + 1,
                    'fator': str(factor),
                    'reduzido_de': str(original),
                    'reduzido_para': str(adjusted),
                })
                payload_entry['applied_rules'] = applied

        if cbhpm_results:
            # Garante que cada código seja verificado, mesmo que duplicado
            codes_to_check = [entry['payload'].get('codigo') for entry in cbhpm_results if entry['payload'].get('codigo')]
            teto_map = _get_teto_map(codes_to_check, operadora_id=operadora_id)
            rol_map = _fetch_tuss_rol_map(codes_to_check)
            for entry in cbhpm_results:
                payload_entry = entry['payload'] # payload_entry é um dicionário único para cada item na lista
                codigo_item = (payload_entry.get('codigo') or '').strip().upper()
                if not codigo_item:
                    continue
                rol_info = rol_map.get(payload_entry.get('codigo')) or rol_map.get(codigo_item) or None
                if rol_info:
                    payload_entry['rol'] = rol_info
                teto_row = teto_map.get(codigo_item)
                if not teto_row:
                    continue
                # O teto é por procedimento, então o valor é unitário
                teto_val = _as_decimal(teto_row.valor_total) 
                calc_total = entry['totals'].get('total_final') or entry['totals'].get('total')
                if calc_total is None:
                    calc_total = _as_decimal(payload_entry.get('total_final') or payload_entry.get('total'))
                excedido = False
                excedente = None
                pct_total = None
                pct_excedente = None
                if teto_val is not None and calc_total is not None:
                    diff = (calc_total - teto_val).quantize(quantize_money, rounding=ROUND_HALF_UP)
                    if diff > Decimal('0'):
                        excedido = True
                        excedente = diff
                    if teto_val != Decimal('0'):
                        pct_total = (calc_total / teto_val * Decimal('100')).quantize(quantize_pct, rounding=ROUND_HALF_UP)
                        if diff > Decimal('0'):
                            pct_excedente = (diff / teto_val * Decimal('100')).quantize(quantize_pct, rounding=ROUND_HALF_UP)
                payload_entry['teto_valor_total'] = _stringify_for_output(teto_val) if teto_val is not None else None
                payload_entry['teto_descricao'] = teto_row.descricao
                payload_entry['teto_excedente'] = _stringify_for_output(excedente) if excedente is not None else None
                payload_entry['teto_excedido'] = excedido
                payload_entry['teto_status'] = 'ULTRAPASSA' if excedido else 'OK'
                payload_entry['teto_pct_total'] = _stringify_for_output(pct_total) if pct_total is not None else None
                payload_entry['teto_pct_excedente'] = _stringify_for_output(pct_excedente) if pct_excedente is not None else None
                if excedido:
                    teto_alerts.append({
                        'codigo': payload_entry.get('codigo'),
                        'descricao': payload_entry.get('descricao'),
                        'total_calculado': _stringify_for_output(calc_total),
                        'teto_valor_total': _stringify_for_output(teto_val),
                        'excedente': _stringify_for_output(excedente),
                        'pct_total': _stringify_for_output(pct_total) if pct_total is not None else None,
                        'pct_excedente': _stringify_for_output(pct_excedente) if pct_excedente is not None else None,
                        'descricao_teto': teto_row.descricao,
                    })

        itens.extend([entry['payload'] for entry in cbhpm_results]) # Adiciona cada resultado individualmente

        for meta in dtp_items:
            val = meta.get('valor')
            if val is None:
                val = d0
            elif not isinstance(val, Decimal):
                val = _as_decimal(val) or d0
            itens.append({
                'codigo': meta.get('codigo'),
                'descricao': meta.get('descricao'),
                'total_porte': '0',
                'total_filme': '0',
                'total_uco': '0',
                'total_porte_an': '0',
                'total_auxiliares': '0',
                'total': str(val),
                'total_original': str(val),
                'total_final': str(val),
                'percentual_aplicado': '100',
                'origem': 'dtp',
                'tabela_origem': meta.get('tabela_nome'),
                'uf_origem': meta.get('uf'),
                'auxiliares_detalhe': [],
                'acomodacao': 'enfermaria',
            })
        sum_porte = sum(to_decimal(item.get('total_porte'), item.get('qtd', 1) if item.get('origem') != 'dtp' else 1) for item in itens)
        sum_filme = sum(to_decimal(item.get('total_filme'), item.get('qtd', 1) if item.get('origem') != 'dtp' else 1) for item in itens)
        sum_uco = sum(to_decimal(item.get('total_uco'), item.get('qtd', 1) if item.get('origem') != 'dtp' else 1) for item in itens)
        sum_an = sum(to_decimal(item.get('total_porte_an'), item.get('qtd', 1) if item.get('origem') != 'dtp' else 1) for item in itens)
        sum_aux = sum(to_decimal(item.get('total_auxiliares'), item.get('qtd', 1) if item.get('origem') != 'dtp' else 1) for item in itens)
        sum_total = sum(to_decimal(item.get('total'), 1) for item in itens) # total já tem qtd
        sum_total_original = sum(to_decimal(item.get('total_original'), 1) for item in itens) # total já tem qtd
        sum_total_final = sum(to_decimal(item.get('total_final') or item.get('total'), 1) for item in itens) # total já tem qtd

        via_out_map = {
            key: str(value)
            for key, value in applied_via_map.items()
            if key != '__default__'
        }

        acomodacoes_utilizadas = [
            (entry['payload'].get('acomodacao') or 'enfermaria').lower()
            for entry in cbhpm_results
            if entry['payload'].get('origem') == 'cbhpm'
        ]
        if acomodacoes_utilizadas:
            unique_acomodacoes = set(acomodacoes_utilizadas)
            if len(unique_acomodacoes) == 1:
                acomodacao_summary = unique_acomodacoes.pop()
            else:
                acomodacao_summary = 'misto'
        else:
            acomodacao_summary = None

        payload_agregado = {
            'itens': itens,
            'total_porte': str(sum_porte),
            'total_filme': str(sum_filme),
            'total_uco': str(sum_uco),
            'total_porte_an': str(sum_an),
            'total_auxiliares': str(sum_aux),
            'total': str(sum_total_final if sum_total_final is not None else sum_total),
            'total_original': str(sum_total_original),
            'total_final': str(sum_total_final if sum_total_final is not None else sum_total),
            'porte_tabela_usada': (porte_tab_name or _resolve_porte_tabela_nome(t_ref.id_operadora, t_ref.uf, porte_hint, None)),
            'porte_an_tabela_usada': (porte_an_tab_name or _resolve_porte_an_tabela_nome(t_ref.id_operadora, t_ref.uf, porte_an_hint, None)),
            'uco_valor': str(t_ref.uco_valor) if getattr(t_ref, 'uco_valor', None) is not None else None,
            'versao_base': versao,
            'ajuste_porte_pct': str(aj_porte_pct),
            'ajuste_porte_an_pct': str(aj_an_pct),
            'via_entrada_pcts': via_out_map,
            'cbhpm_rules_info': rules_meta,
            'teto_alertas': teto_alerts,
            'teto_status': 'ULTRAPASSA' if teto_alerts else 'OK',
            'acomodacao_summary': acomodacao_summary,
            'acomodacao_map': acomodacao_out_map,
        }
        return payload_agregado, 200

    breakdown = compute_cbhpm_breakdown(
        base, t_ref,
        porte_hint=porte_hint, porte_an_hint=porte_an_hint,
        ajuste_porte_pct=aj_porte_pct, ajuste_porte_an_pct=aj_an_pct,
        rules=ruleset_dict
    )
    breakdown = apply_via_entrada(breakdown, codigo)
    breakdown = apply_acomodacao(breakdown, codigo)
    resp = {k: _stringify_for_output(v) for k, v in breakdown.items()}
    resp['acomodacao'] = breakdown.get('acomodacao') or 'enfermaria'
    resp.update({
        'codigo': codigo,
        'descricao': base.procedimento,
        'uco_valor': str(t_ref.uco_valor) if getattr(t_ref, 'uco_valor', None) is not None else None,
        'versao_base': versao,
        'porte_tabela_usada': (porte_tab_name or _resolve_porte_tabela_nome(t_ref.id_operadora, t_ref.uf, porte_hint, base.porte)),
        'porte_an_tabela_usada': (porte_an_tab_name or _resolve_porte_an_tabela_nome(t_ref.id_operadora, t_ref.uf, porte_an_hint, base.porte_anestesico)),
        'ajuste_porte_pct': str(aj_porte_pct),
        'ajuste_porte_an_pct': str(aj_an_pct),
        'via_entrada_pct': str(breakdown.get('via_entrada_pct')) if breakdown.get('via_entrada_pct') is not None else None,
        'via_entrada_pcts': {
            key: str(value)
            for key, value in applied_via_map.items()
            if key != '__default__'
        },
        'acomodacao_map': {
            str(codigo).strip().upper(): resp['acomodacao']
        } if codigo else {},
        'cbhpm_rules_info': rules_meta,
    })

    if codigo:
        rol_lookup = _fetch_tuss_rol_map([codigo])
        rol_single = rol_lookup.get(codigo) or next(iter(rol_lookup.values()), None)
        if rol_single:
            resp['rol'] = rol_single
        teto_row = _get_teto_map([codigo], operadora_id=operadora_id).get(codigo.strip().upper())
        if teto_row:
            teto_val = _as_decimal(teto_row.valor_total)
            calc_total = _as_decimal(resp.get('total_final') or resp.get('total'))
            excedido = False
            excedente = None
            pct_total = None
            pct_excedente = None
            if teto_val is not None and calc_total is not None:
                diff = (calc_total - teto_val).quantize(quantize_money, rounding=ROUND_HALF_UP)
                if diff > Decimal('0'):
                    excedido = True
                    excedente = diff
                if teto_val != Decimal('0'):
                    pct_total = (calc_total / teto_val * Decimal('100')).quantize(quantize_pct, rounding=ROUND_HALF_UP)
                    if diff > Decimal('0'):
                        pct_excedente = (diff / teto_val * Decimal('100')).quantize(quantize_pct, rounding=ROUND_HALF_UP)
            resp['teto_valor_total'] = _stringify_for_output(teto_val) if teto_val is not None else None
            resp['teto_descricao'] = teto_row.descricao
            resp['teto_excedente'] = _stringify_for_output(excedente) if excedente is not None else None
            resp['teto_excedido'] = excedido
            resp['teto_status'] = 'ULTRAPASSA' if excedido else 'OK'
            resp['teto_pct_total'] = _stringify_for_output(pct_total) if pct_total is not None else None
            resp['teto_pct_excedente'] = _stringify_for_output(pct_excedente) if pct_excedente is not None else None
            resp['teto_alertas'] = [{
                'codigo': codigo,
                'descricao': resp.get('descricao'),
                'total_calculado': _stringify_for_output(calc_total) if calc_total is not None else None,
                'teto_valor_total': _stringify_for_output(teto_val) if teto_val is not None else None,
                'excedente': _stringify_for_output(excedente) if excedente is not None else None,
                'pct_total': _stringify_for_output(pct_total) if pct_total is not None else None,
                'pct_excedente': _stringify_for_output(pct_excedente) if pct_excedente is not None else None,
                'descricao_teto': teto_row.descricao,
            }] if excedido else []
        else:
            resp['teto_alertas'] = []
            resp['teto_status'] = 'OK'
    return resp, 200



@app.route('/api/simulacao_cbhpm', methods=['POST'])
@login_required
def api_simulacao_cbhpm():
    data = request.get_json(force=True, silent=True) or {}
    payload, status = _compute_simulacao_cbhpm(data)
    if status == 200:
        restore_payload = {
            'codigo': data.get('codigo') or '',
            'codigos': list(data.get('codigos') or []),
            'dtp_items': list(data.get('dtp_items') or []),
            'uf': data.get('uf') or '',
            'versao': data.get('versao') or '',
            'porte_tab': data.get('porte_tab') or '',
            'porte_an_tab': data.get('porte_an_tab') or '',
            'uco_valor': data.get('uco_valor') or '',
            'filme_valor': data.get('filme_valor') or '',
            'incidencias': data.get('incidencias') or '',
            'via_entrada_pcts': data.get('via_entrada_pcts') or {},
            'via_entrada_pct': data.get('via_entrada_pct') or '',
            'ajuste_porte_pct': data.get('ajuste_porte_pct') or '',
            'ajuste_porte_an_pct': data.get('ajuste_porte_an_pct') or '',
            'operadora_id': data.get('operadora_id') or session.get('operadora_id'),
        }
        label_parts = []
        codigo_label = (restore_payload.get('codigo') or '').strip()
        if codigo_label:
            label_parts.append(codigo_label)
        codes_list = restore_payload.get('codigos') or []
        if codes_list:
            snippet = ', '.join([str(c) for c in codes_list[:2]])
            if len(codes_list) > 2:
                snippet += ', ...'
            label_parts.append(f"codigos {snippet}")
        versao_label = (restore_payload.get('versao') or '').strip()
        if versao_label:
            label_parts.append(versao_label)
        label_raw = ' | '.join(filter(None, label_parts)) or 'Simulacao CBHPM'
        label = unicodedata.normalize('NFKD', label_raw).encode('ascii', 'ignore').decode()
        signature_source = json.dumps({'type': 'cbhpm', 'payload': restore_payload}, sort_keys=True).encode('utf-8')
        entry_id = hashlib.md5(signature_source).hexdigest()[:10]
        _store_history_entry({
            'type': 'cbhpm',
            'id': entry_id,
            'signature': f'cbhpm:{entry_id}',
            'url_fragment': f'sim_hist={entry_id}',
            'label': label[:80],
            'timestamp': datetime.now().strftime('%d/%m %H:%M'),
            'payload': restore_payload,
        })
    return jsonify(payload), status


@app.route('/api/simulacao_cbhpm/pdf', methods=['POST'])
@login_required
def export_simulacao_pdf():
    data = request.get_json(force=True, silent=True) or {}
    payload, status = _compute_simulacao_cbhpm(data)
    if status != 200:
        return jsonify(payload), status

    rules_meta = payload.get('cbhpm_rules_info') or {}
    if not rules_meta:
        _, rules_model = _get_active_cbhpm_ruleset(return_model=True)
        rules_meta = {
            'nome': getattr(rules_model, 'nome', 'Padrao'),
            'versao': getattr(rules_model, 'versao', None),
            'descricao': getattr(rules_model, 'descricao', None),
            'id': getattr(rules_model, 'id', None),
        }

    def fmt_brl(value):
        if value in (None, '', 'None'):
            return '-'
        try:
            val = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return str(value)
        formatted = f"{val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {formatted}"

    def fmt_pct(value):
        if value in (None, '', 'None'):
            return '-'
        try:
            val = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return str(value)
        return f"{val:.2f}%".replace('.', ',')

    def fmt_text(value, placeholder='-'):
        if value in (None, '', 'None'):
            return placeholder
        return str(value)

    logo_max_height = 18 * mm

    def load_logo_bytes(cache: dict[str, bytes] = {}) -> bytes | None:

        def _bitmap_to_bytes(path_logo: str):
            try:
                from PIL import Image as PILImage  # type: ignore
            except ImportError:
                return None
            try:
                with PILImage.open(path_logo) as opened:
                    opened.load()
                    if opened.mode not in ('RGB', 'RGBA'):
                        opened = opened.convert('RGBA')
                    if opened.mode == 'RGBA':
                        white_bg = PILImage.new('RGBA', opened.size, (255, 255, 255, 255))
                        opened = PILImage.alpha_composite(white_bg, opened)
                    buffer = io.BytesIO()
                    opened.convert('RGB').save(buffer, format='PNG')
                    return buffer.getvalue()
            except Exception:
                return None

        def _svg_to_bytes(path_logo: str):
            try:
                from svglib.svglib import svg2rlg  # type: ignore
                from reportlab.graphics import renderPM  # type: ignore
            except ImportError:
                return None
            try:
                drawing = svg2rlg(path_logo)
            except Exception:
                return None
            if not drawing or not getattr(drawing, 'height', None):
                return None
            try:
                scale = logo_max_height / float(drawing.height)
                drawing.scale(scale, scale)
                drawing.width = drawing.width * scale
                drawing.height = drawing.height * scale
                return renderPM.drawToString(drawing, fmt='PNG')
            except Exception:
                return None

        def _smart_to_bytes(path_logo: str):
            try:
                with open(path_logo, 'rb') as fh:
                    head = fh.read(256).lstrip()
            except OSError:
                return None
            if head.startswith(b'<svg'):
                return _svg_to_bytes(path_logo)
            return _bitmap_to_bytes(path_logo)

        candidates = [
            ('logo-pdf.svg', _svg_to_bytes),
            ('logo-pdf.png', _smart_to_bytes),
            ('logo-menu.png', _bitmap_to_bytes),
            ('logo-header.png', _bitmap_to_bytes),
            ('logo-login.png', _bitmap_to_bytes),
        ]

        for filename, loader in candidates:
            path_logo = os.path.join(app.root_path, 'static', filename)
            if not os.path.exists(path_logo):
                continue

            cache_key = os.path.abspath(path_logo)
            if cache_key not in cache:
                cache[cache_key] = loader(path_logo) or b''

            data = cache.get(cache_key) or b''
            if data:
                setattr(load_logo_bytes, 'last_static_name', filename)
                return data

        return None

    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M')
    logo_bytes = load_logo_bytes()
    logo_static_name = getattr(load_logo_bytes, 'last_static_name', None)
    logo_static_path = f'static/{logo_static_name}' if logo_static_name else None
    logo_uri = None
    if logo_bytes:
        import base64
        logo_uri = f"data:image/png;base64,{base64.b64encode(logo_bytes).decode('ascii')}"

    meta_rows = []
    meta_rows.append({'label': 'UF', 'value': fmt_text(data.get('uf') or 'Todos')})
    meta_rows.append({'label': 'Versão referência', 'value': fmt_text(payload.get('versao_base') or data.get('versao') or '-')})
    meta_rows.append({'label': 'Tabela de Porte', 'value': fmt_text(payload.get('porte_tabela_usada'))})
    meta_rows.append({'label': 'Tabela Porte AN', 'value': fmt_text(payload.get('porte_an_tabela_usada'))})
    meta_rows.append({'label': 'Ajuste Porte %', 'value': fmt_pct(payload.get('ajuste_porte_pct'))})
    meta_rows.append({'label': 'Ajuste Porte AN %', 'value': fmt_pct(payload.get('ajuste_porte_an_pct'))})
    meta_rows.append({'label': 'Valor UCO', 'value': fmt_brl(payload.get('uco_valor'))})
    meta_rows.append({'label': 'Incidências', 'value': fmt_text(data.get('incidencias') or '-')})

    requested_codes = []
    if isinstance(data.get('codigos'), list):
        requested_codes.extend([str(c).strip() for c in data.get('codigos') if c])
    if data.get('codigo'):
        requested_codes.append(str(data.get('codigo')).strip())
    codes_from_payload = []
    if isinstance(payload.get('itens'), list):
        codes_from_payload.extend([it.get('codigo') for it in payload['itens'] if it.get('codigo')])
    if payload.get('codigo'):
        codes_from_payload.append(payload.get('codigo'))
    merged_codes = []
    for code in requested_codes + codes_from_payload:
        if code and code not in merged_codes:
            merged_codes.append(code)

    if payload.get('descricao'):
        meta_rows.append({'label': 'Procedimento base', 'value': fmt_text(payload.get('descricao'))})
    if payload.get('itens'):
        meta_rows.append({'label': 'Quantidade de itens', 'value': str(len(payload['itens']))})

    limit = 12
    codes_preview = ', '.join(merged_codes[:limit]) if merged_codes else ''
    codes_extra = max(len(merged_codes) - limit, 0) if merged_codes else 0

    itens = payload.get('itens') or []
    itens_rows = []
    for item in itens:
        aux_raw = item.get('auxiliares_detalhe') or []
        aux_detail = []
        for det in aux_raw:
            aux_detail.append({
                'indice': det.get('indice'),
                'percentual': fmt_pct(det.get('percentual_pct')),
                'valor': fmt_brl(det.get('valor')),
            })
        itens_rows.append({
            'codigo': fmt_text(item.get('codigo')),
            'descricao': fmt_text(item.get('descricao')),
            'total_porte': fmt_brl(item.get('total_porte')),
            'total_filme': fmt_brl(item.get('total_filme')),
            'total_uco': fmt_brl(item.get('total_uco')),
            'total_porte_an': fmt_brl(item.get('total_porte_an')),
            'total_auxiliares': fmt_brl(item.get('total_auxiliares')),
            'total': fmt_brl(item.get('total')),
            'auxiliares_detalhe': aux_detail,
            'auxiliares_qtd': len(aux_detail),
        })

    fallback = {
        'codigo': fmt_text(payload.get('codigo') or data.get('codigo') or '-'),
        'descricao': fmt_text(payload.get('descricao') or ''),
    }

    summary_labels = [
        ('Total Porte', fmt_brl(payload.get('total_porte'))),
        ('Total Filme', fmt_brl(payload.get('total_filme'))),
        ('Total UCO', fmt_brl(payload.get('total_uco'))),
        ('Total Porte AN', fmt_brl(payload.get('total_porte_an'))),
        ('Total Auxiliares', fmt_brl(payload.get('total_auxiliares'))),
        ('Total Geral', fmt_brl(payload.get('total'))),
    ]

    summary_rows = [
        {
            'label': label,
            'value': value,
            'highlight': False,
        }
        for label, value in summary_labels
    ]
    if summary_rows:
        summary_rows[-1]['highlight'] = True

    context = {
        'generated_at': generated_at,
        'logo_data_uri': logo_uri,
        'logo_height_mm': 18,
        'logo_static_path': logo_static_path,
        'meta_rows': meta_rows,
        'codes_preview': codes_preview,
        'codes_extra': codes_extra,
        'itens': itens_rows,
        'fallback': fallback,
        'summary_rows': summary_rows,
        'logo_bytes': logo_bytes,
    }
    context['cbhpm_rules_info'] = rules_meta

    html_output = render_template('simulacao_cbhpm_pdf.html', **context)

    def render_reportlab_pdf(ctx: dict) -> bytes:
        buffer_rl = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer_rl,
            pagesize=A4,
            leftMargin=25 * mm,
            rightMargin=25 * mm,
            topMargin=30 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'PDFTitle',
            parent=styles['Title'],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#0f172a'),
            alignment=2,
        )
        header_info_style = ParagraphStyle(
            'PDFHeaderInfo',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#475569'),
            alignment=2,
        )
        meta_label_style = ParagraphStyle(
            'MetaLabel', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'), leading=12
        )
        meta_value_style = ParagraphStyle(
            'MetaValue', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#111827'), leading=12
        )
        table_header_style = ParagraphStyle(
            'TableHeader', parent=styles['Normal'], alignment=1, fontSize=10, textColor=colors.white, leading=12
        )
        table_text_style = ParagraphStyle(
            'TableText', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#111827'), leading=12
        )

        story = []
        logo_img = None
        if ctx.get('logo_bytes'):
            try:
                logo_img = Image(io.BytesIO(ctx['logo_bytes']))
                scale = logo_max_height / logo_img.imageHeight
                logo_img.drawHeight = logo_max_height
                logo_img.drawWidth = logo_img.imageWidth * scale
                logo_img.hAlign = 'LEFT'
            except Exception:
                logo_img = None

        header_title = Paragraph('<b>Relatório de Simulação CBHPM</b>', title_style)
        header_info = Paragraph(f"Gerado em {html.escape(ctx['generated_at'])}", header_info_style)
        if logo_img:
            logo_width = logo_img.drawWidth + 6
            col_widths = [logo_width, doc.width - logo_width]
        else:
            col_widths = [doc.width * 0.25, doc.width * 0.75]
        header_data = [
            [logo_img if logo_img else '', header_title],
            ['', header_info],
        ]
        header_table = Table(header_data, colWidths=col_widths)
        header_table.setStyle(
            TableStyle([
                ('SPAN', (0, 0), (0, 1)),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('ALIGN', (1, 1), (1, 1), 'RIGHT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
            ])
        )
        story.append(header_table)
        story.append(Spacer(0, 16))

        meta_rows_ctx = ctx.get('meta_rows') or []
        if meta_rows_ctx:
            rows = []
            for i in range(0, len(meta_rows_ctx), 2):
                left = meta_rows_ctx[i]
                right = meta_rows_ctx[i + 1] if i + 1 < len(meta_rows_ctx) else {'label': '', 'value': ''}
                rows.append([
                    Paragraph(f"<b>{html.escape(str(left['label']))}</b>", meta_label_style),
                    Paragraph(html.escape(str(left['value'])), meta_value_style),
                    Paragraph(f"<b>{html.escape(str(right['label']))}</b>", meta_label_style) if right['label'] else '',
                    Paragraph(html.escape(str(right['value'])), meta_value_style) if right['label'] else '',
                ])
            meta_table = Table(rows, colWidths=[doc.width * 0.18, doc.width * 0.32, doc.width * 0.18, doc.width * 0.32])
            meta_table.setStyle(
                TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ])
            )
            story.append(meta_table)
            story.append(Spacer(0, 12))

        if ctx.get('codes_preview'):
            codes_text = f"<b>Códigos selecionados:</b> {html.escape(ctx['codes_preview'])}"
            if ctx.get('codes_extra'):
                codes_text += f"<br/><font size=9 color='#475569'>+{ctx['codes_extra']} código(s) adicional(is) não exibido(s).</font>"
            story.append(Paragraph(codes_text, meta_value_style))
            story.append(Spacer(0, 10))

        itens_ctx = ctx.get('itens') or []
        if itens_ctx:
            table_data = [[
                Paragraph('<b>Código</b>', table_header_style),
                Paragraph('<b>Descrição</b>', table_header_style),
                Paragraph('<b>Total Porte</b>', table_header_style),
                Paragraph('<b>Total Filme</b>', table_header_style),
                Paragraph('<b>Total UCO</b>', table_header_style),
                Paragraph('<b>Total Porte AN</b>', table_header_style),
                Paragraph('<b>Auxiliares</b>', table_header_style),
                Paragraph('<b>Total</b>', table_header_style),
            ]]
            for item in itens_ctx:
                table_data.append([
                    Paragraph(html.escape(item['codigo']), table_text_style),
                    Paragraph(html.escape(item['descricao']), table_text_style),
                    Paragraph(item['total_porte'], table_text_style),
                    Paragraph(item['total_filme'], table_text_style),
                    Paragraph(item['total_uco'], table_text_style),
                    Paragraph(item['total_porte_an'], table_text_style),
                    Paragraph(item['total_auxiliares'], table_text_style),
                    Paragraph(item['total'], table_text_style),
                ])
            col_widths = [
                doc.width * 0.1,
                doc.width * 0.4,
                doc.width * 0.1,
                doc.width * 0.1,
                doc.width * 0.1,
                doc.width * 0.1,
                doc.width * 0.1,
                doc.width * 0.0 + 40,
            ]
            resultados_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            resultados_table.setStyle(
                TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ])
            )
            story.append(resultados_table)
            story.append(Spacer(0, 12))
        else:
            fallback_ctx = ctx.get('fallback') or {}
            story.append(Paragraph(
                f"<b>Código:</b> {html.escape(fallback_ctx.get('codigo', '-'))}",
                meta_value_style,
            ))
            if fallback_ctx.get('descricao'):
                story.append(Paragraph(f"<b>Descrição:</b> {html.escape(fallback_ctx['descricao'])}", meta_value_style))
            story.append(Spacer(0, 10))

        summary_rows_ctx = ctx.get('summary_rows') or []
        if summary_rows_ctx:
            rows = []
            highlight_map = []
            for i in range(0, len(summary_rows_ctx), 2):
                left = summary_rows_ctx[i]
                right = summary_rows_ctx[i + 1] if i + 1 < len(summary_rows_ctx) else {'label': '', 'value': '', 'highlight': False}
                highlight_map.append(bool(left.get('highlight') or right.get('highlight')))
                rows.append([
                    Paragraph(f"<b>{html.escape(left['label'])}</b>", meta_label_style),
                    Paragraph(html.escape(left['value']), meta_value_style),
                    Paragraph(f"<b>{html.escape(right['label'])}</b>", meta_label_style) if right['label'] else '',
                    Paragraph(html.escape(right['value']), meta_value_style) if right['label'] else '',
                ])
            summary_table = Table(rows, colWidths=[doc.width * 0.2, doc.width * 0.3, doc.width * 0.2, doc.width * 0.3])
            summary_table.setStyle(
                TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ecfeff')),
                    ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#0891b2')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#bae6fd')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ])
            )
            for idx, highlight in enumerate(highlight_map):
                if highlight:
                    summary_table.setStyle(
                        TableStyle([
                            ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#0f766e')),
                            ('TEXTCOLOR', (0, idx), (-1, idx), colors.white),
                            ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                        ])
                    )
            story.append(summary_table)

        def _footer(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setStrokeColor(colors.HexColor('#cbd5e1'))
            canvas_obj.setLineWidth(0.5)
            canvas_obj.line(doc_obj.leftMargin, 15, doc_obj.leftMargin + doc_obj.width, 15)
            canvas_obj.setFont('Helvetica', 9)
            canvas_obj.setFillColor(colors.HexColor('#475569'))
            canvas_obj.drawString(doc_obj.leftMargin, 5, f"Sistema de Simulação • Página {doc_obj.page}")
            canvas_obj.restoreState()

        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
        buffer_rl.seek(0)
        return buffer_rl.getvalue()

    try:
        from weasyprint import HTML  # type: ignore

        pdf_bytes = HTML(string=html_output, base_url=app.root_path).write_pdf()
    except Exception as exc:
        app.logger.warning('WeasyPrint indisponível (%s); usando ReportLab fallback.', exc)
        pdf_bytes = render_reportlab_pdf(context)

    buffer = io.BytesIO(pdf_bytes)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='simulacao.pdf')


@app.route('/api/simulacao_cbhpm/xlsx', methods=['POST'])
@login_required
def export_simulacao_xlsx():
    data = request.get_json(force=True, silent=True) or {}
    payload, status = _compute_simulacao_cbhpm(data)
    if status != 200:
        return jsonify(payload), status

    def to_number(value):
        try:
            return float(value)
        except (TypeError, ValueError, InvalidOperation):
            return 0.0

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Simulacao")

    bold = workbook.add_format({'bold': True})
    money = workbook.add_format({'num_format': 'R$ #,##0.00'})

    headers = [
        "Codigo",
        "Descricao",
        "Total Porte",
        "Total Filme",
        "Total UCO",
        "Total Porte AN",
        "Total Auxiliares",
        "Via Entrada",
        "Total",
        "Teto",
        "Excedente",
    ]
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, bold)

    row_idx = 1
    itens = payload.get('itens') or []
    if itens:
        for item in itens:
            worksheet.write(row_idx, 0, item.get('codigo'))
            worksheet.write(row_idx, 1, item.get('descricao') or '')
            worksheet.write_number(row_idx, 2, to_number(item.get('total_porte')), money)
            worksheet.write_number(row_idx, 3, to_number(item.get('total_filme')), money)
            worksheet.write_number(row_idx, 4, to_number(item.get('total_uco')), money)
            worksheet.write_number(row_idx, 5, to_number(item.get('total_porte_an')), money)
            worksheet.write_number(row_idx, 6, to_number(item.get('total_auxiliares')), money)
            worksheet.write(row_idx, 7, item.get('via_entrada_pct') or '-')
            worksheet.write_number(row_idx, 8, to_number(item.get('total')), money)
            teto_value = item.get('teto_valor_total')
            excedente_value = item.get('teto_excedente')
            if teto_value not in (None, '', 'None'):
                worksheet.write_number(row_idx, 9, to_number(teto_value), money)
            else:
                worksheet.write_blank(row_idx, 9, None)
            if excedente_value not in (None, '', 'None'):
                worksheet.write_number(row_idx, 10, to_number(excedente_value), money)
            else:
                worksheet.write_blank(row_idx, 10, None)
            row_idx += 1
    else:
        worksheet.write(row_idx, 0, payload.get('codigo'))
        worksheet.write(row_idx, 1, payload.get('descricao') or '')
        worksheet.write_number(row_idx, 2, to_number(payload.get('total_porte')), money)
        worksheet.write_number(row_idx, 3, to_number(payload.get('total_filme')), money)
        worksheet.write_number(row_idx, 4, to_number(payload.get('total_uco')), money)
        worksheet.write_number(row_idx, 5, to_number(payload.get('total_porte_an')), money)
        worksheet.write_number(row_idx, 6, to_number(payload.get('total_auxiliares')), money)
        worksheet.write(row_idx, 7, payload.get('via_entrada_pct') or (payload.get('via_entrada_summary') or '-'))
        worksheet.write_number(row_idx, 8, to_number(payload.get('total')), money)
        teto_value = payload.get('teto_valor_total')
        excedente_value = payload.get('teto_excedente')
        if teto_value not in (None, '', 'None'):
            worksheet.write_number(row_idx, 9, to_number(teto_value), money)
        if excedente_value not in (None, '', 'None'):
            worksheet.write_number(row_idx, 10, to_number(excedente_value), money)
        row_idx += 1

    worksheet.write(row_idx + 1, 7, 'TOTAL GERAL', bold)
    worksheet.write_number(row_idx + 1, 8, to_number(payload.get('total')), money)

    workbook.close()
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='simulacao.xlsx'
    )

@app.route('/api/simulacao_dtp')
@login_required
def api_simulacao_dtp():
    """Pesquisa itens em 'Diárias, Taxas e Pacotes' por tabela e termo (código ou descrição).
    Parâmetros: tabela_nome (obrig.), q (código ou parte da descrição), uf (opcional), operadora_id (opcional)
    """
    tabela_nome = request.args.get('tabela_nome') or ''
    q = (request.args.get('q') or '').strip()
    uf = (request.args.get('uf') or '').strip() or None

    # Multi-operadora: obter operadora_id do request ou da sessão
    operadora_id = request.args.get('operadora_id')
    if not operadora_id:
        operadora_id = session.get('operadora_id')

    if not tabela_nome:
        return jsonify({'itens': [], 'total': '0'})
    t = Tabela.query.filter_by(nome=tabela_nome, tipo_tabela='diarias_taxas_pacotes').first()
    if not t:
        return jsonify({'itens': [], 'total': '0'})

    query = db.session.query(Procedimento).filter(Procedimento.id_tabela == t.id)

    # Multi-operadora: filtrar por operadora_id se fornecido
    if operadora_id:
        query = query.filter(Procedimento.operadora_id == operadora_id)

    if uf:
        query = query.filter(or_(Procedimento.uf == uf, t.uf == uf))
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Procedimento.codigo == q, Procedimento.codigo.ilike(f"{q}%"), Procedimento.descricao.ilike(like)))
    rows = query.order_by(Procedimento.codigo).limit(200).all()
    itens = [{'codigo': r.codigo, 'descricao': r.descricao, 'valor': (str(r.valor) if r.valor is not None else None)} for r in rows]
    total = sum([_as_decimal(r.valor) or Decimal('0') for r in rows])
    return jsonify({'itens': itens, 'total': str(total)})


@app.route('/api/prestadores_por_codigo')
@login_required
def api_prestadores_por_codigo():
    """Retorna a lista de prestadores que possuem o código informado
    dentro da tabela selecionada e UF opcional.
    Parâmetros: tabela_nome, codigo, uf (opcional), operadora_id (opcional)
    """
    tabela_nome = request.args.get('tabela_nome')
    codigo = (request.args.get('codigo') or '').strip()
    uf = request.args.get('uf')

    # Multi-operadora: obter operadora_id do request ou da sessão
    operadora_id = request.args.get('operadora_id')
    if not operadora_id:
        operadora_id = session.get('operadora_id')

    if not tabela_nome or not codigo:
        return jsonify([])

    # Extrai somente o código caso venha no formato "codigo - descricao"
    if ' - ' in codigo:
        codigo = codigo.split(' - ', 1)[0].strip()

    q = db.session.query(Procedimento.prestador).join(Tabela, Procedimento.id_tabela == Tabela.id)
    q = q.filter(Tabela.nome == tabela_nome)

    # Multi-operadora: filtrar por operadora_id se fornecido
    if operadora_id:
        q = q.filter(Procedimento.operadora_id == operadora_id)

    if uf:
        q = q.filter(or_(Tabela.uf == uf, Procedimento.uf == uf))
    # Match por igualdade ou prefixo
    q = q.filter(or_(Procedimento.codigo == codigo, Procedimento.codigo.ilike(f"{codigo}%")))
    q = q.filter((Procedimento.prestador.isnot(None)) & (Procedimento.prestador != ''))
    prestadores = [r[0] for r in q.distinct().order_by(Procedimento.prestador).all()]
    return jsonify(prestadores)


@app.route('/api/versoes_por_codigo')
@login_required
def api_versoes_por_codigo():
    tabela_nome = request.args.get('tabela_nome')
    codigo = (request.args.get('codigo') or '').strip()
    uf = request.args.get('uf')
    if not codigo:
        return jsonify([])
    if ' - ' in codigo:
        codigo = codigo.split(' - ', 1)[0].strip()
    qv = db.session.query(Tabela.nome).join(CBHPMItem, CBHPMItem.id_tabela == Tabela.id).filter(Tabela.tipo_tabela == 'cbhpm')
    if uf:
        qv = qv.filter(or_(Tabela.uf == uf, CBHPMItem.uf == uf))
    qv = qv.filter(or_(CBHPMItem.codigo == codigo, CBHPMItem.codigo.ilike(f"{codigo}%")))
    versoes = [r[0] for r in qv.distinct().order_by(Tabela.nome).all()]
    return jsonify(versoes)


# ═══════════════════════════════════════════════════════════════════════════
# SMART FILTERS API ENDPOINTS - Para o Consulta & Comparar
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/api/tabela-info/<int:table_id>')
@login_required
def api_tabela_info(table_id):
    """
    Retorna informações sobre a tabela (tipo: CBHPM ou DTP).

    Parâmetros:
    - table_id: ID da tabela

    Resposta:
    {
        'id': int,
        'nome': str,
        'tipo': 'cbhpm' ou 'diarias_taxas_pacotes'
    }
    """
    tabela = Tabela.query.get(table_id)
    if not tabela:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    # Multi-operadora: verificar se usuário tem acesso a esta tabela
    operadora_id = session.get('operadora_id')
    if operadora_id and tabela.id_operadora != operadora_id:
        return jsonify({'error': 'Acesso negado'}), 403

    return jsonify({
        'id': tabela.id,
        'nome': tabela.nome,
        'tipo': tabela.tipo_tabela or 'diarias_taxas_pacotes'
    })


@app.route('/api/prestadores/<int:table_id>')
@login_required
def api_get_prestadores(table_id):
    """
    Retorna lista de prestadores únicos da tabela selecionada.

    Parâmetros Query:
    - uf: (opcional) Filtrar por UF

    Resposta:
    {
        'tabela_id': int,
        'prestadores': [str],
        'total': int
    }
    """
    tabela = Tabela.query.get(table_id)
    if not tabela:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    # Multi-operadora: verificar acesso
    operadora_id = session.get('operadora_id')
    if operadora_id and tabela.id_operadora != operadora_id:
        return jsonify({'error': 'Acesso negado'}), 403

    uf = request.args.get('uf', '').strip()

    # Construir query
    query = db.session.query(Procedimento.prestador)\
        .filter(Procedimento.id_tabela == table_id)\
        .filter(Procedimento.prestador.isnot(None))\
        .filter(Procedimento.prestador != '')

    # Filtrar por UF se fornecido
    if uf:
        query = query.filter(or_(Procedimento.uf == uf, Tabela.uf == uf))

    # Multi-operadora: filtrar por operadora_id
    if operadora_id:
        query = query.filter(Procedimento.operadora_id == operadora_id)

    prestadores = sorted(list(set([r[0] for r in query.all()])))

    return jsonify({
        'tabela_id': table_id,
        'prestadores': prestadores,
        'total': len(prestadores)
    })


@app.route('/api/versoes/<int:table_id>')
@login_required
def api_get_versoes(table_id):
    """
    Retorna lista de versões para tabelas CBHPM.

    Parâmetros:
    - table_id: ID da tabela (apenas para validação)

    Resposta:
    {
        'tabela_id': int,
        'versoes': [str],
        'total': int
    }
    """
    # Validar que a tabela existe
    tabela = Tabela.query.get(table_id)
    if not tabela:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    # Multi-operadora: verificar acesso
    operadora_id = session.get('operadora_id')
    if operadora_id and tabela.id_operadora != operadora_id:
        return jsonify({'error': 'Acesso negado'}), 403

    # Buscar todas as versões CBHPM da operadora
    query = db.session.query(Tabela.nome)\
        .filter(Tabela.tipo_tabela == 'cbhpm')\
        .distinct()

    # Multi-operadora: filtrar por operadora_id
    if operadora_id:
        query = query.filter(Tabela.id_operadora == operadora_id)

    versoes = [v[0] for v in query.order_by(Tabela.nome).all()]

    return jsonify({
        'tabela_id': table_id,
        'versoes': versoes,
        'total': len(versoes)
    })


@app.route('/api/tabelas-list')
@login_required
def api_tabelas_list():
    """
    Retorna lista de tabelas com seus IDs para busca por nome.

    Parâmetros:
    - tipo: tipo de tabela (ex: diarias_taxas_pacotes, cbhpm)
    - nome: nome da tabela (busca exata)

    Resposta:
    {
        "tabelas": [
            {"id": 1, "nome": "DIÁRIAS, TAXAS E PACOTES 2024"}
        ]
    }
    """
    operadora_id = session.get('operadora_id')
    tipo = (request.args.get('tipo') or '').strip()
    nome = (request.args.get('nome') or '').strip()

    query = Tabela.query
    if tipo:
        query = query.filter_by(tipo_tabela=tipo)
    if nome:
        query = query.filter_by(nome=nome)
    if operadora_id:
        query = query.filter_by(id_operadora=operadora_id)

    tabelas = query.all()
    resultado = [
        {'id': t.id, 'nome': t.nome}
        for t in tabelas
    ]

    return jsonify({'tabelas': resultado})


@app.route('/api/dtp-codigos/<int:table_id>')
@login_required
def api_dtp_codigos(table_id):
    """
    Retorna lista de códigos DTP para autocomplete.

    Parâmetros:
    - table_id: ID da tabela DTP
    - q: termo de busca (opcional, para filtro)

    Resposta:
    {
        "codigos": [
            {"codigo": "01.01.01", "descricao": "Diária de UTI"},
            {"codigo": "01.01.02", "descricao": "Diária de Hospital"},
            ...
        ],
        "total": 50
    }
    """
    operadora_id = session.get('operadora_id')

    # Verificar se a tabela existe e pertence à operadora
    tabela = Tabela.query.get(table_id)
    if not tabela:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    if operadora_id and tabela.id_operadora != operadora_id:
        return jsonify({'error': 'Acesso negado'}), 403

    # Verificar se é tabela DTP
    if tabela.tipo_tabela != 'diarias_taxas_pacotes':
        return jsonify({'error': 'Tabela não é do tipo DTP'}), 400

    # Obter termo de busca (opcional)
    q = (request.args.get('q') or '').strip()

    # Query base
    query = Procedimento.query.filter_by(id_tabela=table_id)

    # Filtrar por termo de busca se fornecido
    if q:
        query = query.filter(
            (Procedimento.codigo.ilike(f'{q}%')) |
            (Procedimento.descricao.ilike(f'%{q}%'))
        )

    # Filtrar por operadora se usuário tem operadora específica
    if operadora_id:
        query = query.filter_by(operadora_id=operadora_id)

    # Buscar e ordenar
    procedimentos = query.order_by(
        Procedimento.codigo.asc()
    ).limit(100).all()

    # Montar resposta
    codigos = [
        {
            'codigo': proc.codigo,
            'descricao': proc.descricao or '',
            'prestador': proc.prestador or '',
            'valor': float(proc.valor) if proc.valor else None,
        }
        for proc in procedimentos
    ]

    return jsonify({
        'codigos': codigos,
        'total': len(codigos)
    })


@app.route('/api/dtp-prestadores/<int:table_id>')
@login_required
def api_dtp_prestadores(table_id):
    """
    Retorna lista de prestadores únicos de uma tabela DTP.

    Parâmetros:
    - table_id: ID da tabela DTP

    Resposta:
    {
        "prestadores": [
            "Hospital Central",
            "Hospital Metropolitano",
            "Clínica Médica Plus"
        ],
        "total": 3
    }
    """
    operadora_id = session.get('operadora_id')

    # Verificar se a tabela existe e pertence à operadora
    tabela = Tabela.query.get(table_id)
    if not tabela:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    if operadora_id and tabela.id_operadora != operadora_id:
        return jsonify({'error': 'Acesso negado'}), 403

    # Verificar se é tabela DTP
    if tabela.tipo_tabela != 'diarias_taxas_pacotes':
        return jsonify({'error': 'Tabela não é do tipo DTP'}), 400

    # Buscar prestadores únicos
    query = Procedimento.query.filter_by(id_tabela=table_id)

    # Filtrar por operadora se usuário tem operadora específica
    if operadora_id:
        query = query.filter_by(operadora_id=operadora_id)

    # Buscar prestadores únicos e ordenar
    prestadores = db.session.query(Procedimento.prestador).filter_by(
        id_tabela=table_id
    ).filter(Procedimento.prestador.isnot(None)).filter(
        Procedimento.prestador != ''
    ).distinct().order_by(Procedimento.prestador.asc()).all()

    if operadora_id:
        prestadores = db.session.query(Procedimento.prestador).filter(
            Procedimento.id_tabela == table_id,
            Procedimento.operadora_id == operadora_id,
            Procedimento.prestador.isnot(None),
            Procedimento.prestador != ''
        ).distinct().order_by(Procedimento.prestador.asc()).all()

    prestadores_list = [p[0] for p in prestadores]

    return jsonify({
        'prestadores': prestadores_list,
        'total': len(prestadores_list)
    })


@app.route('/api/procedimentos/suggest')
@login_required
def api_procedimentos_suggest():
    """Sugere procedimentos a partir de um termo livre.

    Aceita parâmetros opcionais:
    - tabela_nome: restringe à tabela selecionada
    - uf: filtra por UF associada à tabela ou ao procedimento
    - limit: quantidade máxima de itens (default=30, máximo 100)
    """
    term = (request.args.get('q') or '').strip()
    if len(term) < 2:
        return jsonify({'items': []})

    tabela_nome = (request.args.get('tabela_nome') or '').strip()
    uf = (request.args.get('uf') or '').strip().upper()
    try:
        limit = int(request.args.get('limit') or 30)
    except (TypeError, ValueError):
        limit = 30
    limit = max(1, min(limit, 100))

    tabela = None
    if tabela_nome:
        tabela = Tabela.query.filter(Tabela.nome == tabela_nome).first()

    tabela_id = tabela.id if tabela else None

    like_term = f"%{term}%"
    code_prefix = f"{term}%"

    items = []
    seen_codes = set()

    if tabela and tabela.tipo_tabela == 'cbhpm':
        query = (
            db.session.query(CBHPMItem.codigo, CBHPMItem.procedimento, Tabela.nome)
            .join(Tabela, CBHPMItem.id_tabela == Tabela.id)
            .filter(CBHPMItem.id_tabela == tabela_id)
        )
        if uf:
            query = query.filter(or_(CBHPMItem.uf == uf, Tabela.uf == uf))
        query = query.filter(or_(CBHPMItem.codigo.ilike(code_prefix), CBHPMItem.procedimento.ilike(like_term)))
        query = query.order_by(CBHPMItem.codigo).limit(limit)
        for codigo, descricao, versao_nome in query.all():
            codigo_norm = (codigo or '').strip()
            if not codigo_norm or codigo_norm in seen_codes:
                continue
            seen_codes.add(codigo_norm)
            items.append({
                'codigo': codigo_norm,
                'descricao': descricao,
                'caminho': versao_nome,
            })
    elif tabela:
        query = (
            db.session.query(Procedimento.codigo, Procedimento.descricao, Tabela.nome)
            .join(Tabela, Procedimento.id_tabela == Tabela.id)
            .filter(Procedimento.id_tabela == tabela_id)
        )
        if uf:
            query = query.filter(or_(Tabela.uf == uf, Procedimento.uf == uf))
        query = query.filter(or_(Procedimento.codigo.ilike(code_prefix), Procedimento.descricao.ilike(like_term)))
        query = query.order_by(Procedimento.codigo).limit(limit)
        for codigo, descricao, tabela_nome_row in query.all():
            codigo_norm = (codigo or '').strip()
            if not codigo_norm or codigo_norm in seen_codes:
                continue
            seen_codes.add(codigo_norm)
            items.append({
                'codigo': codigo_norm,
                'descricao': descricao,
                'caminho': tabela_nome_row,
            })

    # Se nenhuma tabela específica foi pedida (ou não encontrada), faz uma busca genérica
    if not items and tabela is None:
        cbhpm_query = (
            db.session.query(CBHPMItem.codigo, CBHPMItem.procedimento, Tabela.nome)
            .join(Tabela, CBHPMItem.id_tabela == Tabela.id)
            .filter(Tabela.tipo_tabela == 'cbhpm')
        )
        if uf:
            cbhpm_query = cbhpm_query.filter(or_(CBHPMItem.uf == uf, Tabela.uf == uf))
        cbhpm_query = cbhpm_query.filter(or_(CBHPMItem.codigo.ilike(code_prefix), CBHPMItem.procedimento.ilike(like_term)))
        cbhpm_query = cbhpm_query.order_by(CBHPMItem.codigo).limit(limit)
        for codigo, descricao, versao_nome in cbhpm_query.all():
            codigo_norm = (codigo or '').strip()
            if not codigo_norm or codigo_norm in seen_codes:
                continue
            seen_codes.add(codigo_norm)
            items.append({
                'codigo': codigo_norm,
                'descricao': descricao,
                'caminho': versao_nome,
            })
            if len(items) >= limit:
                break

    return jsonify({'items': items[:limit]})


@app.route('/api/cbhpm/detalhe')
@login_required
def api_cbhpm_detalhe():
    codigo = (request.args.get('codigo') or '').strip()
    if not codigo:
        return jsonify({'error': 'Código obrigatório.'}), 400

    uf = (request.args.get('uf') or '').strip().upper() or None
    versoes = [v.strip() for v in request.args.getlist('versoes') if v and v.strip()]
    tabela_nome = (request.args.get('tabela_nome') or '').strip()

    query = (
        db.session.query(
            Tabela.nome.label('versao'),
            CBHPMItem.procedimento,
            CBHPMItem.subtotal,
            CBHPMItem.total_porte,
            CBHPMItem.valor_porte,
            CBHPMItem.total_uco,
            CBHPMItem.uco,
            CBHPMItem.total_filme,
            CBHPMItem.filme
        )
        .join(Tabela, CBHPMItem.id_tabela == Tabela.id)
        .filter(Tabela.tipo_tabela == 'cbhpm')
        .filter(CBHPMItem.codigo == codigo)
    )

    if versoes:
        query = query.filter(Tabela.nome.in_(versoes))
    elif tabela_nome:
        query = query.filter(Tabela.nome == tabela_nome)

    if uf:
        query = query.filter(or_(Tabela.uf == uf, CBHPMItem.uf == uf))

    records = query.order_by(Tabela.nome).all()

    items = []
    numeric_values = []
    for versao, descricao, subtotal, total_porte, valor_porte, total_uco, uco, total_filme, filme in records:
        valor = subtotal or total_porte or valor_porte or total_uco or uco or total_filme or filme
        if valor is not None:
            try:
                numeric = float(valor)
                numeric_values.append(numeric)
            except (TypeError, ValueError):
                pass
        items.append({
            'versao': versao,
            'descricao': descricao,
            'valor': float(valor) if valor is not None else None,
        })

    summary = None
    if numeric_values:
        count = len(numeric_values)
        summary = {
            'min': min(numeric_values),
            'max': max(numeric_values),
            'avg': sum(numeric_values) / count if count else None,
            'count': count,
        }

    return jsonify({'items': items, 'summary': summary})


def _serialize_public_cbhpm_item(item: CBHPMItem, tabela_ref: Tabela) -> dict:
    valor_total = _resolve_cbhpm_valor_total(item, tabela_ref)
    return {
        'codigo': item.codigo,
        'descricao': item.procedimento,
        'versao': tabela_ref.nome,
        'valor_total': _format_money_decimal(valor_total),
        'moeda': 'BRL',
    }


@app.route('/api/v1/cbhpm/procedimentos/<codigo>')
@public_api_key_required
def api_public_cbhpm_detail(codigo: str):
    codigo = (codigo or '').strip()
    if not codigo:
        return _api_error('invalid_input', 'Código obrigatório.', 400)

    tabela = _get_latest_cbhpm_table()
    if not tabela:
        return _api_error('not_found', 'Nenhuma tabela CBHPM disponível.', 404)

    cache_key = f"{tabela.id}:{codigo.lower()}"
    cached = _cbhpm_api_detail_cache.get(cache_key)
    if cached:
        return jsonify(cached)

    item = (
        CBHPMItem.query
        .filter(CBHPMItem.id_tabela == tabela.id)
        .filter(or_(CBHPMItem.codigo == codigo, CBHPMItem.codigo.ilike(codigo)))
        .order_by(CBHPMItem.codigo)
        .first()
    )
    if not item:
        return _api_error('not_found', 'Procedimento não encontrado na CBHPM mais recente.', 404)

    payload = _serialize_public_cbhpm_item(item, tabela)
    _cbhpm_api_detail_cache[cache_key] = payload
    return jsonify(payload)


@app.route('/api/v1/cbhpm/procedimentos')
@public_api_key_required
def api_public_cbhpm_search():
    q = (request.args.get('q') or '').strip()
    if not q:
        return _api_error('invalid_input', 'Parâmetro q é obrigatório.', 400)
    if len(q) > 120:
        q = q[:120]
    limit = _parse_positive_int(request.args.get('limit'), 20, maximum=100)

    tabela = _get_latest_cbhpm_table()
    if not tabela:
        return _api_error('not_found', 'Nenhuma tabela CBHPM disponível.', 404)

    cache_key = f"{tabela.id}:{q.lower()}:{limit}"
    cached = _cbhpm_api_cache.get(cache_key)
    if cached:
        return jsonify(cached)

    like_term = f"%{q}%"
    code_prefix = f"{q}%"

    items: list[dict] = []
    seen_codes: set[str] = set()

    code_rows = (
        db.session.query(CBHPMItem)
        .filter(CBHPMItem.id_tabela == tabela.id)
        .filter(or_(CBHPMItem.codigo == q, CBHPMItem.codigo.ilike(code_prefix)))
        .order_by(CBHPMItem.codigo)
        .limit(limit * 2)
        .all()
    )
    for row in code_rows:
        codigo_norm = (row.codigo or '').strip()
        if not codigo_norm or codigo_norm in seen_codes:
            continue
        seen_codes.add(codigo_norm)
        items.append(_serialize_public_cbhpm_item(row, tabela))
        if len(items) >= limit:
            break

    if len(items) < limit:
        desc_rows = (
            db.session.query(CBHPMItem)
            .filter(CBHPMItem.id_tabela == tabela.id)
            .filter(CBHPMItem.procedimento.ilike(like_term))
            .order_by(CBHPMItem.codigo)
            .limit(limit * 2)
            .all()
        )
        for row in desc_rows:
            if len(items) >= limit:
                break
            codigo_norm = (row.codigo or '').strip()
            if not codigo_norm or codigo_norm in seen_codes:
                continue
            seen_codes.add(codigo_norm)
            items.append(_serialize_public_cbhpm_item(row, tabela))

    if not items:
        return _api_error('not_found', 'Procedimento não encontrado na CBHPM mais recente.', 404)

    payload = {
        'versao': tabela.nome,
        'items': items[:limit],
    }
    _cbhpm_api_cache[cache_key] = payload
    return jsonify(payload)


@app.route('/api/tuss-rol')
@login_required
@feature_required('tuss_rol')
def api_tuss_rol_lookup():
    raw_codigos = []
    raw_codigos.extend(request.args.getlist('codigo'))
    raw_codigos.extend(request.args.getlist('codigos'))
    codigos_extra = (request.args.get('codes') or '').strip()
    if codigos_extra:
        raw_codigos.extend([c.strip() for c in codigos_extra.split(',') if c.strip()])

    normalized = [_normalize_tuss_codigo(c) for c in raw_codigos if _normalize_tuss_codigo(c)]
    # Permite códigos duplicados - não remove duplicatas

    if normalized:
        records = (
            TussRolCorrelacao.query
            .filter(TussRolCorrelacao.codigo.in_(normalized))
            .all()
        )
    else:
        try:
            limit = int(request.args.get('limit', 200) or 200)
        except ValueError:
            limit = 200
        limit = max(1, min(limit, 500))
        records = (
            TussRolCorrelacao.query
            .order_by(TussRolCorrelacao.codigo.asc())
            .limit(limit)
            .all()
        )

    items = [{
        'codigo': rec.codigo,
        'descricao': rec.descricao,
        'consta': bool(rec.consta_rol),
    } for rec in records]
    return jsonify({'items': items})


@app.route('/api/tuss-rol/<codigo>')
@login_required
@feature_required('tuss_rol')
def api_tuss_rol_item(codigo: str):
    codigo_norm = _normalize_tuss_codigo(codigo)
    if not codigo_norm:
        abort(400, description='Código inválido.')
    registro = TussRolCorrelacao.query.filter_by(codigo=codigo_norm).first()
    if not registro:
        return jsonify({'codigo': codigo_norm, 'consta': False, 'descricao': None}), 404
    return jsonify({
        'codigo': registro.codigo,
        'descricao': registro.descricao,
        'consta': bool(registro.consta_rol),
    })


@app.route('/gerenciar-usuarios')
@admin_required
def gerenciar_usuarios():
    # Filtros de busca
    nome_filter = request.args.get('nome', '').strip()
    operadora_filter = request.args.get('operadora_id', '').strip()

    # Query base
    query = Usuario.query

    # Filtro por nome
    if nome_filter:
        query = query.filter(
            or_(
                Usuario.nome.ilike(f'%{nome_filter}%'),
                Usuario.email.ilike(f'%{nome_filter}%')
            )
        )

    # Filtro por operadora
    if operadora_filter:
        try:
            operadora_id = int(operadora_filter)
            # Join com a tabela de relacionamento usuario_operadoras
            query = query.join(Usuario.operadoras).filter(Operadora.id == operadora_id)
        except (TypeError, ValueError):
            pass

    usuarios = query.order_by(Usuario.nome).all()

    # Lista de operadoras para o filtro
    operadoras_list = Operadora.query.filter_by(status='Ativa').order_by(Operadora.nome).all()

    return render_template(
        'gerenciar-usuarios.html',
        usuarios=usuarios,
        operadoras_list=operadoras_list,
        nome_filter=nome_filter,
        operadora_filter=operadora_filter
    )


@app.route('/gerenciar-operadoras')
@admin_required
def gerenciar_operadoras():
    operadoras = Operadora.query.all()
    return render_template('gerenciar-operadoras.html', operadoras=operadoras)


@app.route('/gerenciar-tabelas')
@admin_required
def gerenciar_tabelas():
    tabelas = Tabela.query.all()
    operadoras = Operadora.query.all()
    cbhpm_tabelas = Tabela.query.filter_by(tipo_tabela='cbhpm').order_by(Tabela.nome).all()
    return render_template('gerenciar-tabelas.html', tabelas=tabelas, operadoras=operadoras, UFS=BR_UFS, cbhpm_tabelas=cbhpm_tabelas)


@app.route('/contratos-resumo', methods=['GET', 'POST'])
@login_required
@feature_required('contratos')
def contratos_resumo():
    erro = None
    form_data: dict[str, str | None] = {}

    # Multi-operadora: obter operadora_id do usuário logado
    user_operadora_id = None
    if hasattr(g, 'current_user') and g.current_user and g.current_user.operadoras:
        # Usuário com operadoras específicas - usar primeira operadora
        user_operadora_id = g.current_user.operadoras[0].id

    # Operadora selecionada (para admins) ou do usuário
    selected_operadora_id = request.args.get('operadora_id')
    if selected_operadora_id:
        try:
            selected_operadora_id = int(selected_operadora_id)
        except (TypeError, ValueError):
            selected_operadora_id = None

    # Se usuário tem operadora específica, forçar usar ela
    if user_operadora_id:
        selected_operadora_id = user_operadora_id
    elif not selected_operadora_id:
        selected_operadora_id = session.get('operadora_id', 1)

    if request.method == 'POST':
        form = request.form or {}
        record_id_raw = (form.get('record_id') or '').strip()
        prestador = (form.get('prestador') or '').strip()
        tabela_honorarios = (form.get('tabela_honorarios') or '').strip() or None
        tabela_portes = (form.get('tabela_portes') or '').strip() or None
        valor_uco_raw = (form.get('valor_uco') or '').strip()
        inflator_deflator = (form.get('inflator_deflator') or '').strip() or None
        filme_radiologico = (form.get('filme_radiologico') or '').strip() or None
        observacoes = (form.get('observacoes') or '').strip() or None
        operadora_id_form = (form.get('operadora_id') or '').strip()

        # Validar operadora do formulário
        if operadora_id_form:
            try:
                operadora_id_form = int(operadora_id_form)
            except (TypeError, ValueError):
                operadora_id_form = selected_operadora_id
        else:
            operadora_id_form = selected_operadora_id

        # SEGURANÇA: Se usuário tem operadora específica, forçar usar ela
        if user_operadora_id:
            operadora_id_form = user_operadora_id

        valor_uco = None
        if valor_uco_raw:
            try:
                valor_uco = _parse_money(valor_uco_raw)
            except Exception:
                valor_uco = None

        if not prestador:
            erro = 'Informe o prestador.'
        else:
            try:
                if record_id_raw:
                    record_id = int(record_id_raw)
                    resumo = ContractSummary.query.get(record_id)
                    if not resumo:
                        erro = 'Registro não encontrado.'
                    # SEGURANÇA: Verificar se usuário tem acesso a este contrato
                    elif user_operadora_id and resumo.operadora_id != user_operadora_id:
                        erro = 'Você não tem permissão para editar este contrato.'
                    else:
                        resumo.prestador = prestador
                        resumo.tabela_honorarios = tabela_honorarios
                        resumo.tabela_portes = tabela_portes
                        resumo.valor_uco = valor_uco
                        resumo.inflator_deflator = inflator_deflator
                        resumo.filme_radiologico = filme_radiologico
                        resumo.observacoes = observacoes
                        resumo.operadora_id = operadora_id_form
                        db.session.commit()
                        flash('Resumo atualizado com sucesso.', 'success')
                        return redirect(url_for('contratos_resumo', operadora_id=operadora_id_form))
                else:
                    resumo = ContractSummary(
                        prestador=prestador,
                        tabela_honorarios=tabela_honorarios,
                        tabela_portes=tabela_portes,
                        valor_uco=valor_uco,
                        inflator_deflator=inflator_deflator,
                        filme_radiologico=filme_radiologico,
                        observacoes=observacoes,
                        operadora_id=operadora_id_form,
                    )
                    db.session.add(resumo)
                    db.session.commit()
                    flash('Resumo cadastrado com sucesso.', 'success')
                    return redirect(url_for('contratos_resumo', operadora_id=operadora_id_form))
            except ValueError:
                erro = 'Identificador inválido.'
            except Exception as exc:
                db.session.rollback()
                app.logger.error('Erro ao salvar resumo de contrato: %s', exc)
                erro = 'Não foi possível salvar o resumo. Tente novamente.'

        form_data = {
            'record_id': record_id_raw,
            'prestador': prestador,
            'tabela_honorarios': tabela_honorarios or '',
            'tabela_portes': tabela_portes or '',
            'valor_uco': valor_uco_raw,
            'inflator_deflator': inflator_deflator or '',
            'filme_radiologico': filme_radiologico or '',
            'observacoes': observacoes or '',
        }
    else:
        edit_id = request.args.get('edit')
        if edit_id:
            try:
                resumo = ContractSummary.query.get(int(edit_id))
            except (TypeError, ValueError):
                resumo = None
            # SEGURANÇA: Verificar se usuário tem acesso a este contrato
            if resumo and user_operadora_id and resumo.operadora_id != user_operadora_id:
                flash('Você não tem permissão para editar este contrato.', 'danger')
                resumo = None
            if resumo:
                valor_uco_display = ''
                if resumo.valor_uco is not None:
                    try:
                        valor_uco_display = str(resumo.valor_uco.normalize()).replace('.', ',')
                    except Exception:
                        valor_uco_display = str(resumo.valor_uco)
                form_data = {
                    'record_id': resumo.id,
                    'prestador': resumo.prestador,
                    'tabela_honorarios': resumo.tabela_honorarios or '',
                    'tabela_portes': resumo.tabela_portes or '',
                    'valor_uco': valor_uco_display,
                    'inflator_deflator': resumo.inflator_deflator or '',
                    'filme_radiologico': resumo.filme_radiologico or '',
                    'observacoes': resumo.observacoes or '',
                    'operadora_id': resumo.operadora_id,
                }

    # Multi-operadora: Filtrar registros por operadora
    query = ContractSummary.query
    if selected_operadora_id:
        query = query.filter_by(operadora_id=selected_operadora_id)
    registros = query.order_by(ContractSummary.prestador.asc(), ContractSummary.id.asc()).all()

    # Lista de operadoras (filtrada pelo usuário)
    operadoras_list = _get_user_operadoras_list()

    # Lista de tabelas DTP (com IDs para JavaScript)
    dtp_list = [
        {'id': t.id, 'nome': t.nome}
        for t in Tabela.query.filter_by(tipo_tabela='diarias_taxas_pacotes').order_by(Tabela.nome).all()
    ]

    # ========== DTP SEARCH LOGIC ==========
    dtp_results = []
    dtp_search_params = {}
    if request.args.get('dtp_search') == '1':
        dtp_tabela_nome = (request.args.get('dtp_tabela') or '').strip()
        dtp_prestador = (request.args.get('dtp_prestador') or '').strip()
        dtp_codigo = (request.args.get('dtp_codigo') or '').strip()

        if dtp_tabela_nome:
            # Encontrar a tabela DTP pelo nome
            tabela = Tabela.query.filter_by(
                nome=dtp_tabela_nome,
                tipo_tabela='diarias_taxas_pacotes'
            ).first()

            if tabela:
                # Construir query base
                query = Procedimento.query.filter_by(id_tabela=tabela.id)

                # Filtrar por prestador (case-insensitive, partial match)
                if dtp_prestador:
                    query = query.filter(
                        Procedimento.prestador.ilike(f'%{dtp_prestador}%')
                    )

                # Filtrar por código (case-insensitive, partial match)
                if dtp_codigo:
                    query = query.filter(
                        Procedimento.codigo.ilike(f'%{dtp_codigo}%')
                    )

                # Multi-operadora: filtrar por operadora do usuário
                if selected_operadora_id:
                    query = query.filter_by(operadora_id=selected_operadora_id)

                # Ordenar e limitar resultados
                dtp_results = query.order_by(
                    Procedimento.prestador.asc(),
                    Procedimento.codigo.asc()
                ).limit(500).all()

                # Armazenar parâmetros da busca para exibição no template
                dtp_search_params = {
                    'tabela': dtp_tabela_nome,
                    'prestador': dtp_prestador,
                    'codigo': dtp_codigo,
                    'total': len(dtp_results),
                }
            else:
                erro = 'Tabela DTP não encontrada.'
        else:
            erro = 'Selecione uma tabela DTP para buscar.'

    modal_prefill = {
        'record_id': form_data.get('record_id'),
        'prestador': form_data.get('prestador', ''),
        'tabela_honorarios': form_data.get('tabela_honorarios', ''),
        'tabela_portes': form_data.get('tabela_portes', ''),
        'valor_uco': form_data.get('valor_uco', ''),
        'inflator_deflator': form_data.get('inflator_deflator', ''),
        'filme_radiologico': form_data.get('filme_radiologico', ''),
        'observacoes': form_data.get('observacoes', ''),
        'operadora_id': form_data.get('operadora_id', selected_operadora_id),
    }
    modal_open = False
    if request.method == 'POST' and (erro or form_data.get('record_id')):
        modal_open = True
    if request.args.get('edit') and form_data.get('record_id'):
        modal_open = True

    return render_template(
        'contratos_resumo.html',
        registros=registros,
        form=form_data,
        erro=erro,
        modal_prefill=modal_prefill,
        modal_open=modal_open,
        operadoras_list=operadoras_list,
        selected_operadora_id=selected_operadora_id,
        user_has_specific_operadora=(user_operadora_id is not None),
        dtp_list=dtp_list,
        dtp_results=dtp_results,
        dtp_search_params=dtp_search_params,
    )


@app.route('/contratos-resumo/<int:cid>/excluir', methods=['POST'])
@login_required
@feature_required('contratos')
def contratos_resumo_excluir(cid: int):
    resumo = ContractSummary.query.get_or_404(cid)

    # SEGURANÇA: Verificar se usuário tem acesso a este contrato
    user_operadora_id = None
    if hasattr(g, 'current_user') and g.current_user and g.current_user.operadoras:
        user_operadora_id = g.current_user.operadoras[0].id

    if user_operadora_id and resumo.operadora_id != user_operadora_id:
        flash('Você não tem permissão para excluir este contrato.', 'danger')
        return redirect(url_for('contratos_resumo'))

    operadora_id = resumo.operadora_id
    db.session.delete(resumo)
    db.session.commit()
    flash('Resumo removido com sucesso.', 'success')
    return redirect(url_for('contratos_resumo', operadora_id=operadora_id))


@app.route('/reembolsos', methods=['GET', 'POST'])
@login_required
def reembolsos_index():
    error = None
    if request.method == 'POST':
        upload = request.files.get('arquivo')
        tipo_documento = (request.form.get('tipo_documento') or 'AUTO').strip().upper()
        if not upload or not upload.filename:
            error = 'Selecione um arquivo para upload.'
        elif not _reembolso_allowed_extension(upload.filename):
            error = 'Formato nÃ£o suportado. Envie PDF, PNG ou JPG.'
        else:
            max_bytes = REEMBOLSO_MAX_FILE_MB * 1024 * 1024
            try:
                upload.stream.seek(0, io.SEEK_END)
                file_size = upload.stream.tell()
                upload.stream.seek(0)
            except Exception:
                file_size = None
            if file_size is not None and file_size > max_bytes:
                error = f'Arquivo excede {REEMBOLSO_MAX_FILE_MB} MB.'

        if not error:
            ext = Path(upload.filename).suffix.lower()
            stored_filename = f"{uuid4().hex}{ext}"
            subdir = REEMBOLSO_STORAGE_DIR / datetime.utcnow().strftime('%Y') / datetime.utcnow().strftime('%m')
            subdir.mkdir(parents=True, exist_ok=True)
            storage_path = subdir / stored_filename
            try:
                upload.save(storage_path)
            except Exception as exc:  # noqa: BLE001
                error = f'Falha ao salvar o arquivo: {exc}'

        if not error:
            texto, is_pdf_native, ocr_status, ocr_message = _reembolso_extract_text(storage_path)
            inferred_tipo = _reembolso_infer_tipo(texto)
            final_tipo = inferred_tipo if tipo_documento == 'AUTO' else tipo_documento
            dados_extraidos = _reembolso_extract_fields(texto)
            tuss_codes = _reembolso_extract_tuss_codes(texto)
            if tuss_codes:
                operadora_id = session.get('operadora_id')
                if not operadora_id and hasattr(g, 'current_user') and g.current_user and getattr(g.current_user, 'operadoras', None):
                    operadora_id = g.current_user.operadoras[0].id
                tuss_matches = _reembolso_lookup_tuss_values(tuss_codes, operadora_id)
                dados_extraidos['tuss_codigos'] = tuss_codes
                dados_extraidos['tuss_matches'] = tuss_matches
            doc = ReembolsoDocumento(
                usuario_id=(g.current_user.id if hasattr(g, 'current_user') and g.current_user else None),
                tipo_documento=final_tipo,
                original_filename=upload.filename,
                stored_filename=stored_filename,
                storage_path=str(storage_path),
                mime_type=upload.mimetype,
                is_pdf_native=is_pdf_native,
                texto_extraido=texto,
                dados_extraidos=dados_extraidos,
                status='PENDENTE',
                ocr_status=ocr_status,
                ocr_message=ocr_message,
            )
            try:
                db.session.add(doc)
                db.session.commit()
            except Exception as exc:  # noqa: BLE001
                db.session.rollback()
                error = f'Falha ao registrar o reembolso: {exc}'

        if error:
            flash(error, 'danger')
        else:
            flash('Documento enviado. Revise e valide os campos.', 'success')
            return redirect(url_for('reembolsos_review', doc_id=doc.id))

    query = ReembolsoDocumento.query
    if hasattr(g, 'current_user') and g.current_user and g.current_user.perfil != 'adm':
        query = query.filter(ReembolsoDocumento.usuario_id == g.current_user.id)
    documentos = (
        query.order_by(ReembolsoDocumento.created_at.desc())
        .limit(50)
        .all()
    )

    return render_template(
        'reembolsos_index.html',
        documentos=documentos,
        fields_by_type=REEMBOLSO_FIELDS_BY_TYPE,
    )


def _reembolsos_can_access(doc: ReembolsoDocumento) -> bool:
    if not hasattr(g, 'current_user') or not g.current_user:
        return False
    if g.current_user.perfil == 'adm':
        return True
    return doc.usuario_id == g.current_user.id


@app.route('/reembolsos/<int:doc_id>')
@login_required
def reembolsos_review(doc_id: int):
    documento = ReembolsoDocumento.query.get_or_404(doc_id)
    if not _reembolsos_can_access(documento):
        abort(403)
    values = documento.dados_validado or documento.dados_extraidos or {}
    fields = _reembolso_fields_for_type(documento.tipo_documento)
    return render_template(
        'reembolsos_review.html',
        documento=documento,
        fields=fields,
        values=values,
    )


@app.route('/reembolsos/<int:doc_id>/validar', methods=['POST'])
@login_required
def reembolsos_validar(doc_id: int):
    documento = ReembolsoDocumento.query.get_or_404(doc_id)
    if not _reembolsos_can_access(documento):
        abort(403)

    tipo_documento = (request.form.get('tipo_documento') or documento.tipo_documento).strip().upper()
    status = (request.form.get('status') or 'VALIDADO').strip().upper()
    if status not in {'PENDENTE', 'VALIDADO', 'REJEITADO'}:
        status = 'VALIDADO'
    if tipo_documento not in REEMBOLSO_FIELDS_BY_TYPE:
        tipo_documento = documento.tipo_documento or 'DESCONHECIDO'

    values: dict[str, str | None] = {}
    for key, _label in _reembolso_fields_for_type(tipo_documento):
        value = (request.form.get(key) or '').strip()
        values[key] = value or None

    documento.tipo_documento = tipo_documento
    documento.dados_validado = values
    documento.status = status
    try:
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(f'Falha ao salvar: {exc}', 'danger')
        return redirect(url_for('reembolsos_review', doc_id=doc_id))

    flash('Reembolso atualizado com sucesso.', 'success')
    return redirect(url_for('reembolsos_index'))


@app.route('/reembolsos/<int:doc_id>/arquivo')
@login_required
def reembolsos_arquivo(doc_id: int):
    documento = ReembolsoDocumento.query.get_or_404(doc_id)
    if not _reembolsos_can_access(documento):
        abort(403)
    return send_file(documento.storage_path, as_attachment=False, download_name=documento.original_filename)


@app.route('/admin/tetos')
@admin_required
def admin_tetos():
    per_page = 25
    try:
        page = int(request.args.get('page', 1) or 1)
    except (TypeError, ValueError):
        page = 1
    page = max(page, 1)
    search = (request.args.get('q') or '').strip()
    query = CbhpmTeto.query
    if search:
        like = f"%{search}%"
        query = query.filter(or_(CbhpmTeto.codigo.ilike(like), CbhpmTeto.descricao.ilike(like)))
    total = query.count()
    tetos = (
        query.order_by(CbhpmTeto.codigo.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    pages = max((total + per_page - 1) // per_page, 1) if total else 1

    preview_token = (request.args.get('preview_token') or '').strip()
    preview_payload = None
    if preview_token:
        preview_payload = _load_teto_preview(preview_token)
        if not preview_payload:
            flash('Pré-visualização expirada ou inválida. Envie o arquivo novamente.', 'warning')
            return redirect(url_for('admin_tetos'))

    teto_jobs = _snapshot_teto_jobs()

    # Multi-operadora: buscar lista de operadoras ativas (filtrada pelo usuário)
    operadoras_list = _get_user_operadoras_list()
    current_operadora_id = session.get('operadora_id', 1)

    return render_template(
        'admin_tetos.html',
        tetos=tetos,
        total=total,
        page=page,
        per_page=per_page,
        pages=pages,
        q=search,
        preview=preview_payload,
        format_brl=_format_brl,
        teto_jobs=teto_jobs,
        operadoras_list=operadoras_list,
        current_operadora_id=current_operadora_id,
    )


@app.route('/admin/tetos/import', methods=['POST'])
@admin_required
def admin_tetos_import():
    # Multi-operadora: capturar operadora_id do formulário
    operadora_id = request.form.get('operadora_id')
    if operadora_id:
        try:
            operadora_id = int(operadora_id)
        except (TypeError, ValueError):
            operadora_id = 1
    else:
        operadora_id = 1

    confirm_token = (request.form.get('token') or '').strip()
    if confirm_token:
        preview_payload = _load_teto_preview(confirm_token)
        if not preview_payload or not preview_payload.get('rows'):
            _discard_teto_preview(confirm_token)
            flash('Pré-visualização expirada ou vazia. Envie o arquivo novamente.', 'warning')
            return redirect(url_for('admin_tetos'))
        # Passa operadora_id para o job
        preview_payload['operadora_id'] = operadora_id
        job_id = _start_teto_import_job(preview_payload, confirm_token)
        flash(f'Importação agendada (Job {job_id}). Atualize esta página para acompanhar o status.', 'info')
        return redirect(url_for('admin_tetos'))

    upload = request.files.get('arquivo')
    if not upload or not upload.filename:
        flash('Selecione um arquivo CSV ou XLSX para importar.', 'danger')
        return redirect(url_for('admin_tetos'))

    suffix = Path(upload.filename).suffix or '.csv'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        upload.save(tmp)
        temp_path = Path(tmp.name)
    try:
        parsed = _parse_teto_import_file(temp_path)
    finally:
        temp_path.unlink(missing_ok=True)

    if not parsed['rows']:
        if parsed['errors']:
            for message in parsed['errors'][:10]:
                flash(message, 'warning')
        flash('Nenhum registro válido encontrado no arquivo.', 'warning')
        return redirect(url_for('admin_tetos'))

    meta = {
        'filename': upload.filename,
        'total_input': parsed['total_input'],
        'valid_count': parsed['valid_count'],
        'duplicate_count': parsed['duplicate_count'],
        'error_count': len(parsed['errors']),
        'generated_at': datetime.utcnow().isoformat(),
        'operadora_id': operadora_id,  # Multi-operadora: salvar operadora_id no preview
    }
    token = _store_teto_preview({'rows': parsed['rows'], 'meta': meta, 'errors': parsed['errors'], 'operadora_id': operadora_id})
    if parsed['errors']:
        flash(f"Pré-visualização gerada com {parsed['valid_count']} registro(s) válido(s) e {len(parsed['errors'])} aviso(s).", 'warning')
    else:
        flash(f"{parsed['valid_count']} registro(s) prontos para importação.", 'success')
    return redirect(url_for('admin_tetos', preview_token=token))


@app.route('/admin/tetos/template.csv')
@admin_required
def admin_tetos_template_download():
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['codigo', 'descricao', 'valor_total'])
    writer.writerow(['12345', 'Procedimento exemplo', '1234,56'])
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=teto_template.csv'
    return response


@app.route('/admin/tetos/copy', methods=['POST'])
@admin_required
def admin_tetos_copy():
    """Copia todos os tetos de uma operadora para outra"""
    operadora_origem_id = request.form.get('operadora_origem_id')
    operadora_destino_id = request.form.get('operadora_destino_id')

    if not operadora_origem_id or not operadora_destino_id:
        flash('Selecione as operadoras origem e destino.', 'danger')
        return redirect(url_for('admin_tetos'))

    try:
        operadora_origem_id = int(operadora_origem_id)
        operadora_destino_id = int(operadora_destino_id)
    except (TypeError, ValueError):
        flash('IDs de operadoras inválidos.', 'danger')
        return redirect(url_for('admin_tetos'))

    if operadora_origem_id == operadora_destino_id:
        flash('As operadoras origem e destino devem ser diferentes.', 'warning')
        return redirect(url_for('admin_tetos'))

    # Verificar se operadoras existem
    operadora_origem = Operadora.query.get(operadora_origem_id)
    operadora_destino = Operadora.query.get(operadora_destino_id)

    if not operadora_origem or not operadora_destino:
        flash('Operadora não encontrada.', 'danger')
        return redirect(url_for('admin_tetos'))

    # Buscar todos os tetos da operadora origem
    tetos_origem = CbhpmTeto.query.filter_by(operadora_id=operadora_origem_id).all()

    if not tetos_origem:
        flash(f'Nenhum teto encontrado para a operadora {operadora_origem.nome}.', 'warning')
        return redirect(url_for('admin_tetos'))

    # Copiar tetos para operadora destino (upsert)
    copied = 0
    updated = 0

    for teto in tetos_origem:
        # Verificar se já existe teto com esse código na operadora destino
        existing = CbhpmTeto.query.filter_by(
            codigo=teto.codigo,
            operadora_id=operadora_destino_id
        ).first()

        if existing:
            # Atualizar
            existing.descricao = teto.descricao
            existing.valor_total = teto.valor_total
            updated += 1
        else:
            # Inserir novo
            novo_teto = CbhpmTeto(
                codigo=teto.codigo,
                operadora_id=operadora_destino_id,
                descricao=teto.descricao,
                valor_total=teto.valor_total
            )
            db.session.add(novo_teto)
            copied += 1

    db.session.commit()
    flash(f'Cópia concluída: {copied} tetos copiados, {updated} atualizados de {operadora_origem.nome} para {operadora_destino.nome}.', 'success')
    return redirect(url_for('admin_tetos'))


@app.route('/admin/procedimentos/copy', methods=['POST'])
@admin_required
def admin_procedimentos_copy():
    """Copia todos os procedimentos/DTPs de uma operadora para outra"""
    operadora_origem_id = request.form.get('operadora_origem_id')
    operadora_destino_id = request.form.get('operadora_destino_id')
    tabela_nome = request.form.get('tabela_nome')  # Opcional: copiar apenas de uma tabela específica

    if not operadora_origem_id or not operadora_destino_id:
        flash('Selecione as operadoras origem e destino.', 'danger')
        return redirect(request.referrer or url_for('gerenciar_tabelas'))

    try:
        operadora_origem_id = int(operadora_origem_id)
        operadora_destino_id = int(operadora_destino_id)
    except (TypeError, ValueError):
        flash('IDs de operadoras inválidos.', 'danger')
        return redirect(request.referrer or url_for('gerenciar_tabelas'))

    if operadora_origem_id == operadora_destino_id:
        flash('As operadoras origem e destino devem ser diferentes.', 'warning')
        return redirect(request.referrer or url_for('gerenciar_tabelas'))

    # Verificar se operadoras existem
    operadora_origem = Operadora.query.get(operadora_origem_id)
    operadora_destino = Operadora.query.get(operadora_destino_id)

    if not operadora_origem or not operadora_destino:
        flash('Operadora não encontrada.', 'danger')
        return redirect(request.referrer or url_for('gerenciar_tabelas'))

    # Buscar procedimentos da operadora origem
    query = Procedimento.query.filter_by(operadora_id=operadora_origem_id)

    # Se especificou tabela, filtrar por ela
    if tabela_nome:
        tabela_ids = [t.id for t in Tabela.query.filter_by(nome=tabela_nome, id_operadora=operadora_origem_id).all()]
        if tabela_ids:
            query = query.filter(Procedimento.id_tabela.in_(tabela_ids))

    procedimentos_origem = query.all()

    if not procedimentos_origem:
        flash(f'Nenhum procedimento encontrado para a operadora {operadora_origem.nome}.', 'warning')
        return redirect(request.referrer or url_for('gerenciar_tabelas'))

    # Primeiro, copiar/criar tabelas correspondentes na operadora destino
    tabelas_map = {}  # Mapeia id_tabela_origem -> id_tabela_destino
    for proc in procedimentos_origem:
        tabela_origem = proc.tabela
        if tabela_origem.id not in tabelas_map:
            # Verificar se já existe tabela com mesmo nome na operadora destino
            tabela_destino = Tabela.query.filter_by(
                nome=tabela_origem.nome,
                id_operadora=operadora_destino_id
            ).first()

            if not tabela_destino:
                # Criar nova tabela na operadora destino
                tabela_destino = Tabela(
                    nome=tabela_origem.nome,
                    prestador=tabela_origem.prestador,
                    tipo_tabela=tabela_origem.tipo_tabela,
                    uf=tabela_origem.uf,
                    data_vigencia=tabela_origem.data_vigencia,
                    id_operadora=operadora_destino_id
                )
                db.session.add(tabela_destino)
                db.session.flush()  # Obter ID

            tabelas_map[tabela_origem.id] = tabela_destino.id

    # Copiar procedimentos
    copied = 0
    for proc in procedimentos_origem:
        novo_proc = Procedimento(
            codigo=proc.codigo,
            descricao=proc.descricao,
            valor=proc.valor,
            prestador=proc.prestador,
            uf=proc.uf,
            id_tabela=tabelas_map[proc.id_tabela],
            operadora_id=operadora_destino_id
        )
        db.session.add(novo_proc)
        copied += 1

    db.session.commit()

    tabela_msg = f' da tabela {tabela_nome}' if tabela_nome else ''
    flash(f'Cópia concluída: {copied} procedimentos{tabela_msg} copiados de {operadora_origem.nome} para {operadora_destino.nome}.', 'success')
    return redirect(request.referrer or url_for('gerenciar_tabelas'))


@app.route('/admin/tetos/<codigo>/delete', methods=['POST'])
@admin_required
def admin_tetos_delete(codigo: str):
    codigo_norm = (codigo or '').strip().upper()
    if not codigo_norm:
        flash('Código inválido.', 'danger')
        return redirect(url_for('admin_tetos'))

    # Multi-operadora: obter operadora_id do form ou query param
    operadora_id = request.form.get('operadora_id') or request.args.get('operadora_id')
    if operadora_id:
        try:
            operadora_id = int(operadora_id)
        except (TypeError, ValueError):
            operadora_id = None

    # Buscar por PK composta (codigo, operadora_id)
    if operadora_id:
        row = CbhpmTeto.query.filter_by(codigo=codigo_norm, operadora_id=operadora_id).first()
    else:
        # Fallback: busca qualquer registro com esse código (compatibilidade)
        row = CbhpmTeto.query.filter_by(codigo=codigo_norm).first()

    if not row:
        flash('Registro não encontrado.', 'warning')
        return redirect(url_for('admin_tetos'))

    db.session.delete(row)
    db.session.commit()
    flash(f'Teto {codigo_norm} (operadora {row.operadora.nome}) removido com sucesso.', 'success')
    return redirect(url_for('admin_tetos'))


@app.route('/admin/tuss-rol', methods=['GET', 'POST'])
@admin_required
def admin_tuss_rol():
    stats = {
        'total': db.session.query(func.count(TussRolCorrelacao.id)).scalar() or 0,
        'total_consta': db.session.query(func.count(TussRolCorrelacao.id)).filter(TussRolCorrelacao.consta_rol.is_(True)).scalar() or 0,
        'last_updated': db.session.query(func.max(TussRolCorrelacao.atualizado_em)).scalar(),
    }
    erros_import: list[str] = []
    resumo_import: dict[str, int] | None = None

    if request.method == 'POST':
        upload = request.files.get('arquivo')
        if not upload or not upload.filename:
            flash('Selecione um arquivo CSV ou XLSX para importar.', 'danger')
            return redirect(url_for('admin_tuss_rol'))

        raw_bytes = upload.read() or b''
        if not raw_bytes:
            flash('Arquivo vazio. Nenhum dado processado.', 'warning')
            return redirect(url_for('admin_tuss_rol'))

        suffix = (Path(upload.filename).suffix or '').lower()

        def _read_rows_from_upload(data: bytes, ext: str):
            if ext in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
                try:
                    from openpyxl import load_workbook
                except ImportError as exc:
                    raise ValueError('Biblioteca openpyxl não disponível para ler arquivos XLSX.') from exc
                workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
                try:
                    sheet = workbook.active
                    rows = list(sheet.iter_rows(values_only=True))
                finally:
                    workbook.close()
                if not rows:
                    raise ValueError('Arquivo XLSX sem conteúdo.')
                headers = [str(cell or '').strip() for cell in rows[0]]
                data_rows = []
                for row in rows[1:]:
                    values = [str(cell) if cell is not None else '' for cell in row]
                    data_rows.append(values)
                return headers, data_rows

            # fallback CSV / texto
            try:
                text_data = data.decode('utf-8-sig')
            except UnicodeDecodeError:
                text_data = data.decode('latin-1', errors='ignore')

            sio = io.StringIO(text_data)
            try:
                sample = sio.read(2048)
                sio.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=';,')
            except Exception:
                sio.seek(0)
                dialect = csv.excel

            reader = csv.reader(sio, dialect)
            headers = next(reader, [])
            if not headers:
                raise ValueError('Cabeçalho não encontrado no arquivo.')
            data_rows = [[cell for cell in row] for row in reader]
            return headers, data_rows

        try:
            headers, data_rows = _read_rows_from_upload(raw_bytes, suffix)
        except ValueError as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('admin_tuss_rol'))

        header_map = {_norm_header(h): idx for idx, h in enumerate(headers)}
        codigo_idx = header_map.get('codigo')
        desc_idx = header_map.get('descricao')
        consta_idx = header_map.get('seconsta') or header_map.get('consta')
        if codigo_idx is None or desc_idx is None or consta_idx is None:
            flash('As colunas esperadas "CODIGO", "DESCRIÇÃO" e "SE CONSTA" não foram encontradas.', 'danger')
            return redirect(url_for('admin_tuss_rol'))

        registros: dict[str, dict] = {}
        for offset, row in enumerate(data_rows, start=2):
            if not row or all(not (cell or '').strip() for cell in row):
                continue
            try:
                codigo_raw = row[codigo_idx] if codigo_idx < len(row) else ''
                descricao_raw = row[desc_idx] if desc_idx < len(row) else ''
                consta_raw = row[consta_idx] if consta_idx < len(row) else ''
            except IndexError:
                erros_import.append(f'Linha {offset}: formato incorreto.')
                continue

            codigo_norm = _normalize_tuss_codigo(codigo_raw)
            if not codigo_norm:
                erros_import.append(f'Linha {offset}: código vazio.')
                continue
            consta_val = _parse_sim_nao(consta_raw)
            if consta_val is None:
                erros_import.append(f'Linha {offset}: valor inválido em "Se Consta" ({consta_raw!r}).')
                continue
            descricao_val = (descricao_raw or '').strip()
            registros[codigo_norm] = {
                'codigo': codigo_norm,
                'descricao': descricao_val,
                'consta': consta_val,
            }

        if not registros:
            flash('Nenhum registro válido encontrado no arquivo.', 'warning')
            return redirect(url_for('admin_tuss_rol'))

        codigos = list(registros.keys())
        existentes = {
            item.codigo: item
            for item in TussRolCorrelacao.query.filter(TussRolCorrelacao.codigo.in_(codigos)).all()
        }

        criados = 0
        atualizados = 0
        try:
            for codigo, payload in registros.items():
                row = existentes.get(codigo)
                if row:
                    alterou = False
                    if (row.descricao or '') != payload['descricao']:
                        row.descricao = payload['descricao']
                        alterou = True
                    if bool(row.consta_rol) != bool(payload['consta']):
                        row.consta_rol = bool(payload['consta'])
                        alterou = True
                    if alterou:
                        atualizados += 1
                else:
                    db.session.add(TussRolCorrelacao(
                        codigo=payload['codigo'],
                        descricao=payload['descricao'],
                        consta_rol=bool(payload['consta']),
                    ))
                    criados += 1
            db.session.commit()
            resumo_import = {
                'processados': len(registros),
                'criados': criados,
                'atualizados': atualizados,
                'erros': len(erros_import),
            }
            if criados or atualizados:
                flash(f'Importação concluída: {len(registros)} registro(s), {criados} novo(s), {atualizados} atualizado(s).', 'success')
            else:
                flash('Arquivo processado, mas nenhum registro foi alterado.', 'info')
        except Exception as exc:
            db.session.rollback()
            flash(f'Erro ao salvar importação: {exc}', 'danger')
            app.logger.exception('Falha ao importar TUSS/ROL')
            return redirect(url_for('admin_tuss_rol'))

        if erros_import:
            for msg in erros_import[:20]:
                flash(msg, 'warning')

        stats = {
            'total': db.session.query(func.count(TussRolCorrelacao.id)).scalar() or 0,
            'total_consta': db.session.query(func.count(TussRolCorrelacao.id)).filter(TussRolCorrelacao.consta_rol.is_(True)).scalar() or 0,
            'last_updated': db.session.query(func.max(TussRolCorrelacao.atualizado_em)).scalar(),
        }

    exemplos = (
        TussRolCorrelacao.query
        .order_by(TussRolCorrelacao.codigo.asc())
        .limit(50)
        .all()
    )
    return render_template(
        'admin_tuss_rol.html',
        stats=stats,
        exemplos=exemplos,
        resumo_import=resumo_import,
    )


@app.route('/tuss-rol', methods=['GET', 'POST'])
@login_required
@feature_required('tuss_rol')
def tuss_rol_consulta():
    raw_input = ''
    if request.method == 'POST':
        raw_input = request.form.get('codigos') or ''
    else:
        raw_input = request.args.get('codigos') or request.args.get('q') or ''

    separator_pattern = r'[;\n\r,]+'
    requested_codes: list[str] = []
    for part in re.split(separator_pattern, raw_input):
        original = (part or '').strip()
        if not original:
            continue
        normalized = _normalize_tuss_codigo(original)
        if not normalized:
            continue
        # Permite códigos duplicados - não verifica se já existe na lista
        requested_codes.append(normalized)

    results: list[dict] = []
    summary = {
        'total': 0,
        'consta': 0,
        'nao_consta': 0,
        'nao_encontrado': 0,
    }
    rol_map = _fetch_tuss_rol_map(requested_codes) if requested_codes else {}
    for code in requested_codes:
        entry = rol_map.get(code)
        if entry:
            results.append({
                'codigo': code,
                'descricao': entry.get('descricao') or '',
                'consta': bool(entry.get('consta')),
                'status': 'consta' if entry.get('consta') else 'nao_consta',
            })
            summary['total'] += 1
            if entry.get('consta'):
                summary['consta'] += 1
            else:
                summary['nao_consta'] += 1
        else:
            results.append({
                'codigo': code,
                'descricao': '',
                'consta': None,
                'status': 'nao_encontrado',
            })
            summary['total'] += 1
            summary['nao_encontrado'] += 1

    return render_template(
        'tuss-rol-consulta.html',
        entrada=raw_input,
        resultados=results,
        resumo=summary,
    )


@app.route('/cbhpm/regras')
@admin_required
def cbhpm_rules():
    status = request.args.get('status')
    rulesets = (
        CBHPMRuleSet.query
        .order_by(CBHPMRuleSet.ativo.desc(), CBHPMRuleSet.atualizado_em.desc())
        .all()
    )
    return render_template(
        'cbhpm_rules_list.html',
        rulesets=rulesets,
        status=status,
        default_rules=DEFAULT_CBHPM_RULES
    )


@app.route('/cbhpm/regras/nova', methods=['GET', 'POST'])
@admin_required
def cbhpm_rules_new():
    default_rules = _clone_default_cbhpm_rules()
    error = None
    form_data = {
        'nome': '',
        'versao': '',
        'descricao': '',
        'ativo': False,
        'regras': json.dumps(default_rules, indent=2, ensure_ascii=False),
        'regras_dict': default_rules
    }
    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        versao = (request.form.get('versao') or '').strip()
        descricao = (request.form.get('descricao') or '').strip()
        ativo = request.form.get('ativo') == 'on'
        regras_raw = (request.form.get('regras') or '').strip()
        parsed_rules = None
        try:
            parsed_rules = json.loads(regras_raw or '{}')
            if not isinstance(parsed_rules, dict):
                raise ValueError('Estrutura deve ser um objeto JSON.')
        except Exception as exc:
            error = f'JSON invalido: {exc}'
        regras_json = parsed_rules if isinstance(parsed_rules, dict) else {}
        if not nome:
            error = 'Informe um nome para a regra.'
        if not error:
            try:
                if ativo:
                    CBHPMRuleSet.query.filter(CBHPMRuleSet.ativo.is_(True)).update({'ativo': False}, synchronize_session=False)
                ruleset = CBHPMRuleSet(
                    nome=nome,
                    versao=versao or None,
                    descricao=descricao or None,
                    ativo=ativo,
                    regras=regras_json
                )
                db.session.add(ruleset)
                db.session.commit()
                return redirect(url_for('cbhpm_rules', status='created'))
            except Exception as exc:
                db.session.rollback()
                error = f'Erro ao gravar: {exc}'
        form_data.update({
            'nome': nome,
            'versao': versao,
            'descricao': descricao,
            'ativo': ativo,
            'regras': regras_raw or '',
            'regras_dict': regras_json if isinstance(parsed_rules, dict) else form_data.get('regras_dict')
        })
    return render_template('cbhpm_rules_form.html', ruleset=None, form_data=form_data, error=error)


@app.route('/cbhpm/regras/<int:ruleset_id>/editar', methods=['GET', 'POST'])
@admin_required
def cbhpm_rules_edit(ruleset_id: int):
    ruleset = CBHPMRuleSet.query.get_or_404(ruleset_id)
    error = None
    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        versao = (request.form.get('versao') or '').strip()
        descricao = (request.form.get('descricao') or '').strip()
        ativo = request.form.get('ativo') == 'on'
        regras_raw = (request.form.get('regras') or '').strip()
        parsed_rules = None
        try:
            parsed_rules = json.loads(regras_raw or '{}')
            if not isinstance(parsed_rules, dict):
                raise ValueError('Estrutura deve ser um objeto JSON.')
        except Exception as exc:
            error = f'JSON invalido: {exc}'
        regras_json = parsed_rules if isinstance(parsed_rules, dict) else {}
        if not nome:
            error = 'Informe um nome para a regra.'
        if not error:
            try:
                ruleset.nome = nome
                ruleset.versao = versao or None
                ruleset.descricao = descricao or None
                ruleset.regras = regras_json
                if ativo:
                    CBHPMRuleSet.query.filter(
                        CBHPMRuleSet.id != ruleset.id,
                        CBHPMRuleSet.ativo.is_(True)
                    ).update({'ativo': False}, synchronize_session=False)
                ruleset.ativo = ativo
                db.session.commit()
                return redirect(url_for('cbhpm_rules', status='updated'))
            except Exception as exc:
                db.session.rollback()
                error = f'Erro ao gravar: {exc}'
        form_data = {
            'nome': nome,
            'versao': versao,
            'descricao': descricao,
            'ativo': ativo,
            'regras': regras_raw or '',
            'regras_dict': regras_json if isinstance(parsed_rules, dict) else (ruleset.regras if isinstance(ruleset.regras, dict) else {})
        }
        return render_template('cbhpm_rules_form.html', ruleset=ruleset, form_data=form_data, error=error)
    current_rules = ruleset.regras if isinstance(ruleset.regras, dict) else {}
    form_data = {
        'nome': ruleset.nome,
        'versao': ruleset.versao or '',
        'descricao': ruleset.descricao or '',
        'ativo': bool(ruleset.ativo),
        'regras': json.dumps(current_rules, indent=2, ensure_ascii=False),
        'regras_dict': current_rules
    }
    return render_template('cbhpm_rules_form.html', ruleset=ruleset, form_data=form_data, error=error)


@app.route('/cbhpm/regras/<int:ruleset_id>/ativar', methods=['POST'])
@admin_required
def cbhpm_rules_activate(ruleset_id: int):
    try:
        ruleset = CBHPMRuleSet.query.get_or_404(ruleset_id)
        CBHPMRuleSet.query.filter(
            CBHPMRuleSet.id != ruleset.id,
            CBHPMRuleSet.ativo.is_(True)
        ).update({'ativo': False}, synchronize_session=False)
        ruleset.ativo = True
        db.session.commit()
    except Exception:
        db.session.rollback()
    return redirect(url_for('cbhpm_rules', status='activated'))


# --- 4. APIs básicas (CRUD mínimo) ---

# Operadoras
@app.route('/api/operadoras', methods=['GET', 'POST'])
@admin_required
def api_operadoras():
    if request.method == 'GET':
        data = [
            {"id": o.id, "nome": o.nome, "cnpj": o.cnpj, "status": o.status}
            for o in Operadora.query.all()
        ]
        return jsonify(data)
    payload = request.json or {}
    o = Operadora(nome=payload.get('nome'), cnpj=payload.get('cnpj'), status=payload.get('status', 'Ativa'))
    db.session.add(o)
    db.session.commit()
    return jsonify({"id": o.id}), 201


@app.route('/api/operadoras/<int:oid>', methods=['PUT', 'DELETE'])
@admin_required
def api_operadora_item(oid):
    o = Operadora.query.get_or_404(oid)
    if request.method == 'PUT':
        payload = request.json or {}
        o.nome = payload.get('nome', o.nome)
        o.cnpj = payload.get('cnpj', o.cnpj)
        o.status = payload.get('status', o.status)
        db.session.commit()
        return jsonify({"ok": True})
    db.session.delete(o)
    db.session.commit()
    return jsonify({"ok": True})


# --- 5. Inicialização ---
def ensure_db(max_retries: int = 20, delay_seconds: int = 3):
    """Cria as tabelas com tentativas/retry para aguardar o MySQL.
    Útil quando o container web inicia antes do banco estar pronto.
    """
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            with app.app_context():
                db.create_all()
                try:
                    ContractSummary.__table__.create(bind=db.engine, checkfirst=True)
                except Exception:
                    db.session.rollback()
                # Tentativa de migração leve para acrescentar colunas caso já exista a tabela
                try:
                    db.session.execute(text("ALTER TABLE tabelas ADD COLUMN prestador VARCHAR(255) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE tabelas ADD COLUMN tipo_tabela VARCHAR(50) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE tabelas ADD COLUMN uf VARCHAR(2) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE tabelas ADD COLUMN uco_valor DECIMAL(12,2) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE bras_item ADD COLUMN tipo_preco VARCHAR(50) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE bras_item ADD COLUMN ean VARCHAR(64) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                # Migração leve: acrescentar coluna UF em operadoras
                try:
                    db.session.execute(text("ALTER TABLE operadoras ADD COLUMN uf VARCHAR(2) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                # Migração leve: acrescentar colunas em procedimentos
                try:
                    db.session.execute(text("ALTER TABLE procedimentos ADD COLUMN prestador VARCHAR(255) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE procedimentos ADD COLUMN uf VARCHAR(2) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN id_operadora INT NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE mv_catalogo_vigente_simpro ADD COLUMN codigo_interno VARCHAR(20) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE mv_catalogo_vigente_simpro ADD COLUMN tuss_numero VARCHAR(16) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE mv_catalogo_vigente_simpro ADD COLUMN referencia VARCHAR(120) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE mv_catalogo_vigente_simpro ADD COLUMN status_final VARCHAR(8) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE mv_catalogo_vigente_simpro ADD COLUMN fracionavel VARCHAR(1) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE simpro_item_norm ADD COLUMN referencia VARCHAR(120) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE simpro_item_norm ADD COLUMN fracionavel VARCHAR(1) NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("CREATE INDEX idx_simpro_tuss_numero ON mv_catalogo_vigente_simpro (tuss_numero)"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    _backfill_catalogo_simpro_identifiers()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text(
                        "ALTER TABLE usuarios ADD CONSTRAINT fk_usuarios_operadora FOREIGN KEY (id_operadora) REFERENCES operadoras(id)"
                    ))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN acesso_insumos TINYINT(1) NOT NULL DEFAULT 1"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN acesso_consulta TINYINT(1) NOT NULL DEFAULT 1"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN acesso_contratos TINYINT(1) NOT NULL DEFAULT 1"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN acesso_tuss_rol TINYINT(1) NOT NULL DEFAULT 1"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN must_reset_senha TINYINT(1) NOT NULL DEFAULT 1"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN senha_atualizada_em DATETIME NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN failed_login_attempts INT NOT NULL DEFAULT 0"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN locked_until DATETIME NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("ALTER TABLE usuarios ADD COLUMN last_logout_at DATETIME NULL"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text("UPDATE usuarios SET acesso_insumos = COALESCE(acesso_insumos, 1), acesso_consulta = COALESCE(acesso_consulta, 1), acesso_contratos = COALESCE(acesso_contratos, 1), acesso_tuss_rol = COALESCE(acesso_tuss_rol, 1), must_reset_senha = COALESCE(must_reset_senha, 0), senha_atualizada_em = COALESCE(senha_atualizada_em, CURRENT_TIMESTAMP)"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text(
                        """
                        CREATE TABLE IF NOT EXISTS usuario_operadoras (
                            usuario_id INT NOT NULL,
                            operadora_id INT NOT NULL,
                            PRIMARY KEY (usuario_id, operadora_id),
                            CONSTRAINT fk_usuario_operadoras_usuario FOREIGN KEY (usuario_id) REFERENCES usuarios(id) ON DELETE CASCADE,
                            CONSTRAINT fk_usuario_operadoras_operadora FOREIGN KEY (operadora_id) REFERENCES operadoras(id) ON DELETE CASCADE
                        )
                        """
                    ))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                try:
                    db.session.execute(text(
                        """
                        INSERT INTO usuario_operadoras (usuario_id, operadora_id)
                        SELECT u.id, u.id_operadora
                        FROM usuarios u
                        LEFT JOIN usuario_operadoras rel
                          ON rel.usuario_id = u.id AND rel.operadora_id = u.id_operadora
                        WHERE u.id_operadora IS NOT NULL AND rel.usuario_id IS NULL
                        """
                    ))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                # Garante criação da tabela CBHPM (se ainda não existir)
                db.create_all()
                try:
                    usuarios = Usuario.query.all()
                    changed = False
                    for usuario in usuarios:
                        if usuario and usuario.senha and not _is_password_hashed(usuario.senha):
                            hashed = _hash_password(usuario.senha)
                            usuario.senha = hashed
                            usuario.senha_atualizada_em = usuario.senha_atualizada_em or _now_utc()
                            db.session.add(usuario)
                            changed = True
                            try:
                                _append_password_history(usuario, hashed)
                            except Exception:
                                pass
                    if changed:
                        db.session.commit()
                except Exception:
                    db.session.rollback()
                # Semeia um usuário admin padrão se não existir nenhum usuário
                try:
                    if db.session.query(Usuario).count() == 0:
                        admin_email = os.getenv('ADMIN_EMAIL', 'admin@local')
                        admin_senha = os.getenv('ADMIN_PASSWORD', 'admin123')
                        admin_nome = os.getenv('ADMIN_NAME', 'Administrador')
                        senha_hash = _hash_password(admin_senha)
                        admin = Usuario(nome=admin_nome, email=admin_email, senha=senha_hash, perfil='adm')
                        admin.must_reset_senha = True
                        admin.senha_atualizada_em = _now_utc()
                        db.session.add(admin)
                        db.session.flush()
                        try:
                            _append_password_history(admin, senha_hash)
                        except Exception:
                            pass
                        db.session.commit()
                        print(f"[init] Usuário admin criado: {admin_email} / senha padrão")
                except Exception:
                    db.session.rollback()

                print(f"[init] Banco pronto após {attempt} tentativa(s).")
                try:
                    if db.session.query(CBHPMRuleSet).count() == 0:
                        regras_default = json.loads(json.dumps(DEFAULT_CBHPM_RULES))
                        ruleset = CBHPMRuleSet(
                            nome='CBHPM Padrão',
                            versao='Base',
                            descricao='Criada automaticamente',
                            ativo=True,
                            regras=regras_default
                        )
                        db.session.add(ruleset)
                        db.session.commit()
                except Exception:
                    db.session.rollback()
                return
        except Exception as e:
            last_err = e
            print(f"[init] MySQL indisponível (tentativa {attempt}/{max_retries}). Aguardando {delay_seconds}s...")
            time.sleep(delay_seconds)
    # Se esgotar
    raise last_err


if os.getenv('SKIP_ENSURE_DB', '').strip().lower() not in {'1', 'true', 'yes', 'on'}:
    ensure_db()


# --- 6. Usuários (UI) ---
@app.route('/usuarios/novo', methods=['GET', 'POST'])
@admin_required
def usuario_novo():
    operadoras = Operadora.query.order_by(Operadora.nome).all()
    if request.method == 'POST':
        nome = request.form.get('nome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        perfil = request.form.get('perfil')
        if not all([nome, email, senha, perfil]):
            return render_template('usuario-form.html', erro='Preencha todos os campos', modo='novo', form=request.form, operadoras=operadoras)
        politica = _password_policy_error(senha or '')
        if politica:
            return render_template('usuario-form.html', erro=politica, modo='novo', form=request.form, operadoras=operadoras)
        raw_operadoras = [s.strip() for s in request.form.getlist('operadora_ids') if s and s.strip()]
        operadora_ids: list[int] = []
        for raw in raw_operadoras:
            try:
                oid = int(raw)
            except ValueError:
                return render_template(
                    'usuario-form.html',
                    erro='Operadora selecionada é inválida.',
                    modo='novo',
                    form=request.form,
                    operadoras=operadoras,
                )
            if oid not in operadora_ids:
                operadora_ids.append(oid)
        operadoras_sel: list[Operadora] = []
        if operadora_ids:
            operadoras_sel = Operadora.query.filter(Operadora.id.in_(operadora_ids)).all()
            if len(operadoras_sel) != len(operadora_ids):
                return render_template(
                    'usuario-form.html',
                    erro='Operadora selecionada é inválida.',
                    modo='novo',
                    form=request.form,
                    operadoras=operadoras,
                )
            op_map = {op.id: op for op in operadoras_sel}
            operadoras_sel = [op_map[oid] for oid in operadora_ids if oid in op_map]
        if perfil in ('operadora', 'adm de contrato') and not operadoras_sel:
            return render_template(
                'usuario-form.html',
                erro='Selecione ao menos uma operadora para usuários dos perfis Operadora ou Adm de Contrato.',
                modo='novo',
                form=request.form,
                    operadoras=operadoras,
                )
        if Usuario.query.filter_by(email=email).first():
            return render_template('usuario-form.html', erro='E-mail já cadastrado', modo='novo', form=request.form, operadoras=operadoras)
        acesso_insumos = bool(request.form.get('acesso_insumos'))
        acesso_consulta = bool(request.form.get('acesso_consulta'))
        acesso_contratos = bool(request.form.get('acesso_contratos'))
        acesso_tuss_rol = bool(request.form.get('acesso_tuss_rol'))
        agora = _now_utc()
        senha_hash = _hash_password(senha)
        u = Usuario(
            nome=nome,
            email=email,
            senha=senha_hash,
            perfil=perfil,
            acesso_insumos=acesso_insumos,
            acesso_consulta=acesso_consulta,
            acesso_contratos=acesso_contratos,
            acesso_tuss_rol=acesso_tuss_rol,
        )
        u.must_reset_senha = True
        u.senha_atualizada_em = agora
        u.failed_login_attempts = 0
        u.locked_until = None
        u.operadoras = operadoras_sel
        db.session.add(u)
        try:
            db.session.flush()
            _append_password_history(u, senha_hash)
            _register_audit('user.create', usuario=u, detalhes={'actor_usuario_id': session.get('user_id')})
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            app.logger.error('Erro ao criar usuário: %s', exc)
            return render_template('usuario-form.html', erro='Erro ao criar usuário. Tente novamente.', modo='novo', form=request.form, operadoras=operadoras)
        return redirect(url_for('gerenciar_usuarios'))
    return render_template('usuario-form.html', modo='novo', operadoras=operadoras)


@app.route('/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
@admin_required
def usuario_editar(uid):
    u = Usuario.query.get_or_404(uid)
    operadoras = Operadora.query.order_by(Operadora.nome).all()
    if request.method == 'POST':
        nome = request.form.get('nome') or u.nome
        email = request.form.get('email') or u.email
        perfil_novo = request.form.get('perfil') or u.perfil
        original_perfil = u.perfil
        original_insumos = u.acesso_insumos
        original_consulta = u.acesso_consulta
        original_contratos = getattr(u, 'acesso_contratos', True)
        original_tuss = u.acesso_tuss_rol
        original_operadoras = sorted(op.id for op in u.operadoras)
        raw_operadoras = [s.strip() for s in request.form.getlist('operadora_ids') if s and s.strip()]
        operadora_ids: list[int] = []
        for raw in raw_operadoras:
            try:
                oid = int(raw)
            except ValueError:
                return render_template(
                    'usuario-form.html',
                    erro='Operadora selecionada é inválida.',
                    modo='editar',
                    usuario=u,
                    form=request.form,
                    operadoras=operadoras,
                )
            if oid not in operadora_ids:
                operadora_ids.append(oid)
        operadoras_sel: list[Operadora] = []
        if operadora_ids:
            operadoras_sel = Operadora.query.filter(Operadora.id.in_(operadora_ids)).all()
            if len(operadoras_sel) != len(operadora_ids):
                return render_template(
                    'usuario-form.html',
                    erro='Operadora selecionada é inválida.',
                    modo='editar',
                    usuario=u,
                    form=request.form,
                    operadoras=operadoras,
                )
            op_map = {op.id: op for op in operadoras_sel}
            operadoras_sel = [op_map[oid] for oid in operadora_ids if oid in op_map]
        if perfil_novo in ('operadora', 'adm de contrato') and not operadoras_sel:
            return render_template(
                'usuario-form.html',
                erro='Selecione ao menos uma operadora para usuários dos perfis Operadora ou Adm de Contrato.',
                modo='editar',
                usuario=u,
                form=request.form,
                operadoras=operadoras,
            )
        acesso_insumos = bool(request.form.get('acesso_insumos'))
        acesso_consulta = bool(request.form.get('acesso_consulta'))
        acesso_contratos = bool(request.form.get('acesso_contratos'))
        acesso_tuss_rol = bool(request.form.get('acesso_tuss_rol'))
        new_senha = (request.form.get('senha') or '').strip()
        password_changed = False
        agora = _now_utc()
        if new_senha:
            politica = _password_policy_error(new_senha)
            if politica:
                return render_template(
                    'usuario-form.html',
                    erro=politica,
                    modo='editar',
                    usuario=u,
                    form=request.form,
                    operadoras=operadoras,
                )
            if _password_was_used_recently(u, new_senha):
                return render_template(
                    'usuario-form.html',
                    erro=f'Não reutilize as últimas {PASSWORD_HISTORY_SIZE} senhas.',
                    modo='editar',
                    usuario=u,
                    form=request.form,
                    operadoras=operadoras,
                )
            senha_hash = _hash_password(new_senha)
            u.senha = senha_hash
            u.senha_atualizada_em = agora
            u.failed_login_attempts = 0
            u.locked_until = None
            u.must_reset_senha = not (session.get('user_id') == u.id)
            _append_password_history(u, senha_hash)
            password_changed = True

        u.nome = nome
        u.email = email
        u.perfil = perfil_novo
        u.operadoras = operadoras_sel
        u.acesso_insumos = acesso_insumos
        u.acesso_consulta = acesso_consulta
        u.acesso_contratos = acesso_contratos
        u.acesso_tuss_rol = acesso_tuss_rol

        try:
            if password_changed:
                _register_audit(
                    'password.reset',
                    usuario=u,
                    detalhes={
                        'actor_usuario_id': session.get('user_id'),
                        'self_service': session.get('user_id') == u.id,
                    },
                )
            permissions_changed = (
                original_perfil != u.perfil
                or original_insumos != acesso_insumos
                or original_consulta != acesso_consulta
                or original_contratos != acesso_contratos
                or original_tuss != acesso_tuss_rol
                or original_operadoras != sorted(op.id for op in u.operadoras)
            )
            if permissions_changed:
                _register_audit(
                    'user.permissions_change',
                    usuario=u,
                    detalhes={
                        'actor_usuario_id': session.get('user_id'),
                        'perfil_antes': original_perfil,
                        'perfil_depois': u.perfil,
                        'acesso_insumos_antes': original_insumos,
                        'acesso_insumos_depois': acesso_insumos,
                        'acesso_consulta_antes': original_consulta,
                        'acesso_consulta_depois': acesso_consulta,
                        'acesso_contratos_antes': original_contratos,
                        'acesso_contratos_depois': acesso_contratos,
                        'acesso_tuss_antes': original_tuss,
                        'acesso_tuss_depois': acesso_tuss_rol,
                        'operadoras_antes': original_operadoras,
                        'operadoras_depois': sorted(op.id for op in u.operadoras),
                    },
                )
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            app.logger.error('Erro ao atualizar usuário %s: %s', u.email, exc)
            return render_template(
                'usuario-form.html',
                erro='Erro ao atualizar usuário. Tente novamente.',
                modo='editar',
                usuario=u,
                form=request.form,
                operadoras=operadoras,
            )

        if session.get('user_id') == u.id:
            if password_changed:
                session['login_time'] = agora.isoformat()
                session['password_changed_at'] = agora.isoformat()
            nomes = [op.nome for op in u.operadoras]
            ids = [op.id for op in u.operadoras]
            session['operadora_ids'] = ids
            session['operadora_id'] = ids[0] if ids else None
            session['operadora_nomes'] = nomes
            session['operadora_nome'] = ', '.join(nomes) if nomes else None
            session['feature_insumos'] = acesso_insumos or (session.get('perfil') == 'adm')
            session['feature_consulta'] = acesso_consulta or (session.get('perfil') == 'adm')
            session['feature_contratos'] = acesso_contratos or (session.get('perfil') in {'adm', 'adm de contrato', 'operadora'})
            session['feature_tuss_rol'] = acesso_tuss_rol or (session.get('perfil') == 'adm')
            session['must_change_senha'] = bool(u.must_reset_senha)
        return redirect(url_for('gerenciar_usuarios'))
    return render_template('usuario-form.html', modo='editar', usuario=u, operadoras=operadoras)


@app.route('/admin/audit-trail')
@admin_required
def admin_audit_trail():
    page = max(request.args.get('page', default=1, type=int) or 1, 1)
    per_page = request.args.get('per_page', default=50, type=int) or 50
    per_page = max(10, min(per_page, 200))
    evento = (request.args.get('evento') or '').strip() or None
    email_q = (request.args.get('email') or '').strip()
    ip_q = (request.args.get('ip') or '').strip()
    inicio_str = (request.args.get('inicio') or '').strip()
    fim_str = (request.args.get('fim') or '').strip()

    inicio_dt = None
    fim_dt = None
    if inicio_str:
        try:
            inicio_dt = datetime.strptime(inicio_str, '%Y-%m-%d')
        except ValueError:
            flash('Data inicial inválida. Use o formato AAAA-MM-DD.', 'warning')
            inicio_dt = None
    if fim_str:
        try:
            fim_dt = datetime.strptime(fim_str, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            flash('Data final inválida. Use o formato AAAA-MM-DD.', 'warning')
            fim_dt = None

    query = (
        AuditLog.query
        .outerjoin(Usuario, AuditLog.usuario_id == Usuario.id)
        .options(joinedload(AuditLog.usuario))
        .order_by(AuditLog.id.desc())
    )

    if evento:
        query = query.filter(AuditLog.evento == evento)
    if email_q:
        like_value = f"%{email_q.lower()}%"
        query = query.filter(
            or_(
                func.lower(AuditLog.email_alvo).like(like_value),
                func.lower(Usuario.email).like(like_value),
            )
        )
    if ip_q:
        query = query.filter(AuditLog.ip.ilike(f"%{ip_q}%"))
    if inicio_dt:
        query = query.filter(AuditLog.criado_em >= inicio_dt)
    if fim_dt:
        query = query.filter(AuditLog.criado_em < fim_dt)

    total = query.count()
    pages = max(1, math.ceil(total / per_page)) if total else 1
    if page > pages:
        page = pages
    offset = (page - 1) * per_page
    logs = query.offset(offset).limit(per_page).all()

    parsed_rows: list[tuple[AuditLog, dict | list | None, str | None, Optional[int]]] = []
    actor_ids: set[int] = set()
    for row in logs:
        parsed_candidate = None
        if row.detalhes:
            try:
                parsed_candidate = json.loads(row.detalhes)
            except Exception:
                parsed_candidate = None
        parsed_value: dict | list | None = parsed_candidate if isinstance(parsed_candidate, (dict, list)) else None
        raw_value = None if parsed_value is not None else row.detalhes
        actor_ref: Optional[int] = None
        if isinstance(parsed_value, dict):
            candidate = parsed_value.get('actor_usuario_id')
            if candidate is None:
                candidate = parsed_value.get('actor_user_id')
            if isinstance(candidate, int):
                actor_ref = candidate
                actor_ids.add(candidate)
            elif isinstance(candidate, str):
                try:
                    parsed_int = int(candidate)
                    actor_ref = parsed_int
                    actor_ids.add(parsed_int)
                except ValueError:
                    pass
        parsed_rows.append((row, parsed_value, raw_value, actor_ref))

    actor_map: dict[int, Usuario] = {}
    if actor_ids:
        try:
            actor_map = {
                usuario.id: usuario
                for usuario in Usuario.query.filter(Usuario.id.in_(actor_ids)).all()
            }
        except Exception:
            actor_map = {}

    entries = []
    for row, parsed_value, raw_value, actor_ref in parsed_rows:
        details_lines: list[str] = []
        if isinstance(parsed_value, dict):
            actor = actor_map.get(actor_ref) if actor_ref else None
            if actor_ref:
                if actor:
                    details_lines.append(f"Ação executada por {actor.nome} ({actor.email})")
                else:
                    details_lines.append(f"Ação executada pelo usuário id {actor_ref}")
            self_service = parsed_value.get('self_service')
            if isinstance(self_service, bool):
                details_lines.append('Autoatendimento pelo próprio usuário.' if self_service else 'Executado por outro usuário (administrativo).')
            known_keys = {'actor_usuario_id', 'actor_user_id', 'self_service'}
            for key in sorted(parsed_value.keys()):
                if key in known_keys:
                    continue
                value = parsed_value[key]
                if isinstance(value, (dict, list)):
                    value_repr = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    value_repr = '—'
                else:
                    value_repr = str(value)
                details_lines.append(f"{key}: {value_repr}")
        elif isinstance(parsed_value, list):
            if parsed_value:
                details_lines.append(json.dumps(parsed_value, ensure_ascii=False))

        entries.append({
            'record': row,
            'parsed': parsed_value,
            'raw': raw_value,
            'details_lines': details_lines,
            'usuario_nome': row.usuario.nome if row.usuario else None,
            'usuario_email': row.usuario.email if row.usuario else None,
        })

    eventos_disponiveis = [
        item[0] for item in db.session.query(AuditLog.evento).distinct().order_by(AuditLog.evento).all()
    ]

    return render_template(
        'audit-logs.html',
        logs=entries,
        eventos=eventos_disponiveis,
        page=page,
        pages=pages,
        total=total,
        per_page=per_page,
        filters={
            'evento': evento or '',
            'email': email_q,
            'ip': ip_q,
            'inicio': inicio_str,
            'fim': fim_str,
        },
    )


# --- 7. Operadoras (UI) ---
BR_UFS = [
    'AC','AL','AP','AM','BA','CE','DF','ES','GO','MA','MT','MS','MG','PA','PB','PR','PE','PI','RJ','RN','RS','RO','RR','SC','SP','SE','TO'
]


@app.route('/operadoras/nova', methods=['GET', 'POST'])
@admin_required
def operadora_nova():
    if request.method == 'POST':
        nome = request.form.get('nome')
        uf = request.form.get('uf')
        cnpj = request.form.get('cnpj')
        if not nome or not uf:
            return render_template('operadora-form.html', erro='Nome e UF são obrigatórios', modo='nova', form=request.form, UFS=BR_UFS)
        o = Operadora(nome=nome, uf=uf, cnpj=cnpj, status='Ativa')
        db.session.add(o)
        db.session.commit()
        return redirect(url_for('gerenciar_operadoras'))
    return render_template('operadora-form.html', modo='nova', UFS=BR_UFS)


@app.route('/operadoras/<int:oid>/editar', methods=['GET', 'POST'])
@admin_required
def operadora_editar(oid):
    o = Operadora.query.get_or_404(oid)
    if request.method == 'POST':
        o.nome = request.form.get('nome') or o.nome
        o.uf = request.form.get('uf') or o.uf
        o.cnpj = request.form.get('cnpj') or o.cnpj
        db.session.commit()
        return redirect(url_for('gerenciar_operadoras'))
    return render_template('operadora-form.html', modo='editar', operadora=o, UFS=BR_UFS)


# --- 8. Importação de Tabelas ---
def _norm_header(s: str) -> str:
    s = (s or '').strip()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.lower()
    s = s.replace(' ', '').replace('-', '').replace('_', '')
    return s


def _parse_money(v) -> Decimal:
    if v is None:
        return Decimal('0')
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        return Decimal('0')
    s = s.replace('R$', '').replace(' ', '')
    s = s.replace('.', '')  # milhar
    s = s.replace(',', '.')  # decimal
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _as_decimal(v):
    try:
        if v is None:
            return None
        if isinstance(v, (int, float, Decimal)):
            return Decimal(str(v))
        s = str(v).strip()
        if s == '' or s == '-':
            return None
        # If it uses comma as decimal separator or has currency, treat as money string
        if (',' in s) or ('R$' in s):
            return _parse_money(s)
        # Otherwise, assume dot-decimal and parse directly (to avoid stripping the dot)
        try:
            return Decimal(s)
        except InvalidOperation:
            return _parse_money(s)
    except Exception:
        setattr(load_logo_bytes, 'last_static_name', None)
        return None


def _sum_decimals(values):
    total = Decimal('0')
    found = False
    for v in values:
        dv = _as_decimal(v)
        if dv is not None:
            total += dv
            found = True
    return (total if found else None)


def _stringify_for_output(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, list):
        return [_stringify_for_output(v) for v in value]
    if isinstance(value, dict):
        return {k: _stringify_for_output(v) for k, v in value.items()}
    return value


def _normalize_tuss_codigo(raw: str | None) -> str | None:
    if raw is None:
        return None
    code = str(raw).strip()
    if not code:
        return None
    return code.upper()


def _fetch_tuss_rol_map(codigos: Sequence[str] | None) -> dict[str, dict]:
    codigos = codigos or []
    normalized: dict[str, str] = {}
    for c in codigos:
        norm = _normalize_tuss_codigo(c)
        if not norm:
            continue
        normalized.setdefault(norm, c)
    if not normalized:
        return {}
    records = (
        TussRolCorrelacao.query
        .filter(TussRolCorrelacao.codigo.in_(normalized.keys()))
        .all()
    )
    result: dict[str, dict] = {}
    for item in records:
        key = normalized.get(item.codigo, item.codigo)
        result[key] = {
            'codigo': item.codigo,
            'descricao': item.descricao,
            'consta': bool(item.consta_rol),
        }
    return result


def _parse_sim_nao(raw_value: str | None) -> bool | None:
    if raw_value is None:
        return None
    value = str(raw_value).strip()
    if not value:
        return None
    value_norm = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode().upper()
    if value_norm in {'SIM', 'S', '1', 'TRUE', 'T', 'YES', 'Y'}:
        return True
    if value_norm in {'NAO', 'N', '0', 'FALSE', 'F', 'NO'}:
        return False
    return None


def _aliquota_bp_to_decimal(raw) -> Decimal | None:
    if raw in (None, ''):
        return None
    try:
        return Decimal(str(raw)) / Decimal('100')
    except (InvalidOperation, ValueError, TypeError):
        return None


def _collect_catalogo_filters_bras(item_id: int, item) -> list[list]:
    filters: list[list] = []
    seen: set[tuple] = set()

    def _add(name: str, *values, extra: list | None = None) -> None:
        key = (name,) + tuple(values)
        if key in seen:
            return
        seen.add(key)
        exprs: list = []
        if name == 'item_id':
            exprs.append(CatalogoBrasindice.item_id == values[0])
        elif name == 'produto_apresentacao':
            exprs.append(CatalogoBrasindice.produto_codigo == values[0])
            exprs.append(CatalogoBrasindice.apresentacao_codigo == values[1])
        elif name == 'produto':
            exprs.append(CatalogoBrasindice.produto_codigo == values[0])
        elif name == 'ean':
            exprs.append(CatalogoBrasindice.ean == values[0])
        elif name == 'anvisa':
            exprs.append(CatalogoBrasindice.registro_anvisa == values[0])
        elif extra:
            exprs.extend(extra)
        if extra and name not in {'produto_apresentacao', 'produto', 'ean', 'anvisa', 'item_id'}:
            exprs.extend(extra)
        if exprs:
            filters.append(exprs)

    _add('item_id', item_id)

    produto_codigo = getattr(item, 'produto_codigo', None)
    apresentacao_codigo = getattr(item, 'apresentacao_codigo', None)
    if produto_codigo and apresentacao_codigo:
        _add('produto_apresentacao', produto_codigo, apresentacao_codigo)
    if produto_codigo:
        _add('produto', produto_codigo)

    ean = getattr(item, 'ean', None)
    if ean:
        _add('ean', ean)

    registro_anvisa = getattr(item, 'registro_anvisa', None)
    if registro_anvisa:
        _add('anvisa', registro_anvisa)

    return filters


def _collect_catalogo_filters_simpro(item_id: int, item) -> list[list]:
    filters: list[list] = []
    seen: set[tuple] = set()

    def _add(name: str, *values) -> None:
        key = (name,) + tuple(values)
        if key in seen:
            return
        seen.add(key)
        exprs: list = []
        if name == 'item_id':
            exprs.append(CatalogoSimpro.item_id == values[0])
        elif name == 'codigo':
            exprs.append(CatalogoSimpro.codigo == values[0])
        elif name == 'codigo_alt':
            exprs.append(CatalogoSimpro.codigo_alt == values[0])
        elif name == 'ean':
            exprs.append(CatalogoSimpro.ean == values[0])
        elif name == 'anvisa':
            exprs.append(CatalogoSimpro.anvisa == values[0])
        elif name == 'tuss':
            exprs.append(CatalogoSimpro.tuss_numero == values[0])
        elif name == 'tiss':
            exprs.append(CatalogoSimpro.codigo_alt == values[0])
        if exprs:
            filters.append(exprs)

    _add('item_id', item_id)

    codigo = getattr(item, 'codigo', None)
    if codigo:
        _add('codigo', codigo)

    codigo_alt = getattr(item, 'codigo_alt', None)
    if codigo_alt:
        _add('codigo_alt', codigo_alt)

    ean = getattr(item, 'ean', None)
    if ean:
        _add('ean', ean)

    anvisa = getattr(item, 'anvisa', None)
    if anvisa:
        _add('anvisa', anvisa)

    tuss = getattr(item, 'tuss_numero', None) or getattr(item, 'tuss', None)
    if tuss:
        _add('tuss', tuss)

    tiss = getattr(item, 'tiss', None)
    if tiss:
        _add('tiss', tiss)

    return filters


def _resolve_catalogo_history(
    origem: str,
    *,
    item_id: int,
    item,
    uf_param: str | None,
) -> tuple[CatalogoBrasindice | CatalogoSimpro | None, list]:
    uf_param = (uf_param or '').upper() or None

    if origem == 'BRAS':
        filter_sets = _collect_catalogo_filters_bras(item_id, item)
        order_columns = (CatalogoBrasindice.periodo.desc(), CatalogoBrasindice.uf.asc())
        model_query_factory = lambda: CatalogoBrasindice.query
    else:
        filter_sets = _collect_catalogo_filters_simpro(item_id, item)
        order_columns = (CatalogoSimpro.periodo.desc(), CatalogoSimpro.uf.asc())
        model_query_factory = lambda: CatalogoSimpro.query

    for filter_exprs in filter_sets:
        query = model_query_factory()
        for expr in filter_exprs:
            query = query.filter(expr)
        rows = query.order_by(*order_columns).limit(50).all()
        if not rows:
            continue
        entry = None
        if uf_param:
            entry = next((row for row in rows if (row.uf or '').upper() == uf_param), None)
        if entry is None:
            entry = rows[0]
        return entry, rows

    return None, []


def _split_substitutos(raw) -> list[str]:
    if not raw:
        return []
    if isinstance(raw, (list, tuple)):
        return [str(item).strip() for item in raw if str(item).strip()]
    text = str(raw)
    parts = re.split(r'[;,\n]+', text)
    return [part.strip() for part in parts if part.strip()]


def _serialize_contexto_clinico(
    entry: InsumoContextoClinico,
    detail_payload: dict,
) -> dict:
    base_price = _as_decimal(detail_payload.get('preco_pfb'))
    if base_price is None:
        base_price = _as_decimal(detail_payload.get('preco_pmc'))
    if base_price is None:
        base_price = _as_decimal(detail_payload.get('preco'))

    custo_procedimento = _as_decimal(entry.custo_procedimento)
    frequencia = _as_decimal(entry.frequencia_relativa)

    variacoes = []
    if base_price is not None:
        for pct_raw in (Decimal('0.05'), Decimal('0.10'), Decimal('0.20')):
            delta_unit = (base_price * pct_raw) if base_price is not None else None
            delta_proc = None
            if custo_procedimento is not None:
                delta_proc = custo_procedimento * pct_raw
            variacoes.append({
                'percentual': _decimal_to_string(pct_raw * Decimal('100'), precision=1),
                'delta_unitario': _decimal_to_string(delta_unit),
                'delta_procedimento': _decimal_to_string(delta_proc) if delta_proc is not None else None,
            })

    substitutos = _split_substitutos(entry.substitutos_raw)

    return {
        'drg': entry.drg,
        'procedimento_codigo': entry.procedimento_codigo,
        'procedimento_descricao': entry.procedimento_descricao,
        'frequencia_relativa': _decimal_to_string(frequencia, precision=3) if frequencia is not None else None,
        'custo_procedimento': _decimal_to_string(custo_procedimento),
        'variacoes': variacoes,
        'substitutos': substitutos,
        'narrativa': entry.narrativa,
        'origem': entry.origem,
    }


def _suggest_similar_items(
    origem: str,
    item_model: BrasItemNormalized | SimproItemNormalized | SimproItem,
    *,
    limit: int = 5,
) -> list[dict]:
    item_id = getattr(item_model, 'id', None)
    if item_id is None:
        return []

    safe_limit = max(1, int(limit or 5))

    def _extract_digits(value: str | None) -> str:
        if not value:
            return ''
        return ''.join(ch for ch in str(value) if ch.isdigit())

    def _normalize_description(text: str | None) -> str:
        if not text:
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        normalized = re.sub(r'[^A-Za-z0-9 ]+', ' ', normalized)
        return normalized.upper()

    ranked: dict[int, dict[str, object]] = {}

    def _register(rows: Sequence[InsumoIndex], base_score: int, reason: str | None = None) -> None:
        for row in rows:
            if row is None or row.item_id == item_id:
                continue
            current = ranked.get(row.item_id)
            score = base_score
            if current is None:
                ranked[row.item_id] = {
                    'row': row,
                    'score': score,
                    'updated_at': getattr(row, 'updated_at', None),
                    'reasons': set([reason] if reason else []),
                }
            else:
                if score > current['score']:
                    current['score'] = score
                if reason:
                    current.setdefault('reasons', set()).add(reason)

    # --- Strong matches: EAN -------------------------------------------------
    ean = getattr(item_model, 'ean', None)
    ean = (ean or '').strip()
    if ean:
        if origem == 'BRAS':
            model_cls = BrasItemNormalized
        elif isinstance(item_model, SimproItemCadastro):
            model_cls = SimproItemCadastro
        else:
            model_cls = SimproItemNormalized
        ean_ids = [
            row_id for (row_id,) in (
                db.session.query(model_cls.id)
                .filter(model_cls.ean == ean)
                .limit(50)
                .all()
            )
        ]
        if ean_ids:
            ean_rows = (
                InsumoIndex.query
                .filter(
                    InsumoIndex.origem == origem,
                    InsumoIndex.item_id != item_id,
                    InsumoIndex.item_id.in_(ean_ids),
                )
                .all()
            )
            _register(ean_rows, 500, 'EAN idêntico')

    # --- ANVISA --------------------------------------------------------------
    anvisa = None
    if hasattr(item_model, 'anvisa'):
        anvisa = (item_model.anvisa or '').strip()
    elif hasattr(item_model, 'registro_anvisa'):
        anvisa = (item_model.registro_anvisa or '').strip()
    if anvisa:
        anvisa_rows = (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == origem,
                InsumoIndex.item_id != item_id,
                InsumoIndex.anvisa == anvisa,
            )
            .limit(safe_limit * 4)
            .all()
        )
        _register(anvisa_rows, 360, 'Registro ANVISA compartilhado')

    # --- Code prefixes (TUSS / TISS) ----------------------------------------
    tuss_value = None
    if hasattr(item_model, 'codigo'):
        tuss_value = item_model.codigo
    elif hasattr(item_model, 'produto_codigo'):
        tuss_value = item_model.produto_codigo
    elif hasattr(item_model, 'tuss'):
        tuss_value = item_model.tuss
    tuss_digits = _extract_digits(tuss_value)
    tuss_prefix = tuss_digits[:5] if len(tuss_digits) >= 5 else tuss_digits[:4]
    if tuss_prefix and len(tuss_prefix) >= 3:
        tuss_rows = (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == origem,
                InsumoIndex.item_id != item_id,
                func.replace(func.replace(func.coalesce(InsumoIndex.tuss, ''), '.', ''), '-', '').like(f"{tuss_prefix}%"),
            )
            .limit(safe_limit * 4)
            .all()
        )
        label = f'Prefixo TUSS {tuss_prefix}'
        _register(tuss_rows, 240, label)

    tiss_value = None
    if hasattr(item_model, 'codigo_alt'):
        tiss_value = item_model.codigo_alt
    elif hasattr(item_model, 'apresentacao_codigo'):
        tiss_value = item_model.apresentacao_codigo
    elif hasattr(item_model, 'tiss'):
        tiss_value = item_model.tiss
    tiss_digits = _extract_digits(tiss_value)
    tiss_prefix = tiss_digits[:5] if len(tiss_digits) >= 5 else tiss_digits[:4]
    if tiss_prefix and len(tiss_prefix) >= 3:
        tiss_rows = (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == origem,
                InsumoIndex.item_id != item_id,
                func.replace(func.replace(func.coalesce(InsumoIndex.tiss, ''), '.', ''), '-', '').like(f"{tiss_prefix}%"),
            )
            .limit(safe_limit * 3)
            .all()
        )
        label = f'Prefixo TISS {tiss_prefix}'
        _register(tiss_rows, 220, label)

    # --- Substitutos (contexto clínico) --------------------------------------
    substituto_codes: set[str] = set()
    contexto_rows = (
        InsumoContextoClinico.query
        .with_entities(InsumoContextoClinico.substitutos_raw)
        .filter_by(origem=origem, item_id=item_id)
        .all()
    )
    for (raw_codes,) in contexto_rows:
        for token in _split_substitutos(raw_codes):
            cleaned = token.strip()
            if cleaned:
                substituto_codes.add(cleaned)

    substituto_filters: list = []
    for code in substituto_codes:
        normalized = code.strip().upper()
        digits = _extract_digits(code)
        if normalized:
            substituto_filters.append(func.upper(func.coalesce(InsumoIndex.tuss, '')) == normalized)
            substituto_filters.append(func.upper(func.coalesce(InsumoIndex.tiss, '')) == normalized)
        if digits and digits != normalized:
            substituto_filters.append(func.upper(func.coalesce(InsumoIndex.tuss, '')) == digits)
            substituto_filters.append(func.upper(func.coalesce(InsumoIndex.tiss, '')) == digits)

    if substituto_filters:
        substituto_rows = (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == origem,
                InsumoIndex.item_id != item_id,
                or_(*substituto_filters),
            )
            .limit(safe_limit * 3)
            .all()
        )
        _register(substituto_rows, 440, 'Marcado como substituto clínico')

    # --- Description tokens --------------------------------------------------
    descricao_source = getattr(item_model, 'descricao', None)
    if not descricao_source and isinstance(item_model, BrasItemNormalized):
        descricao_source = ' '.join(
            part for part in [item_model.produto_nome, item_model.apresentacao_descricao] if part
        )
    description_norm = _normalize_description(descricao_source)
    tokens: list[str] = []
    if description_norm:
        seen = set()
        stopwords = {'COM', 'DE', 'PARA', 'COMPOS', 'SEM', 'CAPS', 'CAPSULAS', 'TABLETE', 'ML', 'MG', 'G'}
        for token in description_norm.split():
            if len(token) < 4:
                continue
            if token.isdigit():
                continue
            if token in stopwords:
                continue
            if token in seen:
                continue
            seen.add(token)
            tokens.append(token)
            if len(tokens) >= 6:
                break

    if tokens:
        token_filters = [func.upper(func.coalesce(InsumoIndex.descricao, '')).like(f"%{token}%") for token in tokens]
        token_rows = (
            InsumoIndex.query
            .filter(
                InsumoIndex.origem == origem,
                InsumoIndex.item_id != item_id,
                or_(*token_filters),
            )
            .limit(safe_limit * 5)
            .all()
        )
        for token in tokens:
            token_rows_token = [
                row for row in token_rows
                if row.descricao and token in row.descricao.upper()
            ]
            if token_rows_token:
                _register(token_rows_token, 160, f'Termo relevante: {token}')

    if not ranked:
        return []

    def _sort_key(entry: dict) -> tuple:
        row: InsumoIndex = entry['row']  # type: ignore[assignment]
        descr = (row.descricao or '').strip().upper()
        updated_at = entry.get('updated_at')
        return (-entry['score'], descr, updated_at or datetime.min, row.item_id)

    sorted_rows = sorted(ranked.values(), key=_sort_key)
    sliced = sorted_rows[:safe_limit]
    results: list[dict] = []
    for entry in sliced:
        payload = _serialize_insumo_index(entry['row'])
        reasons = entry.get('reasons')
        if reasons:
            payload['justificativas'] = sorted(reason for reason in reasons if reason)
        payload['similaridade_score'] = entry['score']
        results.append(payload)
    return results


def _lookup_porte_valor(operadora_id, uf, nome_hint, porte_codigo):
    if not porte_codigo:
        return None
    q = Tabela.query.filter(Tabela.tipo_tabela == 'porte', Tabela.id_operadora == operadora_id)
    if uf:
        q = q.filter(Tabela.uf == uf)
    # Se existir alguma com nome próximo ao hint, prefere
    if nome_hint:
        cand = q.filter(Tabela.nome.ilike(f"%{nome_hint}%")).order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
        if cand:
            pv = PorteValorItem.query.filter_by(id_tabela=cand.id, porte=str(porte_codigo)).first()
            if pv:
                return pv.valor
    # Fallback: mais recente
    cand = q.order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
    if cand:
        pv = PorteValorItem.query.filter_by(id_tabela=cand.id, porte=str(porte_codigo)).first()
        if pv:
            return pv.valor
    return None


def _lookup_porte_an_valor(operadora_id, uf, nome_hint, porte_an):
    if not porte_an:
        return None
    q = Tabela.query.filter(Tabela.tipo_tabela == 'porte_anestesico', Tabela.id_operadora == operadora_id)
    if uf:
        q = q.filter(Tabela.uf == uf)
    if nome_hint:
        cand = q.filter(Tabela.nome.ilike(f"%{nome_hint}%")).order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
        if cand:
            pv = PorteAnestesicoValorItem.query.filter_by(id_tabela=cand.id, porte_an=str(porte_an)).first()
            if pv:
                return pv.valor
    cand = q.order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
    if cand:
        pv = PorteAnestesicoValorItem.query.filter_by(id_tabela=cand.id, porte_an=str(porte_an)).first()
        if pv:
            return pv.valor
    return None


def _resolve_porte_tabela_nome(operadora_id, uf, nome_hint, porte_codigo):
    """Resolve o nome da tabela de Porte utilizada seguindo a mesma heurística
    de _lookup_porte_valor (preferindo por hint; fallback mais recente)."""
    if not porte_codigo:
        return None
    q = Tabela.query.filter(Tabela.tipo_tabela == 'porte', Tabela.id_operadora == operadora_id)
    if uf:
        q = q.filter(Tabela.uf == uf)
    if nome_hint:
        cand = q.filter(Tabela.nome.ilike(f"%{nome_hint}%")).order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
        if cand:
            pv = PorteValorItem.query.filter_by(id_tabela=cand.id, porte=str(porte_codigo)).first()
            if pv:
                return cand.nome
    cand = q.order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
    if cand:
        pv = PorteValorItem.query.filter_by(id_tabela=cand.id, porte=str(porte_codigo)).first()
        if pv:
            return cand.nome
    return None


def _resolve_porte_an_tabela_nome(operadora_id, uf, nome_hint, porte_an):
    """Resolve o nome da tabela de Porte Anestésico utilizada seguindo a mesma
    heurística de _lookup_porte_an_valor."""
    if not porte_an:
        return None
    q = Tabela.query.filter(Tabela.tipo_tabela == 'porte_anestesico', Tabela.id_operadora == operadora_id)
    if uf:
        q = q.filter(Tabela.uf == uf)
    if nome_hint:
        cand = q.filter(Tabela.nome.ilike(f"%{nome_hint}%")).order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
        if cand:
            pv = PorteAnestesicoValorItem.query.filter_by(id_tabela=cand.id, porte_an=str(porte_an)).first()
            if pv:
                return cand.nome
    cand = q.order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc()).first()
    if cand:
        pv = PorteAnestesicoValorItem.query.filter_by(id_tabela=cand.id, porte_an=str(porte_an)).first()
        if pv:
            return cand.nome
    return None


def _get_latest_cbhpm_table():
    return (
        Tabela.query
        .filter(Tabela.tipo_tabela == 'cbhpm')
        .order_by(Tabela.data_vigencia.is_(None), Tabela.data_vigencia.desc(), Tabela.id.desc())
        .first()
    )


def _resolve_cbhpm_valor_total(item: CBHPMItem, tabela_ref: Tabela) -> Decimal | None:
    if not item or not tabela_ref:
        return None
    total = None
    try:
        total = compute_cbhpm_total(item, tabela_ref)
    except Exception:
        db.session.rollback()
        total = None
    if total not in (None, Decimal('0')):
        return _as_decimal(total)
    for candidate in (
        getattr(item, 'subtotal', None),
        getattr(item, 'total_porte', None),
        getattr(item, 'total_uco', None),
        getattr(item, 'total_filme', None),
        getattr(item, 'total_porte_anestesico', None),
        getattr(item, 'total_auxiliares', None),
        getattr(item, 'valor_porte', None),
    ):
        val = _as_decimal(candidate)
        if val not in (None, Decimal('0')):
            return val
    return None



def _clone_default_cbhpm_rules():
    return json.loads(json.dumps(DEFAULT_CBHPM_RULES))


def _get_active_cbhpm_ruleset(return_model: bool = False):
    ruleset = None
    try:
        ruleset = (
            CBHPMRuleSet.query
            .filter(CBHPMRuleSet.ativo.is_(True))
            .order_by(CBHPMRuleSet.atualizado_em.desc())
            .first()
        )
        if not ruleset:
            ruleset = (
                CBHPMRuleSet.query
                .order_by(CBHPMRuleSet.atualizado_em.desc())
                .first()
            )
    except Exception:
        db.session.rollback()
        ruleset = None
    data = _clone_default_cbhpm_rules()
    if ruleset and isinstance(ruleset.regras, dict) and ruleset.regras:
        try:
            data = json.loads(json.dumps(ruleset.regras))
        except Exception:
            data = _clone_default_cbhpm_rules()
    if return_model:
        return data, ruleset
    return data


def _apply_ruleset_to_breakdown(item: CBHPMItem, tabela_ref: Tabela, breakdown: dict, rules: dict | None):
    result = dict(breakdown or {})
    applied = []

    total_porte = _as_decimal(result.get('total_porte'))
    if total_porte is not None:
        aux_cfg = (rules or {}).get('auxiliares') or {}
        current_aux = _as_decimal(result.get('total_auxiliares'))
        aux_count_raw = getattr(item, 'numero_auxiliares', None)
        explicit_no_aux = False
        if aux_count_raw is not None:
            try:
                explicit_no_aux = Decimal(str(aux_count_raw)) == Decimal('0')
            except (InvalidOperation, ValueError):
                explicit_no_aux = False
        aux_details = []
        if (current_aux is None or current_aux == Decimal('0')) and aux_cfg.get('percentuais') and not explicit_no_aux:
            percentuais = aux_cfg.get('percentuais') or []
            try:
                aux_count = int(aux_count_raw) if aux_count_raw is not None else None
            except (TypeError, ValueError):
                aux_count = None
            max_por_porte = aux_cfg.get('max_por_porte') or {}
            porte_key = str(getattr(item, 'porte', '') or '').strip()
            max_aux = max_por_porte.get(porte_key, max_por_porte.get('default'))
            try:
                max_aux = int(max_aux) if max_aux is not None else None
            except (TypeError, ValueError):
                max_aux = None
            if aux_count is None:
                aux_count = max_aux if max_aux is not None else len(percentuais)
            elif max_aux is not None:
                aux_count = min(aux_count, max_aux)
            aux_count = max(aux_count or 0, 0)
            if aux_count and percentuais:
                computed = Decimal('0')
                for idx in range(aux_count):
                    perc = percentuais[min(idx, len(percentuais) - 1)]
                    perc = Decimal(str(perc))
                    if perc > 1:
                        perc = perc / Decimal('100')
                    if perc <= 0:
                        continue
                    value_aux = total_porte * perc
                    computed += value_aux
                    perc_display = perc * Decimal('100')
                    aux_details.append({
                        'indice': idx + 1,
                        'percentual_pct': str(perc_display),
                        'valor': value_aux
                    })
                if computed > 0:
                    result['total_auxiliares'] = computed
                    result['auxiliares_detalhe'] = aux_details
                    applied.append({
                        'component': 'auxiliares',
                        'rule': 'percentuais',
                        'quantidade': aux_count
                    })
        elif explicit_no_aux:
            result['total_auxiliares'] = Decimal('0')
            result['auxiliares_detalhe'] = []
    if 'auxiliares_detalhe' not in result or result['auxiliares_detalhe'] is None:
        aux_existing = []
        for idx, attr in enumerate(['total_1_aux', 'total_2_aux', 'total_3_aux', 'total_4_aux'], start=1):
            val = _as_decimal(getattr(item, attr, None))
            if val is not None and val != Decimal('0'):
                aux_existing.append({
                    'indice': idx,
                    'percentual_pct': None,
                    'valor': val
                })
        if aux_existing:
            result['auxiliares_detalhe'] = aux_existing
    result['total_porte'] = _as_decimal(result.get('total_porte'))
    result['total_filme'] = _as_decimal(result.get('total_filme'))
    result['total_uco'] = _as_decimal(result.get('total_uco'))
    result['total_porte_an'] = _as_decimal(result.get('total_porte_an'))
    result['total_auxiliares'] = _as_decimal(result.get('total_auxiliares'))

    multipliers = [
        ('total_porte', (rules or {}).get('porte'), 'porte'),
        ('total_filme', (rules or {}).get('filme'), 'filme'),
        ('total_uco', (rules or {}).get('uco'), 'uco'),
        ('total_porte_an', (rules or {}).get('porte_an'), 'porte_an'),
    ]
    for key, cfg, comp_name in multipliers:
        if not cfg:
            continue
        factor_raw = cfg.get('multiplicador') if isinstance(cfg, dict) else None
        if factor_raw in (None, '', 'None'):
            continue
        try:
            factor = Decimal(str(factor_raw))
        except (InvalidOperation, ValueError):
            continue
        if factor > Decimal('5'):
            factor = factor / Decimal('100')
        if factor < Decimal('0'):
            factor = Decimal('0')
        current = result.get(key)
        if current is None:
            continue
        new_value = current * factor
        if new_value == current:
            continue
        result[key] = new_value
        applied.append({
            'component': comp_name,
            'rule': 'multiplicador',
            'fator': str(factor)
        })

    result['total'] = _sum_decimals([
        result.get('total_porte'),
        result.get('total_filme'),
        result.get('total_uco'),
        result.get('total_porte_an'),
        result.get('total_auxiliares'),
    ])
    if applied:
        result['applied_rules'] = applied
    return result

def compute_cbhpm_total(item: CBHPMItem, tabela_ref: Tabela, porte_hint: str | None = None, porte_an_hint: str | None = None,
                        ajuste_porte_pct: Decimal | None = None, ajuste_porte_an_pct: Decimal | None = None, rules: dict | None = None):
    breakdown = compute_cbhpm_breakdown(
        item,
        tabela_ref,
        porte_hint=porte_hint,
        porte_an_hint=porte_an_hint,
        ajuste_porte_pct=ajuste_porte_pct,
        ajuste_porte_an_pct=ajuste_porte_an_pct,
        rules=rules,
    )
    return breakdown.get('total')


def compute_cbhpm_breakdown(item: CBHPMItem, tabela_ref: Tabela, porte_hint: str | None = None, porte_an_hint: str | None = None,
                            ajuste_porte_pct: Decimal | None = None, ajuste_porte_an_pct: Decimal | None = None, rules: dict | None = None):
    valor_porte = _as_decimal(item.valor_porte)
    if valor_porte is None:
        valor_porte = _lookup_porte_valor(tabela_ref.id_operadora, tabela_ref.uf, (porte_hint or tabela_ref.nome), item.porte)
    fracao_input = getattr(item, '_fracao_input', None)
    fracao = _as_decimal(fracao_input) if fracao_input is not None else _as_decimal(item.fracao_porte)
    if fracao is None or fracao <= Decimal('0'):
        fracao = Decimal('1')
    elif fracao_input is None and fracao < Decimal('1'):
        fracao = Decimal('1')
    total_porte = None
    if valor_porte is not None:
        total_porte = (valor_porte * fracao)
    if total_porte is None:
        total_porte = _as_decimal(item.total_porte)
    if total_porte is not None and ajuste_porte_pct:
        total_porte = total_porte * (Decimal('1') + (ajuste_porte_pct/Decimal('100')))

    total_filme = _as_decimal(item.total_filme)
    if total_filme is None:
        filme = _as_decimal(item.filme)
        incid = _as_decimal(item.incidencias)
        if filme is not None:
            total_filme = (filme * (incid or Decimal('1')))

    total_uco = _as_decimal(item.total_uco)
    if total_uco is None:
        uco_qtd = _as_decimal(item.uco)
        uco_val = _as_decimal(tabela_ref.uco_valor)
        if uco_qtd is not None and uco_val is not None:
            total_uco = (uco_qtd * uco_val)

    valor_an = _as_decimal(item.valor_porte_anestesico)
    if valor_an is None:
        valor_an = _lookup_porte_an_valor(tabela_ref.id_operadora, tabela_ref.uf, (porte_an_hint or tabela_ref.nome), item.porte_anestesico)
    total_an = _as_decimal(item.total_porte_anestesico)
    if total_an is not None and ajuste_porte_an_pct:
        total_an = total_an * (Decimal('1') + (ajuste_porte_an_pct/Decimal('100')))
    if total_an is None and valor_an is not None:
        total_an = valor_an

    total_aux = _as_decimal(item.total_auxiliares)
    if total_aux is None:
        total_aux = _sum_decimals([item.total_1_aux, item.total_2_aux, item.total_3_aux, item.total_4_aux])

    breakdown = {
        'total_porte': total_porte,
        'total_filme': total_filme,
        'total_uco': total_uco,
        'total_porte_an': total_an,
        'total_auxiliares': total_aux,
        'total': _sum_decimals([total_porte, total_filme, total_uco, total_an, total_aux]),
    }
    ruleset_dict = rules or _get_active_cbhpm_ruleset()
    breakdown = _apply_ruleset_to_breakdown(item, tabela_ref, breakdown, ruleset_dict)
    return breakdown

@app.route('/tabelas/importar/diarias-taxas-pacotes', methods=['POST'])
@admin_required
def importar_diarias_taxas_pacotes():
    file = request.files.get('arquivo')
    nome_tabela = request.form.get('nome_tabela')
    prestador = request.form.get('prestador')
    uf = request.form.get('uf')
    data_vigencia = request.form.get('data_vigencia')  # YYYY-MM-DD
    operadora_id = request.form.get('operadora_id')
    substituir = request.form.get('substituir') in ('on', 'true', '1', 'yes', 'sim', 'true')

    if not file or not nome_tabela or not operadora_id:
        return redirect(url_for('gerenciar_tabelas'))

    # A criação de Tabelas ocorrerá após a leitura do arquivo, podendo ser
    # uma por prestador (e UF) quando não informado no formulário.

    filename = secure_filename(file.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    linhas = []
    if ext == 'csv':
        content = file.read().decode('utf-8-sig', errors='ignore').splitlines()
        if not content:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [h.strip() for h in content[0].split(',')]
        keys = [_norm_header(h) for h in headers]
        for row in content[1:]:
            cols = row.split(',')
            item = {keys[i]: (cols[i].strip() if i < len(cols) else '') for i in range(len(keys))}
            linhas.append(item)
    elif ext == 'xlsx':
        try:
            from openpyxl import load_workbook
        except Exception:
            return redirect(url_for('gerenciar_tabelas'))
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [str(h) if h is not None else '' for h in rows[0]]
        keys = [_norm_header(h) for h in headers]
        for r in rows[1:]:
            item = {keys[i]: (r[i] if i < len(keys) else None) for i in range(len(keys))}
            linhas.append(item)
    else:
        db.session.rollback()
        return redirect(url_for('gerenciar_tabelas'))

    # Importação consolidada: cria uma única Tabela e grava o prestador/UF por item
    if True:
        if substituir:
            subq = db.session.query(Tabela.id).filter(
                Tabela.nome == nome_tabela,
                Tabela.id_operadora == int(operadora_id)
            )
            # Multi-operadora: deletar apenas procedimentos da operadora correta
            db.session.query(Procedimento).filter(
                Procedimento.id_tabela.in_(subq),
                Procedimento.operadora_id == int(operadora_id)
            ).delete(synchronize_session=False)
            db.session.query(Tabela).filter(
                Tabela.nome == nome_tabela,
                Tabela.id_operadora == int(operadora_id)
            ).delete(synchronize_session=False)
            db.session.flush()

        tab = Tabela(
            nome=nome_tabela,
            prestador=None,
            tipo_tabela='diarias_taxas_pacotes',
            uf=uf,
            id_operadora=int(operadora_id)
        )
        if data_vigencia:
            try:
                tab.data_vigencia = date.fromisoformat(data_vigencia)
            except Exception:
                pass
        db.session.add(tab)
        db.session.flush()

        for item in linhas:
            codigo = item.get('codigo') or item.get('cod')
            descricao = item.get('descricao') or item.get('descriçao') or item.get('descrição')
            valor = _parse_money(item.get('valor'))
            if not codigo or not descricao:
                continue
            prest_item = item.get('prestador') or item.get('fornecedor') or item.get('credenciado') or prestador
            prest_item = str(prest_item).strip() if prest_item is not None else None
            uf_item = (item.get('uf') or uf)
            uf_item = str(uf_item).strip() if uf_item else None
            db.session.add(Procedimento(
                codigo=str(codigo),
                descricao=str(descricao),
                valor=valor,
                prestador=prest_item or None,
                uf=uf_item or None,
                id_tabela=tab.id,
                operadora_id=int(operadora_id)  # Multi-operadora
            ))

        db.session.commit()
        return redirect(url_for('gerenciar_tabelas'))

    if prestador:
        # Se solicitado, remove tabelas existentes com o mesmo nome/operadora/UF/prestador
        if substituir:
            db.session.query(Procedimento).filter(
                Procedimento.id_tabela.in_(db.session.query(Tabela.id).filter(
                    Tabela.nome == nome_tabela,
                    Tabela.id_operadora == int(operadora_id),
                    (Tabela.uf == uf) if uf else True,
                    Tabela.prestador == prestador,
                ))
            ).delete(synchronize_session=False)
            db.session.query(Tabela).filter(
                Tabela.nome == nome_tabela,
                Tabela.id_operadora == int(operadora_id),
                (Tabela.uf == uf) if uf else True,
                Tabela.prestador == prestador,
            ).delete(synchronize_session=False)
            db.session.flush()
        # Importa tudo em uma única tabela usando o prestador do formulário
        tab = Tabela(
            nome=nome_tabela,
            prestador=prestador,
            tipo_tabela='diarias_taxas_pacotes',
            uf=uf,
            id_operadora=int(operadora_id)
        )
        if data_vigencia:
            try:
                tab.data_vigencia = date.fromisoformat(data_vigencia)
            except Exception:
                pass
        db.session.add(tab)
        db.session.flush()
        for item in linhas:
            codigo = item.get('codigo') or item.get('cod')
            descricao = item.get('descricao') or item.get('descriçao') or item.get('descrição')
            valor = _parse_money(item.get('valor'))
            if not codigo or not descricao:
                continue
            db.session.add(Procedimento(codigo=str(codigo), descricao=str(descricao), valor=valor, id_tabela=tab.id))
    else:
        # Agrupa por prestador (e UF) vindos do arquivo
        grupos = {}
        for item in linhas:
            codigo = item.get('codigo') or item.get('cod')
            descricao = item.get('descricao') or item.get('descriçao') or item.get('descrição')
            valor = _parse_money(item.get('valor'))
            if not codigo or not descricao:
                continue
            prest = item.get('prestador') or item.get('fornecedor') or item.get('credenciado') or ''
            prest = str(prest).strip() if prest is not None else ''
            uf_row = (item.get('uf') or uf or '').strip() if item.get('uf') is not None else (uf or '')
            nome_arq = item.get('tabela') or nome_tabela
            key = (prest or '-'), (uf_row or '')
            bucket = grupos.setdefault(key, {"nome": nome_arq or nome_tabela, "items": []})
            bucket["items"].append((str(codigo), str(descricao), valor))

        # Se solicitado, remove tabelas existentes com os nomes que serão criados
        if substituir and grupos:
            nomes_alvo = {g["nome"] for g in grupos.values()}
            subq = db.session.query(Tabela.id).filter(
                Tabela.id_operadora == int(operadora_id),
                Tabela.nome.in_(list(nomes_alvo))
            )
            db.session.query(Procedimento).filter(Procedimento.id_tabela.in_(subq)).delete(synchronize_session=False)
            db.session.query(Tabela).filter(
                Tabela.id_operadora == int(operadora_id),
                Tabela.nome.in_(list(nomes_alvo))
            ).delete(synchronize_session=False)
            db.session.flush()

        for (prest_key, uf_key), bucket in grupos.items():
            tab = Tabela(
                nome=bucket["nome"],
                prestador=None if prest_key == '-' else prest_key,
                tipo_tabela='diarias_taxas_pacotes',
                uf=(uf_key or None),
                id_operadora=int(operadora_id)
            )
            if data_vigencia:
                try:
                    tab.data_vigencia = date.fromisoformat(data_vigencia)
                except Exception:
                    pass
            db.session.add(tab)
            db.session.flush()
            for codigo, descricao, valor in bucket["items"]:
                db.session.add(Procedimento(codigo=codigo, descricao=descricao, valor=valor, id_tabela=tab.id))

    db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


@app.route('/tabelas/<int:tid>/excluir', methods=['POST'])
@admin_required
def tabela_excluir(tid):
    t = Tabela.query.get_or_404(tid)
    # Remove itens vinculados e depois a tabela
    db.session.query(Procedimento).filter_by(id_tabela=tid).delete(synchronize_session=False)
    db.session.delete(t)
    db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


@app.route('/tabelas/uco/definir', methods=['POST'])
@admin_required
def definir_uco_cbhpm():
    updated = 0
    form = request.form or {}
    for key, value in form.items():
        if not key.startswith('uco_'):
            continue
        try:
            tid = int(key.split('_', 1)[1])
        except Exception:
            continue
        tab = Tabela.query.get(tid)
        if not tab or tab.tipo_tabela != 'cbhpm':
            continue
        v = (value or '').strip()
        tab.uco_valor = _parse_money(v) if v else None
        updated += 1
    if updated:
        db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


@app.route('/tabelas/importar/porte', methods=['POST'])
@admin_required
def importar_porte():
    file = request.files.get('arquivo')
    nome_tabela = request.form.get('nome_tabela')
    uf = request.form.get('uf')
    data_vigencia = request.form.get('data_vigencia')
    operadora_id = request.form.get('operadora_id')
    substituir = request.form.get('substituir') in ('on', 'true', '1', 'yes', 'sim', 'true')

    if not file or not nome_tabela or not operadora_id:
        return redirect(url_for('gerenciar_tabelas'))

    if substituir:
        subq = db.session.query(Tabela.id).filter(Tabela.nome == nome_tabela, Tabela.id_operadora == int(operadora_id), Tabela.tipo_tabela == 'porte')
        db.session.query(PorteValorItem).filter(PorteValorItem.id_tabela.in_(subq)).delete(synchronize_session=False)
        db.session.query(Tabela).filter(Tabela.id.in_(subq)).delete(synchronize_session=False)
        db.session.flush()

    tab = Tabela(nome=nome_tabela, prestador=None, tipo_tabela='porte', uf=uf, id_operadora=int(operadora_id))
    if data_vigencia:
        try:
            tab.data_vigencia = date.fromisoformat(data_vigencia)
        except Exception:
            pass
    db.session.add(tab)
    db.session.flush()

    filename = secure_filename(file.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    linhas = []
    if ext == 'csv':
        content = file.read().decode('utf-8-sig', errors='ignore').splitlines()
        if not content:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [h.strip() for h in content[0].split(',')]
        keys = [_norm_header(h) for h in headers]
        for row in content[1:]:
            cols = row.split(',')
            linhas.append({keys[i]: (cols[i].strip() if i < len(cols) else '') for i in range(len(keys))})
    elif ext == 'xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [str(h) if h is not None else '' for h in rows[0]]
        keys = [_norm_header(h) for h in headers]
        for r in rows[1:]:
            linhas.append({keys[i]: (r[i] if i < len(keys) else None) for i in range(len(keys))})
    else:
        return redirect(url_for('gerenciar_tabelas'))

    for row in linhas:
        porte = row.get('porte') or row.get('portevalor') or row.get('portecodigo')
        if not porte:
            continue
        valor = _parse_money(row.get('valor'))
        db.session.add(PorteValorItem(porte=str(porte), valor=valor, uf=uf, id_tabela=tab.id))

    db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


@app.route('/tabelas/importar/porte-anestesico', methods=['POST'])
@admin_required
def importar_porte_anestesico():
    file = request.files.get('arquivo')
    nome_tabela = request.form.get('nome_tabela')
    uf = request.form.get('uf')
    data_vigencia = request.form.get('data_vigencia')
    operadora_id = request.form.get('operadora_id')
    substituir = request.form.get('substituir') in ('on', 'true', '1', 'yes', 'sim', 'true')

    if not file or not nome_tabela or not operadora_id:
        return redirect(url_for('gerenciar_tabelas'))

    if substituir:
        subq = db.session.query(Tabela.id).filter(Tabela.nome == nome_tabela, Tabela.id_operadora == int(operadora_id), Tabela.tipo_tabela == 'porte_anestesico')
        db.session.query(PorteAnestesicoValorItem).filter(PorteAnestesicoValorItem.id_tabela.in_(subq)).delete(synchronize_session=False)
        db.session.query(Tabela).filter(Tabela.id.in_(subq)).delete(synchronize_session=False)
        db.session.flush()

    tab = Tabela(nome=nome_tabela, prestador=None, tipo_tabela='porte_anestesico', uf=uf, id_operadora=int(operadora_id))
    if data_vigencia:
        try:
            tab.data_vigencia = date.fromisoformat(data_vigencia)
        except Exception:
            pass
    db.session.add(tab)
    db.session.flush()

    filename = secure_filename(file.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    linhas = []
    if ext == 'csv':
        content = file.read().decode('utf-8-sig', errors='ignore').splitlines()
        if not content:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [h.strip() for h in content[0].split(',')]
        keys = [_norm_header(h) for h in headers]
        for row in content[1:]:
            cols = row.split(',')
            linhas.append({keys[i]: (cols[i].strip() if i < len(cols) else '') for i in range(len(keys))})
    elif ext == 'xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [str(h) if h is not None else '' for h in rows[0]]
        keys = [_norm_header(h) for h in headers]
        for r in rows[1:]:
            linhas.append({keys[i]: (r[i] if i < len(keys) else None) for i in range(len(keys))})
    else:
        return redirect(url_for('gerenciar_tabelas'))

    for row in linhas:
        porte_an = row.get('portean') or row.get('porteanestesico') or row.get('porte_an') or row.get('porte an')
        if not porte_an:
            continue
        valor = _parse_money(row.get('valor'))
        db.session.add(PorteAnestesicoValorItem(porte_an=str(porte_an), valor=valor, uf=uf, id_tabela=tab.id))

    db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


@app.route('/tabelas/importar/cbhpm', methods=['POST'])
@admin_required
def importar_cbhpm():
    file = request.files.get('arquivo')
    nome_tabela = request.form.get('nome_tabela')
    uf = request.form.get('uf')
    data_vigencia = request.form.get('data_vigencia')
    operadora_id = request.form.get('operadora_id')
    substituir = request.form.get('substituir') in ('on', 'true', '1', 'yes', 'sim', 'true')

    if not file or not nome_tabela or not operadora_id:
        return redirect(url_for('gerenciar_tabelas'))

    # Substituição
    if substituir:
        subq = db.session.query(Tabela.id).filter(
            Tabela.nome == nome_tabela,
            Tabela.id_operadora == int(operadora_id),
            Tabela.tipo_tabela == 'cbhpm'
        )
        db.session.query(CBHPMItem).filter(CBHPMItem.id_tabela.in_(subq)).delete(synchronize_session=False)
        db.session.query(Tabela).filter(
            Tabela.nome == nome_tabela,
            Tabela.id_operadora == int(operadora_id),
            Tabela.tipo_tabela == 'cbhpm'
        ).delete(synchronize_session=False)
        db.session.flush()

    # Cria Tabela
    tab = Tabela(
        nome=nome_tabela,
        prestador=None,
        tipo_tabela='cbhpm',
        uf=uf,
        id_operadora=int(operadora_id)
    )
    if data_vigencia:
        try:
            tab.data_vigencia = date.fromisoformat(data_vigencia)
        except Exception:
            pass
    db.session.add(tab)
    db.session.flush()

    filename = secure_filename(file.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    # Leitura
    linhas = []
    keys = []
    if ext == 'csv':
        content = file.read().decode('utf-8-sig', errors='ignore').splitlines()
        if not content:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [h.strip() for h in content[0].split(',')]
        keys = [_norm_header(h) for h in headers]
        for row in content[1:]:
            cols = row.split(',')
            item = {keys[i]: (cols[i].strip() if i < len(cols) else '') for i in range(len(keys))}
            linhas.append(item)
    elif ext == 'xlsx':
        try:
            from openpyxl import load_workbook
        except Exception:
            return redirect(url_for('gerenciar_tabelas'))
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect(url_for('gerenciar_tabelas'))
        headers = [str(h) if h is not None else '' for h in rows[0]]
        keys = [_norm_header(h) for h in headers]
        for r in rows[1:]:
            item = {keys[i]: (r[i] if i < len(keys) else None) for i in range(len(keys))}
            linhas.append(item)
    else:
        db.session.rollback()
        return redirect(url_for('gerenciar_tabelas'))

    # Campos esperados (normalizados)
    def g(d, *names):
        for n in names:
            v = d.get(n)
            if v not in (None, ''):
                return v
        return None

    def dec(v):
        return _parse_money(v) if v not in (None, '') else None

    def intval(v):
        try:
            return int(str(v).strip()) if v not in (None, '') else None
        except Exception:
            return None

    for row in linhas:
        codigo = str(g(row, 'codigo')) if g(row, 'codigo') is not None else None
        descricao = str(g(row, 'procedimento', 'descricao')) if g(row, 'procedimento', 'descricao') is not None else ''
        if not codigo:
            continue
        item = CBHPMItem(
            codigo=codigo,
            procedimento=descricao,
            uf=uf,
            porte=str(g(row, 'porte')) if g(row, 'porte') is not None else None,
            fracao_porte=dec(g(row, 'fracaoporte', 'fraçãoporte')),
            valor_porte=dec(g(row, 'valorporte', 'valor_do_porte')),
            total_porte=dec(g(row, 'totalporte')),
            incidencias=str(g(row, 'incidencias', 'incidências')) if g(row, 'incidencias', 'incidências') is not None else None,
            filme=dec(g(row, 'filme')),
            total_filme=dec(g(row, 'totalfilme')),
            uco=dec(g(row, 'uco')),
            total_uco=dec(g(row, 'totaluco')),
            porte_anestesico=str(g(row, 'porteanestesico', 'porteanestésico')) if g(row, 'porteanestesico', 'porteanestésico') is not None else None,
            valor_porte_anestesico=dec(g(row, 'valorporteanestesico', 'valorporteanestésico')),
            total_porte_anestesico=dec(g(row, 'totalporteanestesico', 'totalporteanestésico')),
            numero_auxiliares=intval(g(row, 'numero_de_auxiliares', 'numerodeauxiliares')),
            total_auxiliares=dec(g(row, 'totalauxiliares')),
            total_1_aux=dec(g(row, 'total1oauxiliar', 'total1ºauxiliar', 'total1auxiliar')),
            total_2_aux=dec(g(row, 'total2oauxiliar', 'total2ºauxiliar', 'total2auxiliar')),
            total_3_aux=dec(g(row, 'total3oauxiliar', 'total3ºauxiliar', 'total3auxiliar')),
            total_4_aux=dec(g(row, 'total4oauxiliar', 'total4ºauxiliar', 'total4auxiliar')),
            subtotal=dec(g(row, 'subtotal')),
            id_tabela=tab.id,
        )
        db.session.add(item)

    db.session.commit()
    return redirect(url_for('gerenciar_tabelas'))


# --- 9. Visualização de Itens da Tabela ---
@app.template_filter('brl')
def brl(value):
    try:
        d = Decimal(value)
    except Exception:
        return value
    s = f"{d:,.2f}"
    return f"R$ {s}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.template_filter('date_br')
def date_br(value):
    try:
        return value.strftime('%d/%m/%Y') if value else '-'
    except Exception:
        return str(value) if value else '-'


@app.template_filter('string_format')
def string_format(value, format_str):
    """Format a number with thousand separators (e.g., 1,000)"""
    try:
        num = int(value)
        return format(num, ',')
    except (ValueError, TypeError):
        return str(value)


@app.route('/tabelas/<int:tid>/itens')
@admin_required
def tabela_itens(tid):
    tabela = Tabela.query.get_or_404(tid)
    q = request.args.get('q', '').strip()
    # Se for CBHPM, lista a partir da tabela específica
    if tabela.tipo_tabela == 'cbhpm':
        query = CBHPMItem.query.filter_by(id_tabela=tid)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (CBHPMItem.codigo.ilike(like)) | (CBHPMItem.procedimento.ilike(like))
            )
        rows = query.order_by(CBHPMItem.codigo).all()
        # Mapeia para o formato consumido pelo template (codigo, descricao, valor)
        itens = []
        for r in rows:
            val = r.subtotal
            if val in (None, Decimal('0')):
                val = compute_cbhpm_total(r, tabela)
            itens.append({
                'codigo': r.codigo,
                'descricao': r.procedimento,
                'valor': val,
            })
        return render_template('tabela-itens.html', tabela=tabela, itens=itens, q=q)

    if tabela.tipo_tabela == 'porte':
        query = PorteValorItem.query.filter_by(id_tabela=tid)
        if q:
            like = f"%{q}%"
            query = query.filter(PorteValorItem.porte.ilike(like))
        rows = query.order_by(PorteValorItem.porte).all()
        itens = [{'porte': r.porte, 'valor': r.valor, 'uf': r.uf} for r in rows]
        return render_template('tabela-porte-itens.html', tabela=tabela, itens=itens, q=q, label='Porte')

    if tabela.tipo_tabela == 'porte_anestesico':
        query = PorteAnestesicoValorItem.query.filter_by(id_tabela=tid)
        if q:
            like = f"%{q}%"
            query = query.filter(PorteAnestesicoValorItem.porte_an.ilike(like))
        rows = query.order_by(PorteAnestesicoValorItem.porte_an).all()
        itens = [{'porte': r.porte_an, 'valor': r.valor, 'uf': r.uf} for r in rows]
        return render_template('tabela-porte-itens.html', tabela=tabela, itens=itens, q=q, label='Porte AN')

    # Default: procedimentos comuns
    query = Procedimento.query.filter_by(id_tabela=tid)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Procedimento.codigo.ilike(like)) | (Procedimento.descricao.ilike(like))
        )
    itens = query.order_by(Procedimento.codigo).all()
    return render_template('tabela-itens.html', tabela=tabela, itens=itens, q=q)


@app.route('/insumos/search')
@login_required
@feature_required('insumos')
def insumos_search():
    page = _parse_positive_int(request.args.get('page'), 1, maximum=500)
    per_page = _parse_positive_int(request.args.get('per_page'), 50, maximum=500)

    filters = _extract_insumo_filters(request.args)
    payload = _catalogo_search(filters, page, per_page)
    return jsonify(payload)


@app.route('/api/insumos/suggest')
@login_required
@feature_required('insumos')
def insumos_suggest():
    """
    Endpoint para autocomplete em buscas de insumos.

    Parâmetros:
    - q: query de busca (obrigatório)
    - field: campo a buscar (descricao, fabricante, tuss, tiss, anvisa) (padrão: descricao)
    - limit: número máximo de resultados (padrão: 10, máx: 20)

    Retorna: Lista de sugestões com até 'limit' itens
    """
    query = (request.args.get('q') or '').strip().lower()
    field = (request.args.get('field') or 'descricao').lower()
    limit = _parse_positive_int(request.args.get('limit'), 10, maximum=20)

    # Validação
    if not query or len(query) < 2:
        return jsonify({'suggestions': []})

    if field not in {'descricao', 'fabricante', 'tuss', 'tiss', 'anvisa'}:
        field = 'descricao'

    try:
        suggestions = []

        field_map = {
            'descricao': InsumoIndex.descricao,
            'fabricante': InsumoIndex.fabricante,
            'tuss': InsumoIndex.tuss,
            'tiss': InsumoIndex.tiss,
            'anvisa': InsumoIndex.anvisa,
        }
        column = field_map.get(field, InsumoIndex.descricao)
        pattern = f'{query}%' if field in {'tuss', 'tiss', 'anvisa'} else f'%{query}%'
        rows = (
            InsumoIndex.query
            .filter(column.isnot(None), func.lower(column).ilike(pattern))
            .with_entities(column)
            .distinct()
            .limit(limit)
            .all()
        )
        suggestions.extend([r[0] for r in rows if r[0]])

        # Limita o resultado final
        suggestions = list(dict.fromkeys(suggestions))[:limit]  # Remove duplicatas mantendo ordem

        return jsonify({
            'suggestions': suggestions,
            'field': field,
            'query': query,
            'count': len(suggestions)
        })

    except Exception as e:
        app.logger.exception(f'Erro ao buscar sugestões para {field}={query}')
        return jsonify({'suggestions': [], 'error': str(e)}), 500


@app.route('/insumos/<origem>/<int:item_id>')
@login_required
@feature_required('insumos')
def insumo_detail(origem: str, item_id: int):
    origem = (origem or '').upper()
    if origem not in {'BRAS', 'SIMPRO'}:
        abort(404)

    if origem == 'BRAS':
        item = BrasItemNormalized.query.get(item_id)
    else:
        item = (
            SimproItemCadastro.query.get(item_id)
            or SimproItemNormalized.query.get(item_id)
            or SimproItem.query.get(item_id)
        )
    if item is None:
        abort(404)

    uf_param = (request.args.get('uf') or request.args.get('uf_referencia') or '').strip().upper()

    aliquota_filter: Decimal | None = None
    aliquota_qs = (request.args.get('aliquota') or '').strip()
    if aliquota_qs:
        aliquota_coerced = _coerce_decimal(aliquota_qs.replace(',', '.'))
        if aliquota_coerced:
            try:
                aliquota_filter = Decimal(aliquota_coerced)
            except (InvalidOperation, ValueError, TypeError):
                aliquota_filter = None

    catalog_entry, historico_rows = _resolve_catalogo_history(
        origem,
        item_id=item_id,
        item=item,
        uf_param=uf_param,
    )

    historico: list[dict[str, object]] = []
    if origem == 'BRAS':
        for row in historico_rows:
            aliquota_value = _aliquota_bp_to_decimal(row.aliquota_bp)
            historico.append({
                'versao': row.periodo,
                'uf': row.uf,
                'preco': _stringify_for_output(
                    row.preco_pmc_unit
                    or row.preco_pmc_pacote
                    or row.preco_pfb_unit
                    or row.preco_pfb_pacote
                ),
                'preco_pmc': _stringify_for_output(row.preco_pmc_unit or row.preco_pmc_pacote),
                'preco_pfb': _stringify_for_output(row.preco_pfb_unit or row.preco_pfb_pacote),
                'aliquota': _stringify_for_output(aliquota_value),
                'data_atualizacao': row.imported_at.strftime('%d/%m/%Y') if isinstance(row.imported_at, datetime) else None,
            })
    else:
        for row in historico_rows:
            aliquota_value = _aliquota_bp_to_decimal(row.aliquota_bp)
            preco_hist, preco_pmc_hist, preco_pfb_hist = _split_simpro_prices(
                row.preco1,
                row.preco2,
                row.preco3,
                row.preco4,
            )
            historico.append({
                'versao': row.periodo,
                'uf': row.uf,
                'preco': _stringify_for_output(preco_hist),
                'preco_pmc': _stringify_for_output(preco_pmc_hist),
                'preco_pfb': _stringify_for_output(preco_pfb_hist),
                'aliquota': _stringify_for_output(aliquota_value),
                'data_atualizacao': row.data_ref.strftime('%d/%m/%Y') if isinstance(row.data_ref, date) else None,
            })

    index_query = InsumoIndex.query.filter_by(origem=origem, item_id=item_id)
    if aliquota_filter is not None:
        index_query = index_query.filter(InsumoIndex.aliquota == aliquota_filter)
    if uf_param:
        like_pattern = f"%|{uf_param}|%"
        index_entry = index_query.filter(
            or_(
                func.upper(InsumoIndex.uf_referencia) == uf_param,
                func.upper(func.coalesce(InsumoIndex.uf_referencia, '')).like(like_pattern)
            )
        ).first()
    else:
        index_entry = None
    if index_entry is None:
        index_entry = index_query.order_by(InsumoIndex.updated_at.desc()).first()

    detail_payload = _serialize_insumo_detail(
            origem,
            item,
            index_entry=index_entry,
            catalog_entry=catalog_entry,
            selected_uf=uf_param or None,
        )
    detail_payload['historico'] = historico
    detail_payload['uf_filtro'] = uf_param or None

    detail_payload['similares'] = _suggest_similar_items(origem, item)

    if session.get('perfil') == 'adm':
        contexto_rows = (
            InsumoContextoClinico.query
            .filter_by(origem=origem, item_id=item_id)
            .order_by(InsumoContextoClinico.procedimento_descricao.asc())
            .all()
        )
        if contexto_rows:
            detail_payload['impacto_clinico'] = [
                _serialize_contexto_clinico(row, detail_payload)
                for row in contexto_rows
            ]

    detail_payload['manual_price'] = None
    if origem == 'SIMPRO' and isinstance(item, SimproItemCadastro):
        tu = (request.args.get('target_uf') or '').strip().upper()
        ta_qs = (request.args.get('target_aliquota') or '').strip()
        ta_val: Decimal | None = None
        if ta_qs:
            cs = _coerce_decimal(ta_qs.replace(',', '.'))
            if cs:
                try:
                    ta_val = _br_norm_aliquota(Decimal(cs)) or Decimal(cs)
                except (InvalidOperation, ValueError, TypeError):
                    ta_val = None
        elif tu:
            ta_val = _simpro_piso_aliquota_for_uf(tu)
        if tu and ta_val is not None:
            tol = Decimal('0.02')
            existing_preco_row = (
                SimproItemPreco.query.filter(
                    SimproItemPreco.cadastro_id == item.id,
                    SimproItemPreco.aliquota >= ta_val - tol,
                    SimproItemPreco.aliquota <= ta_val + tol,
                ).first()
            )
            detail_payload['manual_price'] = {
                'show': True,
                'target_uf': tu,
                'target_aliquota': _decimal_to_string(ta_val),
                'uf_matches_aliquota_map': _simpro_aliquota_includes_uf(ta_val, tu),
                'already_has_price': existing_preco_row is not None,
                'cadastro_id': int(item.id),
            }

    return jsonify(detail_payload)


@app.route('/insumos/simpro/<int:cadastro_id>/preco-manual', methods=['POST'])
@login_required
@feature_required('insumos')
def insumos_simpro_preco_manual(cadastro_id: int):
    """Cria ou atualiza linha em ``simpro_item_preco`` e atualiza ``insumos_index`` (preço manual por UF/alíquota)."""
    cad = SimproItemCadastro.query.get_or_404(cadastro_id)
    payload = request.get_json(silent=True) or {}
    uf_req = (payload.get('uf_referencia') or payload.get('uf') or '').strip().upper()
    al_raw = (payload.get('aliquota') or '').strip()
    if not al_raw and uf_req:
        aq_inf = _simpro_piso_aliquota_for_uf(uf_req)
        if aq_inf is not None:
            al_raw = _decimal_to_string(aq_inf) or str(aq_inf)
    if not al_raw:
        return jsonify({'ok': False, 'error': 'Informe a alíquota (ou uma UF mapeada no piso SIMPRO).'}), 400
    al_cs = _coerce_decimal(al_raw.replace(',', '.'))
    if not al_cs:
        return jsonify({'ok': False, 'error': 'Alíquota inválida.'}), 400
    try:
        al_dec = _br_norm_aliquota(Decimal(al_cs)) or Decimal(al_cs)
    except (InvalidOperation, ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Alíquota inválida.'}), 400

    if uf_req and not _simpro_aliquota_includes_uf(al_dec, uf_req):
        return jsonify({
            'ok': False,
            'error': 'Esta UF não corresponde ao piso da alíquota informada no cadastro SIMPRO.',
        }), 400

    def _parse_money(key_a: str, key_b: str) -> Decimal | None:
        raw = (payload.get(key_a) or payload.get(key_b) or '').strip()
        if not raw:
            return None
        cs = _coerce_decimal(raw.replace(',', '.'))
        if not cs:
            return None
        try:
            return Decimal(cs)
        except (InvalidOperation, ValueError, TypeError):
            return None

    pmc = _parse_money('preco_pmc', 'pmc')
    pfb = _parse_money('preco_pfb', 'pfb')
    if (pmc is None or pmc <= 0) and (pfb is None or pfb <= 0):
        return jsonify({'ok': False, 'error': 'Informe PMC ou PFB maior que zero.'}), 400

    tol = Decimal('0.02')
    row = (
        SimproItemPreco.query.filter(
            SimproItemPreco.cadastro_id == cad.id,
            SimproItemPreco.aliquota >= al_dec - tol,
            SimproItemPreco.aliquota <= al_dec + tol,
        ).first()
    )
    label = f"MANUAL_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    if row is None:
        row = SimproItemPreco(
            cadastro_id=cad.id,
            aliquota=al_dec,
            arquivo_fonte=label,
        )
        db.session.add(row)

    if pfb is not None and pfb > 0:
        row.preco1 = pfb
    if pmc is not None and pmc > 0:
        row.preco2 = pmc
    row.arquivo_fonte = (row.arquivo_fonte or '')[:200] or label
    if row.imported_at is None:
        row.imported_at = datetime.utcnow()

    db.session.flush()
    usuario = getattr(g, 'current_user', None)
    _register_audit(
        'simpro_preco_manual',
        usuario=usuario if isinstance(usuario, Usuario) else None,
        detalhes={
            'cadastro_id': cad.id,
            'aliquota': str(al_dec),
            'uf': uf_req or None,
            'preco_pmc': str(pmc) if pmc is not None else None,
            'preco_pfb': str(pfb) if pfb is not None else None,
            'preco_row_id': int(row.id) if row.id is not None else None,
        },
    )

    db.session.commit()

    _sync_simpro_insumo_index_for_preco_ids([int(row.id)])
    _clear_insumo_cache()

    return jsonify({'ok': True, 'preco_id': int(row.id), 'cadastro_id': int(cad.id)})


@app.route('/insumos/<origem>/<int:item_id>/contexto', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumo_contexto_create(origem: str, item_id: int):
    origem = (origem or '').upper()
    if origem not in {'BRAS', 'SIMPRO'}:
        abort(404)

    if origem == 'BRAS':
        item = BrasItemNormalized.query.get(item_id)
    else:
        item = (
            SimproItemCadastro.query.get(item_id)
            or SimproItemNormalized.query.get(item_id)
            or SimproItem.query.get(item_id)
        )
    if item is None:
        abort(404)

    payload = request.get_json(silent=True) or {}

    def _clean_str(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() or None
        return str(value).strip() or None

    procedimento_codigo = _clean_str(payload.get('procedimento_codigo'))
    procedimento_descricao = _clean_str(payload.get('procedimento_descricao'))
    drg = _clean_str(payload.get('drg'))
    narrativa = _clean_str(payload.get('narrativa'))

    if not procedimento_codigo and not procedimento_descricao:
        return jsonify({'error': 'Informe ao menos o procedimento ou sua descrição.'}), 400

    freq_decimal: Decimal | None = None
    freq_input = payload.get('frequencia_percent')
    if freq_input not in (None, ''):
        freq_str = _coerce_decimal(str(freq_input))
        if freq_str is None:
            return jsonify({'error': 'Frequência inválida.'}), 400
        try:
            freq_decimal = Decimal(freq_str) / Decimal('100')
        except (InvalidOperation, ValueError):
            return jsonify({'error': 'Frequência inválida.'}), 400

    custo_decimal: Decimal | None = None
    custo_input = payload.get('custo_procedimento')
    if custo_input not in (None, ''):
        custo_str = _coerce_decimal(str(custo_input))
        if custo_str is None:
            return jsonify({'error': 'Custo inválido.'}), 400
        try:
            custo_decimal = Decimal(custo_str)
        except (InvalidOperation, ValueError):
            return jsonify({'error': 'Custo inválido.'}), 400

    substitutos_raw = payload.get('substitutos_raw') or payload.get('substitutos')
    if isinstance(substitutos_raw, (list, tuple)):
        substitutos_raw = ', '.join(str(part).strip() for part in substitutos_raw if str(part).strip())
    substitutos_raw = _clean_str(substitutos_raw)

    entry = InsumoContextoClinico(
        origem=origem,
        item_id=item_id,
        drg=drg,
        procedimento_codigo=procedimento_codigo,
        procedimento_descricao=procedimento_descricao,
        frequencia_relativa=freq_decimal,
        custo_procedimento=custo_decimal,
        substitutos_raw=substitutos_raw,
        narrativa=narrativa,
    )

    db.session.add(entry)
    db.session.commit()

    return jsonify({'success': True}), 201


@app.route('/insumos')
@login_required
@feature_required('insumos')
def insumos_dashboard():
    bras_summary = _insumo_summary(BrasItemNormalized)
    bras_versions = _insumo_distinct_versions(BrasItemNormalized)
    if _table_exists('simpro_item_cadastro'):
        simpro_summary = _insumo_summary(SimproItemCadastro)
        simpro_versions = _insumo_distinct_versions(SimproItemCadastro)
    else:
        simpro_summary = _insumo_summary(SimproItemNormalized)
        simpro_versions = _insumo_distinct_versions(SimproItemNormalized)
    versions = sorted(set(bras_versions + simpro_versions))

    # Timeout do fetch na página de insumos (ms). Buscas por termo em base grande podem passar de 15s.
    _to = _safe_int_env('INSUMOS_SEARCH_TIMEOUT_MS', 90000)
    insumos_search_timeout_ms = max(15000, min(300000, _to))

    return render_template(
        'insumos_index.html',
        bras_summary=bras_summary,
        simpro_summary=simpro_summary,
        bras_versions=bras_versions,
        simpro_versions=simpro_versions,
        versions=versions,
        is_admin=(session.get('perfil') == 'adm'),
        UFS=BR_UFS,
        insumos_search_timeout_ms=insumos_search_timeout_ms,
    )


def _set_job_metrics(job: ImportJob, metrics: dict | None) -> None:
    params = dict(job.params or {})
    if metrics:
        params['_metrics'] = metrics
    else:
        params.pop('_metrics', None)
    job.params = params


def _serialize_import_job(job: ImportJob) -> dict:
    def _fmt_dt(value):
        return value.isoformat(sep=' ') if isinstance(value, datetime) else None

    uf_values = [uf.strip() for uf in (job.uf_list or '').split(',') if uf and uf.strip()]
    raw_metrics = {}
    if isinstance(job.params, dict):
        raw_metrics = job.params.get('_metrics') or {}
    return {
        'id': job.id,
        'origem': job.origem,
        'arquivo': job.original_filename,
        'status': job.status,
        'message': job.message,
        'versao': job.versao,
        'aliquota': _decimal_to_string(job.aliquota) if job.aliquota is not None else None,
        'uf_list': uf_values,
        'total_linhas': job.total_linhas,
        'linhas_materializadas': job.linhas_materializadas,
        'created_at': _fmt_dt(job.created_at),
        'started_at': _fmt_dt(job.started_at),
        'finished_at': _fmt_dt(job.finished_at),
        'metrics': raw_metrics,
    }


def _run_import_job(job_id: str) -> None:
    file_path: Path | None = None
    with app.app_context():
        job: ImportJob | None = ImportJob.query.get(job_id)
        if job is None:
            app.logger.warning('Import job %s não encontrado.', job_id)
            return

        _set_current_import_job(job_id)
        _raise_if_import_paused(job_id)
        job.status = ImportJobStatus.RUNNING.value
        job.started_at = datetime.utcnow()
        job.message = _job_message_trim('Processando arquivo de importação...')
        db.session.commit()

        params = job.params or {}
        file_path = Path(job.data_path)
        if not file_path.exists():
            job.status = ImportJobStatus.FAILED.value
            job.message = _job_message_trim('Arquivo da importação não encontrado no servidor.')
            job.finished_at = datetime.utcnow()
            db.session.commit()
            return

        origem = job.origem
        uf_values = params.get('uf_values') or []
        uf_default = params.get('uf_default')
        versao = params.get('versao') or ''
        data_ref = params.get('data_ref') or None
        fmt = params.get('fmt') or 'delimited'
        delimiter = params.get('delimiter')
        quotechar = params.get('quotechar')
        lines_terminated = params.get('lines_terminated') or '\n'
        skip_header = bool(params.get('skip_header'))
        encoding = params.get('encoding') or None
        truncate = bool(params.get('truncate'))
        keep_only_latest_raw = params.get('keep_only_latest_version')
        if origem == 'BRAS':
            if isinstance(keep_only_latest_raw, str):
                keep_only_latest_version = keep_only_latest_raw.strip().lower() in {'on', '1', 'true', 'yes'}
            elif keep_only_latest_raw is None:
                keep_only_latest_version = True
            else:
                keep_only_latest_version = bool(keep_only_latest_raw)
        else:
            keep_only_latest_version = False
        map_config = params.get('map_config') or {}
        sequencia_input = params.get('sequencia_input')
        aliquota_raw = params.get('aliquota')
        arquivo_label_override = params.get('arquivo_label')
        aliquota_decimal: Decimal | None = None
        if aliquota_raw not in (None, ''):
            try:
                aliquota_decimal = Decimal(str(aliquota_raw))
            except (InvalidOperation, ValueError) as exc:
                app.logger.warning('Alíquota inválida para job %s (%s).', job_id, exc)

        metrics: dict[str, object] = {
            'timings': {},
            'context': {
                'origem': origem,
                'versao': versao,
                'uf_values': uf_values,
            },
        }
        if arquivo_label_override:
            metrics['context']['arquivo_label'] = arquivo_label_override
        overall_start = time.perf_counter()

        is_bras_delta = origem == 'BRAS' and (params or {}).get('import_kind') == 'bras_delta'
        is_bras_import_precos = origem == 'BRAS' and (params or {}).get('import_kind') == 'bras_import_precos'
        try:
            if is_bras_delta:
                if truncate:
                    try:
                        deleted_n = BrasItemNormalized.query.delete()
                        deleted_raw = BrasRaw.query.delete()
                        db.session.commit()
                        app.logger.info(
                            'Truncate Brasíndice (job delta): %d normalizados, %d raw removidos',
                            deleted_n,
                            deleted_raw,
                        )
                    except Exception as exc:  # noqa: BLE001
                        db.session.rollback()
                        app.logger.warning('Erro ao truncar antes do delta: %s', exc)
                stage_start = time.perf_counter()
                cpath = params.get('catalog_data_path')
                catalog_p: Path | None = None
                if cpath and Path(cpath).exists():
                    catalog_p = Path(cpath)
                d_del = _normalize_delimiter(str(delimiter) if delimiter is not None else ',')
                qchar = (quotechar if quotechar not in (None, '') else '"')
                if isinstance(qchar, str) and not qchar.strip():
                    qchar = '"'
                res_d = _import_bras_delta(
                    file_path=file_path,
                    versao=versao,
                    delimiter=d_del,
                    quotechar=qchar,
                    encoding=encoding,
                    skip_header=skip_header,
                    data_ref=data_ref,
                    uf_default=uf_default,
                    uf_values=uf_values if uf_values else None,
                    aliquota_default=aliquota_decimal,
                    catalog_file=catalog_p,
                    catalog_encoding=(params.get('catalog_encoding') or 'latin-1'),
                    catalog_delimiter=_normalize_delimiter(str(params.get('catalog_delimiter') or ';')),
                    previous_catalog_version=params.get('previous_catalog_version') or None,
                )
                metrics['timings']['import_stage'] = round(time.perf_counter() - stage_start, 4)
                metrics['delta_brasindice'] = {
                    'versao': res_d.get('versao'),
                    'novos': res_d.get('novos'),
                    'novos_importados': res_d.get('novos_importados'),
                    'alterados': res_d.get('alterados'),
                    'total_processado': res_d.get('total_processado'),
                    'catalog_current_version': res_d.get('catalog_current_version'),
                    'catalog_previous_version': res_d.get('catalog_previous_version'),
                    'catalog_new': res_d.get('catalog_new'),
                    'catalog_changed': res_d.get('catalog_changed'),
                    'catalog_removed': res_d.get('catalog_removed'),
                }
                cat_extra = ''
                if res_d.get('catalog_current_version') is not None:
                    cat_extra = (
                        f" Catálogo {res_d['catalog_current_version']}: +{res_d.get('catalog_new', 0)}/~"
                        f"{res_d.get('catalog_changed', 0)}/-{res_d.get('catalog_removed', 0)}."
                    )
                job.status = ImportJobStatus.SUCCESS.value
                job.message = _job_message_trim(
                    f"Delta Brasíndice concluído: {res_d['novos']} novos, {res_d['alterados']} preços alterados, "
                    f"{res_d.get('novos_importados', 0)} materializados.{cat_extra}"
                )
                job.total_linhas = res_d.get('total_processado')
                job.linhas_materializadas = res_d.get('novos_importados', 0)
                job.finished_at = datetime.utcnow()
                job.versao = versao
                job.uf_list = ', '.join(uf_values)
                if aliquota_decimal is not None:
                    job.aliquota = aliquota_decimal
                _set_job_metrics(job, metrics)
                db.session.commit()
                _clear_insumo_cache()
            elif is_bras_import_precos:
                if aliquota_decimal is None:
                    raise ValueError('Alíquota é obrigatória para importação de somente preços.')
                ed = (versao or '').strip()
                if not ed:
                    raise ValueError('Informe a versão/edição (campo versão) para o arquivo de preços.')
                d_del = _normalize_delimiter(str(delimiter) if delimiter is not None else ',')
                qchar = quotechar if quotechar not in (None, '') else '"'
                if isinstance(qchar, str) and not qchar.strip():
                    qchar = '"'
                stage_start = time.perf_counter()
                st_p = _import_bras_somente_precos(
                    file_path=file_path,
                    edicao=ed,
                    aliquota=aliquota_decimal,
                    delimiter=d_del,
                    quotechar=str(qchar),
                    encoding=encoding,
                    skip_header=skip_header,
                    arquivo_fonte=Path(job.original_filename or 'import-precos').name,
                    update_legacy=not bool(params.get('no_legacy')),
                )
                metrics['timings']['import_stage'] = round(time.perf_counter() - stage_start, 4)
                metrics['bras_somente_precos'] = st_p
                job.status = ImportJobStatus.SUCCESS.value
                abm = st_p.get('autobackfill_cadastro') or {}
                ab_text = ''
                if abm:
                    ab_text = (
                        f" Cadastro vazio: backfill executado (+{abm.get('cadastros_unicos', 0)}; "
                        f"linhas n sem alíquota, só identidade: {abm.get('n_sem_aliquota_so_cadastro', 0)})."
                    )
                v_idx = st_p.get('insumos_index_vinculados', 0)
                idx_m = f' índice vinculado: {v_idx}.' if v_idx else ''
                job.message = _job_message_trim(
                    f"Preços atualizados: {st_p.get('atualizados_preco', 0)} linhas; sem cadastro: {st_p.get('sem_cadastro', 0)}; "
                    f"erros: {st_p.get('erros', 0)}; legacy: {st_p.get('legacy_atualizados', 0)}.{ab_text}{idx_m}"
                )
                job.total_linhas = st_p.get('linhas_lidas')
                job.linhas_materializadas = st_p.get('atualizados_preco')
                job.finished_at = datetime.utcnow()
                job.versao = ed
                job.aliquota = aliquota_decimal
                _set_job_metrics(job, metrics)
                db.session.commit()
                _clear_insumo_cache()
            elif origem == 'BRAS':
                stage_start = time.perf_counter()
                result = _import_bras(
                    file_path=file_path,
                    versao=versao,
                    data_ref=data_ref,
                    fmt=fmt,
                    delimiter=delimiter,
                    quotechar=quotechar,
                    line_terminator=lines_terminated,
                    skip_header=skip_header,
                    encoding=encoding,
                    map_config=map_config,
                    truncate=truncate,
                    uf_default=uf_default,
                    uf_values=uf_values,
                    aliquota_default=aliquota_decimal,
                    arquivo_label_override=arquivo_label_override,
                    keep_only_latest_version=keep_only_latest_version,
                )
                metrics['timings']['import_stage'] = round(time.perf_counter() - stage_start, 4)
            else:
                # CORREÇÃO: SIMPRO importa uma vez só, mesmo com múltiplas UFs
                # As UFs selecionadas são armazenadas no campo uf_referencia codificado
                # Exemplo: importar para AP,MG,SP gera 3.000 linhas (não 9.000)
                stage_start = time.perf_counter()
                base_label = arquivo_label_override or (Path(job.original_filename).stem if job.original_filename else None)
                if not base_label:
                    base_label = versao or 'simpro'

                target_ufs = list(dict.fromkeys([*(uf_values or []), *( [uf_default] if uf_default else [] )]))

                # Importa arquivo UMA VEZ, passando todas as UFs
                result = _import_simpro(
                    file_path=file_path,
                    versao=versao,
                    fmt=fmt,
                    map_config=map_config,
                    encoding=encoding,
                    truncate=truncate,
                    uf_default=uf_default,
                    uf_values=target_ufs if target_ufs else None,
                    aliquota_default=aliquota_decimal,
                    arquivo_label_override=base_label,
                    job_id=job_id,
                )

                metrics['timings']['import_stage'] = round(time.perf_counter() - stage_start, 4)

            if not is_bras_delta and not is_bras_import_precos:
                job.status = ImportJobStatus.SUCCESS.value
                job.message = _job_message_trim(
                    f"Importação concluída (arquivo {result['arquivo']} | {result['linhas_raw']} linhas brutas, "
                    f"{result['linhas_materializadas']} materializadas)."
                )
                purge_summary = result.get('purge_summary') or {}
                if origem == 'BRAS' and purge_summary:
                    removed_total = sum(int(v or 0) for v in purge_summary.values())
                    job.message = _job_message_trim(
                        f"{job.message} Versões antigas removidas: {removed_total} registros."
                    )
                    metrics['purge_summary'] = purge_summary
                job.total_linhas = result.get('linhas_raw')
                job.linhas_materializadas = result.get('linhas_materializadas')
                job.finished_at = datetime.utcnow()
                job.versao = versao
                job.uf_list = ', '.join(uf_values)
                if aliquota_decimal is not None:
                    job.aliquota = aliquota_decimal
                metrics['rows'] = {
                    'linhas_raw': result.get('linhas_raw'),
                    'linhas_materializadas': result.get('linhas_materializadas'),
                }
                if result.get('load_strategy'):
                    metrics['load_strategy'] = result.get('load_strategy')
                _set_job_metrics(job, metrics)
                db.session.commit()

                # Limpa o cache de insumos após importação bem-sucedida
                _clear_insumo_cache()

                try:
                    post_start = time.perf_counter()
                    # Consolida catálogo após importação
                    _post_catalog_ingest(
                        origem=origem,
                        arquivo_label=result.get('arquivo'),
                        versao=versao,
                        sequencia_input=sequencia_input,
                        aliquota_value=aliquota_decimal,
                        uf_values=uf_values,
                    )
                    metrics['timings']['post_catalog'] = round(time.perf_counter() - post_start, 4)
                except Exception as exc:  # noqa: BLE001
                    app.logger.warning('Falha ao consolidar catálogo pós-import (job %s): %s', job_id, exc)
                    db.session.rollback()  # Rollback antes de reutilizar a sessão
                    job = ImportJob.query.get(job_id)
                    if job:
                        job.message = _job_message_trim((job.message or '') + ' Consolidação posterior falhou.')
                        _set_job_metrics(job, metrics)
                        db.session.commit()
        except ImportPauseRequested as exc:
            metrics['timings']['total'] = round(time.perf_counter() - overall_start, 4)
            metrics['error'] = str(exc)
            metrics['paused'] = True
            db.session.rollback()
            job = ImportJob.query.get(job_id)
            if job:
                job.status = ImportJobStatus.FAILED.value
                job.message = _job_message_trim(str(exc))
                job.finished_at = datetime.utcnow()
                _set_job_metrics(job, metrics)
                db.session.commit()
            app.logger.info('Import job %s pausado por solicitação do usuário.', job_id)
        except Exception as exc:  # noqa: BLE001
            metrics['timings']['total'] = round(time.perf_counter() - overall_start, 4)
            metrics['error'] = str(exc)
            db.session.rollback()
            job = ImportJob.query.get(job_id)
            if job:
                job.status = ImportJobStatus.FAILED.value
                job.message = _job_message_trim(str(exc))
                job.finished_at = datetime.utcnow()
                _set_job_metrics(job, metrics)
                db.session.commit()
            app.logger.exception('Falha ao executar job de importação %s', job_id, exc_info=exc)
        else:
            metrics['timings']['total'] = round(time.perf_counter() - overall_start, 4)
            job = ImportJob.query.get(job_id)
            if job:
                _set_job_metrics(job, metrics)
                db.session.commit()
                if job.status == ImportJobStatus.SUCCESS.value:
                    _schedule_post_import_analyze_tables()
            app.logger.info(
                'Import job %s concluído em %.2fs (linhas=%s, estratégia=%s).',
                job_id,
                metrics['timings'].get('total', 0.0),
                metrics.get('rows'),
                metrics.get('load_strategy'),
            )
        finally:
            try:
                if file_path is not None:
                    file_path.unlink(missing_ok=True)
            except Exception as exc:  # noqa: BLE001
                app.logger.warning('Não foi possível remover arquivo temporário %s (%s)', file_path, exc)
            try:
                if file_path is not None:
                    shutil.rmtree(file_path.parent, ignore_errors=True)
            except Exception as exc:  # noqa: BLE001
                app.logger.debug('Falha ao limpar diretório temporário %s (%s)', file_path.parent, exc)
            _clear_import_pause_request(job_id)
            _set_current_import_job(None)
            db.session.remove()


def _run_import_worker_loop(*, poll_interval: int, run_once: bool = False) -> None:
    poll_interval = max(1, poll_interval)
    while True:
        try:
            job = (
                db.session.query(ImportJob)
                .filter(ImportJob.status == ImportJobStatus.PENDING.value)
                .order_by(ImportJob.created_at.asc())
                .with_for_update(skip_locked=True)
                .first()
            )
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            app.logger.exception('Falha ao consultar jobs pendentes', exc_info=exc)
            if run_once:
                db.session.remove()
                break
            db.session.remove()
            time.sleep(poll_interval)
            continue

        if not job:
            db.session.commit()
            if run_once:
                db.session.remove()
                break
            db.session.remove()
            time.sleep(poll_interval)
            continue

        job_id = job.id
        db.session.commit()
        _run_import_job(job_id)
        db.session.remove()

        if run_once:
            break

    db.session.remove()


def _spawn_async_import(job_id: str) -> None:
    disable_env = (os.getenv('INSUMO_IMPORT_BACKGROUND_DISABLE') or '').strip().lower()
    if disable_env in {'1', 'true', 'yes', 'on'}:
        return

    def _runner():
        try:
            _run_import_job(job_id)
        except Exception as exc:  # noqa: BLE001
            app.logger.exception('Falha no processamento assíncrono do job %s', job_id, exc_info=exc)

    thread = threading.Thread(target=_runner, name=f'ImportJob-{job_id[:8]}', daemon=True)
    thread.start()



@app.route('/insumos/aliquotas', methods=['GET', 'POST'])
@admin_required
@feature_required('insumos')
def insumos_aliquotas():
    highlight_uf = (request.args.get('highlight') or '').strip().upper()
    today = date.today()

    if request.method == 'POST':
        target_uf = (request.form.get('target_uf') or '').strip().upper()
        if target_uf not in BR_UFS:
            flash('Selecione uma UF válida para atualizar.', 'danger')
            return redirect(url_for('insumos_aliquotas'))

        aliquota_raw = (request.form.get(f'aliquota_{target_uf}') or '').strip()
        aliquota_str = _coerce_decimal(aliquota_raw) if aliquota_raw else None
        if aliquota_str is None:
            flash('Informe uma alíquota válida (use números, ponto ou vírgula).', 'danger')
            return redirect(url_for('insumos_aliquotas', highlight=target_uf))

        aliquota_decimal = Decimal(aliquota_str)
        if aliquota_decimal < 0:
            flash('Alíquota não pode ser negativa.', 'danger')
            return redirect(url_for('insumos_aliquotas', highlight=target_uf))

        valid_from_raw = (request.form.get(f'valid_from_{target_uf}') or '').strip()
        valid_from_value = _coerce_date(valid_from_raw)
        if valid_from_value is None:
            valid_from_value = today

        basis_points = int((aliquota_decimal * Decimal('100')).to_integral_value(rounding=ROUND_HALF_UP))
        now_ts = datetime.utcnow()

        try:
            current_record = (
                UfAliquota.query
                .filter_by(uf=target_uf)
                .order_by(UfAliquota.valid_from.desc())
                .first()
            )

            if current_record and current_record.is_current and valid_from_value <= current_record.valid_from:
                current_record.aliquota_bp = basis_points
                current_record.valid_from = valid_from_value
                current_record.valid_to = None
                current_record.is_current = True
                current_record.updated_at = now_ts
            else:
                if current_record and current_record.is_current:
                    current_record.is_current = False
                    current_record.valid_to = valid_from_value - timedelta(days=1)
                    current_record.updated_at = now_ts

                new_record = UfAliquota(
                    uf=target_uf,
                    valid_from=valid_from_value,
                    valid_to=None,
                    aliquota_bp=basis_points,
                    is_current=True,
                    created_at=now_ts,
                    updated_at=now_ts,
                )
                db.session.add(new_record)

            db.session.commit()
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            app.logger.exception('Falha ao ajustar alíquota para %s', target_uf, exc_info=exc)
            flash('Falha ao ajustar a alíquota. Verifique os logs.', 'danger')
            return redirect(url_for('insumos_aliquotas', highlight=target_uf))

        refresh_failures: list[str] = []
        for fornecedor in ('BRASINDICE', 'SIMPRO'):
            try:
                _refresh_materialized_catalogs(fornecedor)
            except Exception as exc:  # noqa: BLE001
                app.logger.warning('Falha ao atualizar view %s após ajuste %s (%s)', fornecedor, target_uf, exc)
                refresh_failures.append(fornecedor)

        if refresh_failures:
            flash(
                'Alíquota atualizada, mas não foi possível atualizar as views: '
                + ', '.join(sorted(refresh_failures)),
                'warning'
            )
        else:
            flash(
                f'Alíquota de {target_uf} atualizada para {aliquota_decimal}% '
                f'a partir de {valid_from_value.strftime("%d/%m/%Y")}.',
                'success'
            )

        return redirect(url_for('insumos_aliquotas', highlight=target_uf))

    entries: list[dict[str, object]] = []
    today_iso = today.isoformat()
    for uf in BR_UFS:
        record = (
            UfAliquota.query
            .filter_by(uf=uf)
            .order_by(UfAliquota.valid_from.desc())
            .first()
        )
        percent_display = None
        if record and record.aliquota_bp is not None:
            percent_display = _decimal_to_string(_aliquota_bp_to_decimal(record.aliquota_bp))
        entries.append({
            'uf': uf,
            'record': record,
            'percent_display': percent_display,
            'form_value': percent_display or '',
            'default_valid_from': today_iso,
        })

    history = (
        UfAliquota.query
        .order_by(UfAliquota.updated_at.desc())
        .limit(25)
        .all()
    )

    history_rows: list[dict[str, object]] = []
    for row in history:
        percent_display = None
        if row.aliquota_bp is not None:
            percent_display = _decimal_to_string(_aliquota_bp_to_decimal(row.aliquota_bp))
        history_rows.append({
            'uf': row.uf,
            'percent_display': percent_display,
            'valid_from': row.valid_from,
            'valid_to': row.valid_to,
            'is_current': row.is_current,
            'updated_at': row.updated_at,
        })

    return render_template(
        'insumos_aliquotas.html',
        entries=entries,
        history_rows=history_rows,
        highlight_uf=highlight_uf,
    )


@app.route('/insumos/export/xlsx')
@login_required
@feature_required('insumos')
def insumos_export_xlsx():
    filters = _extract_insumo_filters(request.args)
    limit = _parse_positive_int(request.args.get('limit'), 5000, maximum=20000)
    items = _catalogo_fetch_all(filters, limit)
    items.sort(key=lambda entry: (
        (entry.get('descricao') or '').lower(),
        entry.get('origem') or '',
        entry.get('item_id') or 0,
    ))

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Insumos')

    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#0EA5E9', 'font_color': '#ffffff'})
    money_fmt = workbook.add_format({'num_format': '#,##0.0000'})

    headers = ['Origem', 'TUSS', 'TISS', 'ANVISA', 'Descrição', 'PMC', 'PFB', 'Alíquota', 'Fabricante', 'UF', 'Versão', 'Data Atualização', 'Atualizado em']
    for col, title in enumerate(headers):
        worksheet.write(0, col, title, header_fmt)

    def _to_float(value):
        if value in (None, ''):
            return None
        try:
            return float(Decimal(str(value)))
        except (InvalidOperation, ValueError):
            return None

    for row_idx, item in enumerate(items, start=1):
        worksheet.write(row_idx, 0, item.get('origem') or '')
        worksheet.write(row_idx, 1, item.get('tuss') or '')
        worksheet.write(row_idx, 2, item.get('tiss') or '')
        worksheet.write(row_idx, 3, item.get('anvisa') or '')
        worksheet.write(row_idx, 4, item.get('descricao') or '')

        preco_pmc = _to_float(item.get('preco_pmc'))
        preco_pfb = _to_float(item.get('preco_pfb'))
        aliquota = _to_float(item.get('aliquota'))

        if preco_pmc is not None:
            worksheet.write_number(row_idx, 5, preco_pmc, money_fmt)
        else:
            worksheet.write_blank(row_idx, 5, None)

        if preco_pfb is not None:
            worksheet.write_number(row_idx, 6, preco_pfb, money_fmt)
        else:
            worksheet.write_blank(row_idx, 6, None)

        if aliquota is not None:
            worksheet.write_number(row_idx, 7, aliquota, money_fmt)
        else:
            worksheet.write_blank(row_idx, 7, None)

        worksheet.write(row_idx, 8, item.get('fabricante') or '')
        worksheet.write(row_idx, 9, item.get('uf_referencia') or '')
        worksheet.write(row_idx, 10, item.get('versao_tabela') or '')
        worksheet.write(row_idx, 11, item.get('data_atualizacao') or '')
        worksheet.write(row_idx, 12, item.get('updated_at') or '')

    worksheet.autofilter(0, 0, max(len(items), 1), len(headers) - 1)
    worksheet.freeze_panes(1, 0)

    workbook.close()
    output.seek(0)
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'insumos_{stamp}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )

@app.route('/insumos/import', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumos_import():
    return_to = (request.form.get('return_to') or '').strip()
    redirect_endpoint = 'gerenciar_tabelas' if return_to == 'gerenciar_tabelas' else 'insumos_dashboard'
    def _go_back():
        return redirect(url_for(redirect_endpoint))

    is_ajax = request.headers.get('X-Requested-With', '').lower() == 'xmlhttprequest'

    def _fail(message: str, category: str = 'danger', status_code: int = 400):
        if is_ajax:
            return jsonify({
                'status': 'error',
                'message': message,
                'redirect': url_for(redirect_endpoint),
                'inline': False,
            }), status_code
        _safe_flash(message, category)
        return _go_back()

    origem = (request.form.get('origem') or '').upper()
    if origem not in {'BRAS', 'SIMPRO'}:
        return _fail('Origem inválida para importação.')

    upload = request.files.get('arquivo')
    if not upload or not upload.filename:
        return _fail('Selecione um arquivo TXT/CSV para importar.')

    fmt = (request.form.get('format') or 'delimited').lower()
    delimiter = request.form.get('delimiter') or ';'
    quotechar = request.form.get('quotechar') or '"'
    versao = (request.form.get('versao') or '').strip()
    data_ref = (request.form.get('data_atualizacao') or '').strip() or None
    no_header = request.form.get('no_header') == 'on'
    truncate = request.form.get('truncate') == 'on'
    keep_only_latest_raw = (request.form.get('keep_only_latest_version') or '').strip().lower()
    keep_only_latest_version = True if origem == 'BRAS' and keep_only_latest_raw == '' else (
        keep_only_latest_raw in {'on', '1', 'true', 'yes'}
    )
    encoding = (request.form.get('encoding') or '').strip() or None
    arquivo_label_override_raw = (request.form.get('arquivo_label') or '').strip()
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values: list[str] = []
    seen_ufs: set[str] = set()
    for raw in raw_ufs:
        candidate = (raw or '').strip().upper()
        if not candidate:
            continue
        if candidate not in BR_UFS:
            return _fail(f'UF inválida informada: {candidate}')
        if candidate not in seen_ufs:
            uf_values.append(candidate)
            seen_ufs.add(candidate)
    uf_value = uf_values[0] if uf_values else None
    aliquota_input = (request.form.get('aliquota') or '').strip() or None
    sequencia_input = (request.form.get('sequencia') or '').strip() or None
    aliquota_value: Decimal | None = None
    if aliquota_input:
        aliquota_str = _coerce_decimal(aliquota_input)
        if aliquota_str is None:
            return _fail('Informe uma alíquota válida (use números, ponto ou vírgula).')
        aliquota_value = Decimal(aliquota_str)

    if not versao:
        return _fail('Informe a versão de referência da tabela.')

    map_upload = request.files.get('map_config')
    map_config: dict = {}
    if map_upload and map_upload.filename:
        try:
            raw_bytes = map_upload.read()
            text = None
            for codec in ('utf-8', 'latin-1', 'cp1252'):
                try:
                    text = raw_bytes.decode(codec)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                raise UnicodeDecodeError('map', b'', 0, 1, 'não foi possível decodificar o arquivo de mapeamento')
            text = text.strip()
            map_config = _load_json_relaxed(text) if text else {}
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            return _fail(f'Erro ao ler o mapa: {exc}')
        if not isinstance(map_config, dict):
            return _fail('Arquivo de mapeamento deve conter um objeto JSON.')

    lines_terminated = request.form.get('lines_terminated') or '\n'
    if quotechar is not None and not str(quotechar).strip():
        quotechar = None
    skip_header = not no_header

    if origem == 'BRAS':
        if fmt == 'fixed' and not map_config.get('columns'):
            return _fail('Envie um arquivo de mapeamento contendo "columns" para largura fixa.')

        line_cfg = map_config.get('lines_terminated') or map_config.get('line_terminator')
        if line_cfg:
            lines_terminated = line_cfg

        encoding_cfg = map_config.get('encoding')
        if isinstance(encoding_cfg, str) and encoding_cfg.strip():
            encoding = encoding_cfg.strip()

        skip_header_cfg = map_config.get('skip_header') if 'skip_header' in map_config else None
        if skip_header_cfg is not None:
            skip_header = bool(skip_header_cfg)

        if fmt == 'delimited':
            delimiter_cfg = map_config.get('delimiter')
            if delimiter_cfg:
                delimiter = delimiter_cfg
            quote_cfg = map_config.get('quotechar')
            if quote_cfg is not None:
                quotechar = quote_cfg
            if quotechar is not None and not str(quotechar).strip():
                quotechar = None
    else:
        if fmt not in {'fixed', 'json'}:
            return _fail('Importação SIMPRO suporta apenas JSON ou largura fixa.')
        if fmt == 'fixed' and not map_config:
            # Tenta carregar mapa padrão do SIMPRO
            default_map_paths = [
                Path(__file__).parent / 'testes' / 'mapa.json',
                Path(__file__).parent / 'config' / 'simpro_map.json',
                Path(__file__).parent / 'testes' / 'config' / 'simpro_map.json',
            ]
            for default_map_path in default_map_paths:
                if default_map_path.exists():
                    try:
                        map_text = default_map_path.read_text(encoding='utf-8')
                        map_config = _load_json_relaxed(map_text)
                        app.logger.info(f'Usando mapa padrão SIMPRO: {default_map_path}')
                        break
                    except (json.JSONDecodeError, ValueError) as exc:
                        app.logger.warning(f'Falha ao carregar mapa padrão {default_map_path}: {exc}')
            if not map_config:
                return _fail('Envie um arquivo de mapeamento ou configure o mapa padrão em testes/mapa.json.')

    delimiter_value = _normalize_delimiter(delimiter) if fmt == 'delimited' else delimiter

    job_id = uuid4().hex
    job_dir = INSUMO_IMPORT_ASYNC_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    original_name = secure_filename(upload.filename) or f'{origem.lower()}_{job_id}'
    if arquivo_label_override_raw:
        arquivo_label_override = secure_filename(arquivo_label_override_raw)
    else:
        arquivo_label_override = Path(original_name).stem or None
    data_path = job_dir / original_name

    try:
        upload.stream.seek(0)
        upload.save(data_path)
    except Exception as exc:  # noqa: BLE001
        message = f'Falha ao salvar o arquivo de importação: {exc}'
        shutil.rmtree(job_dir, ignore_errors=True)
        return _fail(message)

    params_payload = {
        'fmt': fmt,
        'delimiter': delimiter_value,
        'quotechar': quotechar,
        'lines_terminated': lines_terminated,
        'skip_header': skip_header,
        'encoding': encoding,
        'versao': versao,
        'data_ref': data_ref,
        'truncate': truncate,
        'keep_only_latest_version': keep_only_latest_version,
        'uf_values': uf_values,
        'uf_default': uf_value,
        'aliquota': str(aliquota_value) if aliquota_value is not None else None,
        'sequencia_input': sequencia_input,
        'map_config': map_config,
        'arquivo_label': arquivo_label_override,
    }

    job = ImportJob(
        id=job_id,
        origem=origem,
        original_filename=upload.filename,
        data_path=str(data_path),
        status=ImportJobStatus.PENDING.value,
        versao=versao,
        aliquota=aliquota_value,
        uf_list=', '.join(uf_values),
        params=params_payload,
        message='Aguardando processamento.',
    )

    try:
        db.session.add(job)
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        message = f'Falha ao registrar a importação: {exc}'
        shutil.rmtree(job_dir, ignore_errors=True)
        return _fail(message)

    job_id = job.id
    app.logger.info(
        'Import job criado: id=%s origem=%s arquivo=%s versao=%s uf=%s',
        job_id,
        origem,
        upload.filename,
        versao,
        ','.join(uf_values) or uf_value or '-',
    )
    inline_env = (os.getenv('INSUMO_IMPORT_RUN_INLINE') or '').strip().lower()
    inline = inline_env in {'1', 'true', 'yes', 'on'}
    redirect_url = url_for(redirect_endpoint)
    prefix = job_id[:8]

    if inline:
        _run_import_job(job_id)
        job = ImportJob.query.get(job_id)
        success = bool(job and job.status == ImportJobStatus.SUCCESS.value)
        if success:
            linhas = job.linhas_materializadas or job.total_linhas or 0
            linhas_txt = f'{linhas} linha{"s" if linhas != 1 else ""}' if linhas else 'Sem linhas materializadas'
            message = f'Importação {origem} concluída (protocolo {prefix}). {linhas_txt}.'
            if not is_ajax:
                _safe_flash(message, 'success')
                return _go_back()
            return jsonify({'status': 'ok', 'job_id': job_id, 'prefix': prefix, 'message': message, 'inline': True, 'redirect': redirect_url})

        error_message = (job.message if job else None) or 'Erro não informado.'
        if not is_ajax:
            _safe_flash(f'Importação {origem} falhou (protocolo {prefix}). {error_message}', 'danger')
            return _go_back()
        return jsonify({
            'status': 'error',
            'job_id': job_id,
            'prefix': prefix,
            'message': f'Importação {origem} falhou (protocolo {prefix}). {error_message}',
            'inline': True,
            'redirect': redirect_url,
        }), 500

    _spawn_async_import(job_id)
    if origem == 'SIMPRO':
        split_meta: list[str] = []
        if versao:
            split_meta.append(f'versão {versao}')
        if aliquota_value is not None:
            split_meta.append(f'alíquota {str(aliquota_value)}%')
        split_txt = f" ({', '.join(split_meta)})" if split_meta else ''
        info_message = (
            f'Importação {origem} (split por alíquota){split_txt} agendada em segundo plano '
            f'(protocolo {prefix}). Acompanhe o status em "Importações em andamento".'
        )
    else:
        info_message = (
            f'Importação {origem} agendada em segundo plano (protocolo {prefix}). '
            'Acompanhe o status em "Importações em andamento".'
        )
    if not is_ajax:
        _safe_flash(info_message, 'info')
        return _go_back()
    return jsonify({'status': 'ok', 'job_id': job_id, 'prefix': prefix, 'message': info_message, 'inline': False, 'redirect': redirect_url})


@app.route('/insumos/bras/analyze-delta', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumos_bras_analyze_delta():
    """Analisa diferenças entre arquivo Brasíndice e dados existentes (preview antes de importar)."""
    upload = request.files.get('arquivo')
    if not upload or not upload.filename:
        return jsonify({'status': 'error', 'message': 'Selecione um arquivo para analisar.'}), 400
    
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    delimiter = request.form.get('delimiter') or ','
    encoding = (request.form.get('encoding') or 'latin-1').strip()
    catalog_upload = request.files.get('catalogo')
    versao = (request.form.get('versao') or '').strip()
    catalog_encoding = (request.form.get('catalog_encoding') or 'latin-1').strip()
    catalog_delimiter = request.form.get('catalog_delimiter') or ';'
    catalog_prev_versao = (request.form.get('catalog_prev_versao') or '').strip() or None
    
    # Salvar arquivo temporariamente
    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.txt') as tmp:
        upload.save(tmp.name)
        tmp_path = Path(tmp.name)
    catalog_tmp_path: Path | None = None
    if catalog_upload and catalog_upload.filename:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.txt') as tmp_catalog:
            catalog_upload.save(tmp_catalog.name)
            catalog_tmp_path = Path(tmp_catalog.name)
    
    try:
        result = _analyze_bras_delta(
            file_path=tmp_path,
            delimiter=_normalize_delimiter(delimiter),
            encoding=encoding,
        )
        catalog_result = None
        if catalog_tmp_path:
            if not versao:
                return jsonify({'status': 'error', 'message': 'Informe a versão para analisar o catálogo Brasíndice.'}), 400
            _sync_bras_catalog_snapshot(
                file_path=catalog_tmp_path,
                versao=versao,
                delimiter=_normalize_delimiter(catalog_delimiter),
                encoding=catalog_encoding,
            )
            catalog_result = _analyze_bras_catalog_delta(
                current_version=versao,
                previous_version=catalog_prev_versao,
            )
        return jsonify({
            'status': 'ok',
            'total_arquivo': result['total_arquivo'],
            'total_existente': result['total_existente'],
            'novos': result['novos'],
            'alterados': result['alterados'],
            'inalterados': result['inalterados'],
            'removidos': result['removidos'],
            'detalhes_novos': result['detalhes_novos'][:20],
            'detalhes_alterados': result['detalhes_alterados'][:20],
            'catalog': ({
                'current_version': catalog_result['current_version'],
                'previous_version': catalog_result['previous_version'],
                'current_total': catalog_result['current_total'],
                'previous_total': catalog_result['previous_total'],
                'new_count': catalog_result['new_count'],
                'changed_count': catalog_result['changed_count'],
                'removed_count': catalog_result['removed_count'],
                'sample_new': catalog_result['sample_new'],
                'sample_changed': catalog_result['sample_changed'],
                'sample_removed': catalog_result['sample_removed'],
            } if catalog_result else None),
        })
    except Exception as exc:
        app.logger.exception('Erro ao analisar delta Brasíndice')
        return jsonify({'status': 'error', 'message': str(exc)}), 500
    finally:
        tmp_path.unlink(missing_ok=True)
        if catalog_tmp_path is not None:
            catalog_tmp_path.unlink(missing_ok=True)


@app.route('/insumos/bras/import-delta', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumos_bras_import_delta():
    """Importa apenas itens novos ou alterados da Brasíndice."""
    upload = request.files.get('arquivo')
    if not upload or not upload.filename:
        return jsonify({'status': 'error', 'message': 'Selecione um arquivo para importar.'}), 400
    catalog_upload = request.files.get('catalogo')
    versao = (request.form.get('versao') or '').strip()
    if not versao:
        return jsonify({'status': 'error', 'message': 'Informe a versão da tabela.'}), 400
    
    delimiter = request.form.get('delimiter') or ','
    encoding = (request.form.get('encoding') or 'latin-1').strip()
    skip_header = request.form.get('skip_header') == 'on'
    data_ref = (request.form.get('data_atualizacao') or '').strip() or None
    truncate = request.form.get('truncate') == 'on'
    catalog_encoding = (request.form.get('catalog_encoding') or 'latin-1').strip()
    catalog_delimiter = request.form.get('catalog_delimiter') or ';'
    catalog_prev_versao = (request.form.get('catalog_prev_versao') or '').strip() or None
    
    # Alíquota e UFs (opcional)
    aliquota_input = (request.form.get('aliquota') or '').strip()
    aliquota_value: Decimal | None = None
    if aliquota_input:
        aliquota_str = _coerce_decimal(aliquota_input)
        if aliquota_str:
            aliquota_value = Decimal(aliquota_str)
    
    raw_ufs = request.form.getlist('ufs')
    uf_values: list[str] = []
    for raw in raw_ufs:
        candidate = (raw or '').strip().upper()
        if candidate and candidate in BR_UFS:
            uf_values.append(candidate)
    uf_default = uf_values[0] if uf_values else None

    # Mesmo mecanismo das importações clássicas: job + thread (aparece em "Importações em andamento")
    job_id = uuid4().hex
    job_dir = INSUMO_IMPORT_ASYNC_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    main_name = secure_filename(upload.filename) or f'bras_delta_{job_id}.txt'
    data_path = job_dir / main_name
    try:
        upload.save(str(data_path))
    except Exception as exc:  # noqa: BLE001
        shutil.rmtree(job_dir, ignore_errors=True)
        app.logger.exception('Erro ao salvar arquivo para delta Brasíndice')
        return jsonify({'status': 'error', 'message': f'Falha ao salvar o arquivo: {exc}'}), 500

    catalog_data_path: str | None = None
    if catalog_upload and catalog_upload.filename:
        cat_stem = secure_filename(catalog_upload.filename) or f'catalogo_{job_id}.txt'
        cpath = job_dir / f'catalogo_{cat_stem}'
        try:
            catalog_upload.save(str(cpath))
            catalog_data_path = str(cpath)
        except Exception as exc:  # noqa: BLE001
            shutil.rmtree(job_dir, ignore_errors=True)
            return jsonify({'status': 'error', 'message': f'Falha ao salvar o catálogo: {exc}'}), 500

    label_on = secure_filename(upload.filename) or 'Brasíndice'
    if catalog_upload and catalog_upload.filename:
        label_on = f"{label_on} + {secure_filename(catalog_upload.filename) or 'catálogo'}"

    params_payload = {
        'import_kind': 'bras_delta',
        'versao': versao,
        'delimiter': _normalize_delimiter(delimiter),
        'quotechar': '"',
        'encoding': encoding,
        'skip_header': skip_header,
        'data_ref': data_ref,
        'truncate': truncate,
        'uf_values': uf_values,
        'uf_default': uf_default,
        'aliquota': str(aliquota_value) if aliquota_value is not None else None,
        'catalog_encoding': catalog_encoding,
        'catalog_delimiter': _normalize_delimiter(catalog_delimiter),
        'previous_catalog_version': catalog_prev_versao,
        'catalog_data_path': catalog_data_path,
    }
    job = ImportJob(
        id=job_id,
        origem='BRAS',
        original_filename=label_on,
        data_path=str(data_path),
        status=ImportJobStatus.PENDING.value,
        message='Aguardando processamento do delta...',
        versao=versao,
        aliquota=aliquota_value,
        uf_list=', '.join(uf_values),
        params=params_payload,
    )
    try:
        db.session.add(job)
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'status': 'error', 'message': f'Falha ao registrar o job: {exc}'}), 500

    inline_env = (os.getenv('INSUMO_IMPORT_RUN_INLINE') or '').strip().lower()
    if inline_env in {'1', 'true', 'yes', 'on'}:
        _run_import_job(job_id)
        job = ImportJob.query.get(job_id)
        if not job or job.status != ImportJobStatus.SUCCESS.value:
            err = (job.message if job else None) or 'Falha na importação delta.'
            return jsonify({'status': 'error', 'message': err, 'inline': True}), 500
        m = (job.params or {}).get('_metrics') or {}
        d = m.get('delta_brasindice') or {}
        return jsonify(
            {
                'status': 'ok',
                'message': job.message,
                'inline': True,
                'versao': d.get('versao') or versao,
                'novos': d.get('novos', 0),
                'novos_importados': d.get('novos_importados', 0),
                'alterados': d.get('alterados', 0),
                'total_processado': d.get('total_processado', 0),
                'catalog_current_version': d.get('catalog_current_version'),
                'catalog_previous_version': d.get('catalog_previous_version'),
                'catalog_new': d.get('catalog_new', 0),
                'catalog_changed': d.get('catalog_changed', 0),
                'catalog_removed': d.get('catalog_removed', 0),
            }
        )

    _spawn_async_import(job_id)
    prefix = job_id[:8]
    return jsonify(
        {
            'status': 'ok',
            'message': f'Delta enfileirado (protocolo {prefix}). Acompanhe o status em "Importações em andamento".',
            'job_id': job_id,
            'prefix': prefix,
            'inline': False,
        }
    )


@app.route('/insumos/bras/edicoes-cadastro', methods=['GET'])
@admin_required
@feature_required('insumos')
def insumos_bras_edicoes_cadastro():
    """Lista `edicao` distintas em `bras_item_cadastro` (e fallback em `bras_item_n`) para o usuário alinhar o campo com a col. 14 do TXT."""
    lim = request.args.get('limit', '40')
    try:
        n = int(lim)
    except ValueError:
        n = 40
    n = min(100, max(1, n))
    items: list[str] = []
    try:
        r1 = db.session.execute(
            text(
                'SELECT DISTINCT TRIM(edicao) AS e FROM bras_item_cadastro '
                "WHERE edicao IS NOT NULL AND TRIM(edicao) <> '' ORDER BY e DESC LIMIT :lim"
            ),
            {'lim': n},
        )
        items = [row[0] for row in r1 if row[0]]
    except Exception as exc:  # noqa: BLE001
        app.logger.warning('edicoes-cadastro (split): %s', exc)
    if len(items) < 5:
        try:
            r2 = db.session.execute(
                text(
                    'SELECT DISTINCT TRIM(edicao) AS e FROM bras_item_n '
                    "WHERE edicao IS NOT NULL AND TRIM(edicao) <> '' ORDER BY e DESC LIMIT :lim"
                ),
                {'lim': n},
            )
            extra = [row[0] for row in r2 if row[0]]
            seen = set(items)
            for e in extra:
                if e not in seen:
                    seen.add(e)
                    items.append(e)
        except Exception as exc:  # noqa: BLE001
            app.logger.debug('edicoes-cadastro fallback n: %s', exc)
    return jsonify({'items': items[:n]})


@app.route('/insumos/bras/import-precos', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumos_bras_import_precos():
    """Atualiza somente preços (layout TXT D) para itens já existentes no cadastro split; opcionalmente replica em legacy."""
    upload = request.files.get('arquivo')
    if not upload or not upload.filename:
        return jsonify({'status': 'error', 'message': 'Selecione um arquivo para importar.'}), 400
    versao = (request.form.get('versao') or '').strip()
    if not versao:
        return jsonify({'status': 'error', 'message': 'Informe a versão/edição da tabela (deve bater com o cadastro).'}), 400
    aliquota_input = (request.form.get('aliquota') or '').strip()
    if not aliquota_input:
        return jsonify({'status': 'error', 'message': 'Informe a alíquota (%) deste arquivo.'}), 400
    al_str = _coerce_decimal(aliquota_input)
    if al_str is None:
        return jsonify({'status': 'error', 'message': 'Alíquota inválida.'}), 400
    alq = _br_norm_aliquota(Decimal(str(al_str)))
    if alq is None:
        return jsonify({'status': 'error', 'message': 'Alíquota inválida.'}), 400
    raw_ufs = request.form.getlist('ufs') or request.form.getlist('uf')
    uf_values = _normalize_uf_codes(raw_ufs)
    if not uf_values:
        uf_values = [uf for uf, floor in _UF_PISO_ALIQUOTA_BRAS.items() if floor == alq]
    delimiter = request.form.get('delimiter') or ','
    encoding = (request.form.get('encoding') or 'latin-1').strip()
    skip_header = request.form.get('skip_header') == 'on'
    no_legacy = request.form.get('no_legacy') == 'on'
    job_id = uuid4().hex
    job_dir = INSUMO_IMPORT_ASYNC_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    main_name = secure_filename(upload.filename) or f'bras_precos_{job_id}.txt'
    data_path = job_dir / main_name
    try:
        upload.save(str(data_path))
    except Exception as exc:  # noqa: BLE001
        shutil.rmtree(job_dir, ignore_errors=True)
        app.logger.exception('Erro ao salvar arquivo import-precos Brasíndice')
        return jsonify({'status': 'error', 'message': f'Falha ao salvar o arquivo: {exc}'}), 500
    label_on = secure_filename(upload.filename) or 'Brasíndice preços'
    params_payload = {
        'import_kind': 'bras_import_precos',
        'versao': versao,
        'delimiter': _normalize_delimiter(delimiter),
        'quotechar': '"',
        'encoding': encoding,
        'skip_header': skip_header,
        'aliquota': str(alq),
        'uf_values': uf_values,
        'uf_default': uf_values[0] if uf_values else None,
        'no_legacy': no_legacy,
    }
    job = ImportJob(
        id=job_id,
        origem='BRAS',
        original_filename=label_on,
        data_path=str(data_path),
        status=ImportJobStatus.PENDING.value,
        message='Aguardando importação de preços (split)...',
        versao=versao,
        aliquota=alq,
        uf_list=', '.join(uf_values),
        params=params_payload,
    )
    try:
        db.session.add(job)
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'status': 'error', 'message': f'Falha ao registrar o job: {exc}'}), 500
    inline_env = (os.getenv('INSUMO_IMPORT_RUN_INLINE') or '').strip().lower()
    if inline_env in {'1', 'true', 'yes', 'on'}:
        _run_import_job(job_id)
        job = ImportJob.query.get(job_id)
        if not job or job.status != ImportJobStatus.SUCCESS.value:
            err = (job.message if job else None) or 'Falha na importação de preços.'
            return jsonify({'status': 'error', 'message': err, 'inline': True}), 500
        m = (job.params or {}).get('_metrics') or {}
        stp = m.get('bras_somente_precos') or {}
        return jsonify(
            {
                'status': 'ok',
                'message': job.message,
                'inline': True,
                'bras_somente_precos': stp,
            }
        )
    _spawn_async_import(job_id)
    prefix = job_id[:8]
    return jsonify(
        {
            'status': 'ok',
            'message': f'Importação de preços enfileirada (protocolo {prefix}). Acompanhe em "Importações em andamento".',
            'job_id': job_id,
            'prefix': prefix,
            'inline': False,
        }
    )


@app.route('/insumos/import/jobs')
@admin_required
@feature_required('insumos')
def insumos_import_jobs_list():
    limit = _parse_positive_int(request.args.get('limit'), 20, maximum=100)
    jobs = (
        ImportJob.query
        .order_by(ImportJob.created_at.desc())
        .limit(limit)
        .all()
    )
    return jsonify({'items': [_serialize_import_job(job) for job in jobs]})


@app.route('/insumos/import/jobs/<job_id>')
@admin_required
@feature_required('insumos')
def insumos_import_job_detail(job_id: str):
    job = ImportJob.query.get_or_404(job_id)
    return jsonify(_serialize_import_job(job))


@app.route('/insumos/import/jobs/<job_id>/pause', methods=['POST'])
@admin_required
@feature_required('insumos')
def insumos_import_job_pause(job_id: str):
    job = ImportJob.query.get_or_404(job_id)
    if job.status == ImportJobStatus.SUCCESS.value:
        return jsonify({'status': 'error', 'message': 'Importação já concluída.'}), 409
    if job.status == ImportJobStatus.FAILED.value:
        paused = 'pausad' in (job.message or '').lower()
        msg = 'Importação já está pausada.' if paused else 'Importação já foi finalizada com falha.'
        return jsonify({'status': 'error', 'message': msg}), 409

    if job.status == ImportJobStatus.PENDING.value:
        job.status = ImportJobStatus.FAILED.value
        job.message = _job_message_trim('Importação pausada antes do início.')
        job.finished_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Importação pausada antes do início.', 'job': _serialize_import_job(job)})

    _request_import_pause(job_id)
    job.message = _job_message_trim('Pausa solicitada. Encerrando processamento...')
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Pausa solicitada. O job será interrompido no próximo checkpoint.', 'job': _serialize_import_job(job)})
