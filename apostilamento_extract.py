"""Extração de apostilamentos PDF → Excel (DTP).

Portado do projeto EMERGENCIAL (`extrair_apostilamentos.py`).
Gera XLSX no formato consumido por `scripts/import_apostilamentos_dtp.py`.

Uso típico:
    python apostilamento_extract.py JULHO
    python scripts/extract_apostilamentos.py JULHO
    python scripts/pipeline_apostilamentos.py JULHO --importar
"""
from __future__ import annotations

import argparse
import atexit
import gc
import logging
import os
import re
import tempfile
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import pdfplumber  # type: ignore
except ImportError:  # pragma: no cover - fallback de ambiente
    pdfplumber = None

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover - fallback de ambiente
    PdfReader = None


def _instalar_shutdown_openpyxl_seguro() -> None:
    """Evita PermissionError cosmético do openpyxl ao limpar temp no Windows."""
    try:
        from openpyxl.worksheet import _writer as ws_writer
    except ImportError:
        return

    shutdown_original = getattr(ws_writer, "_openpyxl_shutdown", None)
    if shutdown_original is None:
        return

    try:
        atexit.unregister(shutdown_original)
    except Exception:
        pass

    def _shutdown_seguro() -> None:
        for path in list(getattr(ws_writer, "ALL_TEMP_FILES", [])):
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except OSError:
                # No Windows, antivírus/indexador pode manter o temp aberto no exit.
                pass

    ws_writer._openpyxl_shutdown = _shutdown_seguro  # type: ignore[attr-defined]
    atexit.register(_shutdown_seguro)


_instalar_shutdown_openpyxl_seguro()


BASE_DIR = Path(__file__).resolve().parent
# PDFs de entrada (ex.: entrada_apostilamentos/JULHO/*.pdf)
ENTRADA_DIR = BASE_DIR / "entrada_apostilamentos"
# XLSX de saída no formato consumido por scripts/import_apostilamentos_dtp.py
SAIDA_DIR = BASE_DIR / "APOSTILAMENTOS"
LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
ENTRADA_DIR.mkdir(exist_ok=True)

LOG_FILE = LOG_DIR / "extrair_apostilamentos.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

LOGGER = logging.getLogger("apostilamento_extract")

ITEM_COLUMNS = [
    "arquivo_origem",
    "pagina",
    "anexo",
    "secao",
    "codigo",
    "codigo_hospital",
    "descricao",
    "unidade_cobranca",
    "valor",
    "valor_numero",
    "observacao_linha",
]

CONDICOES_COLUMNS = [
    "arquivo_origem",
    "pagina",
    "tipo",
    "texto",
]

RESUMO_COLUMNS = [
    "arquivo_origem",
    "total_itens_extraidos",
    "total_condicoes_observacoes",
    "paginas_processadas",
    "status",
    "observacoes_processamento",
]

UNIDADES = [
    "DIA",
    "DIÁRIA",
    "DIÁRIA",
    "HORA",
    "TAXA",
    "EVENTO",
    "SESSAO",
    "SESSÃO",
    "USO",
    "INSTALACAO",
    "INSTALAÇÃO",
    "PACOTE",
    "HM",
    "MINUTO",
]

UNIDADES_REGEX = r"(?:DIA|DI[ÁA]RIA|HORA|TAXA|EVENTO|SESS[ÃA]O|USO|INSTALA(?:Ç|C)[ÃA]O|PACOTE|HM|MINUTO)"
# Aceita R$ 1.234,56 e R$ 1234,56 / R$ 1000,00 (sem separador de milhar)
VALOR_REGEX = r"R\$\s*(?:\d{1,3}(?:\.\d{3})+|\d{1,}),\d{2}"
CODIGO_HIFEN_REGEX = r"\d{2}\.\d{2}\.\d{3}-\d"
CODIGO_REGEX = rf"(?:\d{{7,10}}|{CODIGO_HIFEN_REGEX}|\d(?:\.\d{{2}}){{2,4}})"

# Regras ampliadas para códigos pontuados com bloco final de 4 dígitos e unidades no plural.
UNIDADES_REGEX = r"(?:DIA|DI[ÁA]RIA|HORA|TAXA|EVENTO|SESS(?:ÕES|OES|ÃO|AO)|USO|INSTALA(?:Ç|C)[ÃA]O|PACOTES|PACOTE|HM|MINUTO)"
# CBHPM-like: 3.10.02.390 (último bloco com 3 dígitos)
CODIGO_CBHPM_REGEX = r"\d(?:\.\d{2}){2}\.\d{3}"
CODIGO_DOTTED_4_REGEX = r"\d(?:\.\d{2}){2}\.\d{4}"
# Formato AMB/paramédico: 00.01.1200, 17.01.0011
CODIGO_XX_XX_XXXX_REGEX = r"\d{2}\.\d{2}\.\d{4}"
CODIGO_SHORT_DOTTED_REGEX = r"\d{1,2}\.\d{3}"
# Genérico X.XX.XX… só se não houver ponto extra incompleto (evita 0.02.05 de 0.02.05.0)
CODIGO_DOTTED_GENERIC_REGEX = r"\d(?:\.\d{2}){2,4}(?!\.\d)"
# Códigos no formato decimal do hospital (ex.: 0.02041390) — manter como estão
CODIGO_DECIMAL_HOSPITAL_REGEX = r"0\.\d{6,12}"
CODIGO_REGEX = (
    rf"(?:\d{{7,10}}|{CODIGO_HIFEN_REGEX}|{CODIGO_CBHPM_REGEX}|{CODIGO_DOTTED_4_REGEX}|"
    rf"{CODIGO_XX_XX_XXXX_REGEX}|{CODIGO_DECIMAL_HOSPITAL_REGEX}|"
    rf"{CODIGO_SHORT_DOTTED_REGEX}|{CODIGO_DOTTED_GENERIC_REGEX})"
)

# Tabela de materiais (Einstein + Plan-Assiste 7910xxxxx + ANVISA + unidade).
UNIDADES_MATERIAIS_REGEX = (
    r"(?:PE[ÇC]A?|UN|M|TAXA|PAR|PAC(?:OTE)?s?|CX|FR|KG|ML|KIT|HM|DIA|HORA|EVENTO|USO)"
)
CODIGO_PLAN_MATERIAIS_REGEX = r"7910\d{5}"
MATERIAIS_INICIO_RE = re.compile(
    rf"^(?P<codigo_einstein>[A-Z0-9]{{2,16}}?)(?P<codigo_plan>{CODIGO_PLAN_MATERIAIS_REGEX})",
    re.IGNORECASE,
)
MATERIAIS_FIM_RE = re.compile(
    rf"(?P<anvisa>\d{{8,14}})?(?P<unidade>{UNIDADES_MATERIAIS_REGEX})\s*$",
    re.IGNORECASE,
)
ANEXO_RE = re.compile(r"\bANEXO\s+[IVXLCDM0-9]+\b", re.IGNORECASE)
VALOR_RE = re.compile(VALOR_REGEX, re.IGNORECASE)
CODIGO_RE = re.compile(CODIGO_REGEX)
# Textos da coluna Honorário Médico que não são preço (layout Pacote Hospitalar).
HONORARIO_TEXTO_RE = re.compile(
    r"\b(?:"
    r"Tabela\s+acordada|"
    r"Conforme\s+(?:a\s+)?tabela|"
    r"Conforme\s+contrato|"
    r"Honor[aá]rios?\s+m[eé]dicos?|"
    r"HM\s+acordad[oa]|"
    r"CBHPM(?:\s*\d{4})?|"
    r"AMB(?:\s*\d+)?"
    r")\b",
    re.IGNORECASE,
)
CODIGO_CURTO_PONTUADO_RE = re.compile(r"^\d{1,2}\.\d{3}$")
# Código decimal hospitalar completo (não confundir com 0.020 curto)
CODIGO_DECIMAL_HOSPITAL_RE = re.compile(r"^0\.\d{6,12}$")
# Código incompleto do tipo 0.02.05 / 0.02.05.0
CODIGO_INCOMPLETO_RE = re.compile(r"^\d(?:\.\d{1,2}){1,3}\.?$")
CODIGO_MELHOR_NO_RESTO_RE = re.compile(
    rf"^(?P<code>\d{{7,10}}|{CODIGO_HIFEN_REGEX}|{CODIGO_CBHPM_REGEX}|{CODIGO_DOTTED_4_REGEX}|"
    rf"{CODIGO_XX_XX_XXXX_REGEX}|{CODIGO_DECIMAL_HOSPITAL_REGEX}|\d{{4,6}})\b(?P<rest>.*)$"
)
CODIGO_NO_INICIO_DESC_RE = re.compile(
    rf"^(?P<code>\d{{7,10}}|{CODIGO_HIFEN_REGEX}|{CODIGO_CBHPM_REGEX}|{CODIGO_DOTTED_4_REGEX}|"
    rf"{CODIGO_XX_XX_XXXX_REGEX}|{CODIGO_DECIMAL_HOSPITAL_REGEX}|\d{{4,6}})\b\s*(?P<rest>.*)$"
)

SECTION_PATTERNS = [
    r"TABELA DE DI[ÁA]RIAS(?:, TAXAS, SERVI[ÇC]OS(?:, PACOTES)? E OUTROS)?",
    r"TABELA DE VALORES",
    r"TABELA DE MATERIAIS",
    r"DI[ÁA]RIAS?(?: DE UTI)?",
    r"DI[ÁA]RIA APARTAMENTO",
    r"PEDIATR(?:IA|ICA)",
    r"TAXAS?(?: DE SALA)?(?: [A-Z0-9ÇÃÁÉÍÓÚ/ -]+)?",
    r"PROCEDIMENTOS?",
    r"EXTRAS?",
    r"GASES MEDICINAIS",
    r"CONSULTAS?",
    r"PACOTES?",
    r"MATERIAIS?",
    r"MEDICAMENTOS?",
    r"APARELHOS E OU EQUIPAMENTOS",
    r"TAXA DE APARELHOS E EQUIPAMENTOS",
    r"HONOR[ÁA]RIOS PARAM[ÉE]DICOS",
    r"OUTRAS CONDI[ÇC][ÕO]ES OPERACIONAIS",
    r"COMPOSI[ÇC][ÃA]O DAS DI[ÁA]RIAS",
    r"RECUPERA[ÇC][ÃA]O",
]
SECTION_RES = [re.compile(pattern, re.IGNORECASE) for pattern in SECTION_PATTERNS]

CONDITION_KEYWORDS = {
    "inclusoes": re.compile(r"\bINCLUS[ÕO]ES?\b", re.IGNORECASE),
    "exclusoes": re.compile(r"\bEXCLUS[ÕO]ES?\b", re.IGNORECASE),
    "condicoes_operacionais": re.compile(r"\bCONDI[ÇC][ÕO]ES? OPERACIONAIS\b", re.IGNORECASE),
    "tabela_referencia": re.compile(r"\b(CBHPM|SIMPRO|BRAS[ÍI]NDICE|TUSS|AMB|FILME RADIOL[ÓO]GICO)\b", re.IGNORECASE),
    "observacao": re.compile(r"\b(OBSERVA[ÇC][ÕO]ES?|OBS\.?)\b", re.IGNORECASE),
    "regra": re.compile(r"\b(INCLUI|INCLUIS?|INCLUSO|EXCETO|SER[ÁA]O COBRADOS?|COBRADA|COBRADO)\b", re.IGNORECASE),
}

HEADER_GARBAGE_RE = re.compile(
    r"(MINIST[ÉE]RIO P[ÚU]BLICO|PROGRAMA DE SA[ÚU]DE|APOSTILAMENTO|ASSINATURA DIGITAL|"
    r"ACESSE HTTP|VALIDACAODOCUMENTO|P[ÁA]GINA\s+\d+|TERMO DE APOSTILA|ASSINADO COM LOGIN)",
    re.IGNORECASE,
)

CONDICAO_HEADING_RE = re.compile(
    r"\b(DEMAIS CONDI|OUTRAS CONDI|CONDI|INCLUS|EXCLUS|MATERIAIS HOSPITALARES|OUTROS MATERIAIS|"
    r"SERVI|HONOR|MEDICAMENTOS|FILME RADIOL|IMUNO[- ]HEMATOL)",
    re.IGNORECASE,
)


def normalizar_espacos(texto: str) -> str:
    return re.sub(r"\s+", " ", (texto or "").replace("\u00a0", " ")).strip()


def normalizar_ascii(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto or "")
    return "".join(ch for ch in texto if not unicodedata.combining(ch)).upper()


def eh_inicio_bloco_condicoes(texto: str) -> bool:
    texto_ascii = normalizar_ascii(texto)
    marcadores = [
        "DEMAIS CONDI",
        "OUTRAS CONDI",
        "CONDICOES OPERACIONAIS",
        "INCLUSOES",
        "EXCLUSOES",
        "MATERIAIS HOSPITALARES",
        "OUTROS MATERIAIS",
        "SERVICOS MEDICOS",
        "HONORARIOS MEDICOS",
        "MEDICAMENTOS",
        "FILME RADIOL",
        "IMUNO-HEMATOL",
        "IMUNO HEMATOL",
    ]
    return any(marcador in texto_ascii for marcador in marcadores)


def limpar_nome_aba(nome: str) -> str:
    nome = re.sub(r"[:\\/?*\[\]]", "_", nome)
    return nome[:31]


def converter_valor_brl_para_float(valor: str) -> float | None:
    valor_limpo = normalizar_espacos(valor).replace("R$", "").strip()
    if not valor_limpo:
        return None
    try:
        return float(valor_limpo.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def escolher_match_valor(bloco: str) -> re.Match[str] | None:
    """Escolhe o valor monetário da linha.

    Em tabelas com Honorário + Pacote Hospitalar, o pacote costuma ser a
    última coluna com R$. Com 2+ valores, preferimos o último.
    """
    matches = list(VALOR_RE.finditer(bloco or ""))
    if not matches:
        return None
    if len(matches) >= 2:
        return matches[-1]
    return matches[0]


def limpar_texto_honorario(texto: str) -> tuple[str, str]:
    """Remove placeholders de honorário da descrição; devolve (texto, observação)."""
    texto_n = normalizar_espacos(texto or "")
    if not texto_n:
        return "", ""
    encontrados = [m.group(0) for m in HONORARIO_TEXTO_RE.finditer(texto_n)]
    limpo = HONORARIO_TEXTO_RE.sub(" ", texto_n)
    # Asterisco da coluna Honorário Médico (ex.: "*")
    if "*" in limpo:
        encontrados.append("*")
        limpo = limpo.replace("*", " ")
    limpo = normalizar_espacos(limpo)
    return limpo, " | ".join(encontrados)


def promover_codigo_principal(
    codigo: str,
    codigo_hospital: str,
    resto: str,
) -> tuple[str, str, str]:
    """Corrige falso positivo de código curto (ex.: 1.000) e promove código real."""
    resto_n = normalizar_espacos(resto)
    codigo = normalizar_codigo_extraido(codigo)

    # Código decimal hospitalar completo (0.02041390): manter
    if codigo and CODIGO_DECIMAL_HOSPITAL_RE.fullmatch(codigo):
        return codigo, codigo_hospital, resto_n

    if codigo and CODIGO_CURTO_PONTUADO_RE.fullmatch(codigo):
        # 0.020 + 41390 colados → recombina 0.02041390
        m_recombinado = re.match(r"^(?P<digits>\d{5,10})\b(?P<rest>\s+.*)?$", resto_n)
        if m_recombinado and codigo.startswith("0."):
            combinado = f"{codigo}{m_recombinado.group('digits')}"
            if CODIGO_DECIMAL_HOSPITAL_RE.fullmatch(combinado):
                return combinado, codigo_hospital, normalizar_espacos(m_recombinado.group("rest") or "")

        match = CODIGO_MELHOR_NO_RESTO_RE.match(resto_n)
        if match and normalizar_espacos(match.group("rest")):
            return (
                normalizar_codigo_extraido(match.group("code")),
                codigo_hospital or codigo,
                normalizar_espacos(match.group("rest")),
            )
        # 1.000 sozinho sem código melhor → não usar como código do item
        return "", codigo_hospital or codigo, resto_n

    if not codigo:
        match = CODIGO_MELHOR_NO_RESTO_RE.match(resto_n)
        if match and normalizar_espacos(match.group("rest")):
            return (
                normalizar_codigo_extraido(match.group("code")),
                codigo_hospital,
                normalizar_espacos(match.group("rest")),
            )
    return codigo, codigo_hospital, resto_n


def detectar_anexo(linha: str, atual: str) -> str:
    match = ANEXO_RE.search(linha or "")
    if match:
        return normalizar_espacos(match.group(0).upper())
    return atual


def detectar_secao(linha: str, atual: str) -> str:
    linha_limpa = normalizar_espacos(linha)
    if not linha_limpa:
        return atual
    if not eh_linha_secao(linha_limpa):
        return atual
    for regex in SECTION_RES:
        if regex.search(linha_limpa):
            return normalizar_espacos(linha_limpa.upper())
    return atual


def eh_linha_ignorar(linha: str) -> bool:
    linha_limpa = normalizar_espacos(linha)
    if not linha_limpa:
        return True
    if HEADER_GARBAGE_RE.search(linha_limpa):
        return True
    if re.fullmatch(r"[:.\-–—]+", linha_limpa):
        return True
    return False


def eh_cabecalho_tabela(linha: str) -> bool:
    linha_ascii = normalizar_ascii(linha)
    cabecalhos = [
        "CODIGO",
        "DESCRICAO",
        "VALOR",
        "FATOR",
        "CONSULTA MEDICA AMBULATORIAL",
        "TUSS DESCRICAO DO PROCEDIMENTO",
        "DIAGNOSE",
        "HONORARIO MEDICO",
        "HONORARIOS MEDICOS",
        "PACOTE HOSPITALAR",
        "TABELA ACORDADA",
    ]
    return any(chave in linha_ascii for chave in cabecalhos)


def eh_linha_secao(linha: str) -> bool:
    linha_limpa = normalizar_espacos(linha)
    if not linha_limpa:
        return False
    if VALOR_RE.search(linha_limpa):
        return False
    if re.match(rf"^\s*(?:{CODIGO_REGEX})", linha_limpa):
        return False
    if len(linha_limpa) > 90 and linha_limpa != linha_limpa.upper():
        return False
    return any(regex.search(linha_limpa) for regex in SECTION_RES) and linha_limpa == linha_limpa.upper()


def eh_linha_item_inicial(linha: str) -> bool:
    linha_limpa = normalizar_espacos(linha)
    if not linha_limpa or not VALOR_RE.search(linha_limpa):
        # Pode iniciar item sem valor na mesma linha (valor vem na seguinte)
        if MATERIAIS_INICIO_RE.match(linha_limpa):
            return True
        if re.match(rf"^\s*{CODIGO_REGEX}\b", linha_limpa):
            return True
        return False
    if MATERIAIS_INICIO_RE.match(linha_limpa):
        return True
    if re.match(rf"^\s*{CODIGO_REGEX}", linha_limpa):
        return True
    return bool(re.match(r"^[A-ZÀ-Ú0-9/().,\- ]{3,}\s+R\$\s*\d", linha_limpa, re.IGNORECASE))


def descricao_parece_incompleta(texto: str) -> bool:
    """Descrição cortada no meio (ex.: termina em POR, DE, COM)."""
    t = normalizar_espacos(texto)
    if not t:
        return True
    if VALOR_RE.search(t):
        # valor na linha — se termina com conector antes do R$, ainda incompleta?
        antes = VALOR_RE.split(t)[0] if VALOR_RE.search(t) else t
        antes = normalizar_espacos(antes)
    else:
        antes = t
    if re.search(
        r"(?i)(?:,\s*)?\b(POR|DE|COM|PARA|EM|NA|NO|E|OU|/)\s*$",
        antes,
    ):
        return True
    if antes.endswith(",") or antes.endswith("/"):
        return True
    return False


def juntar_linhas_quebradas(texto: str) -> list[str]:
    linhas = [normalizar_espacos(linha) for linha in (texto or "").splitlines()]
    blocos: list[str] = []
    buffer = ""

    for linha in linhas:
        if eh_linha_ignorar(linha):
            continue

        if ANEXO_RE.search(linha):
            if buffer:
                blocos.append(buffer.strip())
                buffer = ""
            blocos.append(linha)
            continue

        if eh_linha_secao(linha) or eh_cabecalho_tabela(linha):
            if buffer:
                blocos.append(buffer.strip())
                buffer = ""
            blocos.append(linha)
            continue

        # Continua descrição incompleta (… POR / USO …) mesmo se a próxima linha
        # também parecer item (sem R$ ainda).
        if buffer and descricao_parece_incompleta(buffer) and not VALOR_RE.search(buffer):
            if not re.match(rf"^\s*{CODIGO_REGEX}\b", linha) or not VALOR_RE.search(linha):
                # junta se não for um item novo completo
                if not (re.match(rf"^\s*{CODIGO_REGEX}\b", linha) and VALOR_RE.search(linha)):
                    buffer = f"{buffer} {linha}".strip()
                    if VALOR_RE.search(buffer) and not descricao_parece_incompleta(buffer):
                        blocos.append(buffer.strip())
                        buffer = ""
                    continue

        if eh_linha_item_inicial(linha) or (
            re.match(rf"^\s*{CODIGO_REGEX}\b", linha) and not buffer
        ):
            if buffer:
                # se o buffer anterior está incompleto, junta em vez de fechar
                if descricao_parece_incompleta(buffer) and not re.match(
                    rf"^\s*{CODIGO_REGEX}\b", linha
                ):
                    buffer = f"{buffer} {linha}".strip()
                    if VALOR_RE.search(buffer) and not descricao_parece_incompleta(buffer):
                        blocos.append(buffer.strip())
                        buffer = ""
                    continue
                if buffer:
                    blocos.append(buffer.strip())
            buffer = linha
            if VALOR_RE.search(linha) and not descricao_parece_incompleta(linha):
                blocos.append(buffer.strip())
                buffer = ""
            continue

        if buffer:
            buffer = f"{buffer} {linha}".strip()
            if VALOR_RE.search(buffer) and not descricao_parece_incompleta(buffer):
                blocos.append(buffer.strip())
                buffer = ""
        else:
            blocos.append(linha)

    if buffer:
        blocos.append(buffer.strip())
    return blocos


def extrair_codigos_inicio(bloco: str) -> tuple[str, str, str]:
    texto = normalizar_espacos(bloco)
    codigo_principal = ""
    codigo_secundario = ""

    # Código float colado (0.02041390APARELHO ou 0.02041390 APARELHO)
    match_float = re.match(r"^(?P<code>0\.\d{6,})(?P<resto>\s*.*)$", texto)
    if match_float and match_float.group("resto").strip():
        return (
            normalizar_codigo_extraido(match_float.group("code")),
            "",
            normalizar_espacos(match_float.group("resto")),
        )

    match_concat_hifen = re.match(
        r"^(?P<codigo1>\d(?:\.\d{2}){2}\.\d{4})(?P<codigo2>\d{2}\.\d{2}\.\d{3}-\d)(?P<resto>.*)$",
        texto,
    )
    if match_concat_hifen:
        return (
            match_concat_hifen.group("codigo1"),
            match_concat_hifen.group("codigo2"),
            normalizar_espacos(match_concat_hifen.group("resto")),
        )

    match_concat_num_hifen = re.match(
        r"^(?P<codigo1>\d{8,10})(?P<codigo2>\d{2}\.\d{2}\.\d{3}-\d)(?P<resto>.*)$",
        texto,
    )
    if match_concat_num_hifen:
        return (
            match_concat_num_hifen.group("codigo1"),
            match_concat_num_hifen.group("codigo2"),
            normalizar_espacos(match_concat_num_hifen.group("resto")),
        )

    match_concat_num_num = re.match(
        r"^(?P<codigo1>\d{8,10})(?P<codigo2>\d{8,10})(?P<resto>.*)$",
        texto,
    )
    if match_concat_num_num:
        return (
            match_concat_num_num.group("codigo1"),
            match_concat_num_num.group("codigo2"),
            normalizar_espacos(match_concat_num_num.group("resto")),
        )

    match_principal = re.match(rf"^(?P<codigo>{CODIGO_REGEX})(?P<resto>.*)$", texto)
    if not match_principal:
        return codigo_principal, codigo_secundario, texto

    codigo_principal = match_principal.group("codigo")
    resto = match_principal.group("resto").strip()

    match_secundario = re.match(rf"^(?P<codigo2>{CODIGO_REGEX})\b(?P<resto2>.*)$", resto)
    if match_secundario:
        codigo_secundario = match_secundario.group("codigo2")
        resto = match_secundario.group("resto2").strip()

    return codigo_principal, codigo_secundario, resto


def normalizar_unidade(unidade: str) -> str:
    unidade_limpa = normalizar_espacos(unidade).upper()
    mapa = {
        "PACOTES": "PACOTE",
        "PAC": "PACOTE",
        "SESSOES": "SESSÃO",
        "SESSÕES": "SESSÃO",
        "SESSAO": "SESSÃO",
        "PEÇA": "PEÇ",
        "PECA": "PEÇ",
        "PEC": "PEÇ",
    }
    return mapa.get(unidade_limpa, unidade_limpa)


def descricao_indica_unidade_semantica(descricao: str) -> bool:
    descricao_ascii = normalizar_ascii(descricao)
    return (
        descricao_ascii.startswith("DIARIA ")
        or descricao_ascii.startswith("DIARIA DE")
        or descricao_ascii.startswith("TAXA ")
        or descricao_ascii.startswith("TAXA DE")
        or descricao_ascii.startswith("TAXA POR")
        or descricao_ascii.startswith("TAXA/")
        or descricao_ascii.startswith("TAXA /")
        or descricao_ascii.startswith("ALUGUEL/TAXA")
        or descricao_ascii.startswith("ALUGUEL / TAXA")
        or descricao_ascii.startswith("HONORARIO")
        or descricao_ascii.startswith("CONSULTA")
        or descricao_ascii.startswith("PACOTE ")
        or descricao_ascii.startswith("PACOTE DE")
        or descricao_ascii.startswith("PACOTES ")
    )


def normalizar_codigo_extraido(codigo: str) -> str:
    """Normaliza o código sem alterar o significado (mantém 0.02041390)."""
    c = (codigo or "").strip()
    if not c:
        return c
    return re.sub(r"\s+", "", c)


def sanitizar_codigo(codigo: str) -> str:
    """Remove lixo de OCR/PDF no código (ex.: parênteses), preserva 0.02041390."""
    c = (codigo or "").strip()
    if not c:
        return c
    c = re.sub(r"[\(\)\[\]\{\}]", "", c)
    # Mantém dígitos, ponto, hífen e letras
    c = re.sub(r"[^\dA-Za-z.\-/]", "", c)
    return c.strip("-/ ")


def unidade_e_continuacao_da_descricao(descricao_sem_unidade: str) -> bool:
    """True se pelear a unidade deixa a descrição incompleta (… POR)."""
    d = normalizar_espacos(descricao_sem_unidade)
    if not d:
        return True
    return bool(
        re.search(
            r"(?i)\b(POR|DE|COM|PARA|EM|NA|NO|A|AO|E|OU|/)\s*$",
            d,
        )
    )


def limpar_descricao_item(descricao: str) -> str:
    """Recompõe descrições quebradas (TAXA consumida como unidade, barras soltas)."""
    d = normalizar_espacos(descricao)
    if not d:
        return d

    # lixo de código quebrado no começo: "( TAXA..." ou "(00 ..."
    d = re.sub(r"^[\(\)\[\]]+\s*", "", d)

    # /FRALDA/DIA ou /FRALDA/ DIA quando "TAXA" foi peledado como unidade
    if d.startswith("/"):
        resto = d.lstrip("/ ").strip()
        if re.match(r"(?i)^FRALDA", resto):
            d = "TAXA/" + resto.replace(" ", "")
            d = re.sub(r"(?i)FRALDA/\s*DIA", "FRALDA/DIA", d)
            d = re.sub(r"(?i)FRALDA/\s*DIA\s*\(", "FRALDA/DIA (", d)
        else:
            d = resto

    # Hífen residual do PDF ("- FISIOTERAPIA")
    d = re.sub(r"^[\-\u2013\u2014]\s*", "", d)

    # Normaliza espaços estranhos em TAXA/FRALDA/ DIA
    d = re.sub(r"(?i)\bTAXA\s*/\s*FRALDA\s*/\s*DIA\b", "TAXA/FRALDA/DIA", d)
    d = re.sub(r"(?i)\bTAXA\s*/\s*FRALDA\s*/\s*DIA\s*\(", "TAXA/FRALDA/DIA (", d)
    return normalizar_espacos(d)


def extrair_codigo_embutido_na_descricao(
    codigo: str,
    descricao: str,
) -> tuple[str, str]:
    """Se o código veio vazio/fraco, puxa 00.01.1200 (etc.) do início da descrição."""
    desc = normalizar_espacos(descricao)
    cod = (codigo or "").strip()

    if cod and not CODIGO_INCOMPLETO_RE.fullmatch(cod) and not CODIGO_CURTO_PONTUADO_RE.fullmatch(cod):
        return cod, limpar_descricao_item(desc)

    match = CODIGO_NO_INICIO_DESC_RE.match(desc)
    if match and normalizar_espacos(match.group("rest")):
        return (
            normalizar_codigo_extraido(match.group("code")),
            limpar_descricao_item(match.group("rest")),
        )

    # Código fraco/incompleto: tenta recuperar do resto da descrição
    if cod and (CODIGO_INCOMPLETO_RE.fullmatch(cod) or CODIGO_CURTO_PONTUADO_RE.fullmatch(cod)):
        # Se a descrição ainda começa com fragmento do código, limpa
        desc2 = re.sub(rf"^{re.escape(cod)}\.?\s*", "", desc)
        match2 = CODIGO_NO_INICIO_DESC_RE.match(desc2)
        if match2 and normalizar_espacos(match2.group("rest")):
            return (
                normalizar_codigo_extraido(match2.group("code")),
                limpar_descricao_item(match2.group("rest")),
            )
        # Descartar código incompleto e deixar descrição íntegra
        if CODIGO_INCOMPLETO_RE.fullmatch(cod):
            # junta código fraco de volta se a desc começar com .0 etc
            if desc.startswith(".") or desc.startswith("0 "):
                return "", limpar_descricao_item(f"{cod}{desc}")
            return "", limpar_descricao_item(desc)

    return cod, limpar_descricao_item(desc)


def parse_item_materiais(
    bloco: str,
    pagina: int,
    arquivo_origem: str,
    anexo: str,
    secao: str,
) -> dict[str, Any] | None:
    """Parser da TABELA DE MATERIAIS: Einstein + 7910xxxxx + descrição + ANVISA + unidade + valor."""
    texto = normalizar_espacos(bloco)
    valor_match = VALOR_RE.search(texto)
    if not valor_match:
        return None

    # Usa só o primeiro valor para evitar itens colados na mesma linha.
    prefixo = normalizar_espacos(texto[: valor_match.start()])
    valor = normalizar_espacos(valor_match.group(0))
    sobra = normalizar_espacos(texto[valor_match.end() :])

    inicio = MATERIAIS_INICIO_RE.match(prefixo)
    if not inicio:
        return None

    codigo_einstein = inicio.group("codigo_einstein")
    codigo_plan = inicio.group("codigo_plan")
    resto = normalizar_espacos(prefixo[inicio.end() :])

    anvisa = ""
    unidade = ""
    descricao = resto

    fim = MATERIAIS_FIM_RE.search(resto)
    if fim:
        anvisa = fim.group("anvisa") or ""
        unidade = normalizar_unidade(fim.group("unidade"))
        descricao = normalizar_espacos(resto[: fim.start()])
    else:
        # Fallback: tenta separar unidade colada no fim mesmo sem ANVISA.
        fim_unidade = re.search(
            rf"(?P<unidade>{UNIDADES_MATERIAIS_REGEX})\s*$",
            resto,
            re.IGNORECASE,
        )
        if fim_unidade:
            unidade = normalizar_unidade(fim_unidade.group("unidade"))
            descricao = normalizar_espacos(resto[: fim_unidade.start()])

    if not descricao:
        descricao = f"ANVISA {anvisa}" if anvisa else "ITEM SEM DESCRICAO"

    observacao_parts = []
    if anvisa and not descricao.startswith(f"ANVISA {anvisa}"):
        observacao_parts.append(f"ANVISA {anvisa}")
    if sobra and not sobra.lower().startswith("assinatura"):
        # Descarta rodapé de assinatura digital; mantém eventual texto útil.
        if "ASSINATURA DIGITAL" not in normalizar_ascii(sobra):
            observacao_parts.append(sobra)

    return {
        "arquivo_origem": arquivo_origem,
        "pagina": pagina,
        "anexo": anexo,
        "secao": secao or "TABELA DE MATERIAIS",
        "codigo": codigo_einstein,
        "codigo_hospital": codigo_plan,
        "descricao": descricao,
        "unidade_cobranca": unidade,
        "valor": valor,
        "valor_numero": converter_valor_brl_para_float(valor),
        "observacao_linha": " | ".join(observacao_parts),
    }


def parse_item_bloco(
    bloco: str,
    pagina: int,
    arquivo_origem: str,
    anexo: str,
    secao: str,
) -> dict[str, Any] | None:
    # Materiais (códigos alfanuméricos + 7910xxxxx) têm layout próprio.
    if MATERIAIS_INICIO_RE.match(normalizar_espacos(bloco)):
        item_materiais = parse_item_materiais(bloco, pagina, arquivo_origem, anexo, secao)
        if item_materiais:
            return item_materiais

    valor_match = escolher_match_valor(bloco)
    if not valor_match:
        return None

    valor = normalizar_espacos(valor_match.group(0))
    # Prefixo até o valor escolhido; remove outros R$ (ex.: honorário numérico).
    prefixo = normalizar_espacos(VALOR_RE.sub(" ", bloco[: valor_match.start()]))
    sufixo = normalizar_espacos(bloco[valor_match.end() :])

    codigo, codigo_hospital, corpo = extrair_codigos_inicio(prefixo)
    codigo, codigo_hospital, corpo = promover_codigo_principal(codigo, codigo_hospital, corpo)
    # Linha inteira como float-código colado: 0.02041390 APARELHO...
    if (not codigo or CODIGO_CURTO_PONTUADO_RE.fullmatch(codigo or "")) and re.match(
        r"^0\.\d{6,}", prefixo
    ):
        m_float = re.match(r"^(?P<code>0\.\d+)(?P<rest>\s+.*)?$", prefixo)
        if m_float:
            cod_fix = normalizar_codigo_extraido(m_float.group("code"))
            rest_float = normalizar_espacos(m_float.group("rest") or "")
            if cod_fix and rest_float:
                codigo = cod_float
                corpo = rest_float

    corpo = re.sub(
        r"\b(PACOTES|PACOTE|SESSÕES|SESSOES|SESSÃO|SESSAO|TAXA|USO|DIA|HORA|EVENTO|MINUTO)(?=[A-ZÁÀÂÃÉÊÍÓÔÕÚÇ]{2,})",
        r"\1 ",
        corpo,
    )

    unidade = ""
    descricao = normalizar_espacos(corpo)

    if not codigo and re.search(r"(CBHPM|SIMPRO|BRAS[ÍI]NDICE|TABELA MEDICA|HONOR[ÁA]RIOS M[ÉE]DICOS)", prefixo, re.IGNORECASE):
        return None

    # Unidade só no início se for palavra isolada (não "TAXA/FRALDA...")
    unidade_inicio_match = re.match(
        rf"^(?P<unidade>{UNIDADES_REGEX})\b(?!\s*/)(?:\s+)(?P<descricao>.+)$",
        corpo,
        re.IGNORECASE,
    )
    if unidade_inicio_match:
        unidade_candidata = normalizar_unidade(unidade_inicio_match.group("unidade"))
        descricao_candidata = normalizar_espacos(unidade_inicio_match.group("descricao"))
        if descricao_candidata and not descricao_indica_unidade_semantica(corpo):
            unidade = unidade_candidata
            descricao = descricao_candidata

    unidade_match = re.search(
        rf"(?P<descricao>.*?)(?:\s+(?P<unidade>{UNIDADES_REGEX}))\s*$",
        corpo,
        re.IGNORECASE,
    )
    if unidade_match and not unidade:
        descricao_candidata = normalizar_espacos(unidade_match.group("descricao"))
        unidade_cand = normalizar_unidade(unidade_match.group("unidade"))
        # Não pelear USO/DIA/HORA que completa "... POR USO"
        if unidade_e_continuacao_da_descricao(descricao_candidata):
            unidade = ""
            descricao = normalizar_espacos(corpo)
        elif not (descricao_candidata.upper().endswith("/") or descricao_candidata.upper().endswith("/T")):
            unidade = unidade_cand
            descricao = normalizar_espacos(corpo)
            if unidade == "HM" and descricao.upper().endswith("SEM"):
                descricao = normalizar_espacos(corpo)
                unidade = ""
            if not descricao_candidata:
                descricao = normalizar_espacos(corpo)
            else:
                if len(descricao_candidata) >= 3:
                    descricao = descricao_candidata
    else:
        descricao = normalizar_espacos(descricao)

    descricao, obs_honorario = limpar_texto_honorario(descricao)
    descricao = limpar_descricao_item(descricao)
    codigo = sanitizar_codigo(normalizar_codigo_extraido(codigo))
    codigo, descricao = extrair_codigo_embutido_na_descricao(codigo, descricao)
    codigo = sanitizar_codigo(codigo)
    descricao = limpar_descricao_item(descricao)

    # Se a unidade ficou e a descrição ainda termina em POR, reanexa (defesa extra)
    if unidade and re.search(r"(?i)\bPOR\s*$", descricao or ""):
        descricao = normalizar_espacos(f"{descricao} {unidade}")
        unidade = ""

    if not codigo and not descricao:
        return None
    # Só valor/código sem texto: não importa linha vazia de descrição
    if codigo and not descricao:
        return None

    observacao_parts = []
    if obs_honorario:
        observacao_parts.append(obs_honorario)
    if sufixo:
        observacao_parts.append(sufixo)
    observacao = " | ".join(p for p in observacao_parts if p)

    if not codigo and len(descricao.split()) > 18:
        return None

    return {
        "arquivo_origem": arquivo_origem,
        "pagina": pagina,
        "anexo": anexo,
        "secao": secao,
        "codigo": codigo,
        "codigo_hospital": codigo_hospital,
        "descricao": descricao,
        "unidade_cobranca": unidade,
        "valor": valor,
        "valor_numero": converter_valor_brl_para_float(valor),
        "observacao_linha": observacao,
    }


def extrair_itens_do_texto(
    texto: str,
    pagina: int,
    arquivo_origem: str,
) -> list[dict[str, Any]]:
    itens: list[dict[str, Any]] = []
    anexo_atual = ""
    secao_atual = ""
    ultimo_item: dict[str, Any] | None = None

    for bloco in juntar_linhas_quebradas(texto):
        anexo_atual = detectar_anexo(bloco, anexo_atual)
        secao_atual = detectar_secao(bloco, secao_atual)

        if eh_cabecalho_tabela(bloco) or eh_linha_secao(bloco) or ANEXO_RE.search(bloco):
            continue

        item = parse_item_bloco(bloco, pagina, arquivo_origem, anexo_atual, secao_atual)
        if item:
            if (
                not item["descricao"]
                and ultimo_item
                and item["codigo"] == ultimo_item["codigo"]
                and item["valor"] == ultimo_item["valor"]
            ):
                continue
            itens.append(item)
            ultimo_item = item

    preencher_descricoes_ausentes(itens, texto)
    return itens


def linha_parece_condicao_ou_nota(linha: str) -> bool:
    linha_ascii = normalizar_ascii(linha)
    prefixos = [
        "SERVICOS MEDICOS",
        "MEDICAMENTOS",
        "MATERIAIS",
        "SADT",
        "OPME",
        "HEMOTERAPIA",
        "FILME",
        "INCLUI",
        "INCLUSOS",
        "EXCLUSOS",
        "OBSERVACAO",
        "ESTAO INCLUSOS",
        "OUTROS MATERIAIS",
        "DEMAIS CONDICOES",
    ]
    return any(linha_ascii.startswith(prefixo) for prefixo in prefixos)


def extrair_descricoes_orfas(texto: str) -> list[str]:
    descricoes: list[str] = []
    encontrou_item_sem_descricao = False
    for linha in texto.splitlines():
        linha_limpa = normalizar_espacos(linha)
        if not linha_limpa:
            continue
        if ANEXO_RE.search(linha_limpa) or eh_cabecalho_tabela(linha_limpa):
            continue
        if linha_parece_condicao_ou_nota(linha_limpa):
            continue

        if re.match(rf"^\s*{CODIGO_REGEX}", linha_limpa):
            if re.match(rf"^\s*{CODIGO_REGEX}\s*{VALOR_REGEX}\s*$", linha_limpa, re.IGNORECASE):
                encontrou_item_sem_descricao = True
            valor_match = re.search(rf"{VALOR_REGEX}(?P<resto>.+)$", linha_limpa, re.IGNORECASE)
            if valor_match:
                resto = normalizar_espacos(valor_match.group("resto"))
                if (
                    encontrou_item_sem_descricao
                    and resto
                    and "CODIGO TUSS" not in normalizar_ascii(resto)
                    and not linha_parece_condicao_ou_nota(resto)
                ):
                    descricoes.append(resto)
            continue

        if not encontrou_item_sem_descricao:
            if eh_linha_secao(linha_limpa):
                continue
            continue

        if re.match(rf"^\s*{VALOR_REGEX}", linha_limpa, re.IGNORECASE):
            linha_limpa = re.sub(rf"^\s*{VALOR_REGEX}", "", linha_limpa, flags=re.IGNORECASE).strip()
            if linha_limpa and not linha_parece_condicao_ou_nota(linha_limpa):
                descricoes.append(linha_limpa)
            continue

        if VALOR_RE.search(linha_limpa):
            continue

        linha_ascii = normalizar_ascii(linha_limpa)
        if linha_ascii == linha_ascii.upper() and len(linha_limpa.split()) <= 5:
            continue

        if len(linha_limpa.split()) >= 4:
            descricoes.append(linha_limpa)

    return descricoes


def preencher_descricoes_ausentes(itens: list[dict[str, Any]], texto: str) -> None:
    faltantes = [item for item in itens if not normalizar_espacos(str(item.get("descricao", "")))]
    if not faltantes:
        return

    descricoes_orfas = extrair_descricoes_orfas(texto)
    if not descricoes_orfas:
        return

    for item, descricao in zip(faltantes, descricoes_orfas):
        item["descricao"] = descricao


def classificar_tipo_condicao(texto: str) -> str:
    for tipo, regex in CONDITION_KEYWORDS.items():
        if regex.search(texto):
            return tipo
    if VALOR_RE.search(texto):
        return "linha_monetaria_nao_classificada"
    return "texto_livre"


def extrair_condicoes(
    texto: str,
    pagina: int,
    arquivo_origem: str,
) -> list[dict[str, Any]]:
    condicoes: list[dict[str, Any]] = []
    blocos = juntar_linhas_quebradas(texto)
    anexo_encontrado = False
    em_bloco_condicoes = False

    for bloco in blocos:
        bloco_limpo = normalizar_espacos(bloco)
        if not bloco_limpo:
            continue
        if ANEXO_RE.search(bloco_limpo):
            anexo_encontrado = True
            continue
        if eh_inicio_bloco_condicoes(bloco_limpo):
            em_bloco_condicoes = True
        elif eh_linha_secao(bloco_limpo):
            em_bloco_condicoes = False
        if eh_cabecalho_tabela(bloco_limpo):
            if em_bloco_condicoes:
                tipo = classificar_tipo_condicao(bloco_limpo)
                condicoes.append(
                    {
                        "arquivo_origem": arquivo_origem,
                        "pagina": pagina,
                        "tipo": tipo,
                        "texto": bloco_limpo,
                    }
                )
            continue
        if eh_linha_secao(bloco_limpo) and "CONDI" not in normalizar_ascii(bloco_limpo):
            continue
        if parse_item_bloco(bloco_limpo, pagina, arquivo_origem, "", ""):
            continue

        if not anexo_encontrado and not em_bloco_condicoes:
            continue
        if not em_bloco_condicoes and not any(regex.search(bloco_limpo) for regex in CONDITION_KEYWORDS.values()):
            continue

        tipo = classificar_tipo_condicao(bloco_limpo)
        if tipo == "texto_livre":
            if len(bloco_limpo.split()) < 4 and not re.search(r"(CBHPM|SIMPRO|BRAS|HEMATOL|FILME)", bloco_limpo, re.IGNORECASE):
                continue
        condicoes.append(
            {
                "arquivo_origem": arquivo_origem,
                "pagina": pagina,
                "tipo": tipo,
                "texto": bloco_limpo,
            }
        )

    return condicoes


def extrair_texto_paginas(caminho_pdf: Path) -> list[str]:
    return list(iterar_texto_paginas(caminho_pdf))


def iterar_texto_paginas(caminho_pdf: Path):
    """Extrai texto página a página, sem carregar o PDF inteiro na memória."""
    caminho_pdf = Path(caminho_pdf)
    usar_pypdf_leve = caminho_pdf.stat().st_size >= 1_000_000 and PdfReader is not None

    if usar_pypdf_leve:
        reader = PdfReader(str(caminho_pdf))
        total = len(reader.pages)
        LOGGER.info("PDF grande detectado (%s páginas). Usando leitura leve página a página.", total)
        for page in reader.pages:
            yield page.extract_text() or ""
        return

    if pdfplumber is not None:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                yield page.extract_text(x_tolerance=2, y_tolerance=3) or ""
        return

    if PdfReader is not None:
        reader = PdfReader(str(caminho_pdf))
        for page in reader.pages:
            yield page.extract_text() or ""
        return

    raise RuntimeError("Nenhuma biblioteca de leitura de PDF disponível. Instale pdfplumber ou pypdf.")


def formatar_paginas_processadas(paginas: list[int]) -> str:
    if not paginas:
        return ""
    if len(paginas) <= 20:
        return ", ".join(str(p) for p in paginas)
    return f"1-{paginas[-1]} ({len(paginas)} páginas)"


def processar_pdf(caminho_pdf: str | Path) -> dict[str, Any]:
    caminho_pdf = Path(caminho_pdf)
    arquivo_origem = caminho_pdf.name
    LOGGER.info("Processando PDF: %s", arquivo_origem)

    resultado = {
        "itens": [],
        "condicoes_observacoes": [],
        "resumo": [],
    }

    observacoes_processamento: list[str] = []
    status = "sucesso"

    try:
        paginas_processadas: list[int] = []

        for indice, texto in enumerate(iterar_texto_paginas(caminho_pdf), start=1):
            paginas_processadas.append(indice)
            blocos_pagina = juntar_linhas_quebradas(texto)
            itens_pagina = extrair_itens_do_texto(texto, indice, arquivo_origem)
            condicoes_pagina = extrair_condicoes(texto, indice, arquivo_origem)

            blocos_monetarios_nao_classificados = [
                bloco
                for bloco in blocos_pagina
                if VALOR_RE.search(bloco)
                and not eh_cabecalho_tabela(bloco)
                and parse_item_bloco(bloco, indice, arquivo_origem, "", "") is None
            ]
            if blocos_monetarios_nao_classificados and len(observacoes_processamento) < 50:
                observacoes_processamento.append(
                    f"Página {indice}: {len(blocos_monetarios_nao_classificados)} blocos monetários foram direcionados para Condicoes_Observacoes."
                )

            resultado["itens"].extend(itens_pagina)
            resultado["condicoes_observacoes"].extend(condicoes_pagina)

            if indice % 100 == 0:
                LOGGER.info(
                    "Progresso %s: página %s | itens=%s | condições=%s",
                    arquivo_origem,
                    indice,
                    len(resultado["itens"]),
                    len(resultado["condicoes_observacoes"]),
                )

        if not resultado["itens"]:
            status = "alerta"
            observacoes_processamento.append("Nenhum item precificado foi extraído.")

        resultado["resumo"].append(
            {
                "arquivo_origem": arquivo_origem,
                "total_itens_extraidos": len(resultado["itens"]),
                "total_condicoes_observacoes": len(resultado["condicoes_observacoes"]),
                "paginas_processadas": formatar_paginas_processadas(paginas_processadas),
                "status": status,
                "observacoes_processamento": " | ".join(observacoes_processamento),
            }
        )
        return resultado

    except Exception as exc:  # pragma: no cover - tratamento batch
        LOGGER.exception("Falha ao processar %s", arquivo_origem)
        resultado["resumo"].append(
            {
                "arquivo_origem": arquivo_origem,
                "total_itens_extraidos": 0,
                "total_condicoes_observacoes": 0,
                "paginas_processadas": "",
                "status": "erro",
                "observacoes_processamento": str(exc),
            }
        )
        return resultado


def processar_e_salvar_pdf_grande(caminho_pdf: Path, caminho_saida: Path) -> None:
    """Extrai e grava página a página (PDFs grandes), evitando MemoryError."""
    from openpyxl import Workbook
    from openpyxl.cell.cell import WriteOnlyCell

    arquivo_origem = caminho_pdf.name
    LOGGER.info("Processando PDF grande com gravação streaming: %s", arquivo_origem)

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(suffix=".xlsx", prefix="apost_", dir=caminho_saida.parent)
    os.close(fd)
    caminho_temp = Path(temp_name)

    total_itens = 0
    total_condicoes = 0
    paginas_processadas: list[int] = []
    observacoes_processamento: list[str] = []
    status = "sucesso"

    try:
        wb = Workbook(write_only=True)
        ws_itens = wb.create_sheet(limpar_nome_aba("Itens"))
        ws_cond = wb.create_sheet(limpar_nome_aba("Condicoes_Observacoes"))
        ws_resumo = wb.create_sheet(limpar_nome_aba("Resumo"))

        def header(ws, colunas: list[str]) -> None:
            cells = []
            for col in colunas:
                cell = WriteOnlyCell(ws, value=col)
                cell.font = Font(bold=True)
                cells.append(cell)
            ws.append(cells)

        header(ws_itens, ITEM_COLUMNS)
        header(ws_cond, CONDICOES_COLUMNS)

        for indice, texto in enumerate(iterar_texto_paginas(caminho_pdf), start=1):
            paginas_processadas.append(indice)
            itens_pagina = extrair_itens_do_texto(texto, indice, arquivo_origem)
            condicoes_pagina = extrair_condicoes(texto, indice, arquivo_origem)

            for item in itens_pagina:
                ws_itens.append([item.get(col) for col in ITEM_COLUMNS])
            for cond in condicoes_pagina:
                ws_cond.append([cond.get(col) for col in CONDICOES_COLUMNS])

            total_itens += len(itens_pagina)
            total_condicoes += len(condicoes_pagina)

            if indice % 100 == 0:
                LOGGER.info(
                    "Progresso %s: página %s | itens=%s | condições=%s",
                    arquivo_origem,
                    indice,
                    total_itens,
                    total_condicoes,
                )

        if total_itens == 0:
            status = "alerta"
            observacoes_processamento.append("Nenhum item precificado foi extraído.")

        header(ws_resumo, RESUMO_COLUMNS)
        ws_resumo.append(
            [
                arquivo_origem,
                total_itens,
                total_condicoes,
                formatar_paginas_processadas(paginas_processadas),
                status,
                " | ".join(observacoes_processamento),
            ]
        )

        wb.save(caminho_temp)
        wb.close()
        gc.collect()

        try:
            os.replace(caminho_temp, caminho_saida)
        except PermissionError as exc:
            raise PermissionError(
                f"Não foi possível gravar '{caminho_saida.name}'. "
                "Feche o arquivo no Excel (ou outro programa) e tente novamente."
            ) from exc

        LOGGER.info(
            "Planilha gerada: %s (itens=%s, condições=%s, páginas=%s)",
            caminho_saida.name,
            total_itens,
            total_condicoes,
            len(paginas_processadas),
        )
    except Exception:
        LOGGER.exception("Falha no processamento streaming de %s", arquivo_origem)
        raise
    finally:
        if caminho_temp.exists():
            try:
                caminho_temp.unlink()
            except OSError:
                pass


def ajustar_layout_excel(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        valores = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        largura = min(max(len(valor) for valor in valores) + 2, 80)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = largura


def _escrever_aba_write_only(workbook, nome_aba: str, colunas: list[str], linhas: list[dict[str, Any]]) -> None:
    from openpyxl.cell.cell import WriteOnlyCell

    ws = workbook.create_sheet(limpar_nome_aba(nome_aba))
    header_cells = []
    for col in colunas:
        cell = WriteOnlyCell(ws, value=col)
        cell.font = Font(bold=True)
        header_cells.append(cell)
    ws.append(header_cells)

    for linha in linhas:
        ws.append([linha.get(col) for col in colunas])


def salvar_excel(resultado: dict[str, Any], caminho_saida: str | Path) -> None:
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    itens = resultado.get("itens", [])
    condicoes = resultado.get("condicoes_observacoes", [])
    resumo = resultado.get("resumo", [])
    total_linhas = len(itens) + len(condicoes)

    fd, temp_name = tempfile.mkstemp(suffix=".xlsx", prefix="apost_", dir=caminho_saida.parent)
    os.close(fd)
    caminho_temp = Path(temp_name)

    try:
        # Arquivos grandes: write_only evita MemoryError do openpyxl.
        if total_linhas >= 3_000:
            from openpyxl import Workbook

            LOGGER.info(
                "Salvando planilha grande em modo streaming (%s linhas)...",
                total_linhas,
            )
            wb = Workbook(write_only=True)
            _escrever_aba_write_only(wb, "Itens", ITEM_COLUMNS, itens)
            _escrever_aba_write_only(wb, "Condicoes_Observacoes", CONDICOES_COLUMNS, condicoes)
            _escrever_aba_write_only(wb, "Resumo", RESUMO_COLUMNS, resumo)
            wb.save(caminho_temp)
            wb.close()
        else:
            df_itens = pd.DataFrame(itens, columns=ITEM_COLUMNS)
            df_condicoes = pd.DataFrame(condicoes, columns=CONDICOES_COLUMNS)
            df_resumo = pd.DataFrame(resumo, columns=RESUMO_COLUMNS)

            with pd.ExcelWriter(caminho_temp, engine="openpyxl") as writer:
                df_itens.to_excel(writer, sheet_name=limpar_nome_aba("Itens"), index=False)
                df_condicoes.to_excel(
                    writer,
                    sheet_name=limpar_nome_aba("Condicoes_Observacoes"),
                    index=False,
                )
                df_resumo.to_excel(writer, sheet_name=limpar_nome_aba("Resumo"), index=False)

                for sheet_name in writer.book.sheetnames:
                    ajustar_layout_excel(writer.book[sheet_name])

        gc.collect()

        try:
            os.replace(caminho_temp, caminho_saida)
        except PermissionError as exc:
            raise PermissionError(
                f"Não foi possível gravar '{caminho_saida.name}'. "
                "Feche o arquivo no Excel (ou outro programa) e tente novamente."
            ) from exc
    finally:
        if caminho_temp.exists():
            try:
                caminho_temp.unlink()
            except OSError:
                pass


def resolver_pasta_entrada(caminho: str | Path | None) -> Path | None:
    """Resolve pasta de PDFs: relativo à raiz ou a entrada_apostilamentos/."""
    if caminho is None:
        return None
    pasta = Path(caminho)
    if pasta.is_absolute():
        return pasta.resolve()

    candidatas = [
        (BASE_DIR / pasta).resolve(),
        (ENTRADA_DIR / pasta).resolve(),
    ]
    for cand in candidatas:
        if cand.exists():
            return cand
    # Preferência: sob entrada_apostilamentos (mesmo que ainda não exista)
    return candidatas[1]


def resolver_pasta_saida(caminho: str | Path | None) -> Path:
    if caminho is None:
        return SAIDA_DIR
    pasta = Path(caminho)
    if not pasta.is_absolute():
        pasta = BASE_DIR / pasta
    return pasta.resolve()


def listar_pdfs_entrada(pasta_entrada: Path | None = None) -> list[Path]:
    if pasta_entrada is not None:
        return sorted(pasta_entrada.glob("*.pdf"))

    pdfs_entrada = sorted(ENTRADA_DIR.glob("*.pdf")) if ENTRADA_DIR.exists() else []
    if pdfs_entrada:
        return pdfs_entrada
    return []


def executar_batch(
    pasta_entrada: str | Path | None = None,
    pasta_saida: str | Path | None = None,
) -> list[Path]:
    """Extrai PDFs → XLSX. Retorna lista de planilhas geradas."""
    entrada = resolver_pasta_entrada(pasta_entrada)
    if pasta_saida is not None:
        saida = resolver_pasta_saida(pasta_saida)
    elif entrada is not None:
        # espelha o nome da pasta mensal em APOSTILAMENTOS/
        saida = SAIDA_DIR / entrada.name
    else:
        saida = SAIDA_DIR

    saida.mkdir(parents=True, exist_ok=True)
    pdfs = listar_pdfs_entrada(entrada)
    gerados: list[Path] = []

    if not pdfs:
        origem = entrada or ENTRADA_DIR
        LOGGER.warning("Nenhum PDF encontrado em %s", origem)
        return gerados

    LOGGER.info("Pasta de entrada: %s", entrada or ENTRADA_DIR)
    LOGGER.info("Pasta de saída: %s", saida)
    LOGGER.info("Total de PDFs encontrados: %s", len(pdfs))

    for caminho_pdf in pdfs:
        try:
            caminho_saida = saida / f"{caminho_pdf.stem}.xlsx"
            # PDFs grandes: extrai e grava em streaming para não corromper por MemoryError.
            if caminho_pdf.stat().st_size >= 1_000_000:
                processar_e_salvar_pdf_grande(caminho_pdf, caminho_saida)
            else:
                resultado = processar_pdf(caminho_pdf)
                salvar_excel(resultado, caminho_saida)
                LOGGER.info("Planilha gerada: %s", caminho_saida.name)
            gerados.append(caminho_saida)
        except Exception:
            LOGGER.exception("Erro ao salvar resultado do arquivo %s", caminho_pdf.name)

    return gerados


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extrai itens e condições de apostilamentos PDF para planilhas Excel "
            "(formato compatível com import_apostilamentos_dtp.py)."
        )
    )
    parser.add_argument(
        "pasta",
        nargs="?",
        default=None,
        help=(
            "Pasta com PDFs (ex.: JULHO ou entrada_apostilamentos/JULHO). "
            "Se omitida, usa entrada_apostilamentos/."
        ),
    )
    parser.add_argument(
        "--saida",
        default=None,
        help="Pasta de saída das planilhas. Padrão: APOSTILAMENTOS/<pasta>.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    executar_batch(pasta_entrada=args.pasta, pasta_saida=args.saida)
