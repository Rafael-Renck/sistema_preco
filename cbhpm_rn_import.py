"""Parser das Resoluções Normativas CNHM (atualizações pós-CBHPM 2022)."""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Sequence

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None

from cbhpm_import import (
    CBHPM_PROCEDIMENTO_FIELDS,
    CBHPM_CODE_RE,
    CbhpmProcedimentoRow,
    _apply_chapter_field_rules,
    _is_diagnostico_codigo,
    _parse_tail,
    _write_dict_rows_xlsx,
    procedimentos_to_csv_rows,
    write_cbhpm_output,
)

CODE_IN_TEXT_RE = re.compile(r'\b(\d\.\d{2}\.\d{2}\.\d{2}-\d)\b')
BLOCK_SPLIT_RE = re.compile(r'(?=\b\d\.\d{2}\.\d{2}\.\d{2}-\d\b)')

DECISAO_INCLUSAO_RE = re.compile(r'inclus[aã]o\s+de\s+procedimento', re.I)
DECISAO_EXCLUSAO_RE = re.compile(r'exclus[aã]o\s+de\s+procedimento', re.I)
DECISAO_OBS_RE = re.compile(r'inclus[aã]o\s+de\s+(?:observa|item)', re.I)
DECISAO_ALTERACAO_RE = re.compile(r'altera', re.I)

DE_PARA_RE = re.compile(
    r'De:\s*(?P<de>[^P\n]+?)\s*Para:\s*(?P<para>[^\nDe]+?)(?=De:|$)',
    re.I | re.S,
)
FIELD_LABELS = (
    ('porte', re.compile(r'porte\s+anest', re.I), 'porte_anestesico'),
    ('porte', re.compile(r'^porte\b', re.I), 'porte'),
    ('auxiliar', re.compile(r'auxiliar', re.I), 'num_auxiliares'),
    ('custo operacional', re.compile(r'custo\s+oper', re.I), 'uco'),
    ('filme', re.compile(r'filme', re.I), 'filme'),
    ('incid', re.compile(r'incid', re.I), 'incidencia'),
)

CBHPM_RN_CHANGE_FIELDS = [
    'codigo', 'decisao', 'rn', 'descricao', 'porte', 'uco',
    'num_auxiliares', 'porte_anestesico', 'filme', 'incidencia', 'layout', 'detalhe',
]


@dataclass
class CbhpmRnChange:
    codigo: str
    decisao: str
    rn_num: int
    descricao: str = ''
    porte: str | None = None
    uco: str | None = None
    num_auxiliares: str | None = None
    porte_anestesico: str | None = None
    filme: str | None = None
    incidencia: str | None = None
    layout: str = ''
    detalhe: str = ''
    patches: dict[str, str] = field(default_factory=dict)


def _normalize_space(text: str) -> str:
    text = unicodedata.normalize('NFKC', text or '')
    text = text.replace('\u2003', ' ').replace('\u00ad', '')
    return re.sub(r'\s+', ' ', text).strip()


def _extract_pdf_text(path: Path) -> str:
    if fitz is None:
        raise ValueError('PyMuPDF não instalado.')
    doc = fitz.open(str(path))
    return _normalize_space('\n'.join(page.get_text() or '' for page in doc))


def _rn_number_from_path(path: Path) -> int:
    m = re.search(r'(\d{3})', path.stem)
    if not m:
        raise ValueError(f'Não foi possível identificar número da RN em {path.name}')
    return int(m.group(1))


def _is_procedure_code(codigo: str) -> bool:
    return bool(CBHPM_CODE_RE.match(codigo))


def _infer_layout(codigo: str, *, uco: str | None, filme: str | None, aux: str | None, anest: str | None) -> str:
    if _is_diagnostico_codigo(codigo):
        return 'diagnostico'
    if aux or anest:
        return 'cirurgico'
    if uco:
        return 'consulta' if codigo.startswith(('1.', '2.')) else 'cirurgico'
    chapter = codigo.split('.')[0] if codigo else ''
    if chapter.isdigit() and int(chapter) >= 3:
        return 'cirurgico'
    return 'consulta'


def _parse_de_para_pairs(text: str) -> dict[str, str]:
    patches: dict[str, str] = {}
    for m in DE_PARA_RE.finditer(text):
        de = _normalize_space(m.group('de'))
        para = _normalize_space(m.group('para'))
        label = f'{de} {para}'.lower()
        field_name = None
        for _key, pattern, target in FIELD_LABELS:
            if pattern.search(de) or pattern.search(label):
                field_name = target
                break
        if field_name is None:
            if 'anest' in de.lower():
                field_name = 'porte_anestesico'
            elif 'aux' in de.lower():
                field_name = 'num_auxiliares'
            elif 'custo' in de.lower() or 'oper' in de.lower():
                field_name = 'uco'
            elif 'filme' in de.lower():
                field_name = 'filme'
            else:
                field_name = 'porte'
        val = re.sub(r'^(?:Porte\s+Anest[eé]sico|Porte|Auxiliar|Custo\s+Operacional|Filme)\s*', '', para, flags=re.I)
        val = _normalize_space(val.replace(':', ''))
        if val:
            patches[field_name] = val
    return patches


RN_PORTE_UCO_FILME_RE = re.compile(
    r'Porte\s+(?P<porte>[\dABC]+)\s*,?\s*'
    r'UCO\s+(?P<uco>[\d,\.]+)\s*'
    r'(?:e\s+)?Filme\s+(?P<filme>[\d,\.]+|–|—|-|\*)',
    re.I,
)
RN_PORTE_UCO_RE = re.compile(
    r'Porte\s+(?P<porte>[\dABC]+)\s*,?\s*UCO\s+(?P<uco>[\d,\.]+)',
    re.I,
)
RN_PORTE_FILME_RE = re.compile(
    r'Porte\s+(?P<porte>[\dABC]+)\s*,?\s*Filme\s+(?P<filme>[\d,\.]+|–|—|-|\*)',
    re.I,
)
AUXILIAR_TOKEN = r'Auxiliar(?:es)?'

RN_PORTE_AUX_ANEST_RE = re.compile(
    r'Porte\s+(?P<porte>[\dABC]+)\s*,?\s*'
    r'(?:'
    rf'(?:(?:N[uú]mero\s+de\s+)?(?P<aux_before>\d+)\s+{AUXILIAR_TOKEN}|'
    r'N[uú]mero\s+de\s+Auxiliar\s+(?P<aux2>\d+)|'
    rf'{AUXILIAR_TOKEN}\s+(?P<aux_after>\d+))'
    r')?\s*'
    r'(?:,?\s*(?:e\s+)?Porte\s+Anest[eé]sico\s+(?P<anest>\d+|\*|[–—-]))?',
    re.I,
)


def _parse_rn_aux(body: str) -> str | None:
    for pattern in (
        rf'(?:N[uú]mero\s+de\s+)?(?P<n>\d+)\s+{AUXILIAR_TOKEN}',
        rf'{AUXILIAR_TOKEN}\s+(?P<n>\d+)',
        r'N[uú]mero\s+de\s+Auxiliar\s+(?P<n>\d+)',
    ):
        m = re.search(pattern, body, re.I)
        if m:
            return m.group('n')
    return None


def _parse_rn_anest(body: str) -> str | None:
    m = re.search(r'Porte\s+Anest[eé]sico\s+(?P<n>\d+|\*|[–—-])', body, re.I)
    return m.group('n') if m else None


def _clean_cbhpm_value(value: str | None) -> str | None:
    if value is None:
        return None
    value = _normalize_space(str(value).rstrip(','))
    return value or None


def _parse_rn_incidencia(body: str) -> str | None:
    m = re.search(r'Incid[eê]ncias?:?\s*(\d+)', body, re.I)
    return m.group(1) if m else None


def _parse_rn_filme(body: str) -> str | None:
    m = re.search(r'Filme:?\s+([\d,\.]+(?:\s*m)?|–|—|-|\*)', body, re.I)
    if m:
        return _clean_cbhpm_value(m.group(1))
    if re.search(r'Sem\s+filme', body, re.I):
        return '–'
    return None


def _clean_inclusion_descricao(codigo: str, desc: str) -> str:
    desc = re.sub(r'^Inclus[aã]o de Procedimento\s*', '', desc, flags=re.I)
    desc = re.sub(rf'^{re.escape(codigo)}\s*', '', desc)
    return _normalize_space(desc.strip(' ,'))


def _parse_rn_explicit_fields(body: str) -> dict[str, str | None]:
    """Padrões textuais das RN: Porte X, UCO Y e Filme Z / Aux / Anest."""
    porte = uco = filme = aux = anest = None
    desc = body

    for pattern, groups in (
        (RN_PORTE_UCO_FILME_RE, ('porte', 'uco', 'filme')),
        (RN_PORTE_UCO_RE, ('porte', 'uco')),
        (RN_PORTE_FILME_RE, ('porte', 'filme')),
        (RN_PORTE_AUX_ANEST_RE, ('porte', 'aux', 'anest')),
    ):
        m = pattern.search(body)
        if not m:
            continue
        desc = _normalize_space(body[:m.start()])
        vals = m.groupdict()
        porte = vals.get('porte') or porte
        uco = vals.get('uco') or uco
        filme = vals.get('filme') or filme
        aux = vals.get('aux_before') or vals.get('aux2') or vals.get('aux_after') or aux
        anest = vals.get('anest') or anest
        break

    if not porte:
        m = re.search(r'Porte\s+([\dABC]+)', body, re.I)
        if m:
            porte = m.group(1)
            if desc == body:
                desc = _normalize_space(body[:m.start()])
    if not aux:
        aux = _parse_rn_aux(body)
    if not anest:
        anest = _parse_rn_anest(body)
    if not uco:
        m = re.search(r'UCO\s+([\d,\.]+)', body, re.I)
        if m:
            uco = m.group(1)
    if not filme:
        filme = _parse_rn_filme(body)
    if not uco:
        m = re.search(r'Custo\s+Operacional:?\s*([\d,\.]+)', body, re.I)
        if m:
            uco = m.group(1)
    if not uco:
        m = re.search(r'Custo\s+Operacional\s+([\d,\.]+)\s+UCO', body, re.I)
        if m:
            uco = m.group(1)
    incidencia = _parse_rn_incidencia(body)

    return {
        'descricao': desc,
        'porte': _clean_cbhpm_value(porte),
        'uco': _clean_cbhpm_value(uco),
        'filme': _clean_cbhpm_value(filme),
        'aux': aux,
        'anest': anest,
        'incidencia': incidencia,
    }


def _parse_inclusion_body(codigo: str, body: str) -> dict[str, str | None]:
    body = _normalize_space(body)
    body = re.sub(rf'^{re.escape(codigo)}\s*', '', body)

    rn_fields = _parse_rn_explicit_fields(body)
    if rn_fields.get('porte'):
        desc = _clean_inclusion_descricao(codigo, rn_fields['descricao'] or body)
        layout = _infer_layout(
            codigo,
            uco=rn_fields.get('uco'),
            filme=rn_fields.get('filme'),
            aux=rn_fields.get('aux'),
            anest=rn_fields.get('anest'),
        )
        return {
            'descricao': desc,
            'porte': rn_fields.get('porte'),
            'uco': rn_fields.get('uco'),
            'num_auxiliares': rn_fields.get('aux'),
            'porte_anestesico': rn_fields.get('anest'),
            'filme': rn_fields.get('filme'),
            'incidencia': rn_fields.get('incidencia'),
            'layout': layout,
        }

    parsed = _parse_tail(body, codigo=codigo, allow_simple=True)
    if parsed:
        incidencia = _clean_cbhpm_value(parsed.get('incidencia')) or _parse_rn_incidencia(body)
        return {
            'descricao': _clean_inclusion_descricao(codigo, parsed.get('desc', body)),
            'porte': parsed.get('porte') or None,
            'uco': _clean_cbhpm_value(parsed.get('uco')),
            'num_auxiliares': parsed.get('aux') or None,
            'porte_anestesico': parsed.get('anest') or None,
            'filme': parsed.get('filme') or None,
            'incidencia': incidencia,
            'layout': parsed.get('layout', ''),
        }

    # Padrões explícitos das RN (fallback cirúrgico)
    desc = body
    porte = uco = aux = anest = filme = None
    m = re.search(
        rf'Porte\s+([\dABC]+)(?:.*?(\d+)\s+{AUXILIAR_TOKEN})?(?:.*?Porte\s+Anest[eé]sico\s+(\d+))?',
        body,
        re.I,
    )
    if m:
        porte = m.group(1)
        aux = m.group(2)
        anest = m.group(3)
        desc = _normalize_space(body[:m.start()])
    m2 = re.search(r'Porte\s+([\dABC]+)\s+e\s+Custo\s+Operacional:?\s*([\d,\.]+)', body, re.I)
    if m2:
        porte, uco = m2.group(1), m2.group(2)
        desc = _normalize_space(body[:m2.start()])
    m3 = re.search(r'Porte\s+([\dABC]+)\s*$', body, re.I)
    if m3 and not porte:
        porte = m3.group(1)
        desc = _normalize_space(body[:m3.start()])
    m4 = re.search(r'Porte\s+[\d,\.]+\s+de\s+([\dABC]+)', body, re.I)
    if m4:
        porte = m4.group(1)
        if not desc or desc == body:
            desc = _normalize_space(body[:m4.start()])
    m5 = re.search(r'Custo\s+Operacional:?\s*([\d,\.]+)', body, re.I)
    if m5:
        uco = m5.group(1)
    m6 = re.search(r'Custo\s+Operacional\s+([\d,\.]+)\s+UCO', body, re.I)
    if m6:
        uco = m6.group(1)
    m7 = re.search(r'Filme:?\s+([\d,\.]+(?:\s*m)?|–|—|-|\*)', body, re.I)
    if m7:
        filme = m7.group(1)
    m8 = re.search(r'Sem\s+filme', body, re.I)
    if m8:
        filme = '–'
    incidencia = _parse_rn_incidencia(body)

    layout = _infer_layout(codigo, uco=uco, filme=filme, aux=aux, anest=anest)
    return {
        'descricao': desc,
        'porte': _clean_cbhpm_value(porte),
        'uco': _clean_cbhpm_value(uco),
        'num_auxiliares': aux,
        'porte_anestesico': anest,
        'filme': _clean_cbhpm_value(filme),
        'incidencia': incidencia,
        'layout': layout,
    }


def _classify_decision(block: str) -> str:
    head = block[:400]
    if DECISAO_OBS_RE.search(head) and not DECISAO_INCLUSAO_RE.search(head):
        return 'observacao'
    if DECISAO_EXCLUSAO_RE.search(head):
        return 'exclusao'
    if DECISAO_INCLUSAO_RE.search(head):
        return 'inclusao'
    if DECISAO_ALTERACAO_RE.search(head):
        return 'alteracao'
    return 'outro'


def _parse_rn_block(codigo: str, block: str, rn_num: int) -> CbhpmRnChange | None:
    if not _is_procedure_code(codigo):
        return None
    block = _normalize_space(block)
    block = re.sub(rf'^{re.escape(codigo)}\s*', '', block)
    decisao = _classify_decision(block)
    if decisao == 'observacao':
        return None
    if decisao == 'outro':
        return None

    change = CbhpmRnChange(codigo=codigo, decisao=decisao, rn_num=rn_num, detalhe=block[:500])

    if decisao == 'exclusao':
        m = DECISAO_EXCLUSAO_RE.search(block)
        change.descricao = _normalize_space(block[m.end():]) if m else block
        return change

    if decisao == 'inclusao':
        data = _parse_inclusion_body(codigo, block)
        change.descricao = data.get('descricao') or ''
        change.porte = data.get('porte')  # type: ignore[assignment]
        change.uco = data.get('uco')  # type: ignore[assignment]
        change.num_auxiliares = data.get('num_auxiliares')  # type: ignore[assignment]
        change.porte_anestesico = data.get('porte_anestesico')  # type: ignore[assignment]
        change.filme = data.get('filme')  # type: ignore[assignment]
        change.incidencia = data.get('incidencia')  # type: ignore[assignment]
        change.layout = data.get('layout') or _infer_layout(
            codigo,
            uco=change.uco,
            filme=change.filme,
            aux=change.num_auxiliares,
            anest=change.porte_anestesico,
        )
        return change

    change.patches = _parse_de_para_pairs(block)
    change.detalhe = block[:300]
    return change


def parse_rn_pdf(path: str | Path) -> list[CbhpmRnChange]:
    path = Path(path)
    rn_num = _rn_number_from_path(path)
    text = _extract_pdf_text(path)
    # Remove cabeçalho até primeira tabela
    text = re.split(r'C[oó]digo\s+Decis[aã]o\s+Descri', text, maxsplit=1, flags=re.I)
    body = text[-1] if text else ''
    changes: list[CbhpmRnChange] = []
    for raw_block in BLOCK_SPLIT_RE.split(body):
        raw_block = raw_block.strip()
        if not raw_block:
            continue
        m = CODE_IN_TEXT_RE.search(raw_block)
        if not m:
            continue
        codigo = m.group(1)
        item = _parse_rn_block(codigo, raw_block, rn_num)
        if item:
            changes.append(item)
    return changes


def parse_rn_directory(rn_dir: str | Path) -> list[CbhpmRnChange]:
    rn_dir = Path(rn_dir)
    all_changes: list[CbhpmRnChange] = []
    for path in sorted(rn_dir.glob('RN_CNHM_*.pdf')):
        all_changes.extend(parse_rn_pdf(path))
    return sorted(all_changes, key=lambda x: (x.rn_num, x.codigo))


def load_catalog_from_xlsx(path: str | Path) -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    catalog: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        item = {headers[i]: ('' if row[i] is None else str(row[i]).strip()) for i in range(len(headers))}
        catalog[item['codigo']] = item
    return catalog


def _merge_field(current: str, new: str) -> str:
    dash = {'', '–', '-', '—'}
    if new and new not in dash:
        return new
    return current


def _catalog_row_from_change(change: CbhpmRnChange, base: dict[str, str] | None = None) -> dict[str, str]:
    row = dict(base or {})
    row['codigo'] = change.codigo
    if change.decisao == 'inclusao':
        row.update({
            'descricao': change.descricao or row.get('descricao', ''),
            'porte': _merge_field(row.get('porte', ''), change.porte or ''),
            'uco': _merge_field(row.get('uco', ''), change.uco or ''),
            'num_auxiliares': _merge_field(row.get('num_auxiliares', ''), change.num_auxiliares or ''),
            'porte_anestesico': _merge_field(row.get('porte_anestesico', ''), change.porte_anestesico or ''),
            'filme': _merge_field(row.get('filme', ''), change.filme or ''),
            'incidencia': _merge_field(row.get('incidencia', ''), change.incidencia or ''),
            'layout': change.layout or row.get('layout', ''),
        })
    elif change.decisao == 'alteracao' and change.patches:
        field_map = {
            'porte': 'porte',
            'uco': 'uco',
            'num_auxiliares': 'num_auxiliares',
            'porte_anestesico': 'porte_anestesico',
            'filme': 'filme',
            'incidencia': 'incidencia',
        }
        for src, dst in field_map.items():
            if src in change.patches:
                row[dst] = change.patches[src]
        if change.descricao:
            row['descricao'] = change.descricao
    row['rn_ultima_alteracao'] = f'RN {change.rn_num:03d}'
    return row


def apply_rn_changes(
    catalog: dict[str, dict[str, str]],
    changes: Sequence[CbhpmRnChange],
) -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    log: list[dict[str, str]] = []
    for change in sorted(changes, key=lambda x: (x.rn_num, x.codigo)):
        log.append({
            'codigo': change.codigo,
            'decisao': change.decisao,
            'rn': f'RN {change.rn_num:03d}',
            'descricao': change.descricao,
            'porte': change.porte or change.patches.get('porte', ''),
            'uco': change.uco or change.patches.get('uco', ''),
            'num_auxiliares': change.num_auxiliares or change.patches.get('num_auxiliares', ''),
            'porte_anestesico': change.porte_anestesico or change.patches.get('porte_anestesico', ''),
            'filme': change.filme or change.patches.get('filme', ''),
            'incidencia': change.incidencia or change.patches.get('incidencia', ''),
            'layout': change.layout,
            'detalhe': change.detalhe[:300],
        })
        if change.decisao == 'exclusao':
            catalog.pop(change.codigo, None)
            continue
        base = catalog.get(change.codigo, {})
        if change.decisao == 'inclusao':
            catalog[change.codigo] = _catalog_row_from_change(change, base)
            continue
        if change.decisao == 'alteracao':
            merged = dict(base)
            merged.update(_catalog_row_from_change(change, base))
            catalog[change.codigo] = merged
    return catalog, log


def catalog_to_procedure_rows(catalog: dict[str, dict[str, str]]) -> list[CbhpmProcedimentoRow]:
    rows: list[CbhpmProcedimentoRow] = []
    for codigo in sorted(catalog):
        item = catalog[codigo]
        rows.append(CbhpmProcedimentoRow(
            codigo=codigo,
            descricao=item.get('descricao', ''),
            porte=item.get('porte') or None,
            uco=item.get('uco') or None,
            num_auxiliares=item.get('num_auxiliares') or None,
            porte_anestesico=item.get('porte_anestesico') or None,
            filme=item.get('filme') or None,
            incidencia=item.get('incidencia') or None,
            layout=item.get('layout') or 'unknown',
            capitulo=item.get('capitulo') or None,
            grupo=item.get('grupo') or None,
            subgrupo=item.get('subgrupo') or None,
            pagina=int(item['pagina']) if str(item.get('pagina', '')).isdigit() else 0,
        ))
    return rows


def write_rn_change_log(changes: Sequence[dict[str, str]], path: str | Path) -> None:
    _write_dict_rows_xlsx(changes, path, fieldnames=CBHPM_RN_CHANGE_FIELDS, sheet_title='Alterações RN')


def write_updated_catalog_xlsx(catalog: dict[str, dict[str, str]], path: str | Path) -> None:
    fields = list(CBHPM_PROCEDIMENTO_FIELDS) + ['rn_ultima_alteracao']
    rows = []
    for codigo in sorted(catalog):
        item = catalog[codigo]
        rows.append({name: item.get(name, '') for name in fields})
    _write_dict_rows_xlsx(rows, path, fieldnames=fields, sheet_title='CBHPM atualizado')
